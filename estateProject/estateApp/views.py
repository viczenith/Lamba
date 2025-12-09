import math
from django.shortcuts import render, get_object_or_404, redirect
from django.contrib.auth import login, authenticate

from .forms import *
from django.contrib import messages
from django.contrib.auth.decorators import login_required, user_passes_test
from .models import *
import random
from types import SimpleNamespace
from django.contrib.auth import authenticate, update_session_auth_hash
from django.contrib.admin.views.decorators import staff_member_required
from django.views.decorators.csrf import csrf_exempt
from django.utils.timezone import now
from django.utils import timezone
import datetime as dt
from datetime import datetime, timedelta
from django.utils.dateparse import parse_date
from itertools import islice


from django.db import DatabaseError, IntegrityError, transaction
import csv
from django.http import Http404, HttpResponse, HttpResponseNotAllowed
from django.core.exceptions import ValidationError
import json
from django.http import JsonResponse, HttpResponseBadRequest, HttpResponseForbidden

from django.db.models import Prefetch, Count, Max, Q, F, Value, Sum, DecimalField, OuterRef, Subquery, Exists, ExpressionWrapper
from django.db.models.functions import Concat, Coalesce

import unicodedata
import re

from django.contrib.auth.mixins import LoginRequiredMixin
from django.views.generic import TemplateView, FormView, View

from django.template.loader import render_to_string, get_template

from reportlab.lib.pagesizes import A4
from reportlab.pdfgen import canvas
from reportlab.lib import colors
from reportlab.platypus import Table, TableStyle


# NOTIFICATIONS
from django.core.mail import send_mail
from django.conf import settings
from django.utils.html import strip_tags
from django.views.generic.edit import FormView 
from django.views.decorators.csrf import csrf_protect
from django.contrib.auth.views import LoginView, LogoutView
from django.urls import reverse_lazy


# MANAGEMENT DASHBOARD
# from weasyprint import HTML, CSS  # Commented out due to Windows compatibility issues
from xhtml2pdf import pisa
from django.urls             import reverse
from django.views.decorators.http import require_http_methods, require_GET, require_POST, require_http_methods
from django.utils.decorators import method_decorator
from django.contrib.auth import get_user_model
import uuid
from io import BytesIO
User = get_user_model()
from decimal import Decimal, ROUND_HALF_UP, InvalidOperation
from dateutil.relativedelta import relativedelta
import logging
from math import ceil
from .services.geoip import lookup_ip_location, extract_client_ip
from DRF.shared_drf.push_service import send_chat_message_push, send_user_notification_push
from .tasks import (
    BATCH_SIZE,
    dispatch_notification_stream,
    dispatch_notification_stream_sync,
    is_celery_worker_available,
)


logger = logging.getLogger(__name__)


SUPPORT_ROLES = ('admin', 'support')


# ============================================================================
# HELPER FUNCTION: Generate URL-safe slugs from user full names
# ============================================================================
def generate_name_slug(full_name):
    """
    Convert full name to URL-safe slug for profile links.
    Example: "Victor Godwin" -> "victor-godwin"
    """
    if not full_name:
        return None
    
    # Convert to lowercase and replace spaces with hyphens
    slug = full_name.strip().lower().replace(' ', '-')
    # Remove any special characters except hyphens
    slug = re.sub(r'[^a-z0-9\-]', '', slug)
    # Remove multiple consecutive hyphens
    slug = re.sub(r'-+', '-', slug)
    # Remove leading/trailing hyphens
    slug = slug.strip('-')
    
    return slug if slug else None


# ============================================================================
# HELPER FUNCTION: Redirect to tenant-aware dashboard
# ============================================================================
def get_tenant_dashboard_redirect(request):
    """
    Helper function to get redirect URL to tenant-aware dashboard.
    Used for redirects from legacy views.
    """
    from django.urls import reverse
    company = getattr(request.user, 'company_profile', None)
    
    if company:
        return redirect(reverse('tenant-dashboard', kwargs={'company_slug': company.slug}))
    else:
        messages.error(request, "You are not assigned to any company!")
        return redirect('login')


def get_user_company_from_slug(request):
    """
    Helper function to get user's company from query parameter or default.
    
    Supports dynamic slug routing:
    - Current: /company-profile/
    - With query param: /company-profile/?company=lamba-real-homes
    
    Returns: Company object or None
    Raises: HttpResponseForbidden if user tries to access different company
    """
    user = request.user
    company = getattr(user, 'company_profile', None)
    
    # Check for ?company=slug parameter
    company_slug = request.GET.get('company')
    if company_slug:
        try:
            company_by_slug = Company.objects.get(slug=company_slug)
            # SECURITY: Verify user has access to this company
            if user.company_profile == company_by_slug:
                company = company_by_slug
            else:
                # User trying to access different company - FORBIDDEN
                raise HttpResponseForbidden("❌ Access Denied: You can only access your own company's profile")
        except Company.DoesNotExist:
            # Invalid slug - return None or could raise 404
            return None
    
    return company


def custom_csrf_failure_view(request, reason=""):
    return render(request, 'csrf_failure.html', {"reason": reason}, status=403)


# HOME VIEW
class HomeView(LoginRequiredMixin, TemplateView):
    template_name = 'home.html'


@login_required
def admin_dashboard(request):
    """
    ⚠️ DEPRECATED: This view is for backward compatibility only.
    Use tenant-aware routes instead: /<company-slug>/dashboard/
    This view redirects to the new Facebook-style routing.
    """
    company = getattr(request.user, 'company_profile', None)
    
    if not company:
        messages.error(request, "You are not assigned to any company!")
        return redirect('login')
    
    # Redirect to new tenant-aware route
    from django.urls import reverse
    new_url = reverse('tenant-dashboard', kwargs={'company_slug': company.slug})
    return redirect(new_url)

@login_required
def add_plotsize(request):
    from django.http import JsonResponse
    from .models import PlotSize, PlotSizeUnits, Estate
    from django.db.models import Count, Q
    
    # SECURITY: Get company context for data isolation
    company = getattr(request, 'company', None)
    
    if request.method == 'POST':
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            action = request.POST.get('action', 'add')
            
            if action == 'add':
                size = request.POST.get('size', '').strip()
                
                if not size:
                    return JsonResponse({'success': False, 'message': 'Plot size is required'})
                
                # SECURITY: Check if plot size already exists for THIS company only
                if PlotSize.objects.filter(size__iexact=size, company=company).exists():
                    return JsonResponse({'success': False, 'message': f'Plot size "{size}" already exists for your company'})
                
                try:
                    PlotSize.objects.create(size=size, company=company)  # SECURITY: Bind to company
                    return JsonResponse({'success': True, 'message': f'Plot size "{size}" added successfully!'})
                except Exception as e:
                    return JsonResponse({'success': False, 'message': str(e)})
            
            elif action == 'update':
                plot_size_id = request.POST.get('plot_size_id')
                new_size = request.POST.get('new_size', '').strip()
                
                if not new_size:
                    return JsonResponse({'success': False, 'message': 'New plot size is required'})
                
                try:
                    # SECURITY: Only get plot sizes from THIS company
                    plot_size = PlotSize.objects.get(id=plot_size_id, company=company)
                    old_size = plot_size.size
                    
                    # SECURITY: Check if new size already exists in THIS company (excluding current)
                    if PlotSize.objects.filter(size__iexact=new_size, company=company).exclude(id=plot_size_id).exists():
                        return JsonResponse({'success': False, 'message': f'Plot size "{new_size}" already exists'})
                    
                    # Update the plot size
                    plot_size.size = new_size
                    plot_size.save()
                    
                    return JsonResponse({
                        'success': True, 
                        'message': f'Plot size updated from "{old_size}" to "{new_size}" across all estates and allocations!'
                    })
                except PlotSize.DoesNotExist:
                    return JsonResponse({'success': False, 'message': 'Plot size not found or does not belong to your company'})
                except Exception as e:
                    return JsonResponse({'success': False, 'message': str(e)})
    
    # SECURITY: Annotate plot sizes with estate count - ONLY for THIS company
    plot_sizes_data = []
    for plot_size in PlotSize.objects.filter(company=company).order_by('size'):
        # Count estates using this plot size through PlotSizeUnits
        estate_count = PlotSizeUnits.objects.filter(plot_size=plot_size).values('estate_plot__estate').distinct().count()
        
        # Get estate names
        estate_names = []
        if estate_count > 0:
            estate_plots = PlotSizeUnits.objects.filter(plot_size=plot_size).select_related('estate_plot__estate').distinct()
            estate_names = [ep.estate_plot.estate.name for ep in estate_plots]
        
        plot_sizes_data.append({
            'id': plot_size.id,
            'size': plot_size.size,
            'estate_count': estate_count,
            'estate_names': estate_names,
            'is_assigned': estate_count > 0
        })
    
    return render(request, "admin_side/add_plotsize.html", {'plot_sizes': plot_sizes_data})

@login_required
def add_plotnumber(request):
    from django.http import JsonResponse
    from .models import PlotNumber, PlotAllocation
    from .ws_utils import broadcast_user_notification
    
    # SECURITY: Get company context for data isolation
    company = getattr(request, 'company', None)
    
    if request.method == 'POST':
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            action = request.POST.get('action', 'add')
            
            if action == 'add':
                number = request.POST.get('number', '').strip()
                
                if not number:
                    return JsonResponse({'success': False, 'message': 'Plot number is required'})
                
                # SECURITY: Check if plot number already exists for THIS company only
                if PlotNumber.objects.filter(number__iexact=number, company=company).exists():
                    return JsonResponse({'success': False, 'message': f'Plot number "{number}" already exists for your company'})
                
                try:
                    PlotNumber.objects.create(number=number, company=company)  # SECURITY: Bind to company
                    return JsonResponse({'success': True, 'message': f'Plot number "{number}" added successfully!'})
                except Exception as e:
                    return JsonResponse({'success': False, 'message': str(e)})
            
            elif action == 'update':
                plot_number_id = request.POST.get('plot_number_id')
                new_number = request.POST.get('number', '').strip()
                
                if not new_number:
                    return JsonResponse({'success': False, 'message': 'New plot number is required'})
                
                try:
                    # SECURITY: Only get plot numbers from THIS company
                    plot_number = PlotNumber.objects.get(id=plot_number_id, company=company)
                    old_number = plot_number.number
                    
                    # SECURITY: Check if new number already exists in THIS company (excluding current)
                    if PlotNumber.objects.filter(number__iexact=new_number, company=company).exclude(id=plot_number_id).exists():
                        return JsonResponse({'success': False, 'message': f'Plot number "{new_number}" already exists'})
                    
                    # Update the plot number
                    plot_number.number = new_number
                    plot_number.save()
                    
                    return JsonResponse({
                        'success': True, 
                        'message': f'Plot number updated from "{old_number}" to "{new_number}" across all estates and client allocations!'
                    })
                except PlotNumber.DoesNotExist:
                    return JsonResponse({'success': False, 'message': 'Plot number not found or does not belong to your company'})
                except Exception as e:
                    return JsonResponse({'success': False, 'message': str(e)})
    
    # SECURITY: Annotate plot numbers with allocation data - ONLY for THIS company
    plot_numbers_data = []
    for plot_number in PlotNumber.objects.filter(company=company).order_by('number'):
        # Get all allocations for this plot number that belong to this company
        allocations = PlotAllocation.objects.filter(
            plot_number=plot_number,
            estate__company=company
        ).select_related('client', 'estate').distinct()
        
        allocation_count = allocations.count()
        
        # Build detailed allocation info: "Client Name - Estate Name"
        allocation_details = []
        if allocation_count > 0:
            for alloc in allocations:
                client_name = alloc.client.full_name if alloc.client else "Unknown Client"
                estate_name = alloc.estate.name if alloc.estate else "Unknown Estate"
                allocation_details.append(f"{client_name} - {estate_name}")
        
        plot_numbers_data.append({
            'id': plot_number.id,
            'number': plot_number.number,
            'allocation_count': allocation_count,
            'allocation_details': allocation_details,
            'is_assigned': allocation_count > 0
        })
    
    return render(request, "admin_side/add_plotnumber.html", {'plot_numbers': plot_numbers_data})

@login_required
def delete_plotsize(request, pk):
    from django.http import JsonResponse
    from .models import PlotSize, PlotSizeUnits
    
    # SECURITY: Get company context
    company = getattr(request, 'company', None)
    
    if request.method == 'POST':
        try:
            # SECURITY: Only allow deleting company's own plot sizes
            plot_size = PlotSize.objects.get(id=pk, company=company)
            size_name = plot_size.size
            
            # Check if plot size is assigned to any estate
            estate_count = PlotSizeUnits.objects.filter(plot_size=plot_size).values('estate_plot__estate').distinct().count()
            
            if estate_count > 0:
                return JsonResponse({
                    'success': False, 
                    'message': f'Cannot delete "{size_name}". It is currently assigned to {estate_count} estate(s). Please use the Edit function to modify it instead.'
                })
            
            plot_size.delete()
            return JsonResponse({'success': True, 'message': f'Plot size "{size_name}" deleted successfully!'})
        except PlotSize.DoesNotExist:
            return JsonResponse({'success': False, 'message': 'Plot size not found'})
        except Exception as e:
            return JsonResponse({'success': False, 'message': f'Error: {str(e)}'})
    
    return JsonResponse({'success': False, 'message': 'Invalid request method'})

@login_required
def delete_plotnumber(request, pk):
    from django.http import JsonResponse
    from .models import PlotNumber, PlotAllocation
    
    # SECURITY: Get company context
    company = getattr(request, 'company', None)
    
    if request.method == 'POST':
        try:
            # SECURITY: Only allow deleting company's own plot numbers
            plot_number = PlotNumber.objects.get(id=pk, company=company)
            number_name = plot_number.number
            
            # Check if plot number is allocated to any client/estate
            allocation_count = PlotAllocation.objects.filter(plot_number=plot_number, estate__company=company).count()
            
            if allocation_count > 0:
                return JsonResponse({
                    'success': False, 
                    'message': f'Cannot delete "{number_name}". It is currently allocated to {allocation_count} client(s). Please use the Edit function to modify it instead.'
                })
            
            plot_number.delete()
            return JsonResponse({'success': True, 'message': f'Plot number "{number_name}" deleted successfully!'})
        except PlotNumber.DoesNotExist:
            return JsonResponse({'success': False, 'message': 'Plot number not found'})
        except Exception as e:
            return JsonResponse({'success': False, 'message': f'Error: {str(e)}'})
    
    return JsonResponse({'success': False, 'message': 'Invalid request method'})

@login_required
def management_dashboard(request):
    """
    ⚠️ DEPRECATED: This view is for backward compatibility only.
    Use tenant-aware routes instead: /<company-slug>/management/
    This view redirects to the new Facebook-style routing.
    """
    company = getattr(request.user, 'company_profile', None)
    
    if not company:
        messages.error(request, "You are not assigned to any company!")
        return redirect('login')
    
    # Redirect to new tenant-aware route
    from django.urls import reverse
    new_url = reverse('tenant-management', kwargs={'company_slug': company.slug})
    return redirect(new_url)


def estate_allocation_data(request):
    # SECURITY: Filter by company to prevent cross-tenant data access
    company = request.user.company_profile
    estates = []
    allocated_data = []
    reserved_data = []
    total_data = []
    
    for estate in Estate.objects.filter(company=company):
        total_allocated = 0
        total_reserved = 0
        
        for size_unit in estate.estate_plots.plotsizeunits.all():
            total_allocated += size_unit.full_allocations
            total_reserved += size_unit.part_allocations
        
        estates.append(estate.name)
        allocated_data.append(total_allocated)
        reserved_data.append(total_reserved)
        total_data.append(total_allocated + total_reserved)
    
    return JsonResponse({
        'estates': estates,
        'allocated': allocated_data,
        'reserved': reserved_data,
        'total': total_data
    })


def get_allocated_plots(request):
    # SECURITY: return allocations only for the current user's company
    company = getattr(request.user, 'company_profile', None)
    if not company:
        return JsonResponse([], safe=False)

    allocated_plots = PlotAllocation.objects.filter(estate__company=company).values(
        'estate_id', 'plot_size_id', 'plot_number_id'
    )
    return JsonResponse(list(allocated_plots), safe=False)


def get_all_marketers_for_company(company_obj):
    """
    Helper function to fetch all marketers for a company.
    Combines marketers from both company_profile and MarketerAffiliation.
    
    Returns a QuerySet of CustomUser objects with role='marketer' with client count annotation.
    
    IMPORTANT: Counts ONLY from ClientMarketerAssignment table (single source of truth).
    Each assignment is counted EXACTLY ONCE per company. No data leakage between companies.
    A marketer can serve multiple clients within a company.
    A client can be served by a marketer in multiple companies they operate in.
    """
    if not company_obj:
        return CustomUser.objects.none()
    
    # Get marketers assigned directly to the company
    marketers_primary = CustomUser.objects.filter(role='marketer', company_profile=company_obj)
    
    # Get marketers affiliated via MarketerAffiliation (added via "Add Existing User" modal)
    marketers_affiliated = CustomUser.objects.none()
    try:
        affiliation_marketer_ids = MarketerAffiliation.objects.filter(
            company=company_obj
        ).values_list('marketer_id', flat=True).distinct()
        if affiliation_marketer_ids:
            marketers_affiliated = CustomUser.objects.filter(
                id__in=affiliation_marketer_ids
            ).exclude(
                id__in=marketers_primary.values_list('pk', flat=True)
            )
    except Exception as e:
        logger.warning(f"Failed to get affiliated marketers for company {company_obj}: {e}")
        pass
    
    # Combine both QuerySets using set union, then filter in a single QuerySet
    all_marketer_ids = set(marketers_primary.values_list('pk', flat=True)) | set(
        marketers_affiliated.values_list('pk', flat=True)
    )
    
    # Count clients for each marketer ONLY from ClientMarketerAssignment table
    # Strict company filtering prevents data leakage and duplicate counting
    from estateApp.models import ClientMarketerAssignment
    
    # Single source of truth: Count distinct clients assigned to each marketer for THIS company ONLY
    client_count_subquery = ClientMarketerAssignment.objects.filter(
        marketer_id=OuterRef('id'),
        company=company_obj
    ).values('marketer_id').annotate(
        count=Count('id', distinct=True)
    ).values('count')
    
    # Return a single QuerySet with all marketers sorted by name
    # Strict company isolation: each count belongs to this company only
    return CustomUser.objects.filter(id__in=all_marketer_ids).annotate(
        client_count=Subquery(client_count_subquery)
    ).annotate(
        # Handle NULL for marketers with no assignments in this company
        client_count=Coalesce('client_count', 0)
    ).order_by('full_name')


@login_required
def api_marketer_client_counts(request):
    """
    API endpoint that returns the current client count for each marketer in the company.
    Used by frontend to dynamically update dropdown without page reload.
    
    Returns JSON: {
        "success": true,
        "marketers": [
            {"id": 15, "client_count": 5},
            {"id": 8, "client_count": 3},
            ...
        ]
    }
    """
    # Get company from request
    company = getattr(request, 'company', None) or getattr(request.user, 'company_profile', None)
    
    if not company:
        return JsonResponse({'success': False, 'error': 'No company context'}, status=400)
    
    # Get all marketers with client counts
    marketers = get_all_marketers_for_company(company)
    
    # Build response
    marketer_data = [
        {
            'id': marketer.id,
            'full_name': marketer.full_name,
            'email': marketer.email,
            'client_count': marketer.client_count
        }
        for marketer in marketers
    ]
    
    return JsonResponse({
        'success': True,
        'marketers': marketer_data,
        'timestamp': now().isoformat()
    })


def user_registration(request):
    # Fetch all users with the 'marketer' role
    # SECURITY: Filter by company to prevent cross-company data leakage
    company = getattr(request, 'company', None) or getattr(request.user, 'company_profile', None)
    
    # Get all marketers for the company (both primary and affiliated)
    marketers = get_all_marketers_for_company(company)


    if request.method == 'POST':
        # Normalize company context: prefer request.company (middleware), fall back to user's company_profile
        company = getattr(request, 'company', None) or getattr(request.user, 'company_profile', None)
        if not company:
            error_msg = "Unable to determine company context for registration. Please ensure you're operating within a company scope."
            if request.headers.get('x-requested-with') == 'XMLHttpRequest':
                return JsonResponse({'success': False, 'message': error_msg})
            messages.error(request, error_msg)
            return render(request, 'admin_side/user_registration.html', {'marketers': marketers})
        source = request.POST.get('source', 'admin_registration')
        # Extract form data
        full_name = request.POST.get('name')
        address = request.POST.get('address')
        phone = request.POST.get('phone')
        email = request.POST.get('email')
        role = request.POST.get('role')
        country = request.POST.get('country')
        date_of_birth = request.POST.get('date_of_birth')
        password = request.POST.get('password')

        # Validate required fields
        if not email or not role or not full_name:
            error_msg = "Email, role, and full name are required."
            if request.headers.get('x-requested-with') == 'XMLHttpRequest':
                return JsonResponse({'success': False, 'message': error_msg})
            messages.error(request, error_msg)
            return render(request, 'admin_side/user_registration.html', {'marketers': marketers})

        # Quick lookup: does a user with this email already exist?
        existing_user = CustomUser.objects.filter(email=email).first()

        # If an existing user id was supplied (frontend may accidentally submit the registration
        # form when an existing user was selected), handle it here by affiliating the user
        # to the company instead of trying to create a new user or requiring a password.
        existing_user_id = request.POST.get('existing_user_id') or request.POST.get('existingId')
        if existing_user_id:
            try:
                # SECURITY: Only admins may add users to company via this path
                if getattr(request.user, 'role', None) != 'admin':
                    error_msg = 'Only admins can add existing users to a company.'
                    if request.headers.get('x-requested-with') == 'XMLHttpRequest':
                        return JsonResponse({'success': False, 'message': error_msg}, status=403)
                    messages.error(request, error_msg)
                    return render(request, 'admin_side/user_registration.html', {'marketers': marketers})

                company = getattr(request, 'company', None) or getattr(request.user, 'company_profile', None)
                user_to_add = CustomUser.objects.get(id=int(existing_user_id), is_active=True, is_deleted=False)

                # Marketer assignment required for clients
                if role == 'client' and not marketer:
                    error_msg = 'Marketer assignment required for client users.'
                    if request.headers.get('x-requested-with') == 'XMLHttpRequest':
                        return JsonResponse({'success': False, 'message': error_msg}, status=400)
                    messages.error(request, error_msg)
                    return render(request, 'admin_side/user_registration.html', {'marketers': marketers})

                # Use transaction to ensure consistent state
                with transaction.atomic():
                    # If client, ensure ClientUser subclass exists and assign marketer and create ClientMarketerAssignment
                    if getattr(user_to_add, 'role', None) == 'client':
                        if not ClientUser.objects.filter(pk=user_to_add.pk).exists():
                            try:
                                ClientUser.objects.create(id=user_to_add.pk, company_profile=company)
                            except Exception:
                                # If multi-table constraints prevent creation, ignore and continue
                                pass
                        # assign marketer on the ClientUser instance if applicable
                        try:
                            client_obj = ClientUser.objects.get(pk=user_to_add.pk)
                        except ClientUser.DoesNotExist:
                            client_obj = CustomUser.objects.get(pk=user_to_add.pk)

                        if marketer and hasattr(client_obj, 'assigned_marketer'):
                            try:
                                client_obj.assigned_marketer = marketer
                                client_obj.save()
                            except Exception:
                                pass

                        # Create company-scoped ClientMarketerAssignment
                        try:
                            ClientMarketerAssignment.objects.get_or_create(
                                client_id=user_to_add.pk,
                                marketer=marketer,
                                company=company
                            )
                        except Exception:
                            pass

                    # For marketers, ensure MarketerUser subclass exists and create MarketerAffiliation
                    if getattr(user_to_add, 'role', None) == 'marketer':
                        if not MarketerUser.objects.filter(pk=user_to_add.pk).exists():
                            try:
                                MarketerUser.objects.create(id=user_to_add.pk, company_profile=company)
                            except Exception:
                                pass
                        try:
                            MarketerAffiliation.objects.get_or_create(marketer_id=user_to_add.pk, company=company)
                        except Exception:
                            pass

                    # Do not overwrite user's primary company_profile if already set; only set if missing
                    if not getattr(user_to_add, 'company_profile', None):
                        try:
                            user_to_add.company_profile = company
                            user_to_add.save()
                        except Exception:
                            pass

                # Return success (AJAX or normal flow)
                success_msg = f"{user_to_add.full_name} successfully added to {company.company_name}"
                if request.headers.get('x-requested-with') == 'XMLHttpRequest':
                    return JsonResponse({'success': True, 'message': success_msg, 'user': {'id': user_to_add.id, 'email': user_to_add.email, 'full_name': user_to_add.full_name}})
                messages.success(request, success_msg)
                return render(request, 'admin_side/user_registration.html', {'marketers': marketers})
            except CustomUser.DoesNotExist:
                error_msg = 'Selected existing user not found.'
                if request.headers.get('x-requested-with') == 'XMLHttpRequest':
                    return JsonResponse({'success': False, 'message': error_msg}, status=404)
                messages.error(request, error_msg)
                return render(request, 'admin_side/user_registration.html', {'marketers': marketers})

        # Only assign a marketer if the role is 'client'
        marketer = None
        if role == 'client':
            marketer_id = request.POST.get('marketer')
            if not marketer_id:
                error_msg = "Please assign a marketer to this client. Marketer assignment is required."
                if request.headers.get('x-requested-with') == 'XMLHttpRequest':
                    return JsonResponse({'success': False, 'message': error_msg})
                messages.error(request, error_msg)
                return render(request, 'admin_side/user_registration.html', {'marketers': marketers})
            try:
                marketer = MarketerUser.objects.get(id=marketer_id)
            except MarketerUser.DoesNotExist:
                error_msg = f"Selected marketer does not exist. Please select a valid marketer."
                if request.headers.get('x-requested-with') == 'XMLHttpRequest':
                    return JsonResponse({'success': False, 'message': error_msg})
                messages.error(request, error_msg)
                return render(request, 'admin_side/user_registration.html', {'marketers': marketers})


        # Dynamic user creation and linking logic (robust multi-company support)
        success_msg = None
        try:
            from estateApp.models import MarketerAffiliation, ClientMarketerAssignment
            import logging
            from django.utils.crypto import get_random_string
            logger = logging.getLogger(__name__)
            if role == 'admin':
                admin_user = AdminUser.objects.filter(email=email).first()
                if admin_user:
                    # Link to this company if not already
                    if admin_user.company_profile != request.company:
                        admin_user.company_profile = request.company
                        admin_user.save()
                    success_msg = f"Admin user <strong>{full_name}</strong> already exists and is linked."
                else:
                    # Auto-generate password when not provided
                    if not password:
                        generated_password = get_random_string(12)
                        password_to_set = generated_password
                        pw_generated = True
                    else:
                        password_to_set = password
                        pw_generated = False

                    admin_user = AdminUser.objects.create(
                        email=email,
                        full_name=full_name,
                        address=address or '',
                        phone=phone,
                        date_of_birth=date_of_birth,
                        country=country or '',
                        company_profile=company
                    )
                    admin_user.set_password(password_to_set)
                    admin_user.save()
                    success_msg = f"Admin user <strong>{admin_user.full_name}</strong> has been successfully registered! (New user)"
                    try:
                        if pw_generated:
                            success_msg += ' A temporary password was generated for this user.'
                    except NameError:
                        pass
            elif role == 'marketer':
                marketer_user = MarketerUser.objects.filter(email=email).first()
                if marketer_user:
                    # Check if marketer is already affiliated with this company
                    affiliation = MarketerAffiliation.objects.filter(marketer=marketer_user, company=request.company).first()
                    already_in_company = bool(affiliation)
                    # Update info if needed
                    updated = False
                    if marketer_user.full_name != full_name:
                        marketer_user.full_name = full_name
                        updated = True
                    if marketer_user.address != address:
                        marketer_user.address = address
                        updated = True
                    if marketer_user.phone != phone:
                        marketer_user.phone = phone
                        updated = True
                    if marketer_user.date_of_birth != date_of_birth:
                        marketer_user.date_of_birth = date_of_birth
                        updated = True
                    if marketer_user.country != country:
                        marketer_user.country = country
                        updated = True
                    # Ensure marketer's company_profile is set to this company if not already
                    if marketer_user.company_profile != request.company:
                        marketer_user.company_profile = request.company
                        updated = True
                    if updated:
                        marketer_user.save()
                    # Link marketer to company if not already linked
                    ma, ma_created = MarketerAffiliation.objects.get_or_create(
                        marketer=marketer_user,
                        company=company
                    )
                    logger.info('MarketerAffiliation get_or_create: marketer=%s company=%s created=%s', getattr(marketer_user, 'email', None), getattr(company, 'company_name', None), ma_created)
                    # Ensure marketer_user has a visible primary company if missing (do not override existing primary)
                    if ma_created and not getattr(marketer_user, 'company_profile', None):
                        try:
                            marketer_user.company_profile = company
                            marketer_user.save()
                            logger.info('Set marketer primary company_profile for %s -> %s', marketer_user.email, getattr(company, 'company_name', None))
                        except Exception:
                            logger.exception('Failed to set marketer.company_profile')
                    if already_in_company:
                        success_msg = (
                            f"User with email <strong>{marketer_user.email}</strong> already exists and is already affiliated with this company as marketer."
                        )
                    elif ma_created:
                        success_msg = (
                            f"User with email <strong>{marketer_user.email}</strong> already exists in the system and has now been affiliated with this company as marketer."
                        )
                    else:
                        success_msg = (
                            f"User with email <strong>{marketer_user.email}</strong> already exists in the system."
                        )
                else:
                    # Auto-generate password when not provided (company admins do not supply passwords)
                    if not password:
                        generated_password = get_random_string(12)
                        password_to_set = generated_password
                        pw_generated = True
                    else:
                        password_to_set = password
                        pw_generated = False

                    marketer_user = MarketerUser.objects.create(
                        email=email,
                        full_name=full_name,
                        address=address or '',
                        phone=phone,
                        date_of_birth=date_of_birth,
                        country=country or '',
                        company_profile=company
                    )
                    marketer_user.set_password(password_to_set)
                    marketer_user.save()
                    # Link marketer to company
                    ma, ma_created = MarketerAffiliation.objects.get_or_create(
                        marketer=marketer_user,
                        company=company
                    )
                    logger.info('MarketerAffiliation created for new marketer: %s -> %s', getattr(marketer_user, 'email', None), getattr(company, 'company_name', None))
                    success_msg = f"Marketer <strong>{marketer_user.full_name}</strong> is now affiliated with this company! (New user)"
                    try:
                        if pw_generated:
                            success_msg += ' A temporary password was generated for this user.'
                    except NameError:
                        pass
            elif role == 'client':
                client_user = ClientUser.objects.filter(email=email).first()
                if client_user:
                    # Check if client is already affiliated with this company
                    affiliation = ClientMarketerAssignment.objects.filter(client=client_user, company=request.company).first()
                    already_in_company = bool(affiliation)
                    # Update info if needed
                    updated = False
                    if client_user.full_name != full_name:
                        client_user.full_name = full_name
                        updated = True
                    if client_user.address != address:
                        client_user.address = address
                        updated = True
                    if client_user.phone != phone:
                        client_user.phone = phone
                        updated = True
                    if client_user.date_of_birth != date_of_birth:
                        client_user.date_of_birth = date_of_birth
                        updated = True
                    if client_user.country != country:
                        client_user.country = country
                        updated = True
                    if client_user.assigned_marketer != marketer:
                        client_user.assigned_marketer = marketer
                        updated = True
                    if updated:
                        client_user.save()
                    # Link client to company/marketer if not already linked
                    cma, cma_created = ClientMarketerAssignment.objects.get_or_create(
                        client=client_user,
                        marketer=marketer,
                        company=company
                    )
                    logger.info('ClientMarketerAssignment get_or_create: client=%s marketer=%s company=%s created=%s', getattr(client_user, 'email', None), getattr(marketer, 'email', None) if marketer else None, getattr(company, 'company_name', None), cma_created)
                    if already_in_company:
                        success_msg = (
                            f"<strong>{client_user.full_name}</strong> already has an account with email <strong>{client_user.email}</strong>. "
                            f"They are now assigned to <strong>{marketer.full_name}</strong> for this company!"
                        )
                    else:
                        success_msg = (
                            f"<strong>{client_user.full_name}</strong> already has an account with email <strong>{client_user.email}</strong>, "
                            f"but was not yet added to your company.\n"
                            f"This user has now been added and assigned to <strong>{marketer.full_name}</strong> for your company."
                        )
                else:
                    # Auto-generate password when not provided (admins can communicate it to the user)
                    if not password:
                        generated_password = get_random_string(12)
                        password_to_set = generated_password
                        pw_generated = True
                    else:
                        password_to_set = password
                        pw_generated = False

                    client_user = ClientUser.objects.create(
                        email=email,
                        full_name=full_name,
                        address=address or '',
                        phone=phone,
                        date_of_birth=date_of_birth,
                        country=country or '',
                        assigned_marketer=marketer,
                        company_profile=company
                    )
                    client_user.set_password(password_to_set)
                    client_user.save()
                    # Link client to company/marketer
                    cma, cma_created = ClientMarketerAssignment.objects.get_or_create(
                        client=client_user,
                        marketer=marketer,
                        company=company
                    )
                    logger.info('ClientMarketerAssignment get_or_create (new client): client=%s marketer=%s company=%s created=%s', getattr(client_user, 'email', None), getattr(marketer, 'email', None) if marketer else None, getattr(company, 'company_name', None), cma_created)
                    success_msg = f"<strong>{client_user.full_name}</strong> is now assigned to <strong>{marketer.full_name}</strong> for this company! (New user)"
                    try:
                        if pw_generated:
                            success_msg += ' A temporary password was generated for this user.'
                    except NameError:
                        pass

            elif role == 'marketer':
                marketer_user = MarketerUser.objects.filter(email=email).first()
                if marketer_user:
                    # Check if marketer is already affiliated with this company (affiliation record must exist AND marketer's company_profile must match)
                    affiliation = MarketerAffiliation.objects.filter(marketer=marketer_user, company=request.company).first()
                    already_in_company = bool(affiliation and marketer_user.company_profile == request.company)
                    # Update info if needed
                    updated = False
                    if marketer_user.full_name != full_name:
                        marketer_user.full_name = full_name
                        updated = True
                    if marketer_user.address != address:
                        marketer_user.address = address
                        updated = True
                    if marketer_user.phone != phone:
                        marketer_user.phone = phone
                        updated = True
                    if marketer_user.date_of_birth != date_of_birth:
                        marketer_user.date_of_birth = date_of_birth
                        updated = True
                    if marketer_user.country != country:
                        marketer_user.country = country
                        updated = True
                    if updated:
                        marketer_user.save()
                    # Link marketer to company if not already linked
                    ma, ma_created = MarketerAffiliation.objects.get_or_create(
                        marketer=marketer_user,
                        company=company
                    )
                    logger.info('MarketerAffiliation get_or_create: marketer=%s company=%s created=%s', getattr(marketer_user, 'email', None), getattr(company, 'company_name', None), ma_created)
                    if ma_created and not getattr(marketer_user, 'company_profile', None):
                        try:
                            marketer_user.company_profile = company
                            marketer_user.save()
                        except Exception:
                            logger.exception('Failed to set marketer.company_profile on affiliation creation')
                    if already_in_company:
                        success_msg = (
                            f"User with email <strong>{marketer_user.email}</strong> already exists and is already affiliated with this company as marketer."
                        )
                    elif ma_created:
                        success_msg = (
                            f"User with email <strong>{marketer_user.email}</strong> already exists in the system and has now been affiliated with this company as marketer."
                        )
                    else:
                        success_msg = (
                            f"User with email <strong>{marketer_user.email}</strong> already exists in the system."
                        )
                else:
                    # Auto-generate password when not provided (company admins do not supply passwords)
                    if not password:
                        generated_password = get_random_string(12)
                        password_to_set = generated_password
                        pw_generated = True
                    else:
                        password_to_set = password
                        pw_generated = False

                    marketer_user = MarketerUser.objects.create(
                        email=email,
                        full_name=full_name,
                        address=address,
                        phone=phone,
                        date_of_birth=date_of_birth,
                        country=country,
                        company_profile=company
                    )
                    marketer_user.set_password(password_to_set)
                    marketer_user.save()
                    # Link marketer to company
                    ma, ma_created = MarketerAffiliation.objects.get_or_create(
                        marketer=marketer_user,
                        company=company
                    )
                    logger.info('MarketerAffiliation created for new marketer: %s -> %s', getattr(marketer_user, 'email', None), getattr(company, 'company_name', None))
                    success_msg = f"Marketer <strong>{marketer_user.full_name}</strong> is now affiliated with this company! (New user)"
                    try:
                        if pw_generated:
                            success_msg += ' A temporary password was generated for this user.'
                    except NameError:
                        pass

            elif role == 'support':
                support_user = SupportUser.objects.filter(email=email).first()
                if support_user:
                    # Update info if needed
                    updated = False
                    if support_user.full_name != full_name:
                        support_user.full_name = full_name
                        updated = True
                    if support_user.address != address:
                        support_user.address = address
                        updated = True
                    if support_user.phone != phone:
                        support_user.phone = phone
                        updated = True
                    if support_user.date_of_birth != date_of_birth:
                        support_user.date_of_birth = date_of_birth
                        updated = True
                    if support_user.country != country:
                        support_user.country = country
                        updated = True
                    if updated:
                        support_user.save()
                    success_msg = f"Support User, <strong>{full_name}</strong> has been successfully registered! (Existing user)"
                else:
                    # Auto-generate password when not provided
                    if not password:
                        generated_password = get_random_string(12)
                        password_to_set = generated_password
                        pw_generated = True
                    else:
                        password_to_set = password
                        pw_generated = False

                    support_user = SupportUser.objects.create(
                        email=email,
                        full_name=full_name,
                        address=address,
                        phone=phone,
                        date_of_birth=date_of_birth,
                        country=country,
                        company_profile=company
                    )
                    support_user.set_password(password_to_set)
                    support_user.save()
                    success_msg = f"Support User, <strong>{full_name}</strong> has been successfully registered! (New user)"
                    try:
                        if pw_generated:
                            success_msg += ' A temporary password was generated for this user.'
                    except NameError:
                        pass

        except Exception as e:
            # Better-format serializer/form validation errors where possible
            try:
                # If it's a dict-like error (e.g., {'password': ['This field cannot be blank.']}), present a cleaner message
                err_obj = e
                if hasattr(e, 'detail'):
                    err_obj = e.detail
                if isinstance(err_obj, dict):
                    # Collate field errors
                    parts = []
                    for k, v in err_obj.items():
                        parts.append(f"{k}: {', '.join(v)}")
                    error_msg = "Registration failed: " + "; ".join(parts)
                else:
                    error_msg = f"Registration failed: {str(e)}"
            except Exception:
                error_msg = f"Registration failed: {str(e)}"

            if request.headers.get('x-requested-with') == 'XMLHttpRequest':
                return JsonResponse({'success': False, 'message': error_msg})
            messages.error(request, error_msg)
            return render(request, 'admin_side/user_registration.html', {'marketers': marketers})

        if request.headers.get('x-requested-with') == 'XMLHttpRequest':
            # If the server auto-generated a password during this request, include it in the JSON
            pw_generated_flag = locals().get('pw_generated', False)
            pw_value = locals().get('password_to_set') if pw_generated_flag else None
            resp = {'success': True, 'message': success_msg}
            if pw_value:
                resp['temporary_password'] = pw_value
            return JsonResponse(resp)

        messages.success(request, success_msg)
        if source == 'company_profile':
            return redirect('company-profile')
        else:
            return redirect('user-registration')

    return render(request, 'admin_side/user_registration.html', {'marketers': marketers})


# ESTATE FUNCTIONS
@login_required
def view_estate(request):
    company = request.user.company_profile
    estates = Estate.objects.filter(company=company).order_by('-date_added')
    context = {
        'estates': estates,  # Changed key from estate_plots to estates
    }
    return render(request, 'admin_side/view_estate.html', context)


@login_required
def update_estate(request, pk):
    # SECURITY: Fetch only company's estate to prevent cross-tenant access
    company = request.user.company_profile
    estate = get_object_or_404(Estate, pk=pk, company=company)

    if request.method == "POST":
        # Retrieve updated values from the form
        name = request.POST.get('name')
        location = request.POST.get('location')
        title_deed = request.POST.get('title_deed')
        estate_size = request.POST.get('estate_size')  # Retrieve estate size

        # Debugging: Print the received values
        print(f"Received Data -> Name: {name}, Location: {location}, Title Deed: {title_deed}, Estate Size: {estate_size}")

        # Ensure required fields are present
        if not name or not location or not title_deed or not estate_size:
            messages.error(request, "All fields are required.")
            return redirect('edit-estate', pk=pk)

        # Update the estate object
        estate.name = name
        estate.location = location
        estate.title_deed = title_deed
        estate.estate_size = estate_size  # Update the estate size
        estate.save()  # Save to the database

        messages.success(request, f"{estate.name} updated successfully.")
        return redirect('view-estate')  # Redirect to estates list or details page

    # Render the form with existing estate data
    context = {
        'estate': estate,
    }
    return render(request, 'admin_side/edit_estate.html', context)

@login_required
def delete_estate(request, pk):
    """Delete an estate and return a success message on the same page."""
    # SECURITY: Only allow deletion of company's own estates
    company = request.user.company_profile
    estate = get_object_or_404(Estate, pk=pk, company=company)
    if request.method == "POST":
        estate.delete()
        messages.success(request, "Estate deleted successfully!")
        return redirect('view-estate')
    messages.error(request, "Invalid request. Please try again.")
    return redirect('view-estate')

@login_required
def add_estate(request):
    if request.method == "POST":
        # SECURITY: Auto-assign company to prevent cross-tenant data creation
        company = request.user.company_profile
        # Handle form submission and save the estate
        estate_name = request.POST.get('name')
        estate_location = request.POST.get('location')
        estate_size = request.POST.get('estate_size')
        estate_title_deed = request.POST.get('title_deed')
        
        # Create the Estate instance and save it
        estate = Estate.objects.create(
            company=company,
            name=estate_name,
            location=estate_location,
            estate_size=estate_size,
            title_deed=estate_title_deed
        )

        # Check if it's an AJAX request
        is_ajax = request.headers.get('X-Requested-With') == 'XMLHttpRequest'
        
        if is_ajax:
            return JsonResponse({'message': f'{estate.name} added successfully!', 'estate_id': estate.id})
        else:
            # Regular form submission - redirect with success message
            messages.success(request, f'<strong>{estate.name}</strong> added successfully!')
            return redirect('view-estate')

    # For GET request, render the form to add an estate
    return render(request, 'admin_side/add_estate.html', {})


# ESTATE PLOTS AND ALLOCATION FUNCTIONS

@login_required
def plot_allocation(request):
    # SECURITY: Get company context for data isolation
    company = request.user.company_profile
    
    if request.method == "POST":
        try:
            with transaction.atomic():
                # SECURITY: Ensure plot_size_unit belongs to company
                plot_size_unit = get_object_or_404(
                    PlotSizeUnits.objects.select_for_update()
                    .filter(estate_plot__estate__company=company),
                    id=request.POST.get('plotSize')
                )
                
                # SECURITY: Ensure client belongs to company
                # Check both company_profile and ClientMarketerAssignment (for users added via add_existing_user_to_company)
                client_id = request.POST.get('clientName')
                client_in_company = ClientUser.objects.filter(
                    id=client_id, company_profile=company
                ).exists() or ClientMarketerAssignment.objects.filter(
                    client_id=client_id, company=company
                ).exists()
                
                if not client_in_company:
                    raise ValidationError("Selected client does not belong to this company")
                
                client = get_object_or_404(ClientUser, id=client_id)
                
                # Calculate allocated + reserved units
                allocated_reserved = plot_size_unit.full_allocations + plot_size_unit.part_allocations
                
                # Check if available units are exhausted
                if allocated_reserved >= plot_size_unit.total_units:
                    raise ValidationError(
                        f"{plot_size_unit.plot_size.size} sqm units are completely allocated"
                    )

                # Create the allocation (note: each allocation is linked to an estate via the plot_size_unit)
                allocation = PlotAllocation(
                    plot_size_unit=plot_size_unit,
                    client=client,
                    estate=plot_size_unit.estate_plot.estate,
                    plot_size=plot_size_unit.plot_size,
                    payment_type=request.POST.get('paymentType'),
                    plot_number_id=request.POST.get('plotNumber')
                )

                allocation.full_clean()
                allocation.save()

                # Optionally, update available_units (if not using a property)
                plot_size_unit.available_units = plot_size_unit.total_units - (allocated_reserved + 1)
                plot_size_unit.save()

                messages.success(request, f"<strong>{client}</strong> Allocation successful")
                return redirect('plot-allocation')
        except ValidationError as e:
            messages.error(request, str(e))
        except Exception as e:
            messages.error(request, f"Allocation failed: {str(e)}")
        return redirect('plot-allocation')
    else:
        # GET request handling
        # SECURITY: Filter by company to prevent cross-tenant data exposure
        # Include clients from both company_profile and ClientMarketerAssignment
        clients_primary = CustomUser.objects.filter(role='client', company_profile=company)
        
        # Get affiliated clients
        affiliation_client_ids = list(
            ClientMarketerAssignment.objects.filter(company=company).values_list('client_id', flat=True).distinct()
        )
        clients_affiliated = CustomUser.objects.filter(id__in=affiliation_client_ids).exclude(
            id__in=clients_primary.values_list('pk', flat=True)
        )
        
        # Combine both querysets
        clients = list(clients_primary) + list(clients_affiliated)
        
        estates = Estate.objects.filter(company=company)
        context = {
            'clients': clients,
            'estates': estates,
            # Include any additional data your template might need
        }
        return render(request, 'admin_side/plot_allocation.html', context)

def load_plots(request):
    estate_id = request.GET.get('estate_id')
    allocation_id = request.GET.get('allocation_id')  # Get allocation ID for update scenarios
    
    plot_size_units = PlotSizeUnits.objects.filter(
        estate_plot__estate_id=estate_id,
        available_units__gt=0
    ).select_related('plot_size').annotate(
        formatted_size=Concat(F('plot_size__size'), Value(' sqm'))
    ).values('id', 'formatted_size', 'plot_size__id', 'available_units')
    
    # Get available plot numbers
    # If allocation_id is provided, include its plot number even if allocated
    if allocation_id:
        try:
            current_allocation = PlotAllocation.objects.get(id=allocation_id)
            current_plot_id = current_allocation.plot_number.id if current_allocation.plot_number else None
            
            # Get unallocated plots OR the current allocation's plot
            if current_plot_id:
                plot_numbers = PlotNumber.objects.filter(
                    estates__estate_id=estate_id
                ).filter(
                    Q(plotallocation__isnull=True) | Q(id=current_plot_id)
                ).distinct().values('id', 'number')
            else:
                plot_numbers = PlotNumber.objects.filter(
                    estates__estate_id=estate_id
                ).exclude(
                    plotallocation__isnull=False
                ).values('id', 'number')
        except PlotAllocation.DoesNotExist:
            plot_numbers = PlotNumber.objects.filter(
                estates__estate_id=estate_id
            ).exclude(
                plotallocation__isnull=False
            ).values('id', 'number')
    else:
        # Original behavior: only unallocated plots
        plot_numbers = PlotNumber.objects.filter(
            estates__estate_id=estate_id
        ).exclude(
            plotallocation__isnull=False
        ).values('id', 'number')
    
    
    return JsonResponse({
        'plot_size_units': list(plot_size_units),
        'plot_numbers': list(plot_numbers)
    })


def check_availability(request, size_id):
    try:
        size = PlotSizeUnits.objects.get(id=size_id)
        allocated_reserved = size.full_allocations + size.part_allocations
        available_units = size.total_units - allocated_reserved
        message = f"{size.plot_size.size} sqm units are available" if available_units > 0 else f"{size.plot_size.size} sqm units are completely allocated"
        return JsonResponse({'available': available_units, 'message': message})
    except PlotSizeUnits.DoesNotExist:
        return JsonResponse({'available': 0, 'message': 'Invalid plot size selected'}, status=404)


def available_plot_numbers(request, estate_id):
    # SECURITY: ensure estate belongs to user's company and only return company-scoped plot numbers
    company = getattr(request.user, 'company_profile', None)
    if not company:
        return JsonResponse({'error': 'Company context missing'}, status=403)

    try:
        estate = Estate.objects.get(id=estate_id, company=company)
    except Estate.DoesNotExist:
        return JsonResponse({'error': 'Estate not found'}, status=404)

    plot_numbers = PlotNumber.objects.filter(
        estates__estate=estate,
        company=company
    ).exclude(
        plotallocation__isnull=False
    ).values('id', 'number')

    if not plot_numbers:
        return JsonResponse({'error': 'No available plot numbers found.'}, status=404)

    return JsonResponse(list(plot_numbers), safe=False)


def get_allocation(request, allocation_id):
    # SECURITY: ensure allocation belongs to user's company
    company = getattr(request.user, 'company_profile', None)
    allocation = get_object_or_404(PlotAllocation, id=allocation_id, estate__company=company)
    return JsonResponse({
        'id': allocation.id,
        'client': allocation.client.full_name,
        'payment_type': allocation.payment_type,
        'plot_number': allocation.plot_number.id if allocation.plot_number else None
    })


def update_allocated_plot(request):
    User = get_user_model()

    if request.method == "POST":
        allocation_id = request.POST.get('allocation_id')
        client_id = request.POST.get('clientName')
        estate_id = request.POST.get('estateName')
        plot_size_unit_id = request.POST.get('plotSize')
        payment_type = request.POST.get('paymentType')
        plot_number_id = request.POST.get('plotNumber', None)

        try:
            if allocation_id:
                # SECURITY: Verify company ownership of allocation
                company = request.user.company_profile
                allocation = PlotAllocation.objects.get(
                    id=allocation_id,
                    estate__company=company
                )

                # Assign new values
                allocation.client = User.objects.get(id=client_id)
                allocation.estate = Estate.objects.get(id=estate_id, company=company)
                allocation.plot_size_unit = PlotSizeUnits.objects.get(id=plot_size_unit_id)
                allocation.plot_size = allocation.plot_size_unit.plot_size
                allocation.payment_type = payment_type

                if payment_type == 'full':
                    if not plot_number_id:
                        raise ValueError("Plot number is required for full payment")
                    # SECURITY: Only get plot numbers from this company
                    company = getattr(request, 'company', None)
                    allocation.plot_number = PlotNumber.objects.get(id=plot_number_id, company=company)
                else:
                    allocation.plot_number = None

                allocation.save()
                messages.success(request, "Allocation updated successfully.")
            else:
                messages.error(request, "Update not successful.")
                return redirect(request.path)

            return redirect(request.path + "?allocation_id=" + str(allocation.id))

        except Exception as e:
            messages.error(request, str(e))
            fallback_url = request.path
            if allocation_id:
                fallback_url += "?allocation_id=" + str(allocation_id)
            return redirect(request.META.get('HTTP_REFERER', fallback_url))

    # GET: Render the update form.
    allocation = None
    allocation_id = request.GET.get('allocation_id')
    if allocation_id:
        try:
            # SECURITY: Verify company ownership
            company = request.user.company_profile
            allocation = PlotAllocation.objects.get(
                id=allocation_id,
                estate__company=company
            )
        except PlotAllocation.DoesNotExist:
            messages.error(request, "Allocation not found.")

    # Only display plot sizes that are NOT completely allocated (i.e. all units allocated with full payment)
    if allocation and allocation.estate:
        qs = PlotSizeUnits.objects.filter(
            estate_plot__estate=allocation.estate
        ).annotate(
            full_alloc_count=Count(
                'allocations',
                filter=Q(allocations__payment_type='full', allocations__plot_number__isnull=False)
            )
        )
        # When editing, always include the currently assigned plot size unit.
        if allocation.plot_size_unit:
            qs = qs.filter(Q(full_alloc_count__lt=F('total_units')) | Q(id=allocation.plot_size_unit.id))
        else:
            qs = qs.filter(full_alloc_count__lt=F('total_units'))
        plot_size_units = qs.distinct()

        # SECURITY: Only show plot numbers from this company
        company = getattr(request, 'company', None)
        plot_numbers = PlotNumber.objects.filter(estates__estate=allocation.estate, company=company).distinct()
    else:
        plot_size_units = []
        plot_numbers = []

    context = {
        'clients': User.objects.filter(role='client', company_profile=company),
        'estates': Estate.objects.filter(company=company),
        'allocation': allocation,
        'plot_size_units': plot_size_units,
        'plot_numbers': plot_numbers,
    }
    return render(request, 'admin_side/update_allocated_plot.html', context)


def get_allocated_plot(request, allocation_id):
    # SECURITY: Verify company ownership before returning data
    company = request.user.company_profile
    allocation = get_object_or_404(
        PlotAllocation,
        id=allocation_id,
        estate__company=company
    )
    data = {
        'client_id': allocation.client.id,
        'estate_id': allocation.estate.id,
        'plot_size_unit_id': allocation.plot_size_unit.id if allocation.plot_size_unit else None,
        'payment_type': allocation.payment_type,
        'plot_number_id': allocation.plot_number.id if allocation.plot_number else None,
    }
    return JsonResponse(data)


@login_required
def delete_allocation(request):
    if request.method == 'POST':
        try:
            # SECURITY: Verify company ownership before deletion
            company = request.user.company_profile
            allocation_id = json.loads(request.body).get('allocation_id')
            allocation = get_object_or_404(
                PlotAllocation,
                id=allocation_id,
                estate__company=company
            )
            
            # Store related objects before deletion
            plot_size_unit = allocation.plot_size_unit
            plot_number = allocation.plot_number
            
            with transaction.atomic():
                # Delete the allocation
                allocation.delete()
                
                # Recalculate plot size unit availability
                plot_size_unit.save()  # Triggers save() which updates available_units
                
                # Free up plot number if it was a full allocation
                if plot_number:
                    plot_number.is_allocated = False
                    plot_number.save()
                    
            return JsonResponse({'success': True})
            
        except Exception as e:
            return JsonResponse({'success': False, 'error': str(e)}, status=400)
    
    return JsonResponse({'success': False, 'error': 'Invalid request'})

def download_allocations(request):
    # SECURITY: Filter allocations by company to prevent cross-tenant data access
    company = request.user.company_profile
    estate_id = request.GET.get('estate_id')
    
    allocations = PlotAllocation.objects.filter(estate__company=company)
    estate_name = "All_Estates"

    if estate_id:
        try:
            estate = Estate.objects.get(id=estate_id, company=company)
            estate_name = estate.name.replace(" ", "_")  # Replace spaces with underscores for filename
            allocations = allocations.filter(estate=estate)
        except Estate.DoesNotExist:
            return HttpResponse("Estate not found", status=404)

    # Prepare CSV response
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = f'attachment; filename="{estate_name}_Allocation.csv"'

    writer = csv.writer(response)
    writer.writerow([
        'Client Name', 'Estate', 'Plot Size', 
        'Payment Type', 'Plot Number', 'Date Allocated'
    ])

    for alloc in allocations:
        writer.writerow([
            alloc.client.full_name,
            alloc.estate.name,
            f"{alloc.plot_size.size}",
            alloc.get_payment_type_display(),
            alloc.plot_number.number if alloc.plot_number else 'Reserved',
            alloc.date_allocated.strftime("%d-%b-%Y")
        ])

    return response


# XXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXX
# XXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXX
@login_required
def view_allocated_plot(request, id):
    # SECURITY: Get company context
    company = getattr(request, 'company', None)
    
    estate = get_object_or_404(
        Estate.objects.prefetch_related(
            Prefetch('estate_plots',
                queryset=EstatePlot.objects.prefetch_related(
                    Prefetch('plot_numbers', 
                        queryset=PlotNumber.objects.filter(company=company).prefetch_related('plotallocation_set')
                    ),
                    Prefetch('plotsizeunits', 
                        queryset=PlotSizeUnits.objects.select_related('plot_size')
                    )
                )
            )
        ),
        id=id
    )

    context = {
        'estate': estate,
        'estate_plots': estate.estate_plots.all(),
    }
    return render(request, 'admin_side/view_allocated_plots.html', context)

@login_required
def delete_estate_plots(request):
    if request.method == 'POST':
        # SECURITY: Verify company ownership before deleting
        company = request.user.company_profile
        selected_ids = request.POST.getlist('selected')
        EstatePlot.objects.filter(
            id__in=selected_ids,
            estate__company=company
        ).delete()
        return redirect('estate_plot_list')

@login_required
def edit_estate_plot(request, id):
    estate_plot = get_object_or_404(EstatePlot, id=id)
    # SECURITY: Only show company's plot sizes
    company = getattr(request, 'company', None)
    plot_sizes = PlotSize.objects.filter(company=company)

    if request.method == 'POST':
        selected_plot_sizes = request.POST.getlist('plot_sizes[]')
        selected_plot_numbers = request.POST.getlist('plot_numbers[]')
        total_units = 0
        selected_units = {}

        for size in plot_sizes:
            if str(size.id) in selected_plot_sizes:
                unit_value = request.POST.get(f'plot_units_{size.id}', '0')
                try:
                    selected_units[size.id] = int(unit_value)
                    total_units += int(unit_value)
                except ValueError:
                    selected_units[size.id] = 0

        if total_units != len(selected_plot_numbers):
            messages.error(
                request,
                'Total plot size units must equal the total plot numbers selected. Please adjust your selection.'
            )
            context = {
                'estate_plot': estate_plot,
                'plot_sizes': plot_sizes,
                'plot_numbers': estate_plot.plot_numbers.all(),
                'selected_plot_sizes': selected_plot_sizes,
                'selected_units': selected_units,
                'selected_plot_numbers': selected_plot_numbers,
            }
            return render(request, 'admin_side/edit_estate_plot.html', context)

        # ---- UPDATE DATABASE ----
        # Update the plot numbers
        estate_plot.plot_numbers.set(selected_plot_numbers)

        # Clear existing plot size relationships
        estate_plot.plotsizeunits.all().delete()

        # Add new plot size relationships
        for size_id, unit_count in selected_units.items():
            plot_size = get_object_or_404(PlotSize, id=size_id)
            estate_plot.plotsizeunits.create(plot_size=plot_size, total_units=unit_count)

        estate_plot.save()

        messages.success(request, 'Estate plot updated successfully!')
        # return redirect('view-allocated-plot')  # Ensure this URL is correctly defined

    # Prepopulate selected data
    selected_plot_sizes = [str(unit.plot_size.id) for unit in estate_plot.plotsizeunits.all()]
    selected_units = {unit.plot_size.id: unit.total_units for unit in estate_plot.plotsizeunits.all()}
    selected_plot_numbers = [str(plot.id) for plot in estate_plot.plot_numbers.all()]

    context = {
        'estate_plot': estate_plot,
        'plot_sizes': plot_sizes,
        'plot_numbers': estate_plot.plot_numbers.all(),
        'selected_plot_sizes': selected_plot_sizes,
        'selected_units': selected_units,
        'selected_plot_numbers': selected_plot_numbers,
    }
    return render(request, 'admin_side/edit_estate_plot.html', context)

def download_estate_pdf(request, estate_id):
    # Fetch estate details and allocations
    # SECURITY: Verify company ownership of estate before generating PDF
    company = request.user.company_profile
    estate = get_object_or_404(Estate, id=estate_id, company=company)
    allocations = PlotAllocation.objects.filter(estate=estate)

    # Create response object
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="{estate.name}_Allocations.pdf"'

    # Create PDF
    pdf = canvas.Canvas(response, pagesize=A4)
    pdf.setTitle(f"{estate.name}\nAllocations Report")

    # Title Styling
    pdf.setFont("Helvetica-Bold", 18)
    pdf.setFillColor(colors.darkblue)
    title_width = pdf.stringWidth(estate.name, "Helvetica-Bold", 18)
    pdf.drawString((A4[0] - title_width) / 2, 800, estate.name)  # Center title on X-axis

    pdf.setFont("Helvetica-Bold", 14)
    pdf.setFillColor(colors.black)
    subtitle_width = pdf.stringWidth("Plot Allocations Report", "Helvetica-Bold", 14)
    pdf.drawString((A4[0] - subtitle_width) / 2, 780, "Plot Allocations Report")  # Center subtitle

    # Table Header
    data = [["#", "Client", "Phone", "Plot Size", "Payment Type", "Plot Number", "Date"]]

    # Add data rows with numbering
    for i, allocation in enumerate(allocations, start=1):
        data.append([
            str(i),  # Numbering column
            allocation.client.full_name,
            str(allocation.plot_size.size),
            allocation.get_payment_type_display(),
            allocation.plot_number.number if allocation.payment_type == "full" else "Reserved",
            allocation.date_allocated.strftime("%b %d, %Y")
        ])

    # Table Styling with Modern Colors
    table = Table(data, colWidths=[30, 120, 80, 100, 80, 80])
    table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#003366")),  # Dark Blue Header
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),  # White Text in Header
        ("ALIGN", (0, 0), (-1, -1), "CENTER"),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTNAME", (0, 1), (-1, -1), "Helvetica"),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.HexColor("#f2f2f2"), colors.HexColor("#d9e1f2")]),  # Alternating Light Grey and Blue
    ]))

    # Draw table
    table.wrapOn(pdf, 500, 600)
    table.drawOn(pdf, 50, 650)

    # Save and return response
    pdf.save()
    return response


@csrf_exempt
def add_estate_plot(request):
    if request.method == "POST":
        try:
            estate_id = request.POST.get('estate')
            if not estate_id:
                return JsonResponse({'error': 'Please select an estate'}, status=400)
            
            # SECURITY: Verify company ownership of estate
            company = request.user.company_profile
            estate = get_object_or_404(Estate, id=estate_id, company=company)
            new_plot_numbers = request.POST.getlist('plot_numbers[]', [])
            new_selected_plot_sizes = request.POST.getlist('plot_sizes[]', [])

            # Check for plot numbers already assigned to other estates
            conflict_plots = EstatePlot.objects.exclude(estate=estate)\
                .filter(plot_numbers__id__in=new_plot_numbers)\
                .values_list('plot_numbers__number', flat=True)
            if conflict_plots.exists():
                return JsonResponse({
                    'error': f'Plot numbers {list(conflict_plots)} already assigned to other estates'
                }, status=400)

            # Get or create the EstatePlot for this estate
            estate_plot, created = EstatePlot.objects.get_or_create(estate=estate)

            # Retrieve any existing PlotSizeUnits for this estate_plot
            existing_units_qs = PlotSizeUnits.objects.filter(estate_plot=estate_plot)
            existing_units = {str(unit.plot_size.id): unit for unit in existing_units_qs}

            # Process the new plot sizes & unit counts submitted
            new_plot_size_units = {}
            total_new_units = 0
            for size_id in new_selected_plot_sizes:
                plot_size = get_object_or_404(PlotSize, id=size_id)
                units_input = request.POST.get(f'plot_units_{size_id}', '0').strip()
                
                try:
                    new_units = int(units_input) if units_input else 0
                except ValueError:
                    return JsonResponse({
                        'error': f'Invalid number of units entered for {plot_size.size}. Please enter a whole number.'
                    }, status=400)
                
                if new_units < 1:
                    return JsonResponse({
                        'error': f'Number of units for {plot_size.size} must be at least 1'
                    }, status=400)
                
                new_plot_size_units[size_id] = new_units
                total_new_units += new_units

            # Validate plot size units against allocations
            for size_id, existing_unit in existing_units.items():
                if existing_unit.allocations.exists():
                    if size_id not in new_plot_size_units:
                        return JsonResponse({
                            'error': f'PLOT SIZE {existing_unit.plot_size.size} has existing allocations and cannot be removed'
                        }, status=400)
                    
                    total_allocated_reserved = existing_unit.full_allocations + existing_unit.part_allocations
                    
                    if new_plot_size_units[size_id] < total_allocated_reserved:
                        return JsonResponse({
                            'error': f'Cannot decrease PLOT SIZE {existing_unit.plot_size.size} below current allocations. '
                                    f'ALLOCATED: {existing_unit.full_allocations}, '
                                    f'RESERVED: {existing_unit.part_allocations}, '
                                    f'MINIMUM UNITS REQUIRED: {total_allocated_reserved}'
                        }, status=400)

            # Validate total units match plot numbers
            if total_new_units != len(new_plot_numbers):
                return JsonResponse({
                    'error': f'Total units ({total_new_units}) must match the number of plot numbers selected ({len(new_plot_numbers)})'
                }, status=400)

            # Check for allocated plot numbers that shouldn't be removed
            allocated_plot_numbers = set()
            for pn in estate_plot.plot_numbers.all():
                if pn.plotallocation_set.exists():
                    allocated_plot_numbers.add(str(pn.id))
            
            for allocated_id in allocated_plot_numbers:
                if allocated_id not in new_plot_numbers:
                    # SECURITY: Only get plot numbers from this company
                    company = getattr(request, 'company', None)
                    allocated_plot = PlotNumber.objects.get(id=allocated_id, company=company)
                    return JsonResponse({
                        'error': f'PLOT NUMBER {allocated_plot.number} IS ALREADY ALLOCATED AND CANNOT BE REMOVED'
                    }, status=400)

            # Update plot numbers - SECURITY: Only add plot numbers from this company
            estate_plot.plot_numbers.clear()
            company = getattr(request, 'company', None)
            estate_plot.plot_numbers.add(*PlotNumber.objects.filter(id__in=new_plot_numbers, company=company))

            # Update or create plot size units
            for size_id, new_units in new_plot_size_units.items():
                if size_id in existing_units:
                    unit = existing_units[size_id]
                    if new_units != unit.total_units:
                        unit.total_units = new_units
                        unit.available_units = new_units - (unit.full_allocations + unit.part_allocations)
                        unit.save()
                else:
                    plot_size = get_object_or_404(PlotSize, id=size_id)
                    PlotSizeUnits.objects.create(
                        estate_plot=estate_plot,
                        plot_size=plot_size,
                        total_units=new_units,
                        available_units=new_units
                    )
            
            # Remove unused plot size units (if no allocations)
            for size_id, existing_unit in existing_units.items():
                if size_id not in new_plot_size_units:
                    if existing_unit.allocations.exists():
                        return JsonResponse({
                            'error': f'PLOT SIZE {existing_unit.plot_size.size} has existing allocations and cannot be removed'
                        }, status=400)
                    existing_unit.delete()

            return JsonResponse({'message': 'Estate plots updated successfully!'})

        except Exception as e:
            error_message = str(e)
            if 'invalid literal for int() with base 10' in error_message:
                return JsonResponse({
                    'error': 'Invalid number entered. Please check all unit values and ensure they are whole numbers.'
                }, status=400)
            return JsonResponse({
                'error': 'An unexpected error occurred. Please try again or contact support.'
            }, status=400)

    # GET request handling
    # SECURITY: Get company context
    company = getattr(request, 'company', None)
    
    allocated_plot_ids = list(EstatePlot.objects.exclude(plot_numbers=None)
                              .values_list('plot_numbers', flat=True).distinct())
    return render(request, 'admin_side/estate-plot.html', {
        'estates': Estate.objects.filter(company=company),
        'plot_sizes': PlotSize.objects.filter(company=company),
        'plot_numbers': PlotNumber.objects.filter(company=company),
        'allocated_plot_ids': allocated_plot_ids,
    })


def get_estate_details(request, estate_id):
    # SECURITY: ensure estate belongs to user's company
    company = getattr(request.user, 'company_profile', None)
    estate = get_object_or_404(Estate, id=estate_id, company=company)
    try:
        estate_plot = EstatePlot.objects.get(estate=estate)
        plot_sizes_units = list(estate_plot.plotsizeunits.values(
            'plot_size__id', 'total_units'
        ))
        current_plot_numbers = list(estate_plot.plot_numbers.values_list('id', flat=True))
    except EstatePlot.DoesNotExist:
        plot_sizes_units = []
        current_plot_numbers = []

    # Get plot numbers allocated to OTHER estates
    allocated_plot_ids = list(EstatePlot.objects.exclude(estate=estate)
                              .values_list('plot_numbers', flat=True).distinct())

    return JsonResponse({
        'plot_sizes': plot_sizes_units,
        'plot_numbers': current_plot_numbers,
        'allocated_plot_ids': allocated_plot_ids,
    })


def allocate_units(estate_plot_id, size_id, units_to_allocate):
    plot_size_unit = PlotSizeUnits.objects.get(
        estate_plot_id=estate_plot_id,
        plot_size_id=size_id
    )
    
    if plot_size_unit.available_units >= units_to_allocate:
        plot_size_unit.available_units -= units_to_allocate
        plot_size_unit.save()
    else:
        raise ValueError("Not enough units available for this plot size")

def allocated_plot(request, estate_id):
    # SECURITY: ensure estate belongs to user's company
    company = getattr(request.user, 'company_profile', None)
    estate = get_object_or_404(Estate, id=estate_id, company=company)
    # Fetch all plot allocations for this estate
    plot_allocations = PlotAllocation.objects.filter(estate=estate)

    allocated_plots = plot_allocations.filter(payment_type='full')
    unallocated_plots = plot_allocations.filter(payment_type='part')

    # Fetch client details for those who have paid full and partial
    clients_allocated = []
    clients_not_allocated = []

    for allocation in allocated_plots:
        plot_number = allocation.plot_number.number if allocation.plot_number else 'Not Assigned'
        clients_allocated.append({
            'client_name': allocation.client.full_name,
            'payment_type': allocation.payment_type,
            'estate_name': allocation.estate.name,
            'plot_size': allocation.plot_size.size,
            'plot_number': plot_number,
            'status': 'Allocated',
            'action_url': f'/edit-client/{allocation.client.id}/'
        })

    for allocation in unallocated_plots:
        plot_number = allocation.plot_number.number if allocation.plot_number else 'Not Assigned'
        clients_not_allocated.append({
            'client_name': allocation.client.full_name,
            'payment_type': allocation.payment_type,
            'estate_name': allocation.estate.name,
            'plot_size': allocation.plot_size.size,
            'plot_number': plot_number,
            'status': 'Not Allocated',
            'action_url': f'/edit-client/{allocation.client.id}/'
        })

    return render(request, 'admin_side/allocated_plot.html', {
        'estate': estate,
        'allocated_plots': allocated_plots,
        'unallocated_plots': unallocated_plots,
        'clients_allocated': clients_allocated,
        'clients_not_allocated': clients_not_allocated
    })


def fetch_plot_data(request):
    # SECURITY: only return data scoped to the user's company
    company = getattr(request.user, 'company_profile', None)
    if not company:
        return JsonResponse({'allocated_plots': [], 'reserved_count': 0})

    # Allocated plots (with plot numbers assigned)
    from django.db.models import F
    allocated_plots = PlotAllocation.objects.filter(estate__company=company, is_allocated=True)
    allocated_plots = allocated_plots.annotate(client_name=F('client__full_name')).values('plot_number', 'client_name')

    # Reserved or unallocated plots (clients without plot numbers)
    reserved_plots_count = PlotAllocation.objects.filter(estate__company=company, plot_number__isnull=True).count()

    data = {
        'allocated_plots': list(allocated_plots),
        'reserved_count': reserved_plots_count
    }
    return JsonResponse(data)


# FLOOR PLAN

def estate_property_list(request):
    return render(request, "admin_side/estate_property_list.html",)

def add_floor_plan(request):
    estate_id = request.GET.get('estate_id')
    estate = get_object_or_404(Estate, id=estate_id)
    
    # SECURITY: Get company context
    company = getattr(request, 'company', None)
    
    # Get plot sizes available for this estate - ONLY from this company
    plot_sizes = PlotSize.objects.filter(
        plotsizeunits__estate_plot__estate=estate,
        company=company
    ).distinct()

    if request.method == "POST":
        # Handle form submission
        plot_size_id = request.POST.get('plot_size')
        floor_plan_image = request.FILES.get('floor_plan_image')
        plan_title = request.POST.get('plan_title')

        try:
            # SECURITY: Only get plot sizes from this company
            company = getattr(request, 'company', None)
            plot_size = PlotSize.objects.get(id=plot_size_id, company=company)
            EstateFloorPlan.objects.create(
                estate=estate,
                plot_size=plot_size,
                floor_plan_image=floor_plan_image,
                plan_title=plan_title
            )
            messages.success(request, f"Floor plan for {plot_size} added successfully!")
            redirect_url = reverse('add_floor_plan') + f"?estate_id={estate_id}"
            return redirect(redirect_url)
        
        except Exception as e:
            messages.error(request, f"Error saving floor plan: {str(e)}")
            return redirect(request.META.get('HTTP_REFERER'))

    context = {
        'estate': estate,
        'plot_sizes': plot_sizes,
        'preselected_estate_id': estate_id,
    }
    return render(request, 'admin_side/add_floor_plan.html', context)


def get_plot_sizes_for_floor_plan(request, estate_id):
    # SECURITY: Only get plot sizes from this company
    company = getattr(request, 'company', None)
    plot_sizes = PlotSize.objects.filter(
        plotsizeunits__estate_plot__estate_id=estate_id,
        company=company
    ).distinct().values('id', 'size')
    
    return JsonResponse(list(plot_sizes), safe=False)


def estate_details(request):
    estate_id = request.GET.get("estate_id")
    if not estate_id:
        messages.error(request, "Estate ID is missing.")
        return redirect("view-estate")
    
    estate = get_object_or_404(Estate, id=estate_id)
    floor_plans = EstateFloorPlan.objects.filter(estate=estate)\
        .select_related("estate", "plot_size")\
        .order_by("-date_uploaded")
    
    prototypes = estate.prototypes.all()
    
    context = {
        "estate": estate,
        "floor_plans": floor_plans,
        "prototypes": prototypes,
    }
    return render(request, "admin_side/estate_details.html", context)



# protoTypes
def add_prototypes(request):
    estate_id = request.GET.get('estate_id')
    
    try:
        estate = Estate.objects.get(id=estate_id)
    except Estate.DoesNotExist:
        messages.error(request, "Estate not found")
        return get_tenant_dashboard_redirect(request)

    # Get plot sizes available for this estate
    plot_sizes = PlotSize.objects.filter(
        plotsizeunits__estate_plot__estate=estate
    ).distinct()

    if request.method == "POST":
        # Handle form submission
        plot_size_id = request.POST.get('plot_size')
        prototype_image = request.FILES.get('prototype_image')
        title = request.POST.get('title')
        description = request.POST.get('description', '')

        try:
            # SECURITY: Only get plot sizes from this company
            company = getattr(request, 'company', None)
            plot_size = PlotSize.objects.get(id=plot_size_id, company=company)
            EstatePrototype.objects.create(
                estate=estate,
                plot_size=plot_size,
                prototype_image=prototype_image,
                Title=title,
                Description=description,
            )
            messages.success(request, f"{plot_size} prototype added successfully!")
            redirect_url = reverse('add_prototypes') + f"?estate_id={estate_id}"
            return redirect(redirect_url)
        
        except Exception as e:
            messages.error(request, f"Error saving prototype: {str(e)}")
            return redirect(request.META.get('HTTP_REFERER'))

    context = {
        'estate': estate,
        'plot_sizes': plot_sizes,
        'preselected_estate_id': estate_id,
    }
    return render(request, 'admin_side/add_prototype.html', context)


def get_plot_sizes_for_prototypes(request, estate_id):
    # SECURITY: only return plot sizes that belong to the same company as the requesting user
    company = getattr(request.user, 'company_profile', None)
    if not company:
        return JsonResponse([], safe=False)

    plot_sizes = PlotSize.objects.filter(
        plotsizeunits__estate_plot__estate_id=estate_id,
        company=company
    ).distinct().values('id', 'size')

    return JsonResponse(list(plot_sizes), safe=False)

# plot allocation
def get_plot_sizes(request, estate_id):
    # SECURITY: ensure estate belongs to user's company before returning plot sizes
    company = getattr(request.user, 'company_profile', None)
    if not company:
        return JsonResponse({"error": "Company context missing"}, status=403)

    try:
        estate = Estate.objects.get(id=estate_id, company=company)
        plot_sizes = estate.plot_sizes.all()
        data = {
            "plot_sizes": [{"id": plot_size.id, "size": plot_size.size} for plot_size in plot_sizes]
        }
        return JsonResponse(data)
    except Estate.DoesNotExist:
        return JsonResponse({"error": "Estate not found"}, status=404)

@login_required
def deallocate_plot(request, allocation_id):
    try:
        with transaction.atomic():
            company = getattr(request.user, 'company_profile', None)
            if not company:
                messages.error(request, "Company context missing")
                return redirect('allocations-list')

            allocation = PlotAllocation.objects.select_for_update().get(id=allocation_id, estate__company=company)
            unit = allocation.plot_size_unit
            allocation.delete()
            if unit:
                unit.available_units += 1
                unit.save()
            messages.success(request, "Deallocation successful")
    except PlotAllocation.DoesNotExist:
        messages.error(request, "Allocation not found or does not belong to your company")
    except Exception as e:
        messages.error(request, str(e))
    return redirect('allocations-list')


# Estate Amenities
def update_estate_amenities(request):
    # Get estate_id from query parameters
    estate_id = request.GET.get('estate_id')
    if not estate_id:
        messages.error(request, "Estate ID is missing.")
        return redirect('appropriate_redirect_view')  # Replace with a valid redirect
    
    estate = get_object_or_404(Estate, id=estate_id)
    amenity_record, created = EstateAmenitie.objects.get_or_create(estate=estate)

    if request.method == "POST":
        form = AmenitieForm(request.POST, instance=amenity_record)
        if form.is_valid():
            amenity_obj = form.save(commit=False)
            amenity_obj.estate = estate
            amenity_obj.save()
            messages.success(request, "Amenities updated successfully.")
            # Redirect with the estate_id in the query string
            redirect_url = reverse('update_estate_amenities') + f'?estate_id={estate.id}'
            return redirect(redirect_url)
        else:
            messages.error(request, "Error updating amenities.")
    else:
        form = AmenitieForm(instance=amenity_record)

    available_amenities = [
        {'code': code, 'name': name, 'icon': AMENITY_ICONS.get(code, '')}
        for code, name in AMENITIES_CHOICES
    ]
    # selected_amenity_codes = amenity_record.amenities.split(',') if amenity_record.amenities else []
    selected_amenity_codes = amenity_record.amenities if amenity_record.amenities else []

    return render(request, "admin_side/estate_amenities.html", {
        "estate": estate,
        "form": form,
        "available_amenities": available_amenities,
        "selected_amenity_codes": selected_amenity_codes,
    })

# Estate Layout
def add_estate_layout(request):
    estate_id = request.GET.get('estate_id')
    if not estate_id:
        messages.error(request, "Estate ID is missing.")
        return get_tenant_dashboard_redirect(request)

    estate = get_object_or_404(Estate, id=estate_id)

    if request.method == "POST":
        layout_image = request.FILES.get('layout_image')
        if layout_image:
            try:
                EstateLayout.objects.create(
                    estate=estate,
                    layout_image=layout_image,
                )
                messages.success(request, "Estate layout uploaded successfully!")
                # Redirect to a details page or any other page you choose
                redirect_url = reverse('add_estate_layout') + f"?estate_id={estate.id}"
                return redirect(redirect_url)
            except Exception as e:
                messages.error(request, f"Error saving estate layout: {str(e)}")
                return redirect(request.META.get('HTTP_REFERER'))
        else:
            messages.error(request, "No layout image provided.")
            return redirect(request.META.get('HTTP_REFERER'))

    context = {
        "estate": estate,
    }
    return render(request, "admin_side/add_estate_layout.html", context)

# Estate Map
def add_estate_map(request):
    estate_id = request.GET.get("estate_id")
    if not estate_id:
        messages.error(request, "Estate ID is missing.")
        return get_tenant_dashboard_redirect(request)
    
    estate = get_object_or_404(Estate, id=estate_id)
    
    # Get or create the estate map record
    estate_map, created = EstateMap.objects.get_or_create(estate=estate)
    
    if request.method == "POST":
        latitude = request.POST.get("latitude")
        longitude = request.POST.get("longitude")
        google_map_link = request.POST.get("google_map_link")

        if latitude and longitude:
            estate_map.latitude = latitude
            estate_map.longitude = longitude
            if google_map_link:
                estate_map.google_map_link = google_map_link
            estate_map.save()
            messages.success(request, "Estate map updated successfully!")
            redirect_url = reverse("add_estate_map") + f"?estate_id={estate.id}"
            return redirect(redirect_url)
        else:
            messages.error(request, "Please provide both latitude and longitude.")
            return redirect(request.META.get("HTTP_REFERER"))
    
    context = {
        "estate": estate,
        "estate_map": estate_map,
    }
    return render(request, "admin_side/add_estate_map.html", context)

# Estate Work Progress
def add_progress_status(request):
    estate_id = request.GET.get('estate_id')
    if not estate_id:
        messages.error(request, "Estate ID is missing.")
        return get_tenant_dashboard_redirect(request)

    estate = get_object_or_404(Estate, id=estate_id)

    if request.method == "POST":
        progress_text = request.POST.get('progress_status')
        if progress_text:
            ProgressStatus.objects.create(
                estate=estate,
                progress_status=progress_text,
            )
            messages.success(request, "Progress status updated successfully!")
            # Redirect back to the same page to show the updated list
            redirect_url = reverse('add_progress_status') + f"?estate_id={estate_id}"
            return redirect(redirect_url)
        else:
            messages.error(request, "Please enter a progress status.")
            return redirect(request.META.get('HTTP_REFERER'))

    # Retrieve progress updates for the estate, ordered by timestamp descending
    progress_list = estate.progress_status.all().order_by("-timestamp")

    context = {
        "estate": estate,
        "progress_list": progress_list,
    }
    return render(request, "admin_side/add_progress_status.html", context)


@login_required
def marketer_list(request):
    """
    Display marketer list with company-scoped data and client counts.
    SECURITY: Strict company isolation prevents cross-company data leakage.
    """
    # Get company context with fallback
    company = getattr(request, 'company', None) or getattr(request.user, 'company_profile', None)
    if not company:
        messages.error(request, "No company context available.")
        return redirect('home')

    # Get marketers from both primary company_profile and MarketerAffiliation
    # Primary marketers (direct company assignment)
    primary_marketers = MarketerUser.objects.filter(company_profile=company)
    
    # Affiliated marketers (added via add_existing_user_to_company)
    affiliation_marketer_ids = list(
        MarketerAffiliation.objects.filter(company=company)
        .values_list('marketer_id', flat=True)
        .distinct()
    )
    affiliated_marketers = MarketerUser.objects.filter(
        id__in=affiliation_marketer_ids
    ).exclude(id__in=primary_marketers.values_list('pk', flat=True))
    
    # Combine and sort by registration date
    marketers = list(primary_marketers) + list(affiliated_marketers)
    marketers.sort(key=lambda u: getattr(u, 'date_registered', None) or timezone.now(), reverse=True)

    # Calculate company-scoped client counts and IDs
    for marketer in marketers:
        # Count unique clients assigned to this marketer within the company
        assign_client_ids = set(
            ClientMarketerAssignment.objects.filter(
                company=company, 
                marketer_id=marketer.id
            ).values_list('client_id', flat=True)
        )
        direct_client_ids = set(
            ClientUser.objects.filter(
                company_profile=company, 
                assigned_marketer_id=marketer.id
            ).values_list('pk', flat=True)
        )
        marketer.client_count = len(assign_client_ids.union(direct_client_ids))
        
        # Get company-specific marketer ID
        try:
            profile = CompanyMarketerProfile.objects.get(
                marketer_id=marketer.id, 
                company=company
            )
            marketer.company_marketer_id = profile.company_marketer_id
            marketer.company_marketer_uid = profile.company_marketer_uid
        except CompanyMarketerProfile.DoesNotExist:
            # Fallback: use marketer's primary key for stable ID generation
            # This ensures the ID never changes regardless of list order
            marketer.company_marketer_id = marketer.id
            prefix = getattr(company, 'company_name', 'CMP')[:3].upper()
            marketer.company_marketer_uid = f"{prefix}MKT{marketer.id:03d}"

    return render(request, 'admin_side/marketer_list.html', {
        'marketers': marketers, 
        'company': company
    })


@login_required
def api_marketer_list(request):
    """
    API endpoint to retrieve a list of marketers for the authenticated user's company.
    Returns JSON data for use in frontend JavaScript.
    """
    import json
    from django.http import JsonResponse
    from django.core.serializers import serialize
    
    user = request.user
    
    # Get the user's company
    company = user.company_profile
    if not company:
        return JsonResponse({'error': 'No company associated with your account'}, status=400)
    
    # Primary queryset: marketers with MarketerUser subclass
    marketers_qs = MarketerUser.objects.filter(company_profile=company)
    
    # ALSO include marketers from MarketerAffiliation (for users added via add_existing_user_to_company)
    affiliation_marketer_ids = list(
        MarketerAffiliation.objects.filter(company=company).values_list('marketer_id', flat=True).distinct()
    )
    affiliation_marketers_qs = MarketerUser.objects.filter(id__in=affiliation_marketer_ids).exclude(
        id__in=marketers_qs.values_list('pk', flat=True)
    )
    
    # Combine both querysets
    combined = list(marketers_qs) + list(affiliation_marketers_qs)
    
    # Sort by registration date descending for consistency
    combined.sort(key=lambda u: getattr(u, 'date_registered', None) or timezone.now(), reverse=True)
    
    # Serialize the data
    marketers_data = []
    for marketer in combined:
        marketers_data.append({
            'id': marketer.id,
            'full_name': marketer.full_name,
            'email': marketer.email,
            'phone': getattr(marketer, 'phone', ''),
            'date_registered': marketer.date_registered.isoformat() if hasattr(marketer, 'date_registered') else None,
        })
    
    return JsonResponse({'marketers': marketers_data})


@login_required
def admin_marketer_profile(request, slug=None, pk=None, company_slug=None):
    """
    Display marketer profile with performance metrics and leaderboard.
    
    Supports multiple URL formats:
    1. Legacy: /admin-marketer/<pk>/ (with or without ?company=slug)
    2. UID-based: /<LPLMKT001>.marketer-profile?company=<slug> or without company
    3. Name-based slug: /<victor-godwin>.marketer-profile?company=<slug>
    4. Company-namespaced: /<company_slug>/marketer/<identifier>/
    
    SECURITY: Enforces strict company isolation - only shows marketer data scoped to their company
    Prevents cross-company data leakage by validating company context in all code paths.
    """
    
    # ============================================================
    # STEP 1: Determine and validate company context (FLEXIBLE)
    # ============================================================
    company = None
    company_slug_param = None
    
    # Priority 1: URL company_slug parameter (most explicit)
    if company_slug:
        company = get_object_or_404(Company, slug=company_slug)
    # Priority 2: Query parameter ?company=slug (modern routing)
    elif request.GET.get('company'):
        company_slug_param = request.GET.get('company').strip()
        if company_slug_param:  # Only use if not empty string
            company = get_object_or_404(Company, slug=company_slug_param)
        else:
            # Empty company parameter provided (?company=) - user must specify a company
            raise Http404("⚠️  Company parameter is empty. Please specify: ?company=<company-slug>")
    
    # Priority 3: User's primary company_profile (fallback for legacy routes ONLY)
    # Use only if NO company parameter was provided at all
    if not company and not request.GET.get('company'):
        if hasattr(request.user, 'company_profile') and request.user.company_profile:
            company = request.user.company_profile
    
    # SECURITY: No company context = deny access (cannot infer company)
    if not company:
        raise Http404("❌ No company context provided. Required: ?company=<company-slug> or URL must include company slug.")
    
    # ============================================================
    # STEP 2: Verify requesting user has access to this company
    # ============================================================
    user_company = getattr(request.user, 'company_profile', None)
    if user_company != company:
        # Restrict to user's primary company to prevent accessing other company's data
        raise HttpResponseForbidden(f"❌ Access Denied: You can only access profiles within your own company context.")
    
    # ============================================================
    # STEP 3: Find marketer within company context (STRICT FILTERING)
    # ============================================================
    marketer = None
    
    if slug:
        # Slug-based lookup supports MULTIPLE formats:
        # 1. Company-specific UID: "LPLMKT001" (company_marketer_uid)
        # 2. Name-based slug: "amaka-njoku" (converted from full_name)
        # 3. Exact name: "Amaka Njoku"
        
        # Priority 1: Try company-specific UID lookup (LPLMKT001)
        # Check BOTH MarketerUser.company_marketer_uid AND CompanyMarketerProfile.company_marketer_uid
        if company:
            # First check direct UID on MarketerUser
            marketer = MarketerUser.objects.filter(
                company_marketer_uid=slug,
                company_profile=company
            ).first()
            
            # If not found, check CompanyMarketerProfile (for affiliated users)
            if not marketer:
                try:
                    profile = CompanyMarketerProfile.objects.get(
                        company_marketer_uid=slug,
                        company=company
                    )
                    marketer = profile.marketer
                except CompanyMarketerProfile.DoesNotExist:
                    pass
        
        # Priority 2: Try name-based slug (e.g., "amaka-njoku" → "Amaka Njoku")
        if not marketer:
            name_from_slug = slug.replace('-', ' ')
            user = MarketerUser.objects.filter(
                full_name__iexact=name_from_slug
            ).first()
            
            # Verify user exists in THIS company (via company_profile OR affiliation)
            if user:
                has_company_access = (
                    user.company_profile == company or
                    MarketerAffiliation.objects.filter(
                        marketer_id=user.id,
                        company=company
                    ).exists()
                )
                
                if has_company_access:
                    marketer = user
                else:
                    raise HttpResponseForbidden(
                        f"❌ Access Denied: Marketer '{slug}' is not part of {company.company_name}."
                    )
        
        # Priority 3: Try exact name match
        if not marketer:
            user = MarketerUser.objects.filter(
                full_name=slug
            ).first()
            
            if user:
                has_company_access = (
                    user.company_profile == company or
                    MarketerAffiliation.objects.filter(
                        marketer_id=user.id,
                        company=company
                    ).exists()
                )
                
                if has_company_access:
                    marketer = user
                else:
                    raise HttpResponseForbidden(
                        f"❌ Access Denied: Marketer '{slug}' is not part of {company.company_name}."
                    )
        
        if not marketer:
            raise Http404(f"❌ Marketer '{slug}' not found in {company.company_name}.")
    
    elif pk:
        # Legacy ID-based lookup: VERIFY marketer belongs to company
        try:
            marketer = MarketerUser.objects.get(id=pk)
        except MarketerUser.DoesNotExist:
            try:
                marketer = CustomUser.objects.get(id=pk, role='marketer')
            except CustomUser.DoesNotExist:
                raise Http404("❌ Marketer not found.")
        
        # SECURITY: Verify marketer belongs to THIS company ONLY
        marketer_in_company = (
            marketer.company_profile == company or
            MarketerAffiliation.objects.filter(
                marketer_id=marketer.id,
                company=company
            ).exists()
        )
        
        if not marketer_in_company:
            raise HttpResponseForbidden(f"❌ Access Denied: Marketer not associated with {company.company_name}.")
    else:
        raise Http404("❌ No marketer identifier provided (slug or pk required).")
    
    # ============================================================
    # STEP 4: Fetch performance data STRICTLY scoped to company
    # ============================================================
    
    now           = timezone.now()
    current_year  = now.year
    year_str      = str(current_year)
    current_month = now.strftime("%Y-%m")
    password_response = None

    # Lifetime stats (STRICT company filter)
    lifetime_closed_deals = Transaction.objects.filter(
        marketer=marketer,
        company=company  # <-- STRICT COMPANY FILTER
    ).count()

    lifetime_commission = MarketerPerformanceRecord.objects.filter(
        marketer=marketer,
        company=company  # ✅ Added company filter for isolation
    ).aggregate(total=Sum('commission_earned'))['total'] or 0

    performance = {
        'closed_deals':      lifetime_closed_deals,
        'total_sales':       0,
        'commission_earned': lifetime_commission,
        'commission_rate':   0,
        'target_achievement': 0,
        'yearly_target_achievement': None,
    }

    # Latest commission rate (STRICT company filter)
    comm = MarketerCommission.objects.filter(
        marketer=marketer,
        company=company  # ✅ Added company filter for isolation
    ).order_by('-effective_date').first()
    if comm:
        performance['commission_rate'] = comm.rate

    # Monthly target % (STRICT company filter)
    mt = MarketerTarget.objects.filter(
        marketer=marketer,
        company=company,  # ✅ Added company filter for isolation
        period_type='monthly',
        specific_period=current_month
    ).first()
    if mt and mt.target_amount:
        performance['target_achievement'] = min(
            100,
            performance['total_sales'] / mt.target_amount * 100
        )

    # Annual target achievement (STRICT company filter)
    at = (
        MarketerTarget.objects.filter(
            marketer=marketer, 
            company=company,  # ✅ Added company filter for isolation
            period_type='annual', 
            specific_period=year_str
        )
        .first()
        or
        MarketerTarget.objects.filter(
            marketer=None, 
            company=company,  # ✅ Added company filter for isolation
            period_type='annual', 
            specific_period=year_str
        )
        .first()
    )
    if at and at.target_amount:
        total_year_sales = Transaction.objects.filter(
            marketer=marketer,
            company=company,  # <-- STRICT COMPANY FILTER
            transaction_date__year=current_year
        ).aggregate(total=Sum('total_amount'))['total'] or 0
        performance['yearly_target_achievement'] = min(
            100,
            total_year_sales / at.target_amount * 100
        )

    # ============================================================
    # STEP 5: Build leaderboard (COMPANY-SCOPED ONLY)
    # ============================================================
    sales_data = []
    
    # Get marketers ONLY from this company (both primary and affiliated)
    company_marketers = MarketerUser.objects.filter(company_profile=company)
    affiliation_marketer_ids = list(
        MarketerAffiliation.objects.filter(company=company).values_list('marketer_id', flat=True).distinct()
    )
    affiliation_marketers = MarketerUser.objects.filter(
        id__in=affiliation_marketer_ids
    ).exclude(id__in=company_marketers.values_list('pk', flat=True))
    
    company_marketers = list(company_marketers) + list(affiliation_marketers)
    
    for m in company_marketers:
        # STRICT company filter in all calculations
        year_sales = Transaction.objects.filter(
            marketer=m,
            company=company,  # <-- STRICT COMPANY FILTER
            transaction_date__year=current_year
        ).aggregate(total=Sum('total_amount'))['total'] or 0

        tgt = (
            MarketerTarget.objects.filter(
                marketer=m, 
                period_type='annual', 
                specific_period=year_str
            ).first()
            or
            MarketerTarget.objects.filter(
                marketer=None, 
                period_type='annual', 
                specific_period=year_str
            ).first()
        )
        target_amt = tgt.target_amount if tgt else None
        pct = (year_sales / target_amt * 100) if target_amt else None

        sales_data.append({'marketer': m, 'total_sales': year_sales, 'target_amt': target_amt, 'pct': pct})

    sales_data.sort(key=lambda x: x['total_sales'], reverse=True)

    top3 = []
    for idx, e in enumerate(sales_data[:3], start=1):
        pct = e['pct']
        if pct is None:
            category = diff = None
        elif pct > 100:
            category, diff = 'Above Target', round(pct - 100, 1)
        elif pct >= 50:
            category, diff = 'On Target', round(pct, 1)
        else:
            category, diff = 'Below Target', round(100 - pct, 1)

        top3.append({
            'rank': idx,
            'marketer': e['marketer'],
            'category': category,
            'diff_pct': diff,
            'has_target': e['target_amt'] is not None,
        })

    user_entry = None
    for idx, e in enumerate(sales_data, start=1):
        if e['marketer'].id == marketer.id:
            pct = e['pct']
            if pct is None:
                category = diff = None
            elif pct > 100:
                category, diff = 'Above Target', round(pct - 100, 1)
            elif pct >= 50:
                category, diff = 'On Target', round(pct, 1)
            else:
                category, diff = 'Below Target', round(100 - pct, 1)

            user_entry = {
                'rank': idx,
                'marketer': marketer,
                'category': category,
                'diff_pct': diff,
                'has_target': e['target_amt'] is not None,
            }
            break

    if request.method == 'POST':
        # Profile update
        if request.POST.get("update_profile"):
            ok = update_profile_data(marketer, request)
            if ok:
                messages.success(request, "Your profile has been updated successfully!")
            else:
                messages.error(request, "Failed to update your profile.")

        # Password change
        elif request.POST.get('change_password'):
            password_response = change_password(request)
            cd = password_response.context_data
            if cd.get('success'):
                messages.success(request, cd['success'])
            elif cd.get('error'):
                messages.error(request, cd['error'])

        return redirect('marketer-profile')



    return render(request, 'admin_side/marketer_profile.html', {
        'password_response': password_response,
        'performance': performance,
        'top3':        top3,
        'user_entry':  user_entry,
        'current_year': current_year,
        'company': company,
    })


# Admin Chat Hub - Entry point without requiring a specific client
@login_required
def admin_chat_hub_view(request):
    """
    Chat hub view that shows the chat interface without requiring a client_id.
    Users can browse and select clients/marketers from the sidebar.
    Also provides API endpoints for searching company users to initiate new chats.
    """
    # Get clients and marketers for sidebar chat list
    company_filter = {'company_profile': request.company} if hasattr(request, 'company') and request.company else {}
    
    # Get clients who have conversations
    sidebar_clients = CustomUser.objects.filter(
        role='client',
        sent_messages__isnull=False,
        **company_filter
    ).distinct().annotate(
        last_message=Max('sent_messages__date_sent')
    ).order_by('-last_message')[:30]
    
    for c in sidebar_clients:
        c.unread_count = Message.objects.filter(
            sender=c, recipient__role__in=SUPPORT_ROLES, is_read=False
        ).count()
        last_msg = Message.objects.filter(
            Q(sender=c, recipient__role__in=SUPPORT_ROLES) | Q(sender__role__in=SUPPORT_ROLES, recipient=c)
        ).order_by('-date_sent').first()
        c.last_message_text = last_msg.content[:50] if last_msg and last_msg.content else ''

    # Get marketers who have conversations
    sidebar_marketers = CustomUser.objects.filter(
        role='marketer',
        sent_messages__isnull=False,
        **company_filter
    ).distinct().annotate(
        last_message=Max('sent_messages__date_sent')
    ).order_by('-last_message')[:30]
    
    for m in sidebar_marketers:
        m.unread_count = Message.objects.filter(
            sender=m, recipient__role__in=SUPPORT_ROLES, is_read=False
        ).count()
        last_msg = Message.objects.filter(
            Q(sender=m, recipient__role__in=SUPPORT_ROLES) | Q(sender__role__in=SUPPORT_ROLES, recipient=m)
        ).order_by('-date_sent').first()
        m.last_message_text = last_msg.content[:50] if last_msg and last_msg.content else ''

    # Get ALL company clients/marketers for "new chat" search functionality
    all_company_clients = CustomUser.objects.filter(
        role='client',
        **company_filter
    ).order_by('full_name')[:100]
    
    all_company_marketers = CustomUser.objects.filter(
        role='marketer',
        **company_filter
    ).order_by('full_name')[:100]

    context = {
        'clients': sidebar_clients,
        'marketers': sidebar_marketers,
        'all_clients': all_company_clients,
        'all_marketers': all_company_marketers,
        'is_hub': True,  # Flag to indicate this is the hub view (no active chat)
        'is_marketer': False,
    }
    return render(request, 'admin_side/chat_interface.html', context)


@login_required
def chat_search_users_api(request):
    """
    API endpoint to search for company users (clients/marketers) to initiate a new chat.
    Returns ALL users matching the search query by name OR email.
    Users are uniquely identified by their email, so we don't filter out duplicates by name.
    """
    query = request.GET.get('q', '').strip()
    user_type = request.GET.get('type', 'all')  # 'client', 'marketer', or 'all'
    
    if len(query) < 2:
        return JsonResponse({'results': []})
    
    company_filter = {'company_profile': request.company} if hasattr(request, 'company') and request.company else {}
    
    results = []
    
    if user_type in ['client', 'all']:
        # Search by name OR email - show ALL matching users (email is unique identifier)
        clients = CustomUser.objects.filter(
            Q(full_name__icontains=query) | Q(email__icontains=query),
            role='client',
            **company_filter
        ).order_by('full_name', 'email')[:50]  # Increased limit to show all matches
        
        for c in clients:
            # Check if there's an existing conversation
            has_conversation = Message.objects.filter(
                Q(sender=c, recipient__role__in=SUPPORT_ROLES) |
                Q(sender__role__in=SUPPORT_ROLES, recipient=c)
            ).exists()
            
            results.append({
                'id': c.id,
                'name': c.full_name,
                'email': c.email,
                'type': 'client',
                'has_conversation': has_conversation,
                'avatar': c.profile_image.url if c.profile_image else None,
                'initials': c.full_name[0].upper() if c.full_name else '?',
            })
    
    if user_type in ['marketer', 'all']:
        # Search by name OR email - show ALL matching users (email is unique identifier)
        marketers = CustomUser.objects.filter(
            Q(full_name__icontains=query) | Q(email__icontains=query),
            role='marketer',
            **company_filter
        ).order_by('full_name', 'email')[:50]  # Increased limit to show all matches
        
        for m in marketers:
            has_conversation = Message.objects.filter(
                Q(sender=m, recipient__role__in=SUPPORT_ROLES) |
                Q(sender__role__in=SUPPORT_ROLES, recipient=m)
            ).exists()
            
            results.append({
                'id': m.id,
                'name': m.full_name,
                'email': m.email,
                'type': 'marketer',
                'has_conversation': has_conversation,
                'avatar': m.profile_image.url if m.profile_image else None,
                'initials': m.full_name[0].upper() if m.full_name else '?',
            })
    
    return JsonResponse({'results': results})


@login_required
def chat_load_conversation_api(request, user_id, user_type):
    """
    API endpoint to load a conversation asynchronously.
    Returns the chat messages HTML and user info without full page reload.
    """
    from django.urls import reverse
    
    if user_type == 'client':
        chat_user = get_object_or_404(CustomUser, id=user_id, role='client')
        chat_url = reverse('admin_chat', args=[user_id])
    elif user_type == 'marketer':
        chat_user = get_object_or_404(CustomUser, id=user_id, role='marketer')
        chat_url = reverse('admin_marketer_chat', args=[user_id])
    else:
        return JsonResponse({'success': False, 'error': 'Invalid user type'}, status=400)
    
    admin_user = request.user
    
    # Mark all unread messages from this user as read
    Message.objects.filter(
        sender=chat_user, 
        recipient__role__in=SUPPORT_ROLES, 
        is_read=False
    ).update(is_read=True, status='read')
    
    # Get conversation messages
    conversation = Message.objects.filter(
        Q(sender=chat_user, recipient__role__in=SUPPORT_ROLES) |
        Q(sender__role__in=SUPPORT_ROLES, recipient=chat_user)
    ).order_by('date_sent')
    
    # Render messages HTML
    messages_html = ""
    for msg in conversation:
        messages_html += render_to_string('admin_side/chat_message.html', {'msg': msg, 'request': request})
    
    # Get last message ID for polling
    last_msg_id = conversation.last().id if conversation.exists() else 0
    
    return JsonResponse({
        'success': True,
        'user': {
            'id': chat_user.id,
            'name': chat_user.full_name,
            'type': user_type,
            'avatar': chat_user.profile_image.url if chat_user.profile_image else None,
            'initials': chat_user.full_name[0].upper() if chat_user.full_name else '?',
        },
        'messages_html': messages_html,
        'last_msg_id': last_msg_id,
        'chat_url': chat_url,
        'is_marketer': user_type == 'marketer',
    })


# Admin Chat
@login_required
def admin_chat_view(request, client_id):
    # try:
    #     client = CustomUser.objects.get(id=client_id, role='client')
    # except CustomUser.DoesNotExist:
    #     messages.error(request, "The client you are trying to chat with does not exist.")
    #     return redirect('admin_client_chat_list')
        
    client = get_object_or_404(CustomUser, id=client_id, role='client')
    admin_user = request.user
    
    # Mark all unread messages from the client as read when ANY admin opens the chat.
    # This ensures unified dashboard - all admins see the same read/unread status
    Message.objects.filter(sender=client, recipient__role__in=SUPPORT_ROLES, is_read=False).update(is_read=True, status='read')
    
    # Build companies list for explorer: companies where client has allocations
    client_company_ids = (
        PlotAllocation.objects.filter(client_id=client.id)
        .values_list('estate__company', flat=True)
        .distinct()
    )
    companies_qs = Company.objects.filter(id__in=[c for c in client_company_ids if c is not None])
    companies = []
    for comp in companies_qs:
        alloc_count = PlotAllocation.objects.filter(client_id=client.id, estate__company=comp).count()
        companies.append({'company': comp, 'allocations': alloc_count})

    # Query messages: by default show full conversation, but allow optional company scoping
    # (polling endpoints may pass company_id to scope results)
    sel_company_id = None
    if request.method == 'GET':
        sel_company_id = request.GET.get('company_id')
    else:
        sel_company_id = request.POST.get('company_id')

    selected_company = None
    if sel_company_id:
        try:
            selected_company = Company.objects.get(id=int(sel_company_id))
        except Exception:
            selected_company = None

    # SECURITY: Ensure selected company is in client's allowed companies
    if selected_company and not companies_qs.filter(id=selected_company.id).exists():
        # SECURITY: Prevent access to unauthorized company - set to None for full conversation
        selected_company = None

    if selected_company:
        conversation = Message.objects.filter(
            (Q(sender=client, recipient__role__in=SUPPORT_ROLES) & Q(company=selected_company)) |
            (Q(sender__role__in=SUPPORT_ROLES, recipient=client) & Q(company=selected_company))
        ).order_by('date_sent')
    else:
        conversation = Message.objects.filter(
            Q(sender=client, recipient__role__in=SUPPORT_ROLES) |
            Q(sender__role__in=SUPPORT_ROLES, recipient=client)
        ).order_by('date_sent')
    
    # POLLING branch: if GET includes 'last_msg'
    if request.method == "GET" and 'last_msg' in request.GET:
        try:
            last_msg_id = int(request.GET.get('last_msg', 0))
        except ValueError:
            last_msg_id = 0
        new_messages = conversation.filter(id__gt=last_msg_id)

        messages_html = ""
        messages_list = []
        for msg in new_messages:
            messages_html += render_to_string('admin_side/chat_message.html', {'msg': msg, 'request': request})
            messages_list.append({'id': msg.id})
        
        # Also return updated statuses for all messages
        updated_statuses = []
        for m in conversation:
            updated_statuses.append({'id': m.id, 'status': m.status})
        
        return JsonResponse({
            'messages': messages_list,
            'messages_html': messages_html,
            'updated_statuses': updated_statuses
        })
    
    # POST: Admin sends a new message
    if request.method == "POST":
        message_content = request.POST.get('message_content', '').strip()
        file_attachment = request.FILES.get('file')
        
        if not message_content and not file_attachment:
            return JsonResponse({'success': False, 'error': 'Please enter a message or attach a file.'})
        
        # Get company - required for encryption
        company = None
        company_id = request.POST.get('company_id') or request.GET.get('company_id')
        if company_id:
            try:
                company = Company.objects.get(id=int(company_id))
            except Exception:
                return JsonResponse({'success': False, 'error': 'Invalid company_id'}, status=400)

            # SECURITY: Ensure client is affiliated with this company
            if not companies_qs.filter(id=company.id).exists():
                return JsonResponse({'success': False, 'error': 'Client is not affiliated with this company'}, status=403)
        else:
            # Fallback: get first company the client is affiliated with (so they can see the message)
            first_company_id = companies_qs.values_list('id', flat=True).first()
            if first_company_id:
                company = Company.objects.filter(id=first_company_id).first()
            
            # If still no company, try admin's company profile
            if not company:
                company = admin_user.company_profile
            
            if not company:
                return JsonResponse({'success': False, 'error': 'Unable to determine company for message encryption.'}, status=400)

        # Admin sends message - recipient is the specific client
        new_message = Message.objects.create(
            sender=admin_user,
            recipient=client,
            message_type="enquiry",
            content=message_content,
            file=file_attachment,
            status='sent',
            company=company,
        )
        message_html = render_to_string('admin_side/chat_message.html', {'msg': new_message, 'request': request})
        return JsonResponse({'success': True, 'message_html': message_html})
    
    # Get clients and marketers for sidebar chat list
    company_filter = {'company_profile': request.company} if hasattr(request, 'company') and request.company else {}
    sidebar_clients = CustomUser.objects.filter(
        role='client',
        sent_messages__isnull=False,
        **company_filter
    ).distinct().annotate(
        last_message=Max('sent_messages__date_sent')
    ).order_by('-last_message')[:20]
    
    for c in sidebar_clients:
        c.unread_count = Message.objects.filter(
            sender=c, recipient__role__in=SUPPORT_ROLES, is_read=False
        ).count()
        last_msg = Message.objects.filter(
            Q(sender=c, recipient__role__in=SUPPORT_ROLES) | Q(sender__role__in=SUPPORT_ROLES, recipient=c)
        ).order_by('-date_sent').first()
        c.last_message_text = last_msg.content[:50] if last_msg and last_msg.content else ''

    sidebar_marketers = CustomUser.objects.filter(
        role='marketer',
        sent_messages__isnull=False,
        **company_filter
    ).distinct().annotate(
        last_message=Max('sent_messages__date_sent')
    ).order_by('-last_message')[:20]
    
    for m in sidebar_marketers:
        m.unread_count = Message.objects.filter(
            sender=m, recipient__role__in=SUPPORT_ROLES, is_read=False
        ).count()
        last_msg = Message.objects.filter(
            Q(sender=m, recipient__role__in=SUPPORT_ROLES) | Q(sender__role__in=SUPPORT_ROLES, recipient=m)
        ).order_by('-date_sent').first()
        m.last_message_text = last_msg.content[:50] if last_msg and last_msg.content else ''

    context = {
        'client': client,
        'messages': conversation,
        'companies': companies,
        'selected_company': selected_company,
        'clients': sidebar_clients,
        'marketers': sidebar_marketers,
        'is_marketer': False,
    }
    return render(request, 'admin_side/chat_interface.html', context)

@login_required
def marketer_chat_view(request):
    if getattr(request.user, 'role', None) != 'marketer':
        return redirect('login')

    admin_user = CustomUser.objects.filter(role__in=SUPPORT_ROLES).first()
    if not admin_user:
        return JsonResponse({'success': False, 'error': 'No admin available to receive messages.'}, status=400)

    initial_unread = Message.objects.filter(
        sender__role__in=SUPPORT_ROLES,
        recipient=request.user,
        is_read=False
    ).count()

    admin_messages_qs = Message.objects.filter(
        sender__role__in=SUPPORT_ROLES,
        recipient=request.user,
        is_read=False
    )
    admin_messages_qs.update(is_read=True, status='read')

    # Build companies list for marketer explorer (companies where marketer has transactions)
    user = request.user
    company_ids = (
        Transaction.objects.filter(marketer=user)
        .values_list('company', flat=True)
        .distinct()
    )
    companies_qs = Company.objects.filter(id__in=[c for c in company_ids if c is not None])
    companies = []
    for comp in companies_qs:
        txn_count = Transaction.objects.filter(marketer=user, company=comp).count()
        companies.append({'company': comp, 'transactions': txn_count})

    # Determine selected company for conversation
    sel_company_id = None
    if request.method == 'GET':
        sel_company_id = request.GET.get('company_id')
    else:
        sel_company_id = request.POST.get('company_id')

    if not sel_company_id and companies_qs.exists():
        sel_company_id = companies_qs.first().id

    selected_company = None
    if sel_company_id:
        try:
            selected_company = Company.objects.get(id=int(sel_company_id))
        except Exception:
            selected_company = None

    # Always show ALL messages between marketer and admins (regardless of company)
    # This ensures replies from admins are always visible
    conversation = Message.objects.filter(
        Q(sender=request.user, recipient__role__in=SUPPORT_ROLES) |
        Q(sender__role__in=SUPPORT_ROLES, recipient=request.user)
    ).order_by('date_sent')

    if request.method == "GET" and 'last_msg' in request.GET:
        try:
            last_msg_id = int(request.GET.get('last_msg', 0))
        except (ValueError, TypeError):
            last_msg_id = 0

        new_messages = conversation.filter(id__gt=last_msg_id)
        messages_html = ""
        messages_list = []
        for msg in new_messages:
            messages_html += render_to_string('marketer_side/chat_message.html', {'msg': msg, 'request': request})
            messages_list.append({'id': msg.id})

        updated_statuses = [{'id': msg.id, 'status': msg.status} for msg in conversation]

        return JsonResponse({
            'messages': messages_list,
            'messages_html': messages_html,
            'updated_statuses': updated_statuses,
        })
    # Build companies list for marketer explorer (companies where marketer has transactions)
    user = request.user
    company_ids = (
        Transaction.objects.filter(marketer=user)
        .values_list('company', flat=True)
        .distinct()
    )
    companies_qs = Company.objects.filter(id__in=[c for c in company_ids if c is not None])
    companies = []
    for comp in companies_qs:
        txn_count = Transaction.objects.filter(marketer=user, company=comp).count()
        companies.append({'company': comp, 'transactions': txn_count})

    if request.method == "POST":
        message_content = request.POST.get('message_content', '').strip()
        file_attachment = request.FILES.get('file')
        company_id = request.POST.get('company_id')
        reply_to_id = request.POST.get('reply_to')
        reply_to = None
        if reply_to_id:
            try:
                reply_to = Message.objects.get(id=reply_to_id)
            except Message.DoesNotExist:
                reply_to = None

        if not message_content and not file_attachment:
            return JsonResponse({'success': False, 'error': 'Please enter a message or attach a file.'}, status=400)

        # SECURITY: Prevent sending messages if marketer is not associated with any company
        if not companies_qs.exists():
            return JsonResponse({'success': False, 'error': 'You are not associated with any company. Chat is disabled until you have an affiliation.'}, status=400)

        # Validate company_id and ensure marketer is affiliated
        company = None
        if not company_id:
            return JsonResponse({'success': False, 'error': 'Missing company_id'}, status=400)
        try:
            company = Company.objects.get(id=int(company_id))
        except Exception:
            return JsonResponse({'success': False, 'error': 'Invalid company_id'}, status=400)

        if not companies_qs.filter(id=company.id).exists():
            return JsonResponse({'success': False, 'error': 'You are not affiliated with this company'}, status=403)

        new_message = Message.objects.create(
            sender=request.user,
            recipient=admin_user,
            message_type="enquiry",
            content=message_content,
            file=file_attachment,
            reply_to=reply_to,
            status='sent',
            company=company,
        )

        message_html = render_to_string('marketer_side/chat_message.html', {
            'msg': new_message,
            'request': request,
        })
        return JsonResponse({'success': True, 'message_html': message_html})

    context = {
        'messages': conversation,
        'unread_chat_count': initial_unread,
        'global_message_count': initial_unread,
        'companies': companies,
        'selected_company': selected_company,
    }
    # Include companies and selected company for marketer template too
    context.update({'companies': companies, 'selected_company': selected_company})
    return render(request, 'marketer_side/chat_interface.html', context)


@login_required
@require_POST
def delete_message(request, message_id):
    try:
        msg = Message.objects.get(id=message_id)
    except Message.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'Message not found'}, status=404)

    if msg.sender != request.user and msg.recipient != request.user:
        return JsonResponse({'success': False, 'error': 'You do not have permission to delete this message'}, status=403)

    msg.delete()
    return JsonResponse({'success': True})


@login_required
def chat_unread_count(request):
    user = request.user

    data = {
        'total_unread': 0,
        'global_message_count': 0,
        'admin_unread_clients': [],
        'client_count': 0,
        'admin_unread_marketers': [],
        'marketer_count': 0,
    }

    if getattr(user, 'role', None) in SUPPORT_ROLES:
        total = Message.objects.filter(
            recipient__role__in=SUPPORT_ROLES,
            is_read=False
        ).count()
        data['total_unread'] = total

        from django.db.models import Max, Count, Subquery, OuterRef

        latest_content_sq = Subquery(
            Message.objects.filter(
                sender=OuterRef('pk'),
                recipient__role__in=SUPPORT_ROLES
            ).order_by('-date_sent').values('content')[:1]
        )
        latest_file_sq = Subquery(
            Message.objects.filter(
                sender=OuterRef('pk'),
                recipient__role__in=SUPPORT_ROLES
            ).order_by('-date_sent').values('file')[:1]
        )

        unread_clients_qs = (CustomUser.objects
            .filter(
                role='client',
                sent_messages__recipient__role__in=SUPPORT_ROLES,
                sent_messages__is_read=False
            )
            .annotate(
                last_message=Max('sent_messages__date_sent'),
                unread_count=Count('sent_messages'),
                last_content=latest_content_sq,
                last_file=latest_file_sq,
            )
            .distinct()
            .order_by('-last_message')
        )

        admin_unread_clients = []
        for c in unread_clients_qs[:5]:
            profile_url = None
            try:
                if getattr(c, 'profile_image', None):
                    profile_url = c.profile_image.url
            except Exception:
                profile_url = None

            last_msg = (Message.objects
                        .filter(sender=c, recipient__role__in=SUPPORT_ROLES)
                        .order_by('-date_sent')
                        .first())
            last_iso = last_msg.date_sent.isoformat() if last_msg else None
            last_file_name = None
            if last_msg and getattr(last_msg, 'file', None):
                try:
                    last_file_name = last_msg.file.name
                except Exception:
                    last_file_name = None

            admin_unread_clients.append({
                'id': c.id,
                'full_name': getattr(c, 'full_name', 'Client'),
                'profile_image': profile_url,
                'unread_count': getattr(c, 'unread_count', 0),
                'last_content': getattr(c, 'last_content', '') or '',
                'last_file': last_file_name,
                'last_message_iso': last_iso,
            })

        data['client_count'] = unread_clients_qs.count()
        data['admin_unread_clients'] = admin_unread_clients

        unread_marketers_qs = (CustomUser.objects
            .filter(
                role='marketer',
                sent_messages__recipient__role__in=SUPPORT_ROLES,
                sent_messages__is_read=False
            )
            .annotate(
                last_message=Max('sent_messages__date_sent'),
                unread_count=Count('sent_messages'),
                last_content=Subquery(
                    Message.objects.filter(
                        sender=OuterRef('pk'),
                        recipient__role__in=SUPPORT_ROLES
                    ).order_by('-date_sent').values('content')[:1]
                ),
                last_file=Subquery(
                    Message.objects.filter(
                        sender=OuterRef('pk'),
                        recipient__role__in=SUPPORT_ROLES
                    ).order_by('-date_sent').values('file')[:1]
                ),
            )
            .distinct()
            .order_by('-last_message')
        )

        admin_unread_marketers = []
        for m in unread_marketers_qs[:5]:
            profile_url = None
            try:
                if getattr(m, 'profile_image', None):
                    profile_url = m.profile_image.url
            except Exception:
                profile_url = None

            last_msg = (Message.objects
                        .filter(sender=m, recipient__role__in=SUPPORT_ROLES)
                        .order_by('-date_sent')
                        .first())
            last_iso = last_msg.date_sent.isoformat() if last_msg else None
            last_file_name = None
            if last_msg and getattr(last_msg, 'file', None):
                try:
                    last_file_name = last_msg.file.name
                except Exception:
                    last_file_name = None

            admin_unread_marketers.append({
                'id': m.id,
                'full_name': getattr(m, 'full_name', 'Marketer'),
                'profile_image': profile_url,
                'unread_count': getattr(m, 'unread_count', 0),
                'last_content': getattr(m, 'last_content', '') or '',
                'last_file': last_file_name,
                'last_message_iso': last_iso,
            })

        data['marketer_count'] = unread_marketers_qs.count()
        data['admin_unread_marketers'] = admin_unread_marketers
    else:
        data['global_message_count'] = Message.objects.filter(
            sender__role__in=SUPPORT_ROLES,
            recipient=user,
            is_read=False
        ).count()

    return JsonResponse(data)

@login_required
def message_detail(request, message_id):
    message = get_object_or_404(Message, id=message_id)
    return render(request, 'message_detail.html', {'message': message})


def verify_company_profile_access(request, company):
    """
    Backend security helper to verify user has proper access to company profile.
    Returns tuple: (has_access: bool, error_message: str or None)
    
    Security checks:
    1. User must be admin role
    2. User must belong to the company
    3. User must either have full_control OR valid session verification
    """
    user = request.user
    
    # Check 1: Must be admin
    if getattr(user, 'role', None) != 'admin':
        return False, 'Access denied: Admin role required'
    
    # Check 2: Must belong to this company
    if user.company_profile != company:
        return False, 'Access denied: Company mismatch'
    
    # Check 3: Master admin always has access
    master_admin = AdminUser.objects.filter(company_profile=company).order_by('date_joined').first()
    is_master = master_admin and user.id == master_admin.id
    if is_master:
        return True, None
    
    # Check 4: User with has_full_control=True has access
    if getattr(user, 'has_full_control', False):
        return True, None
    
    # Check 5: Session-based verification (for users without full control)
    if request.session.get('company_profile_verified'):
        verified_at = request.session.get('company_profile_verified_at')
        verified_company = request.session.get('company_profile_company_id')
        
        if verified_at and verified_company == company.id:
            try:
                expiry_time = timezone.datetime.fromisoformat(verified_at)
                if timezone.now() < expiry_time:
                    return True, None
                else:
                    # Session expired
                    request.session.pop('company_profile_verified', None)
                    request.session.pop('company_profile_verified_at', None)
                    request.session.pop('company_profile_company_id', None)
            except (ValueError, TypeError):
                pass
    
    return False, 'Access denied: Password verification required'


@login_required
def company_profile_view(request):
    """Admin-only page showing the current company's profile and key real estate stats."""
    user = request.user
    if getattr(user, 'role', None) != 'admin':
        return redirect('home')

    try:
        company = get_user_company_from_slug(request)
    except HttpResponseForbidden as e:
        return e
    
    if not company:
        company = getattr(user, 'company_profile', None)
    
    if not company:
        return redirect('home')

    # SECURITY CHECK: Determine if user needs password verification
    from django.utils import timezone
    master_admin = AdminUser.objects.filter(company_profile=company).order_by('date_joined').first()
    is_master = bool(master_admin and user.id == master_admin.id)
    has_full_control = is_master or bool(getattr(user, 'has_full_control', False))
    
    # If user doesn't have full control, check session verification
    if not has_full_control:
        session_verified = False
        if request.session.get('company_profile_verified'):
            verified_at = request.session.get('company_profile_verified_at')
            verified_company = request.session.get('company_profile_company_id')
            if verified_at and verified_company == company.id:
                try:
                    expiry_time = timezone.datetime.fromisoformat(verified_at)
                    if timezone.now() < expiry_time:
                        session_verified = True
                    else:
                        # Session expired, clear it
                        request.session.pop('company_profile_verified', None)
                        request.session.pop('company_profile_verified_at', None)
                        request.session.pop('company_profile_company_id', None)
                except (ValueError, TypeError):
                    pass
        
        # If not verified, redirect to verification page
        if not session_verified:
            return redirect('company-profile-verify')

    # SECURITY: Filter all metrics by company
    # Basic aggregates for dashboard-style overview
    # Include users from company_profile AND from affiliation models (MarketerAffiliation, ClientMarketerAssignment)
    # IMPORTANT: Exclude duplicates - users with company_profile should not be double-counted if also in affiliation tables
    from estateApp.models import MarketerAffiliation, ClientMarketerAssignment
    
    # Get primary client IDs
    primary_client_ids = set(CustomUser.objects.filter(role='client', company_profile=company).values_list('id', flat=True))
    # Get affiliated client IDs that are NOT already primary
    affiliation_client_ids = set(
        ClientMarketerAssignment.objects.filter(company=company)
        .values_list('client_id', flat=True)
        .distinct()
    ) - primary_client_ids
    total_clients = len(primary_client_ids) + len(affiliation_client_ids)
    
    # Get primary marketer IDs
    primary_marketer_ids = set(CustomUser.objects.filter(role='marketer', company_profile=company).values_list('id', flat=True))
    # Get affiliated marketer IDs that are NOT already primary
    affiliation_marketer_ids = set(
        MarketerAffiliation.objects.filter(company=company)
        .values_list('marketer_id', flat=True)
        .distinct()
    ) - primary_marketer_ids
    total_marketers = len(primary_marketer_ids) + len(affiliation_marketer_ids)

    # Estates and allocations (calculated from TRANSACTIONS as per business logic)
    total_estates = Estate.objects.filter(company=company).count() if 'Estate' in globals() else 0
    
    # ALLOCATION IS CALCULATED FROM THE NUMBER OF TRANSACTIONS MADE
    # Full payments = transactions linked to full payment allocations
    # Part payments = transactions linked to part payment allocations
    from estateApp.models import Transaction
    total_full_allocations = Transaction.objects.filter(
        company=company, 
        allocation__payment_type='full'
    ).count()
    total_part_allocations = Transaction.objects.filter(
        company=company, 
        allocation__payment_type='part'
    ).count()
    total_allocations = total_full_allocations + total_part_allocations

    # Registered users - SECURITY: Filter by company with pagination (50 per page)
    # This includes ALL users: Clients, Marketers, Admins, Support, and any other role
    # MultiTable inheritance requires querying parent + children separately, then combining
    from django.core.paginator import Paginator
    from django.utils import timezone
    from datetime import timedelta
    import json
    from django.db.models import Q
    from itertools import chain
    
    # Query includes ALL user roles by getting both CustomUser and all subclasses
    # Get all CustomUser instances (but NOT subclass instances)
    base_users = CustomUser.objects.filter(
        company_profile=company
    ).exclude(
        Q(adminuser__isnull=False) | Q(supportuser__isnull=False)
    ).order_by('-date_joined')
    
    # Get all AdminUser instances
    admin_users_in_table = AdminUser.objects.filter(
        company_profile=company
    ).order_by('-date_joined')
    
    # Get all SupportUser instances
    support_users_in_table = SupportUser.objects.filter(
        company_profile=company
    ).order_by('-date_joined')
    
    # Get affiliated marketers (added via MarketerAffiliation)
    affiliated_marketer_ids = set(
        MarketerAffiliation.objects.filter(
            company=company
        ).values_list('marketer_id', flat=True)
    )
    
    # Get affiliated clients (added via ClientMarketerAssignment)
    affiliated_client_ids = set(
        ClientMarketerAssignment.objects.filter(
            company=company
        ).values_list('client_id', flat=True)
    )
    
    # Collect already-included user IDs from main queries
    included_ids = set()
    included_ids.update(base_users.values_list('id', flat=True))
    included_ids.update(admin_users_in_table.values_list('id', flat=True))
    included_ids.update(support_users_in_table.values_list('id', flat=True))
    
    # Add affiliated users NOT already included
    additional_user_ids = (affiliated_marketer_ids | affiliated_client_ids) - included_ids
    additional_users = CustomUser.objects.filter(id__in=additional_user_ids).order_by('-date_joined') if additional_user_ids else CustomUser.objects.none()
    
    # Combine all users
    all_users_list = list(chain(base_users, admin_users_in_table, support_users_in_table, additional_users))
    # Sort by date_joined descending
    all_users_list.sort(key=lambda x: x.date_joined, reverse=True)
    
    registered_users_total_count = len(all_users_list)
    
    # For pagination, we need to work with the list
    from django.core.paginator import Paginator as DjangoPaginator
    paginator_obj = DjangoPaginator(all_users_list, 50)  # 50 users per page
    page_number = request.GET.get('users_page', 1)
    try:
        registered_users_page = paginator_obj.page(page_number)
    except Exception:
        registered_users_page = paginator_obj.page(1)
    
    registered_users = list(registered_users_page.object_list)
    
    # For analytics, use ALL users including affiliated ones
    # This matches what's displayed in all_users_list
    registered_users_qs = CustomUser.objects.filter(company_profile=company)
    admin_qs = AdminUser.objects.filter(company_profile=company)
    support_qs = SupportUser.objects.filter(company_profile=company)
    
    # Get affiliated users (not yet in company_profile)
    affiliated_marketer_ids = set(
        MarketerAffiliation.objects.filter(
            company=company
        ).values_list('marketer_id', flat=True)
    )
    affiliated_client_ids = set(
        ClientMarketerAssignment.objects.filter(
            company=company
        ).values_list('client_id', flat=True)
    )
    
    # Collect already-included user IDs
    included_ids = set()
    included_ids.update(registered_users_qs.values_list('id', flat=True))
    included_ids.update(admin_qs.values_list('id', flat=True))
    included_ids.update(support_qs.values_list('id', flat=True))
    
    # Get additional affiliated users NOT already included
    additional_user_ids = (affiliated_marketer_ids | affiliated_client_ids) - included_ids
    additional_users = CustomUser.objects.filter(id__in=additional_user_ids) if additional_user_ids else CustomUser.objects.none()
    
    # Combine for analytics (all_qs = all users we care about)
    all_qs = list(chain(registered_users_qs, admin_qs, support_qs, additional_users))
    
    # USER ENGAGEMENT ANALYTICS - Calculate before pagination
    now = timezone.now()
    last_7_days_cutoff = now - timedelta(days=7)
    last_30_days_cutoff = now - timedelta(days=30)
    last_90_days_cutoff = now - timedelta(days=90)
    
    # Engagement metrics for all users (used for analytics)
    # Compare last_login datetime with cutoff datetime
    active_users_30d = sum(1 for u in all_qs if u.last_login and u.last_login >= last_30_days_cutoff and u.is_active)
    recently_active_7d = sum(1 for u in all_qs if u.last_login and u.last_login >= last_7_days_cutoff and u.is_active)
    at_risk_users = sum(1 for u in all_qs if (not u.last_login or u.last_login < last_30_days_cutoff) and u.is_active)
    dormant_users = sum(1 for u in all_qs if not u.last_login and u.is_active)
    
    # Engagement score (0-100) = (active_users / total_users) * 100
    engagement_score = int((active_users_30d / max(registered_users_total_count, 1)) * 100)
    
    # Calculate user statuses for table display
    user_statuses = {}
    for user_obj in all_qs:
        if not user_obj.last_login:
            user_statuses[user_obj.id] = {'status': 'Never Logged In', 'days_since': None, 'engagement': 'Dormant', 'at_risk': False}
        else:
            days_since = (now - user_obj.last_login).days
            if days_since <= 7:
                status_str = 'Active'
                engagement_str = 'Strong'
                at_risk = False
            elif days_since <= 30:
                status_str = 'Idle'
                engagement_str = 'Moderate'
                at_risk = False
            elif days_since <= 60:
                status_str = 'Low Activity'
                engagement_str = 'Weak'
                at_risk = True
            else:
                status_str = 'At-Risk'
                engagement_str = 'Critical'
                at_risk = True
            
            user_statuses[user_obj.id] = {
                'status': status_str,
                'days_since': days_since,
                'engagement': engagement_str,
                'at_risk': at_risk
            }
    
    # Login frequency by day of week (last 30 days) - for heatmap
    from django.db.models.functions import ExtractWeekDay
    from django.db.models import Count
    login_by_dow = {}
    dow_names = ['Mon', 'Tue', 'Wed', 'Thu', 'Fri', 'Sat', 'Sun']
    
    try:
        from estateApp.audit_logging import AuditLog
        login_events = AuditLog.objects.filter(
            company=company,
            action='LOGIN',
            created_at__gte=last_30_days_cutoff
        ).annotate(
            dow=ExtractWeekDay('created_at')
        ).values('dow').annotate(count=Count('id')).order_by('dow')
        
        for item in login_events:
            dow = item['dow']
            # ExtractWeekDay returns 1-7 (Sunday=1, Monday=2, ..., Saturday=7)
            # Convert to 0-6 for our dow_names array
            dow_index = (dow - 2) % 7  # Shift so Monday=0
            if 0 <= dow_index < 7:
                login_by_dow[dow_names[dow_index]] = item['count']
        
        # Fill in missing days with 0
        for day_name in dow_names:
            if day_name not in login_by_dow:
                login_by_dow[day_name] = 0
    except Exception:
        login_by_dow = {day: 0 for day in dow_names}
    
    # User status distribution
    idle_users = sum(1 for u in all_qs if u.last_login and u.last_login >= last_30_days_cutoff and u.last_login < last_7_days_cutoff and u.is_active)
    status_distribution = {
        'Active': recently_active_7d,
        'Idle': idle_users,
        'At-Risk': at_risk_users,
        'Dormant': dormant_users
    }
    
    # GEOGRAPHIC ANALYTICS - Location-based insights
    # Get top login locations from all users
    location_stats = {}
    top_locations = []
    
    try:
        # Get users with login locations (last 30 days)
        location_counts = {}
        for u in all_qs:
            if u.last_login and u.last_login >= last_30_days_cutoff and u.last_login_location and u.last_login_location.strip():
                location_counts[u.last_login_location] = location_counts.get(u.last_login_location, 0) + 1
        
        # Sort by count and get top 10
        sorted_locations = sorted(location_counts.items(), key=lambda x: x[1], reverse=True)[:10]
        users_with_location = [{'last_login_location': loc, 'count': cnt} for loc, cnt in sorted_locations]
        
        for item in users_with_location:
            if isinstance(item, dict):
                location = item['last_login_location']
                count = item['count']
            else:
                location = item[0]
                count = item[1]
            location_stats[location] = count
            top_locations.append({
                'location': location,
                'count': count,
                'percentage': int((count / max(active_users_30d, 1)) * 100)
            })
    except Exception:
        location_stats = {}
        top_locations = []
    
    # Device/Platform Analytics
    device_stats = {'web': 0, 'ios': 0, 'android': 0, 'other': 0}
    
    try:
        from estateApp.models import UserDeviceToken
        device_usage = UserDeviceToken.objects.filter(
            user__company_profile=company,
            created_at__gte=last_30_days_cutoff
        ).values('platform').annotate(
            count=Count('id')
        ).order_by('-count')
        
        for item in device_usage:
            platform = item['platform'].lower()
            if platform == 'web':
                device_stats['web'] = item['count']
            elif platform == 'ios':
                device_stats['ios'] = item['count']
            elif platform == 'android':
                device_stats['android'] = item['count']
            else:
                device_stats['other'] += item['count']
    except Exception:
        device_stats = {'web': 0, 'ios': 0, 'android': 0, 'other': 0}
    
    # Prepare chart data for JSON
    chart_data = {
        'login_frequency': json.dumps(list(login_by_dow.values())),
        'login_days': json.dumps(list(login_by_dow.keys())),
        'status_labels': json.dumps(list(status_distribution.keys())),
        'status_values': json.dumps(list(status_distribution.values())),
        'location_labels': json.dumps([loc['location'] for loc in top_locations]),
        'location_values': json.dumps([loc['count'] for loc in top_locations]),
        'device_labels': json.dumps(list(device_stats.keys())),
        'device_values': json.dumps(list(device_stats.values())),
    }
    
    # Attach status info to each user for display (already paginated in registered_users)
    for user_obj in registered_users:
        if user_obj.id in user_statuses:
            user_obj.status_info = user_statuses[user_obj.id]

    # Legacy compatibility
    active_users_count = active_users_30d
    inactive_users_count = registered_users_total_count - active_users_count

    # Admin and Support users
    # Determine master admin first
    master_admin = AdminUser.objects.filter(company_profile=company).order_by('date_joined').first()
    master_admin_id = master_admin.id if master_admin else None
    
    # Determine if current user is master admin or has full control
    is_current_user_master = bool(master_admin and user.id == master_admin.id)
    has_full_control_attr = bool(getattr(user, 'has_full_control', False))
    user_has_full_control = is_current_user_master or has_full_control_attr
    
    # Order admin_users with master admin first, then by date_joined
    from django.db.models import Case, When, Value, IntegerField
    admin_users = AdminUser.objects.filter(company_profile=company).annotate(
        is_master=Case(
            When(id=master_admin_id, then=Value(1)),
            default=Value(0),
            output_field=IntegerField()
        )
    ).order_by('-is_master', '-date_joined')
    
    support_users = SupportUser.objects.filter(company_profile=company).order_by('-date_joined')

    # Calculate status for support users based on last_login (inactive if > 5 days)
    for user in support_users:
        if user.last_login:
            days_since_login = (timezone.now() - user.last_login).days
            user.status_display = "Not Active" if days_since_login > 5 else ("Active" if user.is_active else "Muted")
        else:
            # No login record, consider inactive
            user.status_display = "Not Active"

    # AdminSupport tables if available
    try:
        from adminSupport.models import StaffRoster, StaffMember
        staff_roster = StaffRoster.objects.select_related('user').all()[:50]
        # SECURITY: Filter staff by company
        staff_members = StaffMember.objects.filter(company=company)[:50]
        staff_roster_count = StaffRoster.objects.filter(company=company).count()
        staff_members_count = StaffMember.objects.filter(company=company).count()
    except Exception:
        staff_roster = []
        staff_members = []
        staff_roster_count = 0
        staff_members_count = 0

    # App metrics (downloads)
    app_metrics = None
    total_downloads = 0
    if company:
        app_metrics = getattr(company, 'app_metrics', None)
        if not app_metrics:
            try:
                app_metrics = AppMetrics.objects.create(company=company)
            except Exception:
                app_metrics = None
        if app_metrics:
            total_downloads = app_metrics.total_downloads

    # Subscription - Built from Company model subscription fields
    subscription = None
    if company:
        try:
            from datetime import datetime, timedelta
            from estateApp.models import SubscriptionPlan
            
            # Get the subscription plan for pricing
            subscription_plan = None
            try:
                subscription_plan = SubscriptionPlan.objects.get(tier=company.subscription_tier)
            except SubscriptionPlan.DoesNotExist:
                pass
            
            # Calculate days remaining
            days_remaining = 0
            if company.subscription_ends_at:
                now = timezone.now()
                if company.subscription_ends_at > now:
                    days_remaining = (company.subscription_ends_at - now).days
            elif company.trial_ends_at:
                now = timezone.now()
                if company.trial_ends_at > now:
                    days_remaining = (company.trial_ends_at - now).days
            
            # Check if in grace period
            is_in_grace_period = False
            if company.grace_period_ends_at:
                now = timezone.now()
                is_in_grace_period = now < company.grace_period_ends_at and days_remaining <= 0
            
            # Warning level for expiring soon
            warning_level = None
            if days_remaining <= 7 and days_remaining > 0:
                warning_level = 'critical'
            elif days_remaining <= 14 and days_remaining > 7:
                warning_level = 'warning'
            elif days_remaining <= 30 and days_remaining > 14:
                warning_level = 'notice'
            
            # Get limits from features
            features = subscription_plan.features if subscription_plan else {}
            max_estates = features.get('estate_properties', company.max_plots) if isinstance(features.get('estate_properties'), int) else company.max_plots
            max_allocations = features.get('allocations', 30) if isinstance(features.get('allocations'), int) else (999999 if features.get('allocations') == 'unlimited' else 30)
            max_clients = features.get('clients', 30) if isinstance(features.get('clients'), int) else (999999 if features.get('clients') == 'unlimited' else 30)
            max_affiliates = features.get('affiliates', 20) if isinstance(features.get('affiliates'), int) else (999999 if features.get('affiliates') == 'unlimited' else 20)
            
            # Build subscription dict
            subscription = {
                'subscription_plan': subscription_plan,
                'plan_name': subscription_plan.name if subscription_plan else company.get_subscription_tier_display(),
                'tier': company.subscription_tier,
                'status': company.subscription_status,
                'amount': subscription_plan.monthly_price if subscription_plan else 0,
                'monthly_price': subscription_plan.monthly_price if subscription_plan else 0,
                'annual_price': subscription_plan.annual_price if subscription_plan else 0,
                'subscription_starts_at': company.subscription_started_at,
                'subscription_ends_at': company.subscription_ends_at or company.trial_ends_at,
                'days_remaining': days_remaining,
                'is_in_grace_period': is_in_grace_period,
                'warning_level': warning_level,
                'get_current_status': company.get_subscription_status_display(),
                'is_trial': company.subscription_status == 'trial',
                'is_active': company.subscription_status == 'active',
                'max_plots': company.max_plots,
                'max_agents': company.max_agents,
                'current_plots_count': company.current_plots_count,
                'current_agents_count': company.current_agents_count,
                'features': features,
                # Limits from plan features
                'max_estates': max_estates,
                'max_allocations': max_allocations,
                'max_clients': max_clients,
                'max_affiliates': max_affiliates,
            }
        except Exception as e:
            subscription = None
    
    # Get all subscription plans for the upgrade/select plan modal
    subscription_plans = []
    subscription_plans_json = '[]'
    try:
        from estateApp.models import SubscriptionPlan
        import json
        plans_qs = SubscriptionPlan.objects.filter(is_active=True).order_by('monthly_price')
        subscription_plans = list(plans_qs)
        
        # Create JSON-serializable list for JavaScript
        plans_data = []
        for plan in plans_qs:
            plans_data.append({
                'id': plan.id,
                'name': plan.name,
                'tier': plan.tier,
                'description': plan.description or '',
                'monthly_price': float(plan.monthly_price) if plan.monthly_price else 0,
                'annual_price': float(plan.annual_price) if plan.annual_price else 0,
                'max_plots': plan.max_plots or 50,
                'max_agents': plan.max_agents or 1,
                'features': plan.features or {},
                'is_popular': plan.tier == 'professional',
            })
        subscription_plans_json = json.dumps(plans_data)
    except Exception:
        pass

    # Get billing history for recent transactions
    billing_history = []
    try:
        from estateApp.subscription_billing_models import SubscriptionBillingModel, BillingHistory
        billing = SubscriptionBillingModel.objects.filter(company=company).first()
        if billing:
            billing_history = list(BillingHistory.objects.filter(billing=billing).order_by('-billing_date')[:10])
    except Exception:
        pass
    
    # EMAIL ENGAGEMENT SYSTEM - Identify users needing re-engagement or onboarding
    users_needing_reengagement = [u for u in all_qs if u.last_login and u.last_login < last_30_days_cutoff and u.is_active]
    users_never_logged_in = [u for u in all_qs if not u.last_login and u.is_active]
    
    # Prepare email engagement data
    engagement_email_stats = {
        'reengagement_count': len(users_needing_reengagement),
        'onboarding_count': len(users_never_logged_in),
        'total_to_contact': len(users_needing_reengagement) + len(users_never_logged_in),
    }

    context = {
        'company': company,
        'total_clients': total_clients,
        'total_marketers': total_marketers,
        'total_estates': total_estates,
        'total_full_allocations': total_full_allocations,
        'total_part_allocations': total_part_allocations,
        'total_allocations': total_allocations,
        'registered_users': registered_users,
        'registered_users_page': registered_users_page,
        'registered_users_total_count': registered_users_total_count,
        'active_users_count': active_users_count,
        'inactive_users_count': inactive_users_count,
        'admin_users': admin_users,
        'support_users': support_users,
        'staff_roster': staff_roster,
        'staff_members': staff_members,
        'staff_roster_count': staff_roster_count,
        'staff_members_count': staff_members_count,
        'app_metrics': app_metrics,
        'total_downloads': total_downloads,
        'master_admin_id': master_admin_id,
        'is_master_admin': is_current_user_master,
        'user_has_full_control': user_has_full_control,
        'subscription': subscription,
        'subscription_plans': subscription_plans,
        'subscription_plans_json': subscription_plans_json,
        'billing_history': billing_history,
        # User engagement analytics
        'active_users_30d': active_users_30d,
        'recently_active_7d': recently_active_7d,
        'at_risk_users': at_risk_users,
        'dormant_users': dormant_users,
        'engagement_score': engagement_score,
        'status_distribution': status_distribution,
        'chart_data': chart_data,
        # Geographic and device analytics
        'top_locations': top_locations,
        'location_stats': location_stats,
        'device_stats': device_stats,
        # Email engagement system
        'users_needing_reengagement': users_needing_reengagement,
        'users_never_logged_in': users_never_logged_in,
        'engagement_email_stats': engagement_email_stats,
    }
    return render(request, 'admin_side/company_profile.html', context)


@login_required
@require_POST
def send_engagement_emails(request):
    """
    API endpoint to send re-engagement and onboarding emails to users
    Requires admin access to the company
    """
    from estateApp.engagement_emails import send_bulk_reengagement_emails, send_bulk_onboarding_emails
    from django.http import JsonResponse
    
    user = request.user
    if getattr(user, 'role', None) != 'admin':
        return JsonResponse({'ok': False, 'error': 'Only admins can send engagement emails'}, status=403)
    
    company = getattr(user, 'company_profile', None)
    if not company:
        return JsonResponse({'ok': False, 'error': 'No linked company'}, status=400)
    
    email_type = request.POST.get('email_type')  # 'reengagement' or 'onboarding'
    
    if email_type not in ['reengagement', 'onboarding']:
        return JsonResponse({'ok': False, 'error': 'Invalid email type'}, status=400)
    
    try:
        # Get users needing emails (same logic as company_profile_view)
        from django.utils import timezone
        from datetime import timedelta
        from itertools import chain
        
        now = timezone.now()
        last_30_days_cutoff = now - timedelta(days=30)
        
        # Get all users
        registered_users_qs = CustomUser.objects.filter(company_profile=company)
        admin_qs = AdminUser.objects.filter(company_profile=company)
        support_qs = SupportUser.objects.filter(company_profile=company)
        
        affiliated_marketer_ids = set(
            MarketerAffiliation.objects.filter(
                company=company
            ).values_list('marketer_id', flat=True)
        )
        affiliated_client_ids = set(
            ClientMarketerAssignment.objects.filter(
                company=company
            ).values_list('client_id', flat=True)
        )
        
        included_ids = set()
        included_ids.update(registered_users_qs.values_list('id', flat=True))
        included_ids.update(admin_qs.values_list('id', flat=True))
        included_ids.update(support_qs.values_list('id', flat=True))
        
        additional_user_ids = (affiliated_marketer_ids | affiliated_client_ids) - included_ids
        additional_users = CustomUser.objects.filter(id__in=additional_user_ids) if additional_user_ids else CustomUser.objects.none()
        
        all_users = list(chain(registered_users_qs, admin_qs, support_qs, additional_users))
        
        # Send appropriate emails
        if email_type == 'reengagement':
            users = [u for u in all_users if u.last_login and u.last_login < last_30_days_cutoff and u.is_active]
            stats = send_bulk_reengagement_emails(users, company)
        else:  # onboarding
            users = [u for u in all_users if not u.last_login and u.is_active]
            stats = send_bulk_onboarding_emails(users, company)
        
        return JsonResponse({
            'ok': True,
            'message': f"Successfully sent {stats['sent']} {email_type} emails",
            'stats': stats,
        })
        
    except Exception as e:
        logger.error(f"Error sending {email_type} emails: {str(e)}")
        return JsonResponse({
            'ok': False,
            'error': f'Failed to send emails: {str(e)}'
        }, status=500)


@login_required
def company_profile_verify_view(request):
    """
    Standalone verification page for company profile access.
    Non-full-control admins are redirected here to verify with master password.
    """
    from django.utils import timezone
    
    if request.user.role != 'admin':
        return redirect('login')
    
    company = request.user.company_profile
    if not company:
        messages.error(request, 'No company profile found')
        return redirect('login')
    
    # Check if user already has access
    master_admin = AdminUser.objects.filter(company_profile=company).order_by('date_joined').first()
    is_master = master_admin and request.user.id == master_admin.id
    has_full_control = is_master or getattr(request.user, 'has_full_control', False)
    
    # If user has full control, redirect directly to company profile
    if has_full_control:
        return redirect('company-profile')
    
    # Check if session is already verified
    if request.session.get('company_profile_verified'):
        verified_at = request.session.get('company_profile_verified_at')
        verified_company = request.session.get('company_profile_company_id')
        if verified_at and verified_company == company.id:
            try:
                expiry_time = timezone.datetime.fromisoformat(verified_at)
                if timezone.now() < expiry_time:
                    return redirect('company-profile')
            except (ValueError, TypeError):
                pass
    
    return render(request, 'admin_side/company_profile_verify.html', {
        'company': company,
        'master_admin_email': master_admin.email if master_admin else None,
    })


@login_required
@require_POST
def verify_master_password(request):
    """
    Verify master admin password and store verification in session.
    Security features:
    - Rate limiting: Max 5 attempts per 15 minutes
    - Session-based verification with 5-minute expiry
    - Logging of failed attempts
    """
    if request.user.role != 'admin':
        return JsonResponse({'ok': False, 'error': 'Unauthorized'}, status=403)

    # Rate limiting check
    attempt_key = f'password_attempts_{request.user.id}'
    lockout_key = f'password_lockout_{request.user.id}'
    
    # Check if user is locked out
    lockout_until = request.session.get(lockout_key)
    if lockout_until:
        lockout_time = timezone.datetime.fromisoformat(lockout_until)
        if timezone.now() < lockout_time:
            remaining = (lockout_time - timezone.now()).seconds // 60
            return JsonResponse({
                'ok': False, 
                'error': f'Too many failed attempts. Try again in {remaining + 1} minutes.',
                'locked': True
            }, status=429)
        else:
            # Lockout expired, reset
            request.session.pop(lockout_key, None)
            request.session.pop(attempt_key, None)

    password = request.POST.get('password')
    if not password:
        return JsonResponse({'ok': False, 'error': 'Password required'})

    company = request.user.company_profile
    if not company:
        return JsonResponse({'ok': False, 'error': 'No company'}, status=400)

    master_admin = AdminUser.objects.filter(company_profile=company).order_by('date_joined').first()
    if not master_admin:
        return JsonResponse({'ok': False, 'error': 'No master admin found'}, status=400)

    if master_admin.check_password(password):
        # Success - store verification in session with expiry timestamp
        verification_expiry = timezone.now() + timedelta(minutes=5)
        request.session['company_profile_verified'] = True
        request.session['company_profile_verified_at'] = verification_expiry.isoformat()
        request.session['company_profile_company_id'] = company.id
        
        # Reset failed attempts on success
        request.session.pop(attempt_key, None)
        request.session.pop(lockout_key, None)
        
        logger.info(f"Company profile access granted to user {request.user.id} for company {company.id}")
        return JsonResponse({'ok': True})
    else:
        # Failed attempt - increment counter
        attempts = request.session.get(attempt_key, 0) + 1
        request.session[attempt_key] = attempts
        
        if attempts >= 5:
            # Lock out for 15 minutes
            lockout_until = timezone.now() + timedelta(minutes=15)
            request.session[lockout_key] = lockout_until.isoformat()
            logger.warning(f"User {request.user.id} locked out after {attempts} failed password attempts")
            return JsonResponse({
                'ok': False, 
                'error': 'Too many failed attempts. Account locked for 15 minutes.',
                'locked': True
            }, status=429)
        
        remaining = 5 - attempts
        logger.warning(f"Failed password attempt by user {request.user.id} ({remaining} attempts remaining)")
        return JsonResponse({
            'ok': False, 
            'error': f'Invalid password. {remaining} attempts remaining.'
        })


@login_required
@require_POST
def toggle_full_control(request):
    """Toggle full control permission for an admin user. Only master admin can do this."""
    if request.user.role != 'admin':
        return JsonResponse({'ok': False, 'error': 'Unauthorized'}, status=403)

    company = request.user.company_profile
    if not company:
        return JsonResponse({'ok': False, 'error': 'No company'}, status=400)
    
    # SECURITY: Backend verification - prevent frontend bypass
    has_access, error = verify_company_profile_access(request, company)
    if not has_access:
        logger.warning(f"Unauthorized toggle_full_control attempt by user {request.user.id}: {error}")
        return JsonResponse({'ok': False, 'error': error}, status=403)

    # Check if current user is master admin
    master_admin = AdminUser.objects.filter(company_profile=company).order_by('date_joined').first()
    if not master_admin or request.user.id != master_admin.id:
        return JsonResponse({'ok': False, 'error': 'Only master admin can grant/revoke full control'}, status=403)

    user_id = request.POST.get('user_id')
    if not user_id:
        return JsonResponse({'ok': False, 'error': 'User ID required'}, status=400)

    try:
        target_user = AdminUser.objects.get(id=user_id, company_profile=company)
    except AdminUser.DoesNotExist:
        return JsonResponse({'ok': False, 'error': 'Admin user not found'}, status=404)

    # Cannot toggle for master admin
    if target_user.id == master_admin.id:
        return JsonResponse({'ok': False, 'error': 'Cannot change master admin permissions'}, status=400)

    # Toggle the permission
    target_user.has_full_control = not target_user.has_full_control
    target_user.save(update_fields=['has_full_control'])

    action = 'granted' if target_user.has_full_control else 'revoked'
    return JsonResponse({
        'ok': True, 
        'has_full_control': target_user.has_full_control,
        'message': f'Full control {action} for {target_user.full_name}'
    })


@login_required
def check_full_control(request):
    """
    Check if current user has full control (master admin or has_full_control=True).
    Also checks if session verification is valid (for non-full-control users).
    """
    if request.user.role != 'admin':
        return JsonResponse({'has_full_control': False, 'is_master': False, 'session_verified': False})

    company = request.user.company_profile
    if not company:
        return JsonResponse({'has_full_control': False, 'is_master': False, 'session_verified': False})

    master_admin = AdminUser.objects.filter(company_profile=company).order_by('date_joined').first()
    is_master = bool(master_admin and request.user.id == master_admin.id)
    has_full_control_attr = bool(getattr(request.user, 'has_full_control', False))
    has_full_control = is_master or has_full_control_attr
    
    # Check session verification status (for audit/security tracking)
    session_verified = False
    if request.session.get('company_profile_verified'):
        verified_at = request.session.get('company_profile_verified_at')
        verified_company = request.session.get('company_profile_company_id')
        if verified_at and verified_company == company.id:
            try:
                expiry_time = timezone.datetime.fromisoformat(verified_at)
                if timezone.now() < expiry_time:
                    session_verified = True
                else:
                    # Session expired, clear it
                    request.session.pop('company_profile_verified', None)
                    request.session.pop('company_profile_verified_at', None)
                    request.session.pop('company_profile_company_id', None)
            except (ValueError, TypeError):
                pass

    return JsonResponse({
        'has_full_control': bool(has_full_control),
        'is_master': bool(is_master),
        'session_verified': bool(session_verified)
    })


@login_required
def company_profile_update(request):
    """Handle modal-based updates to company details. Returns JSON on AJAX."""
    user = request.user
    if getattr(user, 'role', None) != 'admin':
        return JsonResponse({'ok': False, 'error': 'Forbidden'}, status=403)

    # Support dynamic slug parameter
    try:
        company = get_user_company_from_slug(request)
    except HttpResponseForbidden:
        return JsonResponse({'ok': False, 'error': 'Forbidden'}, status=403)
    
    if not company:
        company = getattr(user, 'company_profile', None)
    
    if not company:
        return JsonResponse({'ok': False, 'error': 'No linked company'}, status=400)
    
    # SECURITY: Backend verification - prevent frontend bypass
    has_access, error = verify_company_profile_access(request, company)
    if not has_access:
        logger.warning(f"Unauthorized company_profile_update attempt by user {user.id}: {error}")
        return JsonResponse({'ok': False, 'error': error}, status=403)

    if request.method != 'POST':
        return JsonResponse({'ok': False, 'error': 'Invalid method'}, status=405)

    # Copy POST to avoid validation failures on legacy CEO fields if modal uses separate CEO inputs
    post_data = request.POST.copy()
    # Remove legacy model fields so form will keep instance values when not provided
    for fld in ('ceo_name', 'ceo_dob'):
        if fld in post_data:
            post_data.pop(fld)

    form = CompanyForm(post_data, request.FILES, instance=company)
    if form.is_valid():
        # Save without committing so we can avoid overwriting legacy CEO fields
        # with None when the modal doesn't submit them.
        orig_ceo_name = company.ceo_name
        orig_ceo_dob = company.ceo_dob
        # If legacy dob missing but we have CompanyCeo records (backfilled), use primary CEO dob
        try:
            if not orig_ceo_dob and company.ceos.exists():
                primary = company.ceos.filter(is_primary=True).first() or company.ceos.first()
                if primary and getattr(primary, 'dob', None):
                    orig_ceo_dob = primary.dob
        except Exception:
            pass
        company_instance = form.save(commit=False)

        # If the form didn't provide legacy CEO values, preserve existing ones
        # (be permissive: check the instance values after save(commit=False))
        if not getattr(company_instance, 'ceo_name', None) and orig_ceo_name:
            company_instance.ceo_name = orig_ceo_name
        if not getattr(company_instance, 'ceo_dob', None) and orig_ceo_dob:
            company_instance.ceo_dob = orig_ceo_dob

        company_instance.save()
        company = company_instance

        # Persist CEO entries (primary + additional) if provided by the modal
        try:
            with transaction.atomic():
                # Backfill legacy CEO fields into CompanyCeo if no records exist
                if not company.ceos.exists() and (company.ceo_name or company.ceo_dob):
                    CompanyCeo.objects.create(
                        company=company,
                        name=company.ceo_name or '',
                        dob=company.ceo_dob,
                        is_primary=True
                    )

                # Read primary CEO inputs
                primary_name = (request.POST.get('primary_ceo_name') or '').strip()
                primary_dob_raw = (request.POST.get('primary_ceo_dob') or '').strip()
                primary_dob = parse_date(primary_dob_raw) if primary_dob_raw else None

                other_names = request.POST.getlist('other_ceo_name[]') or request.POST.getlist('other_ceo_name')
                other_dobs = request.POST.getlist('other_ceo_dob[]') or request.POST.getlist('other_ceo_dob')

                # Build new CEO set if any input was provided
                ceo_entries = []
                if primary_name:
                    ceo_entries.append({'name': primary_name, 'dob': primary_dob, 'is_primary': True})

                for idx, nm in enumerate(other_names or []):
                    name = (nm or '').strip()
                    if not name:
                        continue
                    dob_raw = other_dobs[idx] if idx < len(other_dobs) else ''
                    dob = parse_date(dob_raw) if dob_raw else None
                    ceo_entries.append({'name': name, 'dob': dob, 'is_primary': False})

                if ceo_entries:
                    # Replace existing CEO records for this company with submitted set
                    CompanyCeo.objects.filter(company=company).delete()
                    objs = [CompanyCeo(company=company, name=e['name'], dob=e['dob'], is_primary=e['is_primary']) for e in ceo_entries]
                    CompanyCeo.objects.bulk_create(objs)

        except Exception as e:
            logger.exception('Failed saving company CEOs')
            return JsonResponse({'ok': False, 'error': 'Failed saving CEOs: ' + str(e)}, status=500)

        # minimal payload for re-rendering snippet on client side if desired
        return JsonResponse({'ok': True, 'message': 'Company details updated successfully.'})
    else:
        return JsonResponse({'ok': False, 'errors': form.errors}, status=400)


@login_required
@require_POST
def delete_company_ceo(request, ceo_id: int):
    """Delete a single CompanyCeo row via AJAX.

    - Only company admins may delete.
    - CEO must belong to the request user's company.
    - Primary CEOs are protected from deletion.
    """
    user = request.user
    if getattr(user, 'role', None) != 'admin':
        return JsonResponse({'ok': False, 'error': 'Forbidden'}, status=403)

    # Support dynamic slug parameter
    try:
        company = get_user_company_from_slug(request)
    except HttpResponseForbidden:
        return JsonResponse({'ok': False, 'error': 'Forbidden'}, status=403)
    
    if not company:
        company = getattr(user, 'company_profile', None)
    
    if not company:
        return JsonResponse({'ok': False, 'error': 'No linked company'}, status=400)

    try:
        ceo = CompanyCeo.objects.get(id=ceo_id, company=company)
    except CompanyCeo.DoesNotExist:
        return JsonResponse({'ok': False, 'error': 'CEO not found'}, status=404)

    if ceo.is_primary:
        return JsonResponse({'ok': False, 'error': 'Cannot delete primary CEO'}, status=400)

    try:
        ceo.delete()
        return JsonResponse({'ok': True, 'message': 'CEO removed'})
    except Exception as e:
        logger.exception('Failed deleting CompanyCeo %s', ceo_id)
        return JsonResponse({'ok': False, 'error': 'Failed to delete CEO'}, status=500)


@login_required
@require_POST
def admin_toggle_mute(request, user_id: int):
    """Mute or unmute an admin or support account. Muted => is_active=False (cannot log in)."""
    if getattr(request.user, 'role', None) != 'admin':
        return JsonResponse({'ok': False, 'error': 'Forbidden'}, status=403)

    if request.user.id == user_id:
        return JsonResponse({'ok': False, 'error': "You can't mute your own account."}, status=400)

    try:
        # SECURITY: Verify user belongs to same company
        # Support dynamic slug parameter
        try:
            company = get_user_company_from_slug(request)
        except HttpResponseForbidden:
            return JsonResponse({'ok': False, 'error': 'Forbidden'}, status=403)
        
        if not company:
            company = request.user.company_profile
        
        # SECURITY: Backend verification - prevent frontend bypass
        has_access, error = verify_company_profile_access(request, company)
        if not has_access:
            logger.warning(f"Unauthorized admin_toggle_mute attempt by user {request.user.id}: {error}")
            return JsonResponse({'ok': False, 'error': error}, status=403)
        
        try:
            target = AdminUser.objects.get(id=user_id, company_profile=company)
        except AdminUser.DoesNotExist:
            try:
                target = SupportUser.objects.get(id=user_id, company_profile=company)
            except SupportUser.DoesNotExist:
                return JsonResponse({'ok': False, 'error': 'User not found'}, status=404)
    except Exception:
        return JsonResponse({'ok': False, 'error': 'User not found'}, status=404)

    if target.role not in ('admin', 'support'):
        return JsonResponse({'ok': False, 'error': 'Action allowed only on admin/support users'}, status=400)

    desired = request.POST.get('desired')  # 'mute' or 'unmute'
    if desired not in ('mute', 'unmute'):
        # Fallback to toggle behavior
        desired = 'mute' if target.is_active else 'unmute'

    # Prevent locking out all admins (must leave at least one active admin in company)
    if target.role == 'admin' and desired == 'mute' and target.is_active:
        remaining_active = AdminUser.objects.filter(is_active=True, company_profile=company).exclude(id=target.id).count()
        if remaining_active == 0:
            return JsonResponse({'ok': False, 'error': 'Cannot mute the last active admin.'}, status=400)

    # Prevent muting the Master admin (company-registered primary admin)
    try:
        master_admin = AdminUser.objects.filter(company_profile=company).order_by('date_joined').first()
        if master_admin and target.id == master_admin.id:
            return JsonResponse({'ok': False, 'error': 'Cannot mute the Master admin.'}, status=400)
    except Exception:
        pass

    # Prevent muting the master admin (company-registered primary admin)
    try:
        master_admin = AdminUser.objects.filter(company_profile=company).order_by('date_joined').first()
        if master_admin and target.id == master_admin.id:
            return JsonResponse({'ok': False, 'error': 'Cannot mute the Master admin.'}, status=400)
    except Exception:
        pass

    target.is_active = (desired == 'unmute')
    if desired == 'unmute':
        target.deletion_reason = ''
    target.save(update_fields=['is_active', 'deletion_reason'])

    return JsonResponse({'ok': True, 'status': 'active' if target.is_active else 'muted'})


@login_required
@require_POST
def admin_delete_admin(request, user_id: int):
    """Delete an admin or support user. Supports soft delete (default) or permanent delete."""
    if getattr(request.user, 'role', None) != 'admin':
        return JsonResponse({'ok': False, 'error': 'Forbidden'}, status=403)

    if request.user.id == user_id:
        return JsonResponse({'ok': False, 'error': "You can't delete your own account."}, status=400)

    try:
        # SECURITY: Verify user belongs to same company
        # Support dynamic slug parameter
        try:
            company = get_user_company_from_slug(request)
        except HttpResponseForbidden:
            return JsonResponse({'ok': False, 'error': 'Forbidden'}, status=403)
        
        if not company:
            company = request.user.company_profile
        
        # SECURITY: Backend verification - prevent frontend bypass
        has_access, error = verify_company_profile_access(request, company)
        if not has_access:
            logger.warning(f"Unauthorized admin_delete_admin attempt by user {request.user.id}: {error}")
            return JsonResponse({'ok': False, 'error': error}, status=403)
        
        # Verify current user is Master admin
        master_admin = AdminUser.objects.filter(company_profile=company).order_by('date_joined').first()
        if not master_admin or request.user.id != master_admin.id:
            return JsonResponse({'ok': False, 'error': 'Only the Master admin can delete other admins.'}, status=403)
        
        try:
            target = AdminUser.objects.get(id=user_id, company_profile=company)
        except AdminUser.DoesNotExist:
            try:
                target = SupportUser.objects.get(id=user_id, company_profile=company)
            except SupportUser.DoesNotExist:
                return JsonResponse({'ok': False, 'error': 'User not found'}, status=404)
    except Exception:
        return JsonResponse({'ok': False, 'error': 'User not found'}, status=404)

    if target.role not in ('admin', 'support'):
        return JsonResponse({'ok': False, 'error': 'Action allowed only on admin/support users'}, status=400)

    # Ensure at least one other active admin remains in company
    if target.role == 'admin':
        remaining_active = AdminUser.objects.filter(is_active=True, company_profile=company).exclude(id=target.id).count()
        if remaining_active == 0:
            return JsonResponse({'ok': False, 'error': 'Cannot delete the last active admin.'}, status=400)

    # Prevent deleting the Master admin
    if master_admin and target.id == master_admin.id:
        return JsonResponse({'ok': False, 'error': 'Cannot delete the Master admin account.'}, status=400)

    reason = request.POST.get('reason', '').strip()
    permanent = request.POST.get('permanent', 'false').lower() == 'true'
    
    if permanent:
        # Permanent deletion - remove from database entirely
        target_name = target.full_name
        target.delete()
        return JsonResponse({'ok': True, 'message': f'{target_name} permanently deleted.', 'status': 'deleted'})
    else:
        # Soft delete - mark as deleted but keep in database
        target.is_deleted = True
        target.is_active = False
        target.deleted_at = timezone.now()
        if reason:
            target.deletion_reason = reason
        target.save(update_fields=['is_deleted', 'is_active', 'deleted_at', 'deletion_reason'])
        return JsonResponse({'ok': True, 'message': 'Admin deactivated successfully.', 'status': 'deleted'})



@login_required
def admin_client_chat_list(request):
    # Get all clients who have sent at least one message (to ANY admin)
    # SECURITY: Filter by company to prevent cross-company data leakage
    company_filter = {'company_profile': request.company} if hasattr(request, 'company') and request.company else {}
    clients = CustomUser.objects.filter(
        role='client',
        sent_messages__isnull=False,
        **company_filter
    ).distinct().annotate(
        last_message=Max('sent_messages__date_sent')
    ).order_by('-last_message')
    
    for client in clients:
        # Count unread messages from this client to ANY admin (unified dashboard)
        client.unread_count = Message.objects.filter(
            sender=client,
            recipient__role__in=SUPPORT_ROLES,
            is_read=False
        ).count()
    
    # Marketers section: fetch marketers with any messages
    # SECURITY: Filter by company
    marketers = CustomUser.objects.filter(
        role='marketer',
        sent_messages__isnull=False,
        **company_filter
    ).distinct().annotate(
        last_message=Max('sent_messages__date_sent')
    ).order_by('-last_message')

    for m in marketers:
        m.unread_count = Message.objects.filter(
            sender=m,
            recipient__role__in=SUPPORT_ROLES,
            is_read=False
        ).count()

    # Total unread counts for sidebar notifications
    total_unread_count = Message.objects.filter(sender__role='client', recipient__role__in=SUPPORT_ROLES, is_read=False).count()
    marketers_unread_count = Message.objects.filter(sender__role='marketer', recipient__role__in=SUPPORT_ROLES, is_read=False).count()
    
    context = {
        'clients': clients,
        'marketers': marketers,
        'total_unread_count': total_unread_count,
        'marketers_unread_count': marketers_unread_count,
    }
    return render(request, 'admin_side/chat_list.html', context)


@login_required
def admin_marketer_chat_view(request, marketer_id):
    marketer = get_object_or_404(CustomUser, id=marketer_id, role='marketer')
    admin_user = request.user

    # Mark all unread messages from this marketer to ANY admin as read
    Message.objects.filter(sender=marketer, recipient__role__in=SUPPORT_ROLES, is_read=False).update(is_read=True, status='read')

    # Full conversation between this marketer and ANY admin
    # Build companies list for explorer: companies where marketer has transactions
    company_ids = (
        Transaction.objects.filter(marketer=marketer)
        .values_list('company', flat=True)
        .distinct()
    )
    companies_qs = Company.objects.filter(id__in=[c for c in company_ids if c is not None])
    companies = []
    for comp in companies_qs:
        txn_count = Transaction.objects.filter(marketer=marketer, company=comp).count()
        companies.append({'company': comp, 'transactions': txn_count})

    # Conversation: optionally scope to selected company
    sel_company_id = None
    if request.method == 'GET':
        sel_company_id = request.GET.get('company_id')
    else:
        sel_company_id = request.POST.get('company_id')

    selected_company = None
    if sel_company_id:
        try:
            selected_company = Company.objects.get(id=int(sel_company_id))
        except Exception:
            selected_company = None

    if selected_company:
        conversation = Message.objects.filter(
            (Q(sender=marketer, recipient__role__in=SUPPORT_ROLES) & Q(company=selected_company)) |
            (Q(sender__role__in=SUPPORT_ROLES, recipient=marketer) & Q(company=selected_company))
        ).order_by('date_sent')
    else:
        conversation = Message.objects.filter(
            Q(sender=marketer, recipient__role__in=SUPPORT_ROLES) |
            Q(sender__role__in=SUPPORT_ROLES, recipient=marketer)
        ).order_by('date_sent')

    # Polling branch
    if request.method == "GET" and 'last_msg' in request.GET:
        try:
            last_msg_id = int(request.GET.get('last_msg', 0))
        except ValueError:
            last_msg_id = 0
        new_messages = conversation.filter(id__gt=last_msg_id)

        messages_html = ""
        messages_list = []
        for msg in new_messages:
            messages_html += render_to_string('admin_side/chat_message.html', {'msg': msg, 'request': request})
            messages_list.append({'id': msg.id})

        updated_statuses = [{'id': m.id, 'status': m.status} for m in conversation]

        return JsonResponse({
            'messages': messages_list,
            'messages_html': messages_html,
            'updated_statuses': updated_statuses
        })

    # Sending branch
    if request.method == "POST":
        message_content = request.POST.get('message_content', '').strip()
        file_attachment = request.FILES.get('file')
        if not message_content and not file_attachment:
            return JsonResponse({'success': False, 'error': 'Please enter a message or attach a file.'})
        
        # Get company - required for encryption
        company = None
        company_id = request.POST.get('company_id') or request.GET.get('company_id')
        if company_id:
            try:
                company = Company.objects.get(id=int(company_id))
            except Exception:
                return JsonResponse({'success': False, 'error': 'Invalid company_id'}, status=400)

            # ensure marketer has transactions with this company
            if not Transaction.objects.filter(marketer=marketer, company=company).exists():
                return JsonResponse({'success': False, 'error': 'Marketer is not affiliated with this company'}, status=403)
        else:
            # Fallback: get a company the marketer has transactions with (so they can see the message)
            marketer_company_id = Transaction.objects.filter(marketer=marketer).values_list('company', flat=True).first()
            if marketer_company_id:
                company = Company.objects.filter(id=marketer_company_id).first()
            
            # If still no company, try admin's company profile
            if not company:
                company = admin_user.company_profile
            
            if not company:
                return JsonResponse({'success': False, 'error': 'Unable to determine company for message encryption.'}, status=400)

        new_message = Message.objects.create(
            sender=admin_user,
            recipient=marketer,
            message_type="enquiry",
            content=message_content,
            file=file_attachment,
            status='sent',
            company=company,
        )
        message_html = render_to_string('admin_side/chat_message.html', {'msg': new_message, 'request': request})
        return JsonResponse({'success': True, 'message_html': message_html})

    # Get clients and marketers for sidebar chat list
    company_filter = {'company_profile': request.company} if hasattr(request, 'company') and request.company else {}
    sidebar_clients = CustomUser.objects.filter(
        role='client',
        sent_messages__isnull=False,
        **company_filter
    ).distinct().annotate(
        last_message=Max('sent_messages__date_sent')
    ).order_by('-last_message')[:20]
    
    for c in sidebar_clients:
        c.unread_count = Message.objects.filter(
            sender=c, recipient__role__in=SUPPORT_ROLES, is_read=False
        ).count()
        last_msg = Message.objects.filter(
            Q(sender=c, recipient__role__in=SUPPORT_ROLES) | Q(sender__role__in=SUPPORT_ROLES, recipient=c)
        ).order_by('-date_sent').first()
        c.last_message_text = last_msg.content[:50] if last_msg and last_msg.content else ''

    sidebar_marketers = CustomUser.objects.filter(
        role='marketer',
        sent_messages__isnull=False,
        **company_filter
    ).distinct().annotate(
        last_message=Max('sent_messages__date_sent')
    ).order_by('-last_message')[:20]
    
    for m in sidebar_marketers:
        m.unread_count = Message.objects.filter(
            sender=m, recipient__role__in=SUPPORT_ROLES, is_read=False
        ).count()
        last_msg = Message.objects.filter(
            Q(sender=m, recipient__role__in=SUPPORT_ROLES) | Q(sender__role__in=SUPPORT_ROLES, recipient=m)
        ).order_by('-date_sent').first()
        m.last_message_text = last_msg.content[:50] if last_msg and last_msg.content else ''

    context = {
        'client': marketer,   # Reuse template expecting 'client'
        'messages': conversation,
        'is_marketer': True,
        'companies': companies,
        'selected_company': selected_company,
        'clients': sidebar_clients,
        'marketers': sidebar_marketers,
    }
    return render(request, 'admin_side/chat_interface.html', context)


@login_required
def search_clients_api(request):
    """API endpoint to search for clients to start chat with"""
    query = request.GET.get('q', '').strip()
    
    if len(query) < 2:
        return JsonResponse({'clients': []})
    
    # Search clients by name or email
    # SECURITY: Filter clients by company to prevent cross-tenant data exposure
    company = request.user.company_profile
    clients = CustomUser.objects.filter(
        role='client',
        company_profile=company,
        full_name__icontains=query
    ) | CustomUser.objects.filter(
        role='client',
        company_profile=company,
        email__icontains=query
    )
    
    clients = clients.distinct()[:10]  # Limit to 10 results
    
    clients_data = []
    for client in clients:
        clients_data.append({
            'id': client.id,
            'full_name': client.full_name,
            'email': client.email if client.email else '',
        })
    
    return JsonResponse({'clients': clients_data})


@login_required
def search_marketers_api(request):
    """API endpoint to search for marketers to start chat with"""
    query = request.GET.get('q', '').strip()

    if len(query) < 2:
        return JsonResponse({'marketers': []})

    # SECURITY: Filter marketers by company to prevent cross-tenant data exposure
    company = request.user.company_profile
    
    # Get marketers from company_profile
    marketers_primary = CustomUser.objects.filter(
        role='marketer',
        company_profile=company,
        full_name__icontains=query
    ) | CustomUser.objects.filter(
        role='marketer',
        company_profile=company,
        email__icontains=query
    )
    
    # ALSO get marketers from MarketerAffiliation (for users added via add_existing_user_to_company)
    affiliation_marketer_ids = list(
        MarketerAffiliation.objects.filter(company=company).values_list('marketer_id', flat=True).distinct()
    )
    marketers_affiliated = CustomUser.objects.filter(
        id__in=affiliation_marketer_ids,
        role='marketer'
    ).filter(
        Q(full_name__icontains=query) | Q(email__icontains=query)
    ).exclude(
        id__in=marketers_primary.values_list('pk', flat=True)
    )
    
    # Combine both querysets
    marketers = (marketers_primary | marketers_affiliated).distinct()[:10]

    marketers_data = []
    for m in marketers:
        marketers_data.append({
            'id': m.id,
            'full_name': m.full_name,
            'email': m.email if m.email else '',
        })

    return JsonResponse({'marketers': marketers_data})


# ============================================================================
# CROSS-COMPANY USER DISCOVERY AND ADDITION
# ============================================================================

@login_required
@require_http_methods(["GET"])
def search_existing_users_api(request):
    """
    API endpoint for company admins to search for existing users (registered via signup)
    to add them to their company roster.
    
    Searches globally across all users with role='client' or 'marketer'
    who have NOT been assigned to this company yet.
    
    Query params:
    - q: email or name to search
    - role: 'client' or 'marketer' (required)
    """
    query = request.GET.get('q', '').strip()
    role = request.GET.get('role', '').strip()

    # Allow role to be optional: if provided and valid, search that role only,
    # otherwise search both 'client' and 'marketer'. This helps UI callers
    # that may not know the exact role or when admins want a broader search.
    if role and role not in ['client', 'marketer']:
        return JsonResponse({'error': 'Invalid role'}, status=400)

    if len(query) < 2:
        return JsonResponse({'users': []})
    
    # Get current company
    company = request.user.company_profile
    if not company:
        return JsonResponse({'error': 'User not assigned to any company'}, status=403)
    
    # Search for users matching query (email, name, phone)
    # Find users NOT YET in this company
    qs = CustomUser.objects.filter(is_active=True, is_deleted=False)
    if role:
        qs = qs.filter(role=role)
    else:
        qs = qs.filter(role__in=['client', 'marketer'])

    # Build a robust search expression: exact email, partial email, tokenized name match, phone
    search_q = Q(email__iexact=query) | Q(email__icontains=query) | Q(full_name__icontains=query) | Q(phone__icontains=query)
    # Tokenize query and require all tokens to appear in full_name when multiple words provided
    tokens = [t.strip() for t in query.split() if t.strip()]
    if tokens and len(tokens) > 1:
        name_tokens_q = None
        for tok in tokens:
            if name_tokens_q is None:
                name_tokens_q = Q(full_name__icontains=tok)
            else:
                name_tokens_q &= Q(full_name__icontains=tok)
        if name_tokens_q is not None:
            search_q |= name_tokens_q

    # Get a broader candidate set then perform normalized matching in Python
    candidates = list(qs.filter(search_q).exclude(company_profile=company).distinct()[:200])

    def _strip_diacritics(s):
        if not s:
            return ''
        s = str(s)
        s = unicodedata.normalize('NFD', s)
        s = ''.join(ch for ch in s if not unicodedata.combining(ch))
        return s.lower()

    def _normalize_email_for_matching(email):
        if not email:
            return ''
        email = email.strip().lower()
        if '@' not in email:
            return email
        local, domain = email.split('@', 1)
        # strip plus-tags from local part
        if '+' in local:
            local = local.split('+', 1)[0]
        # For Gmail-style addresses, dots in local part are ignored
        if domain in ('gmail.com', 'googlemail.com'):
            local = local.replace('.', '')
        return f"{local}@{domain}"

    query_norm = _strip_diacritics(query)
    query_is_email = '@' in query
    query_norm_email = _normalize_email_for_matching(query) if query_is_email else None

    matches = []
    seen_ids = set()

    def try_add(u):
        if u.id in seen_ids:
            return
        seen_ids.add(u.id)
        matches.append(u)

    # First pass: exact normalized email matches
    if query_is_email:
        for u in candidates:
            if _normalize_email_for_matching(u.email) == query_norm_email:
                try_add(u)

    # Second pass: diacritic-normalized name match, email contains, or phone contains
    for u in candidates:
        if u.id in seen_ids:
            continue
        uname = _strip_diacritics(u.full_name or '')
        uemail = _strip_diacritics(u.email or '')
        uphone = (u.phone or '')

        if query_norm and query_norm in uname:
            try_add(u)
            continue
        if query_norm and query_norm in uemail:
            try_add(u)
            continue
        if query and uphone and query in uphone:
            try_add(u)
            continue

    # If still no matches and the caller specified a role, broaden to both roles and repeat
    if not matches and role:
        qs2 = CustomUser.objects.filter(is_active=True, is_deleted=False, role__in=['client', 'marketer'])
        candidates2 = list(qs2.filter(search_q).exclude(company_profile=company).distinct()[:400])
        for u in candidates2:
            if u.id in seen_ids:
                continue
            if query_is_email and _normalize_email_for_matching(u.email) == query_norm_email:
                try_add(u)
                continue
            uname = _strip_diacritics(u.full_name or '')
            uemail = _strip_diacritics(u.email or '')
            uphone = (u.phone or '')
            if query_norm and query_norm in uname:
                try_add(u)
                continue
            if query_norm and query_norm in uemail:
                try_add(u)
                continue
            if query and uphone and query in uphone:
                try_add(u)
                continue

    users = matches[:20]

    # If debug flag is present, include diagnostic info to help troubleshoot
    if request.GET.get('debug') in ('1', 'true', 'True'):
        debug_info = {
            'query': query,
            'query_norm': query_norm,
            'query_is_email': query_is_email,
            'query_norm_email': query_norm_email,
            'candidate_count': len(candidates),
            'matched_count': len(matches),
            'matched_ids': [u.id for u in matches[:50]],
            'matched_emails': [u.email for u in matches[:50]],
        }
        users_data = []
        for user in users:
            users_data.append({
                'id': user.id,
                'email': user.email,
                'full_name': user.full_name,
                'phone': user.phone,
                'role': getattr(user, 'role', None),
            })
        return JsonResponse({'users': users_data, 'debug': debug_info})
    
    users_data = []
    for user in users:
        users_data.append({
            'id': user.id,
            'email': user.email,
            'full_name': user.full_name,
            'phone': user.phone,
            # Note: date_registered is not the user's DOB; we only include if there's an explicit date_of_birth field
            'date_registered': user.date_registered.strftime('%Y-%m-%d %H:%M'),
            'date_of_birth': getattr(user, 'date_of_birth', None).isoformat() if getattr(user, 'date_of_birth', None) else None,
            'address': getattr(user, 'address', '') or '',
            'country': getattr(user, 'country', '') or '',
            'role': getattr(user, 'role', None),
            'is_already_in_company': user.company_profile is not None,
            'is_in_requester_company': True if (user.company_profile and company and user.company_profile.id == company.id) else False,
            'current_company': user.company_profile.company_name if user.company_profile else None,
        })
    
    return JsonResponse({'users': users_data})


@login_required
@require_http_methods(["POST"])
def add_existing_user_to_company(request):
    """
    Add an already-registered user (from signup or another company) to current company.
    
    POST params:
    - user_id: ID of user to add
    - marketer_id: (Optional, only for client role) ID of marketer to assign
    
    User becomes part of this company without replacing their existing company relationships.
    """
    import json
    
    try:
        company = request.user.company_profile
        if not company:
            return JsonResponse({'error': 'User not assigned to any company'}, status=403)
        
        # Check if user is admin
        if request.user.role != 'admin':
            return JsonResponse({'error': 'Only admins can add users to company'}, status=403)
        
        data = json.loads(request.body)
        user_id = data.get('user_id')
        marketer_id = data.get('marketer_id')
        
        if not user_id:
            return JsonResponse({'error': 'user_id required'}, status=400)
        
        with transaction.atomic():
            # Get the user
            user = CustomUser.objects.get(id=user_id, is_active=True, is_deleted=False)
            
            # Check if user already belongs to this company
            # For clients: check ClientMarketerAssignment
            # For marketers: check MarketerAffiliation
            # For other roles: check company_profile
            user_already_in_company = False
            if user.role == 'client':
                user_already_in_company = ClientMarketerAssignment.objects.filter(
                    client_id=user.pk, company=company
                ).exists()
            elif user.role == 'marketer':
                user_already_in_company = MarketerAffiliation.objects.filter(
                    marketer_id=user.pk, company=company
                ).exists()
            else:  # admin, support, other
                user_already_in_company = user.company_profile and user.company_profile.id == company.id
            
            if user_already_in_company:
                return JsonResponse({
                    'error': f'{user.email} already exists as {user.role} in {company.company_name}.',
                    'existing': True,
                    'role': getattr(user, 'role', None),
                    'company': company.company_name if company else None,
                    'user': {
                        'id': user.id,
                        'email': user.email,
                        'full_name': user.full_name,
                    }
                }, status=400)
            
            # For clients, ensure marketer assignment and subclass linkage
            if user.role == 'client':
                if not marketer_id:
                    return JsonResponse({
                        'error': 'Marketer assignment required for client users'
                    }, status=400)

                # Resolve marketer record (ensure MarketerUser subclass exists)
                try:
                    # Prefer MarketerUser if available
                    marketer_obj = None
                    if MarketerUser.objects.filter(pk=marketer_id).exists():
                        marketer_obj = MarketerUser.objects.get(pk=marketer_id)
                        # SECURITY: Verify marketer belongs to company (either via company_profile or MarketerAffiliation)
                        if marketer_obj.company_profile != company:
                            if not MarketerAffiliation.objects.filter(marketer_id=marketer_id, company=company).exists():
                                raise CustomUser.DoesNotExist("Marketer not in company")
                    else:
                        # Fallback: ensure parent CustomUser exists and create MarketerUser subclass
                        # Check both company_profile AND MarketerAffiliation
                        parent_marketer = None
                        try:
                            parent_marketer = CustomUser.objects.get(id=marketer_id, role='marketer', company_profile=company)
                        except CustomUser.DoesNotExist:
                            # Check if marketer belongs via MarketerAffiliation
                            if MarketerAffiliation.objects.filter(marketer_id=marketer_id, company=company).exists():
                                parent_marketer = CustomUser.objects.get(id=marketer_id, role='marketer')
                            else:
                                raise CustomUser.DoesNotExist("Marketer not found in company")
                        
                        if not MarketerUser.objects.filter(pk=parent_marketer.pk).exists():
                            MarketerUser.objects.create(id=parent_marketer.pk, company_profile=company)
                        marketer_obj = MarketerUser.objects.get(pk=parent_marketer.pk)
                except CustomUser.DoesNotExist:
                    return JsonResponse({
                        'error': 'Selected marketer not found in company'
                    }, status=404)

                # Ensure the user has a ClientUser subclass so directory queries include them
                if not ClientUser.objects.filter(pk=user.pk).exists():
                    # Create ClientUser row referencing existing CustomUser (multi-table inheritance)
                    # For multi-table inheritance, we use the user's pointer to create the subclass
                    try:
                        # Get or create the ClientUser subclass linked to this CustomUser
                        client_user = ClientUser.objects.get_or_create(
                            customuser_ptr_id=user.id,
                            defaults={'company_profile': company}
                        )[0]
                    except Exception as e:
                        logger.warning(f"Failed to get_or_create ClientUser: {str(e)}")
                        # Fallback: just mark the user as added to company
                        pass

                client_obj = ClientUser.objects.get(pk=user.id)
                # Assign marketer on the ClientUser instance and create company-specific assignment
                client_obj.assigned_marketer = marketer_obj
                client_obj.save()
                try:
                    ClientMarketerAssignment.objects.get_or_create(client=client_obj, marketer=marketer_obj, company=company)
                except Exception:
                    # non-fatal
                    pass
            
            # Add user to company - ALWAYS save to ensure user is added to this company
            # Even if user already has a company_profile, they need to be affiliated with this company too
            # Users can belong to multiple companies via ClientMarketerAssignment/MarketerAffiliation
            if not user.company_profile:
                # User has no primary company yet, set this as their primary
                user.company_profile = company
            user.save()  # Always save - ensures user is persisted to this company

            # Ensure marketer subclass exists for marketers so marketer directory will list them
            if user.role == 'marketer':
                if not MarketerUser.objects.filter(pk=user.pk).exists():
                    # Create marketer subclass tied to this company so marketer_list picks them up
                    try:
                        # Get or create the MarketerUser subclass linked to this CustomUser
                        marketer_user = MarketerUser.objects.get_or_create(
                            customuser_ptr_id=user.id,
                            defaults={'company_profile': company}
                        )[0]
                    except Exception as e:
                        logger.warning(f"Failed to get_or_create MarketerUser: {str(e)}")
                        # Fallback: just mark the user as added to company
                        pass
                
                # Create MarketerAffiliation so marketer can be found in this company
                try:
                    MarketerAffiliation.objects.get_or_create(
                        marketer_id=user.pk,
                        company=company
                    )
                except Exception as e:
                    logger.warning(f"Failed to create MarketerAffiliation for {user.pk} in {company.pk}: {str(e)}")
            
            return JsonResponse({
                'success': True,
                'message': f'{user.email} ({user.role}) successfully added to {company.company_name}',
                'user': {
                    'id': user.id,
                    'email': user.email,
                    'full_name': user.full_name,
                    'role': user.role,
                    'company': company.company_name
                }
            })
    
    except CustomUser.DoesNotExist:
        logger.error(f'CustomUser not found: user_id={user_id}')
        return JsonResponse({'error': 'User not found'}, status=404)
    except json.JSONDecodeError:
        logger.error('Invalid JSON in request body')
        return JsonResponse({'error': 'Invalid JSON'}, status=400)
    except Exception as e:
        logger.error(f'Error adding user to company: {str(e)}', exc_info=True)
        import traceback
        traceback.print_exc()
        return JsonResponse({'error': str(e)}, status=500)


# ============================================================================
# CLIENT CROSS-COMPANY PORTFOLIO SUPPORT
# ============================================================================

@login_required
def client_dashboard_cross_company(request):
    """
    Client dashboard showing all properties across all companies the client is registered with.
    
    Features:
    - Display all properties from all companies
    - Company toggle to filter by specific company
    - Show company list in profile
    """
    user = request.user
    
    # Get all companies this client is registered with
    # A client can have:
    # 1. Primary company_profile (from when added to company)
    # 2. Multiple company relationships via portfolios/transactions
    
    # Find all estates client has properties in
    client_estates = Estate.objects.filter(
        transaction__client=user
    ).distinct().select_related('company').order_by('company', '-date_added')
    
    # Get unique companies from these estates
    client_companies = set()
    for estate in client_estates:
        if estate.company:
            client_companies.add(estate.company)
    
    # Also check if user has direct company_profile
    if user.company_profile:
        client_companies.add(user.company_profile)
    
    client_companies = sorted(list(client_companies), key=lambda x: x.company_name)
    
    # Get selected company from request (for filtering)
    selected_company_id = request.GET.get('company_id')
    
    # Fetch transactions/properties for this client
    if selected_company_id:
        try:
            selected_company = Company.objects.get(id=selected_company_id)
            # Verify client is in this company
            if selected_company not in client_companies:
                selected_company = None
        except Company.DoesNotExist:
            selected_company = None
    else:
        selected_company = None
    
    # Get client's transactions/properties
    transactions = Transaction.objects.filter(client=user).select_related(
        'estate',
        'estate__company',
        'assigned_unit'
    ).order_by('-date_added')
    
    if selected_company:
        transactions = transactions.filter(estate__company=selected_company)
    
    # Prepare context
    context = {
        'client': user,
        'transactions': transactions,
        'client_companies': client_companies,
        'selected_company': selected_company,
        'total_companies': len(client_companies),
    }
    
    return render(request, 'client_side/dashboard_cross_company.html', context)

    
# XXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXX
# CLIENT SIDE
# XXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXX

@login_required
def client(request):
    """
    Display client list with company-scoped data.
    SECURITY: Strict company isolation prevents cross-company data leakage.
    """
    # Get all clients with their assigned marketers using select_related for optimization
    # SECURITY: Filter by company to prevent cross-company data leakage
    company = getattr(request, 'company', None) or getattr(request.user, 'company_profile', None)
    if not company:
        messages.error(request, "No company context available.")
        return redirect('home')

    # Primary queryset: clients represented by ClientUser subclass
    clients_qs = ClientUser.objects.filter(role='client', company_profile=company).select_related('assigned_marketer').order_by('-date_registered')

    # ALSO include clients from ClientMarketerAssignment (for users added via add_existing_user_to_company)
    assignment_client_ids = list(
        ClientMarketerAssignment.objects.filter(company=company).values_list('client_id', flat=True).distinct()
    )
    assignment_clients_qs = ClientUser.objects.filter(id__in=assignment_client_ids).exclude(
        id__in=clients_qs.values_list('pk', flat=True)
    ).select_related('assigned_marketer').order_by('-date_registered')
    
    # Combine both querysets
    clients_qs = list(clients_qs) + list(assignment_clients_qs)

    # Fallback: include any CustomUser(client) rows that do not yet have a ClientUser subclass
    client_ids = [c.pk if hasattr(c, 'pk') else c for c in clients_qs]
    fallback_clients = list(CustomUser.objects.filter(role='client', company_profile=company).exclude(id__in=client_ids).order_by('-date_registered'))

    # Resolve assigned_marketer for fallback clients so templates can show marketer column
    for fc in fallback_clients:
        try:
            # 1) If a ClientUser subclass exists somewhere (unlikely here), prefer its assignment
            if ClientUser.objects.filter(pk=fc.pk).exists():
                cu = ClientUser.objects.get(pk=fc.pk)
                fc.assigned_marketer = cu.get_assigned_marketer(company=company) if hasattr(cu, 'get_assigned_marketer') else getattr(cu, 'assigned_marketer', None)
            else:
                # 2) Try company-scoped ClientMarketerAssignment (may not exist without ClientUser)
                assign = ClientMarketerAssignment.objects.filter(client_id=fc.pk, company=company).select_related('marketer').first()
                if assign:
                    fc.assigned_marketer = assign.marketer
                else:
                    # 3) Fallback: look for recent Transaction by this client in the company and use its marketer
                    txn = Transaction.objects.filter(client_id=fc.pk, company=company, marketer__isnull=False).order_by('-transaction_date').select_related('marketer').first()
                    fc.assigned_marketer = txn.marketer if txn else None
        except Exception:
            # Non-fatal: ensure attribute exists and is None when unresolved
            fc.assigned_marketer = None

    # Combine into a single list
    combined = list(clients_qs) + fallback_clients

    # Sort by registration date descending for display
    combined.sort(key=lambda u: getattr(u, 'date_registered', None) or timezone.now(), reverse=True)

    # Compute per-company sequential client IDs for display (company_uid)
    try:
        ordered_for_ids = sorted(combined, key=lambda u: getattr(u, 'date_registered', None) or timezone.now())
        id_map = {u.pk: idx + 1 for idx, u in enumerate(ordered_for_ids)}
        for c in combined:
            c.company_uid = id_map.get(c.pk, c.pk)
            
            # Fetch the actual CompanyClientProfile for this company
            try:
                profile = CompanyClientProfile.objects.get(client_id=c.pk, company=company)
                c.company_client_id = profile.company_client_id
                c.company_client_uid = profile.company_client_uid
            except CompanyClientProfile.DoesNotExist:
                # Fallback: Use company_uid if no profile exists
                c.company_client_id = c.company_uid
                try:
                    prefix = company._company_prefix() if company else 'CMP'
                except Exception:
                    prefix = getattr(company, 'company_name', 'CMP')[:3].upper()
                c.company_client_uid = f"{prefix}CLT{int(c.company_client_id):03d}"
            except Exception:
                # Final fallback
                c.company_client_id = c.company_uid
                c.company_client_uid = f"CLT{int(getattr(c, 'company_client_id', 0) or 0):03d}"
    except Exception:
        for c in combined:
            c.company_uid = getattr(c, 'pk', None)
            try:
                if not getattr(c, 'company_client_id', None):
                    c.company_client_id = c.company_uid
            except Exception:
                c.company_client_id = c.company_uid

    return render(request, 'admin_side/client.html', {'clients': combined, 'company': company})


@login_required
@require_http_methods(["POST"])
def client_soft_delete(request, pk):
    """Soft delete a client"""
    import json
    try:
        # SECURITY: Verify client belongs to same company (either via company_profile or ClientMarketerAssignment)
        company = request.user.company_profile
        client = CustomUser.objects.get(pk=pk, role='client')
        
        # Check if client belongs to company via company_profile OR ClientMarketerAssignment
        in_company_profile = client.company_profile == company
        in_affiliation = ClientMarketerAssignment.objects.filter(client_id=pk, company=company).exists()
        
        if not (in_company_profile or in_affiliation):
            return JsonResponse({
                'success': False,
                'message': 'Client not found in your company.'
            }, status=404)
        
        # Get deletion reasons from request
        data = json.loads(request.body)
        reasons = data.get('reasons', [])
        
        # Mark as deleted
        client.is_deleted = True
        client.deleted_at = timezone.now()
        client.deletion_reason = ', '.join(reasons)
        client.save()
        
        return JsonResponse({
            'success': True,
            'message': f'Client {client.full_name} has been marked as deleted.'
        })
    except CustomUser.DoesNotExist:
        return JsonResponse({
            'success': False,
            'message': 'Client not found.'
        }, status=404)
    except Exception as e:
        return JsonResponse({
            'success': False,
            'message': str(e)
        }, status=500)


@login_required
@require_http_methods(["POST"])
def client_restore(request, pk):
    """Restore a soft-deleted client"""
    try:
        # SECURITY: Verify client belongs to same company (either via company_profile or ClientMarketerAssignment)
        company = request.user.company_profile
        client = CustomUser.objects.get(pk=pk, role='client')
        
        # Check if client belongs to company via company_profile OR ClientMarketerAssignment
        in_company_profile = client.company_profile == company
        in_affiliation = ClientMarketerAssignment.objects.filter(client_id=pk, company=company).exists()
        
        if not (in_company_profile or in_affiliation):
            return JsonResponse({
                'success': False,
                'message': 'Client not found in your company.'
            }, status=404)
        
        # Restore the client
        client.is_deleted = False
        client.deleted_at = None
        client.deletion_reason = None
        client.save()
        
        return JsonResponse({
            'success': True,
            'message': f'Client {client.full_name} has been restored.'
        })
    except CustomUser.DoesNotExist:
        return JsonResponse({
            'success': False,
            'message': 'Client not found.'
        }, status=404)
    except Exception as e:
        return JsonResponse({
            'success': False,
            'message': str(e)
        }, status=500)


@login_required
@require_http_methods(["POST"])
def marketer_soft_delete(request, pk):
    """Soft delete a marketer"""
    import json
    try:
        # SECURITY: Verify marketer belongs to same company (either via company_profile or MarketerAffiliation)
        company = request.user.company_profile
        marketer = CustomUser.objects.get(pk=pk, role='marketer')
        
        # Check if marketer belongs to company via company_profile OR MarketerAffiliation
        in_company_profile = marketer.company_profile == company
        in_affiliation = MarketerAffiliation.objects.filter(marketer_id=pk, company=company).exists()
        
        if not (in_company_profile or in_affiliation):
            return JsonResponse({
                'success': False,
                'message': 'Marketer not found in your company.'
            }, status=404)
        
        # Get deletion reasons from request
        data = json.loads(request.body)
        reasons = data.get('reasons', [])
        
        # Mark as deleted
        marketer.is_deleted = True
        marketer.deleted_at = timezone.now()
        marketer.deletion_reason = ', '.join(reasons)
        marketer.save()
        
        return JsonResponse({
            'success': True,
            'message': f'Marketer {marketer.full_name} has been marked as deleted.'
        })
    except CustomUser.DoesNotExist:
        return JsonResponse({
            'success': False,
            'message': 'Marketer not found.'
        }, status=404)
    except Exception as e:
        return JsonResponse({
            'success': False,
            'message': str(e)
        }, status=500)


@login_required
@require_http_methods(["POST"])
def marketer_restore(request, pk):
    """Restore a soft-deleted marketer"""
    try:
        # SECURITY: Verify marketer belongs to same company (either via company_profile or MarketerAffiliation)
        company = request.user.company_profile
        marketer = CustomUser.objects.get(pk=pk, role='marketer')
        
        # Check if marketer belongs to company via company_profile OR MarketerAffiliation
        in_company_profile = marketer.company_profile == company
        in_affiliation = MarketerAffiliation.objects.filter(marketer_id=pk, company=company).exists()
        
        if not (in_company_profile or in_affiliation):
            return JsonResponse({
                'success': False,
                'message': 'Marketer not found in your company.'
            }, status=404)
        
        # Restore the marketer
        marketer.is_deleted = False
        marketer.deleted_at = None
        marketer.deletion_reason = None
        marketer.save()
        
        return JsonResponse({
            'success': True,
            'message': f'Marketer {marketer.full_name} has been restored.'
        })
    except CustomUser.DoesNotExist:
        return JsonResponse({
            'success': False,
            'message': 'Marketer not found.'
        }, status=404)
    except Exception as e:
        return JsonResponse({
            'success': False,
            'message': str(e)
        }, status=500)


def calculate_portfolio_metrics(transactions):
    appreciation_total = Decimal(0)
    growth_rates = []
    highest_growth_rate = Decimal(0)
    highest_growth_property = ""
    
    for transaction in transactions:
        # Fetch the latest PropertyPrice for this estate + plot size, preferring the company's price
        try:
            property_price_qs = PropertyPrice.objects.filter(
                estate=transaction.allocation.estate,
                plot_unit__plot_size=transaction.allocation.plot_size,
                company=transaction.company
            ).order_by('-effective')

            property_price = property_price_qs.first()
            if property_price:
                current_value = property_price.current
            else:
                # Fallback to any matching price regardless of company (defensive)
                property_price_alt = PropertyPrice.objects.filter(
                    estate=transaction.allocation.estate,
                    plot_unit__plot_size=transaction.allocation.plot_size
                ).order_by('-effective').first()
                current_value = property_price_alt.current if property_price_alt else transaction.total_amount
        except Exception:
            current_value = transaction.total_amount
        
        appreciation = current_value - transaction.total_amount
        appreciation_total += appreciation
        
        if transaction.total_amount > 0:
            growth_rate = (appreciation / transaction.total_amount) * 100
        else:
            growth_rate = Decimal(0)
        
        growth_rates.append(growth_rate)
        
        if growth_rate > highest_growth_rate:
            highest_growth_rate = growth_rate
            highest_growth_property = transaction.allocation.estate.name
            
        transaction.current_value = current_value
        transaction.appreciation = appreciation
        transaction.growth_rate = growth_rate
        transaction.abs_appreciation = abs(appreciation)
        transaction.abs_growth_rate = abs(growth_rate)
    
    return {
        'transactions': transactions,
        'properties_count': transactions.count(),
        'total_value': sum(t.total_amount for t in transactions),
        'current_value': sum(t.current_value for t in transactions),
        'appreciation_total': appreciation_total,
        'average_growth': sum(growth_rates) / len(growth_rates) if growth_rates else 0,
        'highest_growth_rate': highest_growth_rate,
        'highest_growth_property': highest_growth_property,
    }


@login_required
def client_profile(request, slug=None, pk=None, company_slug=None):
    """
    Display client profile with portfolio/transactions.
    
    Supports multiple URL formats:
    1. Legacy: /client_profile/<pk>/ (with or without ?company=slug)
    2. UID-based: /<LPLCLT002>.client-profile?company=<slug> or without company
    3. Name-based slug: /<victor-godwin>.client-profile?company=<slug>
    4. Company-namespaced: /<company_slug>/client/<identifier>/
    
    SECURITY: Enforces strict company isolation - only shows client and transactions for their company.
    Prevents cross-company data leakage by validating company context in all code paths.
    """
    
    # ============================================================
    # STEP 1: Determine and validate company context (FLEXIBLE)
    # ============================================================
    company = None
    company_slug_param = None
    
    # Priority 1: URL company_slug parameter (most explicit)
    if company_slug:
        company = get_object_or_404(Company, slug=company_slug)
    # Priority 2: Query parameter ?company=slug (modern routing)
    elif request.GET.get('company'):
        company_slug_param = request.GET.get('company').strip()
        if company_slug_param:  # Only use if not empty string
            company = get_object_or_404(Company, slug=company_slug_param)
        else:
            # Empty company parameter provided (?company=) - user must specify a company
            raise Http404("⚠️  Company parameter is empty. Please specify: ?company=<company-slug>")
    
    # Priority 3: User's primary company_profile (fallback for legacy routes ONLY)
    # Use only if NO company parameter was provided at all
    if not company and not request.GET.get('company'):
        if hasattr(request.user, 'company_profile') and request.user.company_profile:
            company = request.user.company_profile
    
    # SECURITY: No company context = deny access (cannot infer company)
    if not company:
        raise Http404("❌ No company context provided. Required: ?company=<company-slug> or URL must include company slug.")
    
    # ============================================================
    # STEP 2: Verify requesting user has access to this company
    # ============================================================
    user_company = getattr(request.user, 'company_profile', None)
    if user_company != company:
        # Check if user has affiliation with this company (e.g., supports multiple companies)
        # For now, restrict to user's primary company to prevent admin from viewing other company's data
        raise HttpResponseForbidden(f"❌ Access Denied: You can only access profiles within your own company context.")
    
    # ============================================================
    # STEP 3: Find client within company context (STRICT FILTERING)
    # ============================================================
    client = None
    
    if slug:
        # Slug-based lookup supports MULTIPLE formats:
        # 1. Company-specific UID: "LPLCLT005" (company_client_uid)
        # 2. Name-based slug: "victor-godwin" (converted from full_name)
        # 3. Exact name: "Victor Godwin"
        
        # Priority 1: Try company-specific UID lookup (LPLCLT005)
        # Check BOTH ClientUser.company_client_uid AND CompanyClientProfile.company_client_uid
        if company:
            # First check direct UID on ClientUser
            client = ClientUser.objects.filter(
                company_client_uid=slug,
                company_profile=company
            ).first()
            
            # If not found, check CompanyClientProfile (for affiliated users)
            if not client:
                try:
                    profile = CompanyClientProfile.objects.get(
                        company_client_uid=slug,
                        company=company
                    )
                    client = profile.client
                except CompanyClientProfile.DoesNotExist:
                    pass
        
        # Priority 2: Try name-based slug (e.g., "victor-godwin" → "Victor Godwin")
        if not client:
            name_from_slug = slug.replace('-', ' ')
            user = ClientUser.objects.filter(
                full_name__iexact=name_from_slug
            ).first()
            
            # Verify user exists in THIS company (via company_profile OR affiliation)
            if user:
                has_company_access = (
                    user.company_profile == company or
                    ClientMarketerAssignment.objects.filter(
                        client_id=user.id,
                        company=company
                    ).exists()
                )
                
                if has_company_access:
                    client = user
                else:
                    raise HttpResponseForbidden(
                        f"❌ Access Denied: User '{slug}' is not part of {company.company_name}."
                    )
        
        # Priority 3: Try exact name match
        if not client:
            user = ClientUser.objects.filter(
                full_name=slug
            ).first()
            
            if user:
                has_company_access = (
                    user.company_profile == company or
                    ClientMarketerAssignment.objects.filter(
                        client_id=user.id,
                        company=company
                    ).exists()
                )
                
                if has_company_access:
                    client = user
                else:
                    raise HttpResponseForbidden(
                        f"❌ Access Denied: User '{slug}' is not part of {company.company_name}."
                    )
        
        if not client:
            raise Http404(f"❌ Client '{slug}' not found in {company.company_name}.")
    
    elif pk:
        # Legacy ID-based lookup: VERIFY client belongs to company
        try:
            client = ClientUser.objects.get(id=pk)
        except ClientUser.DoesNotExist:
            raise Http404("❌ Client not found.")
        
        # SECURITY: Verify client belongs to THIS company ONLY
        client_in_company = (
            client.company_profile == company or
            ClientMarketerAssignment.objects.filter(
                client_id=client.id,
                company=company
            ).exists()
        )
        
        if not client_in_company:
            raise HttpResponseForbidden(f"❌ Access Denied: Client not associated with {company.company_name}.")
    else:
        raise Http404("❌ No client identifier provided (slug or pk required).")
    
    # ============================================================
    # STEP 4: Fetch transactions STRICTLY scoped to company
    # ============================================================
    # CRITICAL: All transactions must be filtered by company to prevent data leakage
    transactions = Transaction.objects.filter(
        client_id=client.id,
        company=company  # <-- STRICT COMPANY FILTER
    ).select_related(
        'allocation__estate',
        'allocation__plot_size'
    ).order_by('-transaction_date')
    
    # ============================================================
    # STEP 5: Calculate portfolio metrics (company-scoped)
    # ============================================================
    appreciation_total = Decimal(0)
    growth_rates = []
    highest_growth_rate = Decimal(0)
    highest_growth_property = ""
    
    for transaction in transactions:
        # Get current price for this property (WITHIN company context)
        try:
            property_price = PropertyPrice.objects.filter(
                estate__company=company,  # <-- STRICT COMPANY FILTER
                plot_unit__plot_size=transaction.allocation.plot_size
            ).get(estate=transaction.allocation.estate)
            current_value = property_price.current
        except PropertyPrice.DoesNotExist:
            current_value = transaction.total_amount 
        
        # Calculate appreciation
        appreciation = current_value - transaction.total_amount
        appreciation_total += appreciation
        
        # Calculate growth rate
        if transaction.total_amount > 0:
            growth_rate = (appreciation / transaction.total_amount) * 100
        else:
            growth_rate = Decimal(0)
            
        growth_rates.append(growth_rate)
        
        # Track highest growth property
        if growth_rate > highest_growth_rate:
            highest_growth_rate = growth_rate
            highest_growth_property = transaction.allocation.estate.name
            
        # Add dynamic properties to transaction
        transaction.current_value = current_value
        transaction.appreciation = appreciation
        transaction.growth_rate = growth_rate
        transaction.abs_appreciation = abs(appreciation)
        transaction.abs_growth_rate = abs(growth_rate)
    
    # Calculate averages
    properties_count = transactions.count()
    average_growth = sum(growth_rates) / len(growth_rates) if growth_rates else 0
    
    context = {
        'client': client,
        'company': company,
        'transactions': transactions,
        'properties_count': properties_count,
        'total_value': sum(t.total_amount for t in transactions) if transactions else Decimal(0),
        'current_value': sum(t.current_value for t in transactions) if transactions else Decimal(0),
        'appreciation_total': appreciation_total,
        'average_growth': average_growth,
        'highest_growth_rate': highest_growth_rate,
        'highest_growth_property': highest_growth_property,
    }
    
    return render(request, 'admin_side/client_profile.html', context)


@login_required
def user_profile(request):
    admin = request.user
    context = {'admin': admin}

    if request.method == 'POST':
        # Handle profile data update
        if 'update_profile' in request.POST:
            if update_profile_data(admin, request):
                messages.success(request, 'Your profile has been updated successfully!')

        # Handle password change
        elif 'change_password' in request.POST:
            password_response = change_password(request)
            if hasattr(password_response, 'success') and password_response.success:
                messages.success(request, 'Your password has been changed successfully!')
            else:
                messages.error(request, 'There was an error changing your password.')
            context.update({'password_response': password_response})

    return render(request, 'admin_side/admin_profile.html', context)


from django.db.models import Avg, Count, F, ExpressionWrapper, DecimalField
from django.db.models.functions import TruncMonth
from django.shortcuts import get_object_or_404
from django.views.generic import ListView, DetailView

class EstateListView(ListView):
    model = Estate
    template_name = "client_side/promo_estates_list.html"
    context_object_name = "estates"
    paginate_by = 12

    def get_queryset(self):
        # SECURITY: Filter estates by company
        company = self.request.user.company_profile
        qs = Estate.objects.filter(company=company).prefetch_related(
            "property_prices",
            "promotional_offers",
            "estate_plots__plotsizeunits"
        )
        qs = qs.order_by("-date_added")
        q = self.request.GET.get("q")
        if q:
            qs = qs.filter(name__icontains=q)
        return qs

    def get_plots_json(self, estate_id):
        # SECURITY: Verify company ownership before returning estate data
        company = self.request.user.company_profile
        try:
            estate = Estate.objects.filter(company=company).prefetch_related(
                Prefetch('estate_plots__plotsizeunits__plot_size'),
                Prefetch('promotional_offers'),
                Prefetch('property_prices', queryset=PropertyPrice.objects.select_related('plot_unit__plot_size'))
            ).get(pk=estate_id)
        except Estate.DoesNotExist:
            raise Http404("Estate not found")

        # Active promo (choose highest discount if multiple active)
        today = timezone.localdate()
        active_promos = estate.promotional_offers.filter(start__lte=today, end__gte=today)
        best_promo = active_promos.order_by('-discount').first()
        discount_pct = int(best_promo.discount) if best_promo else None

        sizes_out = []
        seen_sizes = set()

        for estate_plot in estate.estate_plots.all():
            for psu in estate_plot.plotsizeunits.all():
                try:
                    size_label = psu.plot_size.size
                except Exception:
                    try:
                        size_label = str(psu.plot_size)
                    except Exception:
                        size_label = None

                if size_label in seen_sizes:
                    continue
                seen_sizes.add(size_label)

                price_obj = estate.property_prices.filter(plot_unit=psu).first()
                if not price_obj:
                    price_obj = estate.property_prices.filter(
                        plot_unit__plot_size__size=str(size_label)
                    ).first()

                raw_amount = None
                if price_obj and getattr(price_obj, 'current', None) is not None:
                    try:
                        raw_amount = Decimal(price_obj.current)
                    except (InvalidOperation, TypeError):
                        raw_amount = None

                if raw_amount is None or raw_amount == Decimal('0'):
                    amount_value = None
                    amount_label = "NO AMOUNT SET"
                else:
                    amount_value = float(raw_amount.quantize(Decimal('0.01')))
                    if raw_amount == raw_amount.quantize(Decimal('1')):
                        amount_label = f"₦{int(raw_amount):,}"
                    else:
                        amount_label = f"₦{raw_amount:,}"

                discounted_value = None
                if amount_value is not None and discount_pct:
                    try:
                        discounted = (Decimal(amount_value) * (Decimal(100 - discount_pct) / Decimal(100))).quantize(Decimal('0.01'))
                        discounted_value = float(discounted)
                    except Exception:
                        discounted_value = None

                sizes_out.append({
                    "size": size_label,
                    "amount": amount_value,
                    "amount_label": amount_label,
                    "discounted": discounted_value,
                    "discount_pct": discount_pct,
                })

        if not sizes_out and estate.property_prices.exists():
            for pp in estate.property_prices.all():
                try:
                    size_label = pp.plot_unit.plot_size.size
                except Exception:
                    try:
                        size_label = str(pp.plot_unit.plot_size)
                    except Exception:
                        size_label = None

                if size_label in seen_sizes:
                    continue
                seen_sizes.add(size_label)

                raw_amount = None
                if getattr(pp, 'current', None) is not None:
                    try:
                        raw_amount = Decimal(pp.current)
                    except (InvalidOperation, TypeError):
                        raw_amount = None

                if raw_amount is None or raw_amount == Decimal('0'):
                    amount_value = None
                    amount_label = "NO AMOUNT SET"
                else:
                    amount_value = float(raw_amount.quantize(Decimal('0.01')))
                    if raw_amount == raw_amount.quantize(Decimal('1')):
                        amount_label = f"₦{int(raw_amount):,}"
                    else:
                        amount_label = f"₦{raw_amount:,}"

                discounted_value = None
                if amount_value is not None and discount_pct:
                    try:
                        discounted = (Decimal(amount_value) * (Decimal(100 - discount_pct) / Decimal(100))).quantize(Decimal('0.01'))
                        discounted_value = float(discounted)
                    except Exception:
                        discounted_value = None

                sizes_out.append({
                    "size": size_label,
                    "amount": amount_value,
                    "amount_label": amount_label,
                    "discounted": discounted_value,
                    "discount_pct": discount_pct,
                })

        response = {
            "estate_id": estate.id,
            "estate_name": estate.name,
            "promo": {
                "active": bool(best_promo),
                "name": best_promo.name if best_promo else None,
                "discount_pct": discount_pct
            },
            "sizes": sizes_out
        }
        return JsonResponse(response)

    
    def get(self, request, *args, **kwargs):
        """
        If the request has `estate_id` (or `format=json`) and looks like an XHR,
        return JSON for that estate's sizes/prices. Otherwise behave like ListView.
        """
        estate_id = request.GET.get('estate_id')
        fmt = request.GET.get('format')
        is_xhr = request.headers.get('x-requested-with') == 'XMLHttpRequest'

        if estate_id and (is_xhr or fmt == 'json'):
            return self.get_plots_json(estate_id)

        return super().get(request, *args, **kwargs)


class PromotionListView(ListView):
    model = PromotionalOffer
    template_name = "client_side/promotions_list.html"
    context_object_name = "promotions"
    paginate_by = 8

    def get_client_company_ids(self):
        """Get all company IDs the client is connected to"""
        from estateApp.models import CompanyClientProfile
        client_company_ids = set()
        
        # 1. Primary company
        if hasattr(self.request.user, 'company_profile') and self.request.user.company_profile:
            client_company_ids.add(self.request.user.company_profile.id)
        
        # 2. Affiliations
        affiliated_ids = CompanyClientProfile.objects.filter(
            client=self.request.user
        ).values_list('company_id', flat=True)
        client_company_ids.update(affiliated_ids)
        
        # 3. Allocations (companies from estates they've purchased)
        allocation_ids = Estate.all_objects.filter(
            plotallocation__client=self.request.user
        ).exclude(company__isnull=True).values_list('company_id', flat=True).distinct()
        client_company_ids.update(allocation_ids)
        
        return {cid for cid in client_company_ids if cid is not None}

    def get_queryset(self):
        # Get promotions from all companies the client is connected to
        client_company_ids = self.get_client_company_ids()
        today = timezone.localdate()
        
        if client_company_ids:
            qs = PromotionalOffer.objects.filter(
                estates__company__in=client_company_ids
            ).prefetch_related("estates", "estates__company").distinct().order_by("-created_at")
        else:
            # Fallback: show all promotions
            qs = PromotionalOffer.objects.prefetch_related("estates", "estates__company").order_by("-created_at")
        
        flt = self.request.GET.get('filter', '').lower()
        if flt == 'active':
            qs = qs.filter(start__lte=today, end__gte=today)
        elif flt == 'past':
            qs = qs.filter(end__lt=today)
        return qs
    
    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        today = timezone.localdate()
        client_company_ids = self.get_client_company_ids()
        
        if client_company_ids:
            active_promos = PromotionalOffer.objects.filter(
                estates__company__in=client_company_ids,
                start__lte=today, 
                end__gte=today
            ).prefetch_related("estates", "estates__company").distinct().order_by('-created_at')

            past_promos = PromotionalOffer.objects.filter(
                estates__company__in=client_company_ids,
                end__lt=today
            ).prefetch_related("estates", "estates__company").distinct().order_by('-end')
        else:
            active_promos = PromotionalOffer.objects.filter(
                start__lte=today, end__gte=today
            ).prefetch_related("estates", "estates__company").order_by('-created_at')
            
            past_promos = PromotionalOffer.objects.filter(
                end__lt=today
            ).prefetch_related("estates", "estates__company").order_by('-end')

        # Group promotions by company
        from collections import defaultdict
        promos_by_company = defaultdict(list)
        
        for promo in ctx.get('promotions', []):
            # Get the first estate's company
            first_estate = promo.estates.first()
            if first_estate and first_estate.company:
                company = first_estate.company
                promos_by_company[company.id].append({
                    'promo': promo,
                    'company_name': company.company_name,
                    'company_logo': company.logo.url if company.logo else None,
                    'company_id': company.id,
                })
            else:
                promos_by_company[0].append({
                    'promo': promo,
                    'company_name': 'General',
                    'company_logo': None,
                    'company_id': 0,
                })
        
        # Convert to list format for template
        grouped_promos = []
        companies = Company.objects.filter(id__in=client_company_ids) if client_company_ids else Company.objects.all()
        for company in companies:
            company_promos = promos_by_company.get(company.id, [])
            if company_promos:
                grouped_promos.append({
                    'company_id': company.id,
                    'company_name': company.company_name,
                    'company_logo': company.logo.url if company.logo else None,
                    'promos': [p['promo'] for p in company_promos],
                    'active_count': sum(1 for p in company_promos if p['promo'].start <= today <= p['promo'].end),
                })
        
        ctx['promos_by_company'] = grouped_promos
        ctx['active_promotions'] = active_promos[:6]
        ctx['past_promotions'] = past_promos[:12]
        ctx['current_filter'] = self.request.GET.get('filter', 'all').lower()
        ctx['total_active'] = active_promos.count()
        ctx['total_past'] = past_promos.count()
        
        return ctx


class PromotionDetailView(DetailView):
    model = PromotionalOffer
    template_name = "client_side/promo_detail.html"
    context_object_name = "promo"

    def get_queryset(self):
        return PromotionalOffer.objects.prefetch_related("estates")


def price_update_json(request, pk):
    try:
        ph = PriceHistory.objects.select_related(
            'price__estate', 'price__plot_unit__plot_size'
        ).get(pk=pk)
    except PriceHistory.DoesNotExist:
        return JsonResponse({"error": "Price update not found"}, status=404)

    percent_change = None
    try:
        prev = ph.previous
        curr = ph.current
        if prev not in (None, 0) and curr is not None:
            percent_change = float(((curr - prev) / prev) * 100)
    except Exception:
        percent_change = None

    data = {
        "id": ph.id,
        "estate_id": ph.price.estate.id if ph.price and ph.price.estate else None,
        "estate_name": ph.price.estate.name if ph.price and ph.price.estate else None,
        "size": getattr(ph.price.plot_unit.plot_size, 'size', None),
        "previous": float(ph.previous) if ph.previous is not None else None,
        "current": float(ph.current) if ph.current is not None else None,
        "percent_change": round(percent_change, 2) if percent_change is not None else None,
        "effective": ph.effective.isoformat() if ph.effective else None,
        "notes": ph.notes or "",
        "recorded_at": ph.recorded_at.isoformat() if ph.recorded_at else None,
    }
    return JsonResponse(data)


@login_required
def client_dashboard(request):
    allocations = PlotAllocation.objects.filter(client=request.user)
    total_properties = allocations.count()
    fully_paid_allocations = allocations.filter(payment_type="full").count()
    not_fully_paid_allocations = allocations.exclude(payment_type="full").count()

    client_estates = Estate.objects.filter(plotallocation__client=request.user).distinct()

    today_local = timezone.localdate()

    retention_days = getattr(settings, "PRICE_HISTORY_RETENTION_DAYS", 90)
    cutoff = today_local - timedelta(days=retention_days)

    all_updates = (
        PriceHistory.objects
        .select_related('price__estate', 'price__plot_unit__plot_size', 'price')
        .filter(effective__gte=cutoff)
        .order_by('-effective', '-recorded_at')
    )

    seen = set()
    latest_value = []
    for upd in all_updates:
        estate_id = getattr(upd.price, 'estate_id', None) if getattr(upd, 'price', None) else getattr(upd.price, 'estate_id', None)
        plot_unit_id = getattr(upd.price, 'plot_unit_id', None) if getattr(upd, 'price', None) else getattr(upd.price, 'plot_unit_id', None)
        key = f"{estate_id}_{plot_unit_id}"
        if key not in seen:
            seen.add(key)
            latest_value.append(upd)

    for upd in latest_value:
        try:
            live_current = None
            if getattr(upd, 'price', None) and getattr(upd.price, 'current', None) is not None:
                try:
                    live_current = Decimal(upd.price.current)
                except Exception:
                    live_current = None

            if live_current is not None:
                upd.current = live_current
            else:
                if getattr(upd, 'current', None) is not None:
                    try:
                        upd.current = Decimal(upd.current)
                    except Exception:
                        pass

            if getattr(upd, 'previous', None) is not None:
                try:
                    upd.previous = Decimal(upd.previous)
                except Exception:
                    pass

            if upd.previous not in (None, 0) and upd.current is not None:
                try:
                    upd.percent_change = float(((Decimal(upd.current) - Decimal(upd.previous)) / Decimal(upd.previous)) * 100)
                except Exception:
                    upd.percent_change = 0.0
            else:
                upd.percent_change = 0.0
        except Exception:
            upd.percent_change = 0.0

    latest_value = latest_value[:12]

    # Group latest_value by company for accordion display
    from collections import OrderedDict
    latest_value_by_company = OrderedDict()
    for upd in latest_value:
        company = None
        company_id = None
        company_name = "Other"
        company_logo = None
        if getattr(upd, 'price', None) and getattr(upd.price, 'estate', None):
            company = getattr(upd.price.estate, 'company', None)
            if company:
                company_id = company.id
                company_name = company.company_name
                company_logo = company.logo.url if company.logo else None
        
        if company_id not in latest_value_by_company:
            latest_value_by_company[company_id] = {
                'company_id': company_id,
                'company_name': company_name,
                'company_logo': company_logo,
                'updates': [],
                'has_new': False,  # Will be set if any update is within 7 days
            }
        
        # Check if this update is "new" (effective within 7 days)
        is_new = False
        if getattr(upd, 'effective', None):
            days_diff = (today_local - upd.effective).days
            if days_diff <= 7:
                is_new = True
                latest_value_by_company[company_id]['has_new'] = True
        upd.is_new = is_new
        
        latest_value_by_company[company_id]['updates'].append(upd)
    
    # Convert to list for template iteration
    latest_value_by_company_list = list(latest_value_by_company.values())

    estate_ids = {e.id for e in client_estates}
    estate_ids.update({getattr(u.price.estate, 'id') for u in latest_value if getattr(u.price, 'estate', None)})
    estate_ids = {eid for eid in estate_ids if eid is not None}

    best_promo_by_estate = {}
    if estate_ids:
        promo_qs = (
            PromotionalOffer.objects
            .filter(estates__in=estate_ids, start__lte=today_local, end__gte=today_local)
            .order_by('-discount')
            .prefetch_related('estates')
        )
        for promo in promo_qs:
            for est in promo.estates.all():
                if est.id in estate_ids and est.id not in best_promo_by_estate:
                    best_promo_by_estate[est.id] = promo

    for upd in latest_value:
        try:
            est_id = getattr(upd.price.estate, 'id', None)
            promo = best_promo_by_estate.get(est_id)
            if promo and upd.current is not None:
                current_dec = Decimal(upd.current)
                discount_pct = Decimal(promo.discount)
                upd.promo_price = (current_dec * (Decimal(100) - discount_pct) / Decimal(100)).quantize(Decimal('0.01'))
                upd.promo = {'id': promo.id, 'discount': int(promo.discount), 'name': promo.name}
            else:
                upd.promo_price = None
                upd.promo = None
        except Exception:
            upd.promo_price = None
            upd.promo = None

    base_qs = (
        PriceHistory.objects
        .select_related('price__estate', 'price__plot_unit__plot_size', 'price')
        .filter(price__estate__in=client_estates, effective__gte=cutoff)
        .order_by('-effective', '-recorded_at')
    )

    promo_estate_ids = (
        PromotionalOffer.objects
        .filter(estates__in=client_estates, start__lte=today_local, end__gte=today_local)
        .values_list('estates', flat=True)
        .distinct()
    )
    promo_estate_ids = list(promo_estate_ids)

    promo_qs = PriceHistory.objects.none()
    if promo_estate_ids:
        promo_qs = (
            PriceHistory.objects
            .select_related('price__estate', 'price__plot_unit__plot_size', 'price')
            .filter(price__estate__in=promo_estate_ids)
            .order_by('-effective', '-recorded_at')
        )

    promo_map_for_client = {}
    if client_estates.exists():
        promos_for_client = (
            PromotionalOffer.objects
            .filter(estates__in=client_estates, start__lte=today_local, end__gte=today_local)
            .order_by('-discount')
            .prefetch_related('estates')
        )
        for promo in promos_for_client:
            for est in promo.estates.all():
                if est.id not in promo_map_for_client:
                    promo_map_for_client[est.id] = promo

    seen = set()
    recent_value_updates = []
    recent_value_updates_ph_objects = []

    def append_ph(ph, forced_by_promo=False):
        estate_id = getattr(ph.price.estate, 'id', None)
        plot_unit_id = getattr(ph.price.plot_unit, 'id', None)
        key = f"{estate_id}_{plot_unit_id}"
        if key in seen:
            return
        seen.add(key)

        # prefer live price for display here as well
        try:
            live_current = None
            if getattr(ph, 'price', None) and getattr(ph.price, 'current', None) is not None:
                try:
                    live_current = Decimal(ph.price.current)
                except Exception:
                    live_current = None

            display_current = live_current if live_current is not None else (ph.current if ph.current is not None else None)
            display_prev = ph.previous if ph.previous is not None else None

            percent_change = None
            if display_prev not in (None, 0) and display_current is not None:
                try:
                    percent_change = float(((Decimal(display_current) - Decimal(display_prev)) / Decimal(display_prev)) * 100)
                except Exception:
                    percent_change = None
        except Exception:
            percent_change = None

        # status based on local today
        if ph.effective and ph.effective > today_local:
            status = "Yet Active"
        else:
            status = "Active"

        promo_obj = promo_map_for_client.get(estate_id) or best_promo_by_estate.get(estate_id)
        promo = None
        promo_price = None
        if promo_obj and ph.current is not None:
            try:
                current_dec = Decimal(ph.current)
                discount_pct = Decimal(promo_obj.discount)
                promo_price = (current_dec * (Decimal(100) - discount_pct) / Decimal(100)).quantize(Decimal('0.01'))
                promo = {
                    'id': promo_obj.id,
                    'name': promo_obj.name,
                    'discount': int(promo_obj.discount),
                    'start': promo_obj.start,
                    'end': promo_obj.end,
                }
            except Exception:
                promo_price = None
                promo = None

        recent_value_updates.append({
            'ph': ph,
            'percent_change': percent_change,
            'status': status,
            'promo': promo,
            'promo_price': promo_price,
            'promo_forced': bool(forced_by_promo),
        })
        recent_value_updates_ph_objects.append(ph)

    for ph in base_qs:
        append_ph(ph, forced_by_promo=False)
    for ph in promo_qs:
        append_ph(ph, forced_by_promo=True)

    # Get all companies the client is connected to (primary + affiliations)
    client_company_ids = set()
    
    # 1. Primary company from client's company_profile
    if hasattr(request.user, 'company_profile') and request.user.company_profile:
        client_company_ids.add(request.user.company_profile.id)
    
    # 2. Companies from CompanyClientProfile (affiliations)
    from estateApp.models import CompanyClientProfile
    affiliated_company_ids = CompanyClientProfile.objects.filter(
        client=request.user
    ).values_list('company_id', flat=True)
    client_company_ids.update(affiliated_company_ids)
    
    # 3. Companies from client's allocations (estates they've purchased from)
    allocation_company_ids = Estate.all_objects.filter(
        plotallocation__client=request.user
    ).exclude(company__isnull=True).values_list('company_id', flat=True).distinct()
    client_company_ids.update(allocation_company_ids)
    
    # Remove None if present
    client_company_ids = {cid for cid in client_company_ids if cid is not None}
    
    # Get active promotions from all connected companies
    if client_company_ids:
        active_promotions = PromotionalOffer.objects.filter(
            start__lte=today_local, 
            end__gte=today_local,
            estates__company__in=client_company_ids
        ).prefetch_related('estates', 'estates__company').distinct().order_by('-discount')[:6]
    else:
        # Fallback: show general active promotions
        active_promotions = PromotionalOffer.objects.filter(
            start__lte=today_local, end__gte=today_local
        ).prefetch_related('estates', 'estates__company').order_by('-discount')[:3]

    recent_transactions = PaymentRecord.objects.filter(
        transaction__allocation__client=request.user
    ).order_by('-payment_date')[:5]

    # value trend
    value_trend_data = []
    for i in range(6, -1, -1):
        month = timezone.now() - timedelta(days=30 * i)
        month_start = month.replace(day=1)
        avg_price = PriceHistory.objects.filter(
            price__estate__in=client_estates,
            effective__year=month_start.year,
            effective__month=month_start.month
        ).aggregate(avg_price=Avg('current'))['avg_price'] or 0
        value_trend_data.append({'date': month_start, 'value': float(avg_price)})

    context = {
        'total_properties': total_properties,
        'fully_paid_allocations': fully_paid_allocations,
        'not_fully_paid_allocations': not_fully_paid_allocations,
        'latest_value_updates': recent_value_updates_ph_objects,
        'recent_value_updates': recent_value_updates,
        'active_promotions': active_promotions,
        'recent_transactions': recent_transactions,
        'value_trend_data': value_trend_data,
        'client_estates': client_estates,
        'latest_value': latest_value,
        'latest_value_by_company': latest_value_by_company_list,
        'today_local': today_local,
    }

    return render(request, 'client_side/client_side.html', context)


@login_required
def my_client_profile(request):
    try:
        client = ClientUser.objects.select_related('assigned_marketer').get(id=request.user.id)
    except ClientUser.DoesNotExist:
        # Fallback: some client accounts are stored on the main CustomUser model
        # If the authenticated user has role 'client', use request.user as the client object
        user = request.user
        if getattr(user, 'role', None) == 'client':
            client = user
        else:
            messages.error(request, 'Client profile not found.')
            return redirect('login')

    # Use client_id so queries work regardless of ClientUser subclass presence
    client_id = getattr(client, 'id', client)
    transactions = Transaction.objects.filter(client_id=client_id).select_related(
        'allocation__estate',
        'allocation__plot_size'
    )

    metrics = calculate_portfolio_metrics(transactions)
    context = {'client': client, **metrics}

    if request.method == 'POST':
        action = request.POST.get('action')

        if not action:
            if request.POST.get('update_profile'):
                action = 'update_profile'
            elif request.POST.get('change_password'):
                action = 'change_password'

        if action == "update_profile":
            ok = update_profile_data(client, request)
            if not ok:
                messages.error(request, "Failed to update your profile.")
            return redirect(request.path + '#profile-edit')

        elif action == "change_password":
            password_response = change_password(request)

            if isinstance(password_response, dict):
                cd = password_response.get('context_data', {})
            else:
                cd = getattr(password_response, 'context_data', {})

            if cd.get('success'):
                messages.success(request, cd['success'])
            elif cd.get('error'):
                messages.error(request, cd['error'])
                
            return redirect(request.path + '#profile-change-password')

    return render(request, 'client_side/client_profile.html', context)


@login_required
def client_new_property_request(request):
    
    # SECURITY: Filter estates by company
    company = request.user.company_profile
    estates = Estate.objects.filter(company=company)
    context = {
        "estates": estates,
    }
    return render(request, 'client_side/new_property_request.html', context)


@login_required
def chat_view(request):
    """Client chat interface with company selection"""
    if not request.user.is_authenticated:
        return redirect('login')
    
    if request.user.role != 'client':
        messages.error(request, "Access denied.")
        return redirect('login')
    
    client = request.user
    
    # Get all companies where client has plot allocations
    companies = Company.objects.filter(
        estates__plotallocation__client=client
    ).distinct().order_by('company_name')
    
    # If no allocations found, fall back to user's direct company
    if not companies.exists():
        if client.company_profile:
            companies = Company.objects.filter(id=client.company_profile.id)
        else:
            companies = Company.objects.none()
    
    # Get selected company from query params or default to first company
    selected_company_id = request.GET.get('company_id')
    if selected_company_id:
        try:
            selected_company = companies.get(id=selected_company_id)
        except Company.DoesNotExist:
            messages.error(request, "Selected company not found.")
            selected_company = companies.first() if companies.exists() else None
    else:
        selected_company = companies.first() if companies.exists() else None
    
    # Handle message sending (POST request)
    if request.method == 'POST':
        # Support both 'content' and 'message_content' from frontend
        content = (request.POST.get('content') or request.POST.get('message_content') or '').strip()
        file = request.FILES.get('file')
        message_type = request.POST.get('message_type', 'enquiry')
        # Resolve company from POST to ensure correct scoping
        post_company_id = request.POST.get('company_id')
        if post_company_id:
            try:
                selected_company = companies.get(id=post_company_id)
            except Company.DoesNotExist:
                selected_company = None
        
        if not content and not file:
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                return JsonResponse({
                    'success': False,
                    'error': 'Please provide either a message or attach a file.'
                })
            messages.error(request, 'Please provide either a message or attach a file.')
            return redirect('chat')
        
        # Get company admin as recipient
        recipient = None
        if selected_company:
            recipient = CustomUser.objects.filter(
                role__in=['admin', 'support'],
                company_profile=selected_company
            ).first()
        
        if not recipient:
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                return JsonResponse({
                    'success': False,
                    'error': 'No admin available to receive messages.'
                })
            messages.error(request, 'No admin available to receive messages.')
            return redirect('chat')
        
        # Create the message
        message = Message.objects.create(
            sender=client,
            recipient=recipient,
            company=selected_company,
            content=content,
            file=file,
            message_type=message_type,
            status='sent'
        )
        
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            # Generate HTML for the new message
            message_html = render_to_string('client_side/chat_message.html', {
                'msg': message,
                'request': request,
            })
            
            return JsonResponse({
                'success': True,
                'message': {
                    'id': message.id,
                    'content': message.content,
                    'timestamp': message.date_sent.strftime('%Y-%m-%d %H:%M:%S'),
                    'sender': message.sender.username,
                    'recipient': message.recipient.username,
                    'message_type': message.message_type,
                    'file': message.file.url if message.file else None,
                },
                'message_html': message_html
            })
        
        messages.success(request, 'Message sent successfully.')
        return redirect('chat')
    
    # Handle GET request - check for history/polling
    
    # HISTORY branch: Full conversation load (when switching companies)
    if request.method == "GET" and 'history' in request.GET:
        # Always get ALL messages for this client (no company filter)
        all_messages = Message.objects.filter(
            Q(sender=client) | Q(recipient=client)
        ).select_related('sender', 'recipient', 'company').order_by('date_sent')
        
        # Generate HTML for all messages
        messages_html = ''
        for message in all_messages:
            messages_html += render_to_string('client_side/chat_message.html', {
                'msg': message,
                'request': request,
            })
        
        return JsonResponse({
            'messages_html': messages_html,
            'count': all_messages.count(),
        })
    
    # POLLING branch: if GET includes 'last_msg'
    if request.method == "GET" and 'last_msg' in request.GET:
        try:
            last_msg_id = int(request.GET.get('last_msg', 0))
        except (ValueError, TypeError):
            last_msg_id = 0
        
        # Always get ALL new messages for this client (no company filter)
        new_messages = Message.objects.filter(
            id__gt=last_msg_id
        ).filter(
            Q(sender=client) | Q(recipient=client)
        ).select_related('sender', 'recipient', 'company').order_by('date_sent')
        
        # Generate HTML for new messages
        messages_html = ''
        if new_messages.exists():
            for message in new_messages:
                messages_html += render_to_string('client_side/chat_message.html', {
                    'msg': message,
                    'request': request,
                })
        
        return JsonResponse({
            'messages': list(new_messages.values(
                'id', 'content', 'date_sent', 'sender__email', 
                'recipient__email', 'message_type', 'file', 'status'
            )),
            'messages_html': messages_html,
            'updated_statuses': []  # Could be implemented for read receipts
        })
    
    # Regular GET request - display chat interface
    # Always show ALL messages for this client (regardless of company)
    # This ensures admin replies are always visible
    messages_list = Message.objects.filter(
        Q(sender=client) | Q(recipient=client)
    ).select_related('sender', 'recipient', 'company').order_by('date_sent')
    
    context = {
        'companies': companies,
        'selected_company': selected_company,
        'client': client,
        'messages': messages_list,
    }
    
    return render(request, 'client_side/chat_interface.html', context)


@login_required
def view_all_requests(request):
    # Retrieve all property requests
    # SECURITY: Filter property requests by company
    company = request.user.company_profile
    property_requests = PropertyRequest.objects.filter(company=company)
    return render(request, 'client_side/requests_table.html', {'requests': property_requests})


def chat_dashboard_view(request):
    """
    Chat response tracking dashboard for management oversight.
    Shows SLA compliance and response metrics.
    """
    if not request.user.is_authenticated:
        return redirect('login')
    
    # Security validation
    if not validate_chat_access(request):
        messages.error(request, "Access denied. Please log in to continue.")
        return redirect('login')
    
    # Get user's company
    company = get_user_company(request.user)
    if not company:
        messages.error(request, "No company assigned. Please contact support.")
        return redirect('dashboard')
    
    # Check if user has permission to view dashboard
    if request.user.role not in ['admin', 'support']:
        messages.error(request, "You don't have permission to view this dashboard.")
        return redirect('chat')
    
    # Get tracking data for the last 30 days
    from datetime import date, timedelta
    end_date = date.today()
    start_date = end_date - timedelta(days=30)
    
    tracking_data = ChatResponseTracking.objects.filter(
        company=company,
        date__range=[start_date, end_date]
    ).order_by('-date')[:30]
    
    # Calculate summary statistics
    total_chats = sum(data.total_chats for data in tracking_data)
    responded_chats = sum(data.responded_chats for data in tracking_data)
    breached_chats = sum(data.breached_chats for data in tracking_data)
    
    avg_response_time = None
    if tracking_data:
        response_times = [data.average_response_time for data in tracking_data if data.average_response_time]
        if response_times:
            avg_response_time = sum(response_times, timedelta()) / len(response_times)
    
    # Get current SLA settings
    sla_settings = ChatSLA.objects.filter(company=company)
    
    context = {
        'company': company,
        'tracking_data': tracking_data,
        'total_chats': total_chats,
        'responded_chats': responded_chats,
        'breached_chats': breached_chats,
        'avg_response_time': avg_response_time,
        'sla_settings': sla_settings,
        'page_title': f'Chat Dashboard - {company.name}'
    }
    
    return render(request, 'admin_side/chat_dashboard.html', context)


@login_required
def property_list(request):
    # Fetch only the allocations that belong to the logged-in client.
    allocations = PlotAllocation.objects.filter(client=request.user).order_by('-date_allocated')
    context = {
        "allocations": allocations,
    }
    return render(request, 'client_side/property_list.html', context)


@login_required
def my_companies(request):
    """Show companies where the authenticated client has allocations/purchases."""
    user = request.user
    # Support cases where client is a subclass or base user
    client_id = getattr(user, 'id', user)

    # Get company ids from allocations
    company_ids = (
        PlotAllocation.objects.filter(client_id=client_id)
        .values_list('estate__company', flat=True)
        .distinct()
    )

    companies = Company.objects.filter(id__in=[c for c in company_ids if c is not None])

    # Build simple metrics per company
    company_list = []
    for comp in companies:
        alloc_count = PlotAllocation.objects.filter(client_id=client_id, estate__company=comp).count()
        total_invested = (
            Transaction.objects.filter(client_id=client_id, company=comp)
            .aggregate(total=Coalesce(Sum('total_amount'), Value(0, output_field=DecimalField())))['total']
        )
        # Attach metrics directly onto the Company instance so templates can use company.id, company.slug, etc.
        try:
            setattr(comp, 'allocations', alloc_count)
            setattr(comp, 'total_invested', total_invested)
        except Exception:
            # Fallback to dict structure if attribute assignment fails
            company_list.append({'company': comp, 'allocations': alloc_count, 'total_invested': total_invested})
            continue

        company_list.append(comp)

    return render(request, 'client_side/client_company/my_companies.html', {'companies': company_list})


@login_required
def my_company_portfolio(request, company_id=None, company_slug=None):
    """Show allocations and transactions for this client scoped to a specific company.

    Accepts either `company_id` or `company_slug` (one must be provided).
    """
    from estateApp.models import ClientMarketerAssignment
    
    user = request.user
    client_id = getattr(user, 'id', user)

    if company_slug and not company_id:
        company = get_object_or_404(Company, slug=company_slug)
    else:
        company = get_object_or_404(Company, id=company_id)

    # SECURITY: Verify client has allocations in this company (prevent unauthorized access)
    if not PlotAllocation.objects.filter(client_id=client_id, estate__company=company).exists():
        from django.http import Http404
        raise Http404("You don't have any allocations in this company.")

    # Annotate each allocation with the latest transaction id (if any) for robust client-side wiring
    latest_tx_subq = (
        Transaction.objects.filter(allocation=OuterRef('pk'))
        .order_by('-transaction_date')
        .values('id')[:1]
    )

    allocations = (
        PlotAllocation.objects.filter(client_id=client_id, estate__company=company)
        .select_related('estate', 'plot_size', 'plot_number')
        .annotate(
            total_paid=Coalesce(Sum('transactions__total_amount'), Value(0, output_field=DecimalField())),
            latest_tx_id=Subquery(latest_tx_subq),
            # Get the latest known current price for this estate+plot_size via PropertyPrice
            latest_price_current=Subquery(
                PropertyPrice.objects.filter(
                    estate=OuterRef('estate_id'),
                    plot_unit__plot_size=OuterRef('plot_size_id'),
                    company=company
                ).order_by('-effective').values('current')[:1],
                output_field=DecimalField()
            )
        )
        .order_by('-date_allocated')
    )

    transactions = (
        Transaction.objects.filter(client_id=client_id, company=company)
        .select_related('allocation__estate', 'allocation__plot_size')
        .order_by('-transaction_date')
    )

    total_invested = transactions.aggregate(total=Coalesce(Sum('total_amount'), Value(0, output_field=DecimalField())))['total']

    # Compute portfolio metrics (current_value, appreciation, growth rates) using existing helper
    try:
        metrics = calculate_portfolio_metrics(transactions)
        # calculate_portfolio_metrics attaches `current_value`, `appreciation`, `growth_rate` attrs
        transactions = metrics.get('transactions', transactions)
        appreciation_total = metrics.get('appreciation_total', Decimal(0))
        average_growth = metrics.get('average_growth', 0)
        highest_growth_rate = metrics.get('highest_growth_rate', Decimal(0))
        highest_growth_property = metrics.get('highest_growth_property', '')
    except Exception:
        # Fallbacks if any unexpected error occurs while computing metrics
        appreciation_total = Decimal(0)
        average_growth = 0
        highest_growth_rate = Decimal(0)
        highest_growth_property = ''
    # Build a quick map of allocation_id -> latest transaction (transactions ordered by -transaction_date)
    alloc_tx_map = {}
    for tx in transactions:
        alloc_id = getattr(tx.allocation, 'id', None)
        if alloc_id and alloc_id not in alloc_tx_map:
            alloc_tx_map[alloc_id] = tx

    # Attach latest_transaction attribute to each allocation for template convenience
    for alloc in allocations:
        try:
            setattr(alloc, 'latest_transaction', alloc_tx_map.get(alloc.id))
        except Exception:
            # If allocation is not a model instance or other error, skip
            pass

    # Number of estates in this company
    estates_count = Estate.objects.filter(company=company).count()


    # Calculate total current value (sum of latest_price_current for all allocations, fallback to latest_transaction.current_value)
    total_current_value = 0
    for alloc in allocations:
        if getattr(alloc, 'latest_price_current', None):
            total_current_value += alloc.latest_price_current or 0
        elif getattr(alloc, 'latest_transaction', None) and getattr(alloc.latest_transaction, 'current_value', None):
            total_current_value += alloc.latest_transaction.current_value or 0

    total_value_increased = total_current_value - (total_invested or 0)

    # Get the client's assigned marketer for this specific company (company-scoped)
    # SECURITY: Only fetch marketer if client has this company assignment
    try:
        assignment = ClientMarketerAssignment.objects.filter(
            client_id=client_id,
            company=company
        ).select_related('marketer').first()
        marketer = assignment.marketer if assignment else None
    except Exception:
        marketer = None

    context = {
        'company': company,
        'allocations': allocations,
        'transactions': transactions,
        'total_invested': total_invested,
        'total_current_value': total_current_value,
        'total_value_increased': total_value_increased,
        'estates_count': estates_count,
        'appreciation_total': appreciation_total,
        'average_growth': average_growth,
        'highest_growth_rate': highest_growth_rate,
        'highest_growth_property': highest_growth_property,
        'marketer': marketer,
    }
    return render(request, 'client_side/client_company/my_company_portfolio.html', context)


@login_required
def marketer_my_companies(request):
    """Show companies where the authenticated marketer has affiliations, profiles, or transactions.
    SECURITY: Strict company isolation prevents data leakage between companies.
    """
    user = request.user
    if getattr(user, 'role', None) != 'marketer':
        return redirect('login')
    
    # 1. Companies where marketer has transactions
    transaction_company_ids = set(
        Transaction.objects.filter(marketer=user)
        .values_list('company', flat=True)
        .distinct()
    )
    
    # 2. Companies where marketer has affiliations (MarketerAffiliation)
    affiliation_company_ids = set(
        MarketerAffiliation.objects.filter(marketer=user)
        .values_list('company_id', flat=True)
        .distinct()
    )
    
    # 3. Companies where marketer has CompanyMarketerProfile records
    profile_company_ids = set(
        CompanyMarketerProfile.objects.filter(marketer=user)
        .values_list('company_id', flat=True)
        .distinct()
    )
    
    # 4. The marketer's own company_profile (primary company)
    own_company_id = user.company_profile_id if hasattr(user, 'company_profile_id') and user.company_profile_id else None
    
    # Combine all company IDs from all sources
    all_company_ids = transaction_company_ids | affiliation_company_ids | profile_company_ids
    if own_company_id:
        all_company_ids.add(own_company_id)
    
    # Remove None values
    all_company_ids = [c for c in all_company_ids if c is not None]

    companies = Company.objects.filter(id__in=all_company_ids)

    company_list = []
    current_year = datetime.now().year
    year_str = str(current_year)
    
    for comp in companies:
        # Count ALL transactions FOR THIS COMPANY ONLY - STRICT isolation
        # This matches the logic from section3_marketers_performance.html
        closed_deals = Transaction.objects.filter(
            marketer=user, 
            company=comp  # ✅ Ensures data is scoped to this company only
        ).count()
        
        # Get company-specific commission rate from MarketerCommission (COMPANY-SCOPED)
        # Use the same logic as PerformanceDataAPI for consistency
        today = datetime.now().date()
        
        # Get the most recent commission rate for this marketer in this specific company
        # COMMISSIONS ARE STRICTLY COMPANY-BASED AND MARKETER-SPECIFIC - NO GLOBAL COMMISSIONS
        commission = (
            MarketerCommission.objects
            .filter(
                marketer=user,  # ✅ Only marketer-specific commissions
                company=comp,   # ✅ Company-scoped
                effective_date__lte=today
            )
            .order_by('-effective_date')
            .first()
        )
        
        # Only use marketer-specific commissions - NO GLOBAL COMMISSION FALLBACK
        commission_rate = commission.rate if commission and commission.rate > 0 else None
        
        # Calculate yearly target achievement for this company
        # Use the same logic as PerformanceDataAPI for consistency
        yearly_target_achievement = None
        total_year_sales = 0
        
        # Get the target for this marketer in this company (marketer-specific OR global)
        specific_tgt = (
            MarketerTarget.objects.filter(
                marketer=user,
                company=comp,
                period_type='annual',
                specific_period=year_str
            ).first()
        )
        
        global_tgt = (
            MarketerTarget.objects.filter(
                marketer=None,
                company=comp,
                period_type='annual',
                specific_period=year_str
            ).first()
        )
        
        # Use marketer-specific if available, otherwise use global target
        yearly_target = specific_tgt or global_tgt
        
        if yearly_target and yearly_target.target_amount:
            total_year_sales = Transaction.objects.filter(
                marketer=user,
                company=comp,  # <-- STRICT COMPANY FILTER
                transaction_date__year=current_year
            ).aggregate(total=Sum('total_amount'))['total'] or 0
            
            # Calculate achievement percentage (even if sales are 0)
            if yearly_target.target_amount > 0:
                yearly_target_achievement = min(
                    100,
                    (total_year_sales / yearly_target.target_amount) * 100
                )
        
        # Get target information for all periods
        today = datetime.now().date()
        current_month = today.strftime('%Y-%m')
        current_quarter = f"{today.year}-Q{math.ceil(today.month / 3)}"
        current_year_str = str(current_year)
        
        # Get monthly target
        monthly_target = (
            MarketerTarget.objects.filter(
                marketer=user,
                company=comp,
                period_type='monthly',
                specific_period=current_month
            ).first() or
            MarketerTarget.objects.filter(
                marketer=None,
                company=comp,
                period_type='monthly',
                specific_period=current_month
            ).first()
        )
        
        # Get quarterly target
        quarterly_target = (
            MarketerTarget.objects.filter(
                marketer=user,
                company=comp,
                period_type='quarterly',
                specific_period=current_quarter
            ).first() or
            MarketerTarget.objects.filter(
                marketer=None,
                company=comp,
                period_type='quarterly',
                specific_period=current_quarter
            ).first()
        )
        
        # Calculate monthly sales
        monthly_sales = Transaction.objects.filter(
            marketer=user,
            company=comp,
            transaction_date__year=today.year,
            transaction_date__month=today.month
        ).aggregate(total=Sum('total_amount'))['total'] or 0
        
        # Calculate quarterly sales
        quarter_start_month = (math.ceil(today.month / 3) - 1) * 3 + 1
        quarterly_sales = Transaction.objects.filter(
            marketer=user,
            company=comp,
            transaction_date__year=today.year,
            transaction_date__month__gte=quarter_start_month,
            transaction_date__month__lte=quarter_start_month + 2
        ).aggregate(total=Sum('total_amount'))['total'] or 0
        
        # Calculate target achievements
        monthly_achievement = None
        quarterly_achievement = None
        
        if monthly_target and monthly_target.target_amount > 0:
            monthly_achievement = min(100, (monthly_sales / monthly_target.target_amount) * 100)
            
        if quarterly_target and quarterly_target.target_amount > 0:
            quarterly_achievement = min(100, (quarterly_sales / quarterly_target.target_amount) * 100)
        
        # Get client count for this company (clients the marketer has dealt with)
        client_ids = Transaction.objects.filter(
            marketer=user,
            company=comp
        ).values_list('client_id', flat=True).distinct()
        client_count = len([c for c in client_ids if c is not None])
        
        # Check if marketer has any transactions with this company
        has_transactions = closed_deals > 0
        
        company_list.append({
            'company': comp, 
            'closed_deals': closed_deals, 
            'commission_rate': commission_rate,
            'yearly_target_achievement': yearly_target_achievement,
            'yearly_target': yearly_target,
            'total_year_sales': total_year_sales,
            'monthly_target': monthly_target,
            'quarterly_target': quarterly_target,
            'monthly_sales': monthly_sales,
            'quarterly_sales': quarterly_sales,
            'monthly_achievement': monthly_achievement,
            'quarterly_achievement': quarterly_achievement,
            'current_month': current_month,
            'current_quarter': current_quarter,
            'client_count': client_count,
            'has_transactions': has_transactions,
        })

    return render(request, 'marketer_side/my_companies.html', {
        'companies': company_list,
        'current_year': current_year
    })


@login_required
def marketer_company_portfolio(request, company_id=None):
    """Show marketer's portfolio for a specific company: their transactions and list of clients they handle within that company."""
    user = request.user
    if getattr(user, 'role', None) != 'marketer':
        return redirect('login')

    company = get_object_or_404(Company, id=company_id)
    
    # SECURITY: Check if marketer is affiliated with this company
    # Marketer can access if: 1) has transactions, 2) has affiliation, 3) has profile, 4) is their company_profile
    has_transactions = Transaction.objects.filter(marketer=user, company=company).exists()
    has_affiliation = MarketerAffiliation.objects.filter(marketer=user, company=company).exists()
    has_profile = CompanyMarketerProfile.objects.filter(marketer=user, company=company).exists()
    is_own_company = (hasattr(user, 'company_profile_id') and user.company_profile_id == company.id)
    
    if not (has_transactions or has_affiliation or has_profile or is_own_company):
        messages.error(request, "You don't have access to this company's portfolio.")
        return redirect('marketer-my-companies')
    
    current_year = timezone.now().year
    year_str = str(current_year)

    # Transactions by this marketer for this company
    transactions = (
        Transaction.objects.filter(marketer=user, company=company)
        .select_related('client', 'allocation__estate', 'allocation__plot_size')
        .order_by('-transaction_date')
    )

    # Get distinct client IDs from the transactions
    client_ids = transactions.values_list('client_id', flat=True).distinct()
    client_ids = [c for c in client_ids if c is not None]

    # Use ClientUser with prefetch_related for proper transaction grouping
    # We need to filter the prefetched transactions by company and marketer
    # Use 'company_transactions' to avoid conflict with the existing 'transactions' related field
    clients = (
        ClientUser.objects
        .filter(id__in=client_ids)
        .prefetch_related(
            Prefetch(
                'transactions',
                queryset=Transaction.objects.filter(
                    marketer=user,
                    company=company
                ).select_related(
                    'allocation__estate',
                    'allocation__plot_size',
                    'allocation__plot_number'
                ).order_by('-transaction_date'),
                to_attr='company_transactions'
            )
        )
    )

    # Calculate total transaction value
    total_value = transactions.aggregate(total=Coalesce(Sum('total_amount'), Value(0, output_field=DecimalField())))['total']
    
    # Calculate commission earned based on marketer's commission rate for this company
    commission_rate = Decimal('0')
    try:
        commission_obj = MarketerCommission.objects.filter(
            marketer=user,
            company=company
        ).order_by('-effective_date').first()
        if commission_obj and commission_obj.rate:
            commission_rate = Decimal(str(commission_obj.rate))
    except Exception:
        pass
    
    # Commission = total_value * (commission_rate / 100)
    commission_earned = (total_value * commission_rate / Decimal('100')) if commission_rate > 0 else Decimal('0')

    # === PERFORMANCE DATA FOR THIS COMPANY ===
    # Closed deals count
    closed_deals = transactions.count()
    
    # Yearly sales for this company
    yearly_sales = Transaction.objects.filter(
        marketer=user,
        company=company,
        transaction_date__year=current_year
    ).aggregate(total=Coalesce(Sum('total_amount'), Value(0, output_field=DecimalField())))['total']
    
    # Get yearly target for this company
    yearly_target = (
        MarketerTarget.objects.filter(
            marketer=user, company=company, period_type='annual', specific_period=year_str
        ).first()
        or
        MarketerTarget.objects.filter(
            marketer=None, company=company, period_type='annual', specific_period=year_str
        ).first()
    )
    
    yearly_target_amount = yearly_target.target_amount if yearly_target else None
    yearly_target_achievement = None
    if yearly_target_amount and yearly_target_amount > 0:
        yearly_target_achievement = (yearly_sales / yearly_target_amount) * 100

    # === LEADERBOARD FOR THIS COMPANY ===
    # Get all marketers affiliated with this company
    company_marketers = list(MarketerUser.objects.filter(company_profile=company))
    affiliation_marketer_ids = list(
        MarketerAffiliation.objects.filter(company=company).values_list('marketer_id', flat=True).distinct()
    )
    affiliation_marketers = list(MarketerUser.objects.filter(id__in=affiliation_marketer_ids).exclude(
        id__in=[m.id for m in company_marketers]
    ))
    all_marketers = company_marketers + affiliation_marketers
    
    # Build sales data for leaderboard
    sales_data = []
    for m in all_marketers:
        m_year_sales = Transaction.objects.filter(
            marketer=m,
            company=company,
            transaction_date__year=current_year
        ).aggregate(total=Sum('total_amount'))['total'] or 0

        m_target = (
            MarketerTarget.objects.filter(marketer=m, company=company, period_type='annual', specific_period=year_str).first()
            or
            MarketerTarget.objects.filter(marketer=None, company=company, period_type='annual', specific_period=year_str).first()
        )
        target_amt = m_target.target_amount if m_target else None
        pct = (m_year_sales / target_amt * 100) if target_amt else None

        sales_data.append({'marketer': m, 'total_sales': m_year_sales, 'target_amt': target_amt, 'pct': pct})

    sales_data.sort(key=lambda x: x['total_sales'], reverse=True)

    # Build top 3
    top3 = []
    for idx, e in enumerate(sales_data[:3], start=1):
        pct = e['pct']
        if pct is None:
            category = diff = None
        elif pct > 100:
            category, diff = 'Above Target', round(pct - 100, 1)
        elif pct >= 50:
            category, diff = 'On Target', round(pct, 1)
        else:
            category, diff = 'Below Target', round(100 - pct, 1)

        top3.append({
            'rank': idx,
            'marketer': e['marketer'],
            'category': category,
            'diff_pct': diff,
            'has_target': e['target_amt'] is not None,
        })

    # Find current user's entry
    user_entry = None
    for idx, e in enumerate(sales_data, start=1):
        if e['marketer'].id == user.id:
            pct = e['pct']
            if pct is None:
                category = diff = None
            elif pct > 100:
                category, diff = 'Above Target', round(pct - 100, 1)
            elif pct >= 50:
                category, diff = 'On Target', round(pct, 1)
            else:
                category, diff = 'Below Target', round(100 - pct, 1)

            user_entry = {
                'rank': idx,
                'marketer': user,
                'category': category,
                'diff_pct': diff,
                'has_target': e['target_amt'] is not None,
            }
            break

    context = {
        'company': company,
        'transactions': transactions,
        'clients': clients,
        'total_value': total_value,
        'commission_earned': commission_earned,
        'commission_rate': commission_rate,
        # Performance data
        'closed_deals': closed_deals,
        'yearly_sales': yearly_sales,
        'yearly_target': yearly_target,
        'yearly_target_amount': yearly_target_amount,
        'yearly_target_achievement': yearly_target_achievement,
        'current_year': current_year,
        # Leaderboard data
        'top3': top3,
        'user_entry': user_entry,
    }
    return render(request, 'marketer_side/my_company_portfolio.html', context)

@login_required
def view_client_estate(request, estate_id, plot_size_id):
    estate = get_object_or_404(Estate, id=estate_id)

    plot_size = get_object_or_404(PlotSize, id=plot_size_id)
    
    # Get floor plans for this specific plot size in the estate
    floor_plans = EstateFloorPlan.objects.filter(
        estate=estate, 
        plot_size=plot_size
    )
    
    context = {
        'estate': estate,
        'plot_size': plot_size,
        'floor_plans': floor_plans,
    }
    return render(request, 'client_side/client_estate_detail.html', context)


# XXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXX
# MARKETER SIDE
# XXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXX

@login_required
def marketer_dashboard(request):
    user = request.user

    # 1) Totals
    total_transactions = Transaction.objects.filter(marketer=user, company=user.company_profile).count()
    total_estates_sold = Transaction.objects.filter(marketer=user, allocation__payment_type='full', company=user.company_profile).count()
    
    # Count clients assigned to this marketer - include both direct assignments and ClientMarketerAssignment records
    direct_clients = ClientUser.objects.filter(assigned_marketer=user, company_profile=user.company_profile).count()
    assigned_clients = ClientMarketerAssignment.objects.filter(marketer=user, company=user.company_profile).values_list('client_id', flat=True).distinct().count()
    number_clients = direct_clients + (assigned_clients if assigned_clients > 0 else 0)

    # Helper to build a list of (label, transaction_count, estate_count, new_client_count)
    def build_series(start, step, buckets, date_field='transaction_date'):
        labels = []
        tx_counts = []
        est_counts = []
        cli_counts = []

        current = start
        for _ in range(buckets):
            # label for this bucket:
            labels.append(current.strftime(step['fmt']))

            # window start/end
            window_start = current
            window_end = current + step['delta']

            # transactions in window
            tx_qs = Transaction.objects.filter(
                marketer=user,
                company=user.company_profile,
                **{f"{date_field}__gte": window_start},
                **{f"{date_field}__lt": window_end}
            )
            tx_counts.append(tx_qs.count())

            # full-payment estates sold
            est_qs = tx_qs.filter(allocation__payment_type='full')
            est_counts.append(est_qs.count())

            # new clients assigned in this window
            cli_qs = ClientUser.objects.filter(
                assigned_marketer=user,
                company_profile=user.company_profile,
                date_registered__gte=window_start,
                date_registered__lt=window_end
            )
            cli_counts.append(cli_qs.count())

            current = window_end

        return labels, tx_counts, est_counts, cli_counts

    today = date.today()

    # Weekly: last 7 days
    weekly_start = today - relativedelta(days=6)
    weekly_step = { 'delta': relativedelta(days=1), 'fmt': '%d %b' }
    weekly_labels, weekly_tx, weekly_est, weekly_cli = build_series(weekly_start, weekly_step, 7)

    # Monthly: last 6 months
    monthly_start = (today - relativedelta(months=5)).replace(day=1)
    monthly_step = { 'delta': relativedelta(months=1), 'fmt': '%b %Y' }
    monthly_labels, monthly_tx, monthly_est, monthly_cli = build_series(monthly_start, monthly_step, 6)

    # Yearly: last 5 years
    yearly_start = today.replace(month=1, day=1) - relativedelta(years=4)
    yearly_step = { 'delta': relativedelta(years=1), 'fmt': '%Y' }
    yearly_labels, yearly_tx, yearly_est, yearly_cli = build_series(yearly_start, yearly_step, 5)

    # All-Time: monthly buckets from first transaction month until now
    first_tx = Transaction.objects.filter(marketer=user, company=user.company_profile).order_by('transaction_date').first()
    if first_tx:
        first_month = first_tx.transaction_date.replace(day=1)
    else:
        first_month = today.replace(day=1)
    months = (today.year - first_month.year) * 12 + (today.month - first_month.month) + 1
    all_step = { 'delta': relativedelta(months=1), 'fmt': '%b %Y' }
    all_labels, all_tx, all_est, all_cli = build_series(first_month, all_step, months)

    return render(request, 'marketer_side/marketer_side.html', {
        'total_transactions': total_transactions,
        'total_estates_sold': total_estates_sold,
        'number_clients': number_clients,
        'weekly': {
            'labels': weekly_labels, 'tx': weekly_tx,
            'est': weekly_est, 'cli': weekly_cli
        },
        'monthly': {
            'labels': monthly_labels, 'tx': monthly_tx,
            'est': monthly_est, 'cli': monthly_cli
        },
        'yearly': {
            'labels': yearly_labels, 'tx': yearly_tx,
            'est': yearly_est, 'cli': yearly_cli
        },
        'alltime': {
            'labels': all_labels, 'tx': all_tx,
            'est': all_est, 'cli': all_cli
        }
    })

@login_required
def marketer_profile(request):
    marketer      = request.user
    # Ensure company context is always defined to avoid UnboundLocalError
    company = getattr(request.user, 'company_profile', None)
    now           = timezone.now()
    current_year  = now.year
    year_str      = str(current_year)
    current_month = now.strftime("%Y-%m")
    password_response = None


    lifetime_closed_deals = Transaction.objects.filter(
        marketer=marketer,
        company=marketer.company_profile
    ).count()

    lifetime_commission = MarketerPerformanceRecord.objects.filter(
        marketer=marketer,
        company=company,  # ✅ Added company filter for isolation
        period_type='monthly'
    ).aggregate(total=Sum('commission_earned'))['total'] or 0


    performance = {
        'closed_deals':      lifetime_closed_deals,
        'total_sales':       0,
        'commission_earned': lifetime_commission,
        'commission_rate':   0,
        'target_achievement': 0,
        'yearly_target_achievement': None,
    }

    # Latest commission rate (STRICT company filter)
    comm = MarketerCommission.objects.filter(
        marketer=marketer,
        company=company  # ✅ Added company filter for isolation
    ).order_by('-effective_date').first()
    if comm:
        performance['commission_rate'] = comm.rate

    # Monthly target % (STRICT company filter)
    mt = MarketerTarget.objects.filter(
        marketer=marketer,
        company=company,  # ✅ Added company filter for isolation
        period_type='monthly',
        specific_period=current_month
    ).first()
    if mt and mt.target_amount:
        performance['target_achievement'] = min(
            100,
            performance['total_sales'] / mt.target_amount * 100
        )

    # Annual target achievement (STRICT company filter)
    at = (
        MarketerTarget.objects.filter(
            marketer=marketer, 
            company=company,  # ✅ Added company filter for isolation
            period_type='annual', 
            specific_period=year_str
        )
        .first()
        or
        MarketerTarget.objects.filter(
            marketer=None, 
            company=company,  # ✅ Added company filter for isolation
            period_type='annual', 
            specific_period=year_str
        )
        .first()
    )
    if at and at.target_amount:
        total_year_sales = Transaction.objects.filter(
            marketer=marketer,
            company=company,
            transaction_date__year=current_year
        ).aggregate(total=Sum('total_amount'))['total'] or 0
        performance['yearly_target_achievement'] = min(
            100,
            total_year_sales / at.target_amount * 100
        )

    # — Build leaderboard —
    sales_data = []
    # SECURITY: Filter by company to prevent cross-tenant leakage  
    company = request.user.company_profile
    
    # Get marketers from both company_profile and MarketerAffiliation
    company_marketers = MarketerUser.objects.filter(company_profile=company) if company else MarketerUser.objects.none()
    affiliation_marketer_ids = list(
        MarketerAffiliation.objects.filter(company=company).values_list('marketer_id', flat=True).distinct()
    ) if company else []
    affiliation_marketers = MarketerUser.objects.filter(id__in=affiliation_marketer_ids).exclude(
        id__in=company_marketers.values_list('pk', flat=True)
    ) if affiliation_marketer_ids else MarketerUser.objects.none()
    
    company_marketers = list(company_marketers) + list(affiliation_marketers)
    
    for m in company_marketers:
        year_sales = Transaction.objects.filter(
            marketer=m,
            company=company,
            transaction_date__year=current_year
        ).aggregate(total=Sum('total_amount'))['total'] or 0

        tgt = (
            MarketerTarget.objects.filter(marketer=m, company=company, period_type='annual', specific_period=year_str).first()
            or
            MarketerTarget.objects.filter(marketer=None, company=company, period_type='annual', specific_period=year_str).first()
        )
        target_amt = tgt.target_amount if tgt else None
        pct = (year_sales / target_amt * 100) if target_amt else None

        sales_data.append({'marketer': m, 'total_sales': year_sales, 'target_amt': target_amt, 'pct': pct})

    sales_data.sort(key=lambda x: x['total_sales'], reverse=True)

    top3 = []
    for idx, e in enumerate(sales_data[:3], start=1):
        pct = e['pct']
        if pct is None:
            category = diff = None
        elif pct > 100:
            category, diff = 'Above Target', round(pct - 100, 1)
        elif pct >= 50:
            category, diff = 'On Target', round(pct, 1)
        else:
            category, diff = 'Below Target', round(100 - pct, 1)

        top3.append({
            'rank': idx,
            'marketer': e['marketer'],
            'category': category,
            'diff_pct': diff,
            'has_target': e['target_amt'] is not None,
        })

    user_entry = None
    for idx, e in enumerate(sales_data, start=1):
        if e['marketer'].id == marketer.id:
            pct = e['pct']
            if pct is None:
                category = diff = None
            elif pct > 100:
                category, diff = 'Above Target', round(pct - 100, 1)
            elif pct >= 50:
                category, diff = 'On Target', round(pct, 1)
            else:
                category, diff = 'Below Target', round(100 - pct, 1)

            user_entry = {
                'rank': idx,
                'marketer': marketer,
                'category': category,
                'diff_pct': diff,
                'has_target': e['target_amt'] is not None,
            }
            break

    if request.method == 'POST':
        # Profile update
        if request.POST.get("update_profile"):
            ok = update_profile_data(marketer, request)
            if ok:
                messages.success(request, "Your profile has been updated successfully!")
            else:
                messages.error(request, "Failed to update your profile.")

        # Password change
        elif request.POST.get('change_password'):
            password_response = change_password(request)
            cd = password_response.context_data
            if cd.get('success'):
                messages.success(request, cd['success'])
            elif cd.get('error'):
                messages.error(request, cd['error'])

        return redirect('marketer-profile')



    return render(request, 'marketer_side/marketer_profile.html', {
        'password_response': password_response,
        'performance': performance,
        'top3':        top3,
        'user_entry':  user_entry,
        'current_year': current_year,
    })

@login_required
def client_records(request):
    marketer = request.user
    
    clients = (
        ClientUser.objects
        .filter(assigned_marketer=marketer)
        .prefetch_related(
            'transactions__allocation__estate',
            'transactions__allocation__plot_size',
            'transactions__allocation__plot_number',
        )
    )

    return render(request, 'marketer_side/client_records.html', {
        'clients': clients,
    })


@login_required
def marketer_notification(request):
    return render(request, 'marketer_side/notification.html')

@login_required
def client_notification(request):
    return render(request, 'client_side/notification.html')


# XXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXX
# OTHER CONFIGURATIONS
# XXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXX

@login_required
def dashboard(request):
    user = request.user
    if user.role == 'admin':
        return render(request, 'admin_side/index.html')
    elif user.role == 'client':
        return render(request, 'client_side/client_side.html')
    elif user.role == 'marketer':
        return render(request, 'marketer_side/marketer_side.html')
    elif user.role == 'support':
        return render(request, 'adminSide/birthday_message.html')
    
    return redirect('login')


@login_required
def user_profile(request):
    user = request.user
    if user.role == 'admin':
        return render(request, 'admin_side/admin_profile.html')
    elif user.role == 'client':
        return render(request, 'client_side/client_profile.html')
    elif user.role == 'marketer':
        return render(request, 'marketer_side/marketer_profile.html')
    
    return redirect('dashboard')


class CustomLoginView(LoginView):
    """
    Unified Login View for all user types:
    - Company Admin (primary and secondary admins)
    - Client
    - Marketer
    - AdminSupport
    
    Security Features:
    - Honeypot field for bot protection
    - Dynamic tenant slug support for multi-tenant isolation
    - Client IP tracking and optional GeoIP lookup
    - Rate limiting hooks (server-side enforcement recommended)
    - CSRF protection (built-in Django)
    """
    form_class = CustomAuthenticationForm
    template_name = 'auth/unified_login.html'

    def get_context_data(self, **kwargs):
        """
        Inject security-related context into template:
        - honeypot_field: name of hidden honeypot field for bot protection
        - login_slug: optional dynamic tenant slug from URL kwargs
        """
        context = super().get_context_data(**kwargs)
        # Provide honeypot field name; use settings value or default to 'honeypot'
        context['honeypot_field'] = getattr(settings, 'HONEYPOT_FIELD_NAME', 'honeypot')
        # Provide dynamic slug if present (from URL kwargs)
        context['login_slug'] = self.kwargs.get('login_slug', None)
        return context

    def post(self, request, *args, **kwargs):
        # Check if this is a role selection POST request
        if 'selected_user_id' in request.POST:
            return self.handle_role_selection(request)
        
        # Normal login form processing
        return super().post(request, *args, **kwargs)
    
    def handle_role_selection(self, request):
        """Handle role selection from the multiple users modal."""
        selected_user_id = request.POST.get('selected_user_id')
        user_email = request.POST.get('user_email')
        
        if not selected_user_id or not user_email:
            messages.error(request, "Invalid role selection.")
            return redirect('login')
        
        try:
            # Get the specific user
            user = CustomUser.objects.get(id=selected_user_id, email=user_email)
            
            # SECURITY: Check if user is a system admin
            if getattr(user, 'is_system_admin', False):
                messages.error(request, "Access Denied: You are not allowed to login through this interface.")
                return redirect('login')
            
            # SECURITY: Tenancy Validation
            login_slug = self.kwargs.get('login_slug', None)
            if login_slug:
                try:
                    company = Company.objects.get(slug=login_slug)
                    user_company = getattr(user, 'company_profile', None)
                    if user_company and user_company.id != company.id:
                        messages.error(request, "❌ You do not have permission to login through this tenant portal.")
                        return redirect('login')
                except Company.DoesNotExist:
                    messages.error(request, "❌ Invalid company portal URL.")
                    return redirect('login')
            
            # Log the user in
            login(request, user)
            
            # Capture login details
            try:
                ip = extract_client_ip(request)
                try:
                    from .services.geoip import is_private_ip
                    client_ip_from_form = request.POST.get('client_public_ip')
                    if (not ip or is_private_ip(ip)) and client_ip_from_form:
                        ip = client_ip_from_form.strip()
                except Exception:
                    pass
                
                if hasattr(user, 'last_login_ip'):
                    user.last_login_ip = ip
                    location = lookup_ip_location(ip)
                    if hasattr(user, 'last_login_location') and location:
                        user.last_login_location = location
                        user.save(update_fields=['last_login_ip', 'last_login_location'])
                    else:
                        user.save(update_fields=['last_login_ip'])
            except Exception:
                pass
            
            # Set session expiry
            import time
            request.session['_session_expiry'] = time.time() + 300
            request.session.save()
            
            # Role-specific welcome messages
            company_name = user.company_profile.company_name if user.company_profile else "your company"
            role_messages = {
                'admin': f"Welcome back to {company_name}!",
                'client': f"Welcome back, {user.full_name}!",
                'marketer': f"Welcome back, {user.full_name}!",
                'support': f"Welcome back, {user.full_name}! {company_name}!"
            }
            messages.success(request, role_messages.get(user.role, "Login successful!"))
            
            return redirect(self.get_success_url())
            
        except CustomUser.DoesNotExist:
            messages.error(request, "Selected user not found.")
            return redirect('login')
        except Exception as e:
            messages.error(request, f"Login failed: {str(e)}")
            return redirect('login')
    def form_valid(self, form):
        """
        Override form_valid to handle MultipleUserMatch objects.
        When multiple users are found, show the role selection modal instead of logging in.
        """
        from .backends import MultipleUserMatch
        
        # Check if authentication returned multiple users
        auth_result = form.get_user()
        
        if isinstance(auth_result, MultipleUserMatch):
            # Multiple users found - show role selection modal
            context = self.get_context_data(form=form)
            context['multiple_users'] = auth_result.users
            context['user_email'] = auth_result.email
            from django.shortcuts import render
            return render(self.request, self.template_name, context)
        
        # Single user authentication - proceed normally
        user = auth_result
        
        # SECURITY: Check if user is a system admin - they cannot use this interface
        if getattr(user, 'is_system_admin', False):
            messages.error(
                self.request, 
                "Access Denied: You are not allowed to login through this interface. "
                "System administrators must use the dedicated system administration portal."
            )
            logger.warning(
                f"SECURITY: System admin '{user.email}' attempted to access unified login. "
                f"Access denied. IP: {extract_client_ip(self.request)}"
            )
            return self.form_invalid(form)
        
        # SECURITY: Tenancy Validation - Ensure user belongs to the company slug if provided
        login_slug = self.kwargs.get('login_slug', None)
        if login_slug:
            # User is logging in via tenant-specific URL (e.g., /company-name/login/)
            try:
                company = Company.objects.get(slug=login_slug)
                # Verify user belongs to this company
                user_company = getattr(user, 'company_profile', None)
                if user_company and user_company.id != company.id:
                    # User's company doesn't match the slug - SECURITY VIOLATION
                    messages.error(
                        self.request,
                        "❌ You do not have permission to login through this tenant portal. "
                        "Please use your assigned company's login link."
                    )
                    logger.warning(
                        f"SECURITY: Cross-company login attempt detected. "
                        f"User '{user.email}' (Company: {user_company.company_name}) "
                        f"attempted to access '{company.company_name}' (Slug: {login_slug}). "
                        f"IP: {extract_client_ip(self.request)}"
                    )
                    return self.form_invalid(form)
            except Company.DoesNotExist:
                # Invalid company slug
                messages.error(self.request, "❌ Invalid company portal URL.")
                logger.warning(f"SECURITY: Login attempt with non-existent slug '{login_slug}'.")
                return self.form_invalid(form)
        
        response = super().form_valid(form)
        # Capture last login IP and (optionally) location
        try:
            user = self.request.user
            ip = extract_client_ip(self.request)
            # If we only have a private/localhost IP, prefer the client-provided public IP (if present)
            try:
                from .services.geoip import is_private_ip
                client_ip_from_form = self.request.POST.get('client_public_ip')
                if (not ip or is_private_ip(ip)) and client_ip_from_form:
                    ip = client_ip_from_form.strip()
            except Exception:
                pass
            if hasattr(user, 'last_login_ip'):
                user.last_login_ip = ip
                # Attempt a lightweight GeoIP lookup; ignore failures
                location = lookup_ip_location(ip)
                if hasattr(user, 'last_login_location') and location:
                    user.last_login_location = location
                    user.save(update_fields=['last_login_ip', 'last_login_location'])
                else:
                    user.save(update_fields=['last_login_ip'])
        except Exception:
            pass
        
        # Set session expiry for sliding expiration (5 minutes)
        import time
        self.request.session['_session_expiry'] = time.time() + 300  # 5 minutes
        self.request.session.save()
        
        # Role-specific welcome messages
        company_name = user.company_profile.company_name if user.company_profile else "your company"
        role_messages = {
            'admin': f"Welcome back to {company_name}!",
            'client': f"Welcome back, {user.full_name}!",
            'marketer': f"Welcome back, {user.full_name}!",
            'support': f"Welcome back, {user.full_name}! {company_name}!"
        }
        messages.success(self.request, role_messages.get(user.role, "Login successful!"))
        return response

    def form_invalid(self, form):
        messages.error(self.request, "Invalid email or password. Please try again.")
        return self.render_to_response(self.get_context_data(form=form))

    def get_success_url(self):
        redirect_to = self.get_redirect_url()
        if redirect_to:
            return redirect_to

        user = self.request.user
        
        # SECURITY: System Master Admin cannot use unified login interface
        # They must use separate tenant admin login at /tenant-admin/login/
        if user.role == 'admin' and getattr(user, 'admin_level', None) == 'system':
            # Log this security incident
            logger.warning(
                f"SECURITY: System Master Admin '{user.email}' attempted to access unified login. "
                f"Redirecting to tenant admin panel. IP: {extract_client_ip(self.request)}"
            )
            messages.info(self.request, "System Master Admin must use the admin panel. Redirecting...")
            return reverse_lazy('tenant-admin-dashboard')  # Redirect to proper system admin area
        
        # Role-based routing (company admin and secondary_admin both route to admin dashboard)
        if user.role in ('admin', 'secondary_admin'):
            # Verify admin_level to ensure company-scoped admin
            admin_level = getattr(user, 'admin_level', 'company')
            if admin_level != 'system':
                # Use new tenant-aware routing
                company = user.company_profile
                if company:
                    from django.urls import reverse
                    return reverse('tenant-dashboard', kwargs={'company_slug': company.slug})
                else:
                    # Fallback if no company assigned
                    messages.error(self.request, "You are not assigned to any company!")
                    return reverse_lazy('login')
            else:
                # Fallback: should not reach here due to check above
                return reverse_lazy('tenant-admin-dashboard')
        elif user.role == 'client':
            return reverse_lazy('client-dashboard')
        elif user.role == 'marketer':
            return reverse_lazy('marketer-dashboard')
        elif user.role == 'support':
            return reverse_lazy('adminsupport:support_dashboard')
        else:
            return super().get_success_url()


@csrf_protect
def company_registration(request):
    """
    Company Registration - Creates Company and Admin User
    Companies become the "Company Admin" role with full management access
    
    Security:
    - Validates strict tenancy rules with slug cybersecurity config
    - Admin password stored separately from primary admin
    - System admin cannot register through this interface (redirect enforced)
    - Atomic transaction ensures no partial data creation
    """
    if request.method == 'POST':
        try:
            # SECURITY: Verify user is not already a system master admin
            if request.user.is_authenticated:
                if request.user.role == 'admin' and getattr(request.user, 'admin_level', None) == 'system':
                    messages.error(request, "❌ System Master Admin cannot register companies through this interface. Use admin panel.")
                    return get_tenant_dashboard_redirect(request)
            
            # Extract form data
            company_name = request.POST.get('company_name')
            registration_number = request.POST.get('registration_number')
            registration_date = request.POST.get('registration_date')
            location = request.POST.get('location')
            ceo_name = request.POST.get('ceo_name')
            ceo_dob = request.POST.get('ceo_dob')
            email = request.POST.get('email')
            phone = request.POST.get('phone')
            
            # CRITICAL: Get subscription tier - REQUIRED and MUST be valid
            subscription_tier = request.POST.get('subscription_tier', '').strip()
            if not subscription_tier or subscription_tier not in ['starter', 'professional', 'enterprise']:
                messages.error(
                    request, 
                    "❌ Subscription plan is REQUIRED! Please select a plan: Starter, Professional, or Enterprise."
                )
                return redirect('login')
            
            # Handle both old and new password field names
            password = request.POST.get('password') or request.POST.get('secondary_admin_password')
            confirm_password = request.POST.get('confirm_password') or request.POST.get('secondary_admin_confirm_password')
            
            # Secondary admin details (optional)
            secondary_admin_email = request.POST.get('secondary_admin_email')
            secondary_admin_phone = request.POST.get('secondary_admin_phone')
            secondary_admin_name = request.POST.get('secondary_admin_name')

            # Validation
            if password != confirm_password:
                messages.error(request, "Passwords do not match!")
                return redirect('login')

            if len(password) < 8:
                messages.error(request, "Password must be at least 8 characters long!")
                return redirect('login')

            # Check if company already exists (strict isolation check)
            if Company.objects.filter(company_name=company_name).exists():
                messages.error(request, f"A company with the name '{company_name}' already exists!")
                return redirect('login')

            if Company.objects.filter(registration_number=registration_number).exists():
                messages.error(request, "This registration number is already in use!")
                return redirect('login')

            if Company.objects.filter(email=email).exists():
                messages.error(request, "This company email is already registered!")
                return redirect('login')

            # Check if user email already exists for admin role only
            if CustomUser.objects.filter(email=email, role='admin').exists():
                messages.error(request, "A user with this email already exists as an admin!")
                return redirect('login')

            # Create company with transaction (atomic: all or nothing)
            with transaction.atomic():
                # Calculate trial end date (14 days from now)
                trial_end = timezone.now() + timedelta(days=14)
                
                # Create the company with strict isolation metadata
                company = Company.objects.create(
                    company_name=company_name,
                    registration_number=registration_number,
                    registration_date=registration_date,
                    location=location,
                    ceo_name=ceo_name,
                    ceo_dob=ceo_dob,
                    email=email,
                    phone=phone,
                    is_active=True,
                    subscription_status='trial',
                    subscription_tier=subscription_tier,  # Use selected subscription tier
                    trial_ends_at=trial_end
                )

                # Create the primary admin user (Company Admin - NOT system admin)
                admin_user = CustomUser.objects.create_user(
                    email=email,
                    full_name=ceo_name,
                    phone=phone,
                    password=password,
                    role='admin',  # Company Admin role
                    admin_level='company',  # CRITICAL: Company-level admin, NOT system
                    company_profile=company,
                    address=location,
                    date_of_birth=ceo_dob,
                    is_staff=True,
                    is_superuser=False,  # Explicitly NOT system admin
                    is_active=True
                )

                # Set additional fields
                admin_user.company = company_name
                admin_user.save()
                
                # CRITICAL: Create subscription billing record with 14-day free trial
                from estateApp.subscription_billing_models import SubscriptionBillingModel
                
                trial_starts = timezone.now()
                trial_ends = trial_starts + timedelta(days=14)
                
                billing = SubscriptionBillingModel.objects.create(
                    company=company,
                    status='trial',
                    payment_method='free_trial',
                    trial_started_at=trial_starts,
                    trial_ends_at=trial_ends,
                    billing_cycle='monthly',
                    auto_renew=False,
                    monthly_amount=Decimal('0.00'),
                    annual_amount=Decimal('0.00'),
                )
                
                # Create secondary admin user if provided (optional but separate credentials)
                if secondary_admin_email and secondary_admin_name:
                    if CustomUser.objects.filter(email=secondary_admin_email).exists():
                        messages.warning(request, f"Secondary admin email {secondary_admin_email} already exists. You can add them later from admin panel.")
                    else:
                        # Create secondary admin with same role but different account
                        secondary_admin = CustomUser.objects.create_user(
                            email=secondary_admin_email,
                            full_name=secondary_admin_name,
                            phone=secondary_admin_phone or phone,
                            password=password,  # Use same password as primary initially
                            role='admin',  # Also company admin
                            admin_level='company',  # CRITICAL: Company-level, NOT system
                            company_profile=company,
                            is_staff=True,
                            is_superuser=False,
                            is_active=True
                        )
                        secondary_admin.company = company_name
                        secondary_admin.save()

                messages.success(
                    request,
                    f"🎉 Welcome to Lamba! {company_name} registered successfully! "
                    f"Your 14-day free trial starts now. Login to access your dashboard."
                )
                
                # Send welcome email
                try:
                    send_mail(
                        subject=f'Welcome to Lamba Real Estate Management - {company_name}',
                        message=f'Dear {ceo_name},\n\n'
                                f'Congratulations! Your company "{company_name}" has been successfully registered on Lamba.\n\n'
                                f'🎁 You now have a 14-day FREE TRIAL to explore all features!\n\n'
                                f'Login Details:\n'
                                f'Email: {email}\n'
                                f'Dashboard: Company Admin Portal\n\n'
                                f'What you can do:\n'
                                f'✅ Manage clients and marketers\n'
                                f'✅ Create and allocate properties\n'
                                f'✅ Track commissions and payments\n'
                                f'✅ Generate reports and analytics\n\n'
                                f'Login now: https://lamba.com/login\n\n'
                                f'Thank you for choosing Lamba!\n\n'
                                f'Best regards,\n'
                                f'The Lamba Team\n'
                                f'Transforming Nigerian Real Estate',
                        from_email=settings.DEFAULT_FROM_EMAIL,
                        recipient_list=[email],
                        fail_silently=True,
                    )
                except Exception:
                    pass

                return redirect('login')

        except IntegrityError as e:
            messages.error(request, f"Registration failed: A record with this information already exists!")
            return redirect('login')
        except Exception as e:
            messages.error(request, f"An error occurred during registration: {str(e)}")
            logger.error(f"Company registration error: {str(e)}")
            return redirect('login')

    # GET request - redirect to login page
    return redirect('login')


@csrf_protect
def client_registration(request):
    """
    Client Self-Registration
    Clients can register independently to manage their properties from multiple companies
    """
    if request.method == 'POST':
        try:
            # Extract form data
            first_name = request.POST.get('first_name')
            last_name = request.POST.get('last_name')
            email = request.POST.get('email')
            phone = request.POST.get('phone')
            address = request.POST.get('address')
            date_of_birth_str = request.POST.get('date_of_birth')
            
            # Parse date_of_birth if provided
            date_of_birth = None
            if date_of_birth_str:
                try:
                    from datetime import datetime
                    date_of_birth = datetime.strptime(date_of_birth_str, '%Y-%m-%d').date()
                except ValueError:
                    messages.error(request, "Invalid date format for date of birth!")
                    return redirect('login')
            password = request.POST.get('password')
            confirm_password = request.POST.get('confirm_password')

            # Validation
            if password != confirm_password:
                messages.error(request, "Passwords do not match!")
                return redirect('login')

            if len(password) < 8:
                messages.error(request, "Password must be at least 8 characters long!")
                return redirect('login')

            # Check if user email already exists for client role only
            if CustomUser.objects.filter(email=email, role='client').exists():
                messages.error(request, "A user with this email already exists as a client!")
                return redirect('login')

            # Create client user
            with transaction.atomic():
                client_user = CustomUser.objects.create_user(
                    email=email,
                    full_name=f"{first_name} {last_name}",
                    phone=phone,
                    address=address,
                    password=password,
                    role='client',
                    company_profile=None,  # Clients are not bound to a company initially
                    date_of_birth=date_of_birth,
                    is_active=True,
                    is_staff=False,
                    is_superuser=False
                )

                messages.success(
                    request,
                    f"🎉 Welcome to Lamba, {first_name}! Your client account is ready. "
                    f"Login to view and manage properties from all companies."
                )
                
                # Send welcome email
                try:
                    send_mail(
                        subject='Welcome to Lamba - Your Client Account is Ready!',
                        message=f'Dear {first_name} {last_name},\n\n'
                                f'Welcome to Lamba Real Estate Management System!\n\n'
                                f'Your client account has been successfully created.\n\n'
                                f'Login Details:\n'
                                f'Email: {email}\n'
                                f'Dashboard: Client Portal\n\n'
                                f'What you can do:\n'
                                f'✅ View all your properties in one place\n'
                                f'✅ Track payment schedules\n'
                                f'✅ Manage properties from multiple companies\n'
                                f'✅ Download property documents\n\n'
                                f'Login now: https://lamba.com/login\n\n'
                                f'Thank you for choosing Lamba!\n\n'
                                f'Best regards,\n'
                                f'The Lamba Team',
                        from_email=settings.DEFAULT_FROM_EMAIL,
                        recipient_list=[email],
                        fail_silently=True,
                    )
                except Exception:
                    pass

                # Check if this is an AJAX request
                if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                    return JsonResponse({
                        'success': True,
                        'message': f"🎉 Welcome to Lamba, {first_name}! Your client account is ready. Login to view and manage properties from all companies.",
                        'redirect_url': reverse('login')
                    })

                return redirect('login')

        except Exception as e:
            messages.error(request, f"An error occurred during registration: {str(e)}")
            logger.error(f"Client registration error: {str(e)}")
            
            # Check if this is an AJAX request
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                return JsonResponse({
                    'success': False,
                    'message': f"An error occurred during registration: {str(e)}"
                })
            
            return redirect('login')

    # GET request - redirect to login page
    return redirect('login')


@csrf_protect
def marketer_registration(request):
    """
    Marketer Self-Registration
    Marketers can register to affiliate with multiple companies and manage commissions
    """
    if request.method == 'POST':
        try:
            # Extract form data
            first_name = request.POST.get('first_name')
            last_name = request.POST.get('last_name')
            email = request.POST.get('email')
            phone = request.POST.get('phone')
            address = request.POST.get('address')
            date_of_birth_str = request.POST.get('date_of_birth')
            
            # Parse date_of_birth if provided
            date_of_birth = None
            if date_of_birth_str:
                try:
                    from datetime import datetime
                    date_of_birth = datetime.strptime(date_of_birth_str, '%Y-%m-%d').date()
                except ValueError:
                    messages.error(request, "Invalid date format for date of birth!")
                    return redirect('login')
            password = request.POST.get('password')
            confirm_password = request.POST.get('confirm_password')

            # Validation
            if password != confirm_password:
                messages.error(request, "Passwords do not match!")
                return redirect('login')

            if len(password) < 8:
                messages.error(request, "Password must be at least 8 characters long!")
                return redirect('login')

            # Check if user email already exists for marketer role only
            if CustomUser.objects.filter(email=email, role='marketer').exists():
                messages.error(request, "A user with this email already exists as a marketer!")
                return redirect('login')

            # Create marketer user
            with transaction.atomic():
                marketer_user = CustomUser.objects.create_user(
                    email=email,
                    full_name=f"{first_name} {last_name}",
                    phone=phone,
                    address=address,
                    password=password,
                    role='marketer',
                    company_profile=None,  # Marketers can work with multiple companies
                    date_of_birth=date_of_birth,
                    is_active=True,
                    is_staff=False,
                    is_superuser=False
                )

                messages.success(
                    request,
                    f"🎉 Welcome to Lamba, {first_name}! Your marketer account is ready. "
                    f"Login to start affiliating with companies and earning commissions."
                )
                
                # Send welcome email
                try:
                    send_mail(
                        subject='Welcome to Lamba - Your Marketer Account is Active!',
                        message=f'Dear {first_name} {last_name},\n\n'
                                f'Welcome to Lamba Real Estate Management System!\n\n'
                                f'Your marketer account has been successfully created.\n\n'
                                f'Login Details:\n'
                                f'Email: {email}\n'
                                f'Dashboard: Marketer Portal\n\n'
                                f'What you can do:\n'
                                f'✅ Affiliate with multiple real estate companies\n'
                                f'✅ Track your commissions in real-time\n'
                                f'✅ Manage client referrals\n'
                                f'✅ Access marketing materials and resources\n\n'
                                f'Login now: https://lamba.com/login\n\n'
                                f'Start earning today!\n\n'
                                f'Best regards,\n'
                                f'The Lamba Team',
                        from_email=settings.DEFAULT_FROM_EMAIL,
                        recipient_list=[email],
                        fail_silently=True,
                    )
                except Exception:
                    pass

                # Check if this is an AJAX request
                if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                    return JsonResponse({
                        'success': True,
                        'message': f"🎉 Welcome to Lamba, {first_name}! Your marketer account is ready. Login to start affiliating with companies and earning commissions.",
                        'redirect_url': reverse('login')
                    })

                return redirect('login')

        except Exception as e:
            messages.error(request, f"An error occurred during registration: {str(e)}")
            logger.error(f"Marketer registration error: {str(e)}")
            
            # Check if this is an AJAX request
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                return JsonResponse({
                    'success': False,
                    'message': f"An error occurred during registration: {str(e)}"
                })
            
            return redirect('login')

    # GET request - redirect to login page
    return redirect('login')


@login_required
def change_password(request):
    context = {}

    if request.method == 'POST':
        current = request.POST.get('currentPassword')
        new_pw  = request.POST.get('newPassword')
        confirm = request.POST.get('renewPassword')

        if not request.user.check_password(current):
            context['error'] = "Current password is incorrect."
        elif new_pw != confirm:
            context['error'] = "New password and confirmation password do not match."
        elif len(new_pw) < 8:
            context['error'] = "Password must be at least 8 characters long."
        else:
            # All good: update password
            user = request.user
            user.set_password(new_pw)
            user.save()
            # Keep the session
            update_session_auth_hash(request, user)
            context['success'] = "Your password has been successfully updated!"

    return SimpleNamespace(context_data=context)


def update_profile_data(user, request):
    if request.method == 'POST':
        # Get the posted form data
        about = request.POST.get('about')
        company = request.POST.get('company')
        job = request.POST.get('job')
        country = request.POST.get('country')
        profile_image = request.FILES.get('profile_image')

        # Validate the input fields if necessary
        # if not about or not company or not job or not country:
        #     messages.error(request, "Please fill out all fields.")
        #     return False

        # Update the user fields
        user.about = about
        user.company = company
        user.job = job
        user.country = country

        # If a new profile image was uploaded, update it
        if profile_image:
            user.profile_image = profile_image

        # Save the updated user information
        try:
            user.save()
            messages.success(request, 'Your profile has been updated successfully!')
            return True
        except ValidationError as e:
            messages.error(request, f"Error updating profile: {e}")
            return False
    return False


@login_required
def send_message(request):
    return render(request, 'send_message.html')

from django.shortcuts import render, get_object_or_404


@login_required
def client_message(request):
    """
    Handle client message submission and display client-specific messages.
    """
    if request.method == "POST":
        message_type = request.POST.get('message_type')
        message_content = request.POST.get('message_content')

        if not message_type or not message_content:
            messages.error(request, "All fields are required.")
            return redirect('client-message')

        # Save the client's message
        Message.objects.create(
            sender=request.user,
            recipient=None,  # Indicates a message for the admin
            type=message_type,
            content=message_content,
            date_sent=timezone.now(),
        )

        messages.success(request, "Your message has been sent to the admin dashboard.")
        return redirect('client-message')

    user_messages = Message.objects.filter(sender=request.user).order_by('-date_sent')
    return render(request, 'client_side/contact.html', {'messages': user_messages})


# NOTIFICATIONS
def announcement_form(request):
    form = NotificationForm() 
    return render(request, 'notifications/emails/admin_notification_form.html', {'form': form})


@require_http_methods(["POST"])
def send_announcement(request):
    try:
        # Authentication check
        if not request.user.is_authenticated or not request.user.is_staff:
            return JsonResponse({
                'status': 'error',
                'message': 'Authentication required'
            }, status=401)

        # Data validation
        notification_type = request.POST.get('notification_type')
        title = request.POST.get('title')
        message = request.POST.get('message')

        if not all([notification_type, title, message]):
            raise ValidationError("All fields are required")

        # Create notification
        notification = Notification.objects.create(
            notification_type=notification_type,
            title=title,
            message=message
        )

        # Determine recipients
        roles = []
        if notification_type == Notification.ANNOUNCEMENT:
            roles = ['client', 'marketer']
        elif notification_type == Notification.CLIENT_ANNOUNCEMENT:
            roles = ['client']
        elif notification_type == Notification.MARKETER_ANNOUNCEMENT:
            roles = ['marketer']
        else:
            raise ValidationError("Invalid notification type")

        recipients = CustomUser.objects.filter(role__in=roles)
        if not recipients.exists():
            raise ValidationError("No recipients found for this notification type")

        # Bulk create with batch processing
        batch_size = 100
        user_notifications = [
            UserNotification(user=user, notification=notification)
            for user in recipients.iterator()
        ]
        UserNotification.objects.bulk_create(user_notifications, batch_size=batch_size)

        return JsonResponse({
            'status': 'success',
            'message': f'Notification sent to {len(recipients)} recipients'
        })

    except ValidationError as e:
        return JsonResponse({
            'status': 'error',
            'message': 'Validation error',
            'errors': e.messages
        }, status=400)
        
    except Exception as e:
        return JsonResponse({
            'status': 'error',
            'message': str(e)
        }, status=500)


def send_notification_email(client, title, message):
    context = {
        'user': client,
        'message': message,
        'site_url': settings.SITE_URL
    }
    
    html_content = render_to_string(
        'notifications/emails/announcement.html',
        context
    )
    
    send_mail(
        subject=f"REMS Notification: {title}",
        message=strip_tags(html_content),
        from_email=settings.DEFAULT_FROM_EMAIL,
        recipient_list=[client.email],
        html_message=html_content
    )



# MANAGEMENT DASHBOARD
# Sales Volume Tab

def sales_volume_metrics(request):
    """
    GET params:
      - period: 'monthly', 'quarterly', or 'yearly'
      - specific_period: 'YYYY-MM', 'YYYY-Qn', or 'YYYY'
    Returns JSON with current & previous metrics, including Outstanding Payments
    calculated as (sum of all part‑plan totals) – (sum of all part‑plan payments).
    """

    today  = timezone.now().date()
    period = request.GET.get('period', 'monthly')
    sp     = request.GET.get('specific_period')

    # 1) Determine start/end of the selected period
    if period == 'monthly' and sp:
        year, month = map(int, sp.split('-'))
        start = dt.date(year, month, 1)
        end   = start + relativedelta(months=1)

    elif period == 'quarterly' and sp:
        y_str, q_str = sp.split('-Q')
        y, q = int(y_str), int(q_str)
        start = dt.date(y, 3*(q-1) + 1, 1)
        end   = start + relativedelta(months=3)

    elif period == 'yearly' and sp:
        y = int(sp)
        start = dt.date(y, 1, 1)
        end   = dt.date(y+1, 1, 1)

    else:
        # default to current month
        start = today.replace(day=1)
        end   = start + relativedelta(months=1)

    # helper: sum total_amount in [s,e)
    def sum_amt(s, e):
        return (Transaction.objects
                .filter(transaction_date__gte=s, transaction_date__lt=e)
                .aggregate(total=Sum('total_amount'))['total'] or 0)

    # 2) Monthly Closed Deals: last month in [start,end)
    last_month_start = (end - relativedelta(months=1)).replace(day=1)
    next_month_start = last_month_start + relativedelta(months=1)

    monthly_closed = Transaction.objects.filter(
        transaction_date__gte=last_month_start,
        transaction_date__lt= next_month_start
    ).count()

    prev_month_start = last_month_start - relativedelta(months=1)
    prev_month_end   = last_month_start

    prev_monthly_closed = Transaction.objects.filter(
        transaction_date__gte=prev_month_start,
        transaction_date__lt= prev_month_end
    ).count()

    # 3) Quarterly Sales Volume
    quarterly_sales = sum_amt(start, end)

    prev_quarter_start = start - (end - start)
    prev_quarter_end   = start

    prev_quarterly_sales = sum_amt(prev_quarter_start, prev_quarter_end)

    # 4) Annual Transactions & Average Deal
    annual_transactions = Transaction.objects.filter(
        transaction_date__gte=start,
        transaction_date__lt=end
    ).count()

    average_deal = (quarterly_sales / annual_transactions) if annual_transactions else 0

    prev_annual_transactions = Transaction.objects.filter(
        transaction_date__gte=prev_quarter_start,
        transaction_date__lt=prev_quarter_end
    ).count()

    prev_average_deal = (
        (prev_quarterly_sales / prev_annual_transactions)
        if prev_annual_transactions else 0
    )

    # 5) Active Installments (all‑time)
    active_installments = (
        Transaction.objects
        .filter(allocation__payment_type='part')
        .annotate(
            paid=Coalesce(
                Sum('payment_records__amount_paid'),
                Value(0),
                output_field=DecimalField()
            )
        )
        .filter(paid__lt=F('total_amount'))
        .count()
    )

    # 6) Outstanding Payments (all part‑plans)
    #    Sum all part‑plan totals, then subtract all part‑plan payments
    part_totals = (
        Transaction.objects
        .filter(allocation__payment_type='part')
        .aggregate(grand_total=Sum('total_amount'))['grand_total'] or 0
    )
    paid_sum = (
        PaymentRecord.objects
        .filter(transaction__allocation__payment_type='part')
        .aggregate(paid_total=Sum('amount_paid'))['paid_total'] or 0
    )
    total_outstanding = part_totals - paid_sum

    # 7) Overdue Payments (all‑time): count part‑plans whose due_date < today and still unpaid
    overdue_count = 0
    for tx in Transaction.objects.filter(allocation__payment_type='part'):
        if tx.due_date and tx.due_date < today and tx.total_amount > tx.total_paid:
            overdue_count += 1

    # 8) Total Paid YTD
    y_start = today.replace(month=1, day=1)
    part_paid_ytd = (
        PaymentRecord.objects
        .filter(payment_date__gte=y_start)
        .aggregate(total=Sum('amount_paid'))['total'] or 0
    )
    full_paid_ytd = (
        Transaction.objects
        .filter(allocation__payment_type='full', transaction_date__gte=y_start)
        .aggregate(total=Sum('total_amount'))['total'] or 0
    )
    total_paid = part_paid_ytd + full_paid_ytd

    # Return JSON
    return JsonResponse({
        'monthly_closed':           monthly_closed,
        'prev_monthly_closed':      prev_monthly_closed,
        'quarterly_sales':          float(quarterly_sales),
        'prev_quarterly_sales':     float(prev_quarterly_sales),
        'annual_transactions':      annual_transactions,
        'prev_annual_transactions': prev_annual_transactions,
        'average_deal':             float(average_deal),
        'prev_average_deal':        float(prev_average_deal),
        'active_installments':      active_installments,
        'total_outstanding':        float(total_outstanding),
        'overdue_count':            overdue_count,
        'total_paid':               float(total_paid),
    })


# <!-- Land Plot Transactions -->
@login_required
@require_http_methods(["POST"])
def add_transaction(request):
    # Company isolation - get user's company
    user_company = getattr(request.user, 'company_profile', None)
    if not user_company:
        return JsonResponse({'error': 'No company access'}, status=403)
    
    cid      = request.POST.get("client", "")
    aid      = request.POST.get("allocation", "")
    txn_date = request.POST.get("transaction_date", "")
    total_s  = request.POST.get("total_amount", "0")
    notes    = request.POST.get("special_notes", "")
    plan     = request.POST.get("installment_plan", "")
    pdur     = request.POST.get("payment_duration", "")
    cdur     = request.POST.get("custom_duration", "")
    fp_s     = request.POST.get("first_percent", "")
    sp_s     = request.POST.get("second_percent", "")
    tp_s     = request.POST.get("third_percent", "")
    pm = request.POST.get("payment_method", "")

    # Company isolation - ensure allocation belongs to user's company
    allocation = get_object_or_404(
        PlotAllocation.objects.filter(estate__company=user_company),
        pk=aid
    )
    
    # Verify client belongs to user's company
    client = get_object_or_404(
        CustomUser.objects.filter(company_profile=user_company, role='client'),
        pk=cid
    )
    
    # Check presale price is set before allowing transaction
    presale_exists = PropertyPrice.objects.filter(
        estate=allocation.estate,
        plot_unit=allocation.plot_size_unit,
        company=user_company,
        presale__gt=0
    ).exists()
    
    if not presale_exists:
        return JsonResponse({
            'error': f'Cannot add transaction: Presale price not set for {allocation.estate.name} - {allocation.plot_size_unit.plot_size.size}. Please set presale price in Value Regulation first.'
        }, status=400)
    
    existing   = Transaction.objects.filter(client_id=cid, allocation=allocation, company=user_company).first()
    txn        = existing or Transaction(client_id=cid, allocation=allocation)

    # --- 1) Set company from user's company (CRITICAL for multi-tenant isolation) ---
    txn.company = user_company

    # --- 2) Core fields ---
    txn.transaction_date = txn_date
    txn.total_amount     = Decimal(total_s)
    txn.special_notes    = notes
    txn.installment_plan = plan or None

    # --- 3) Part-payment logic ---
    if allocation.payment_type == "part" and txn.total_amount > 0:
        # 3a) Handle durations
        if pdur and pdur != "custom":
            txn.payment_duration = int(pdur)
            txn.custom_duration  = None
        elif pdur == "custom" and cdur:
            txn.payment_duration = None
            txn.custom_duration  = int(cdur)
        else:
            txn.payment_duration = txn.custom_duration = None

        # 3b) Handle percents & installments
        if plan and plan != "custom":
            pcts = list(map(int, plan.split("-")))
            txn.first_percent, txn.second_percent, txn.third_percent = pcts
        else:
            txn.first_percent  = int(fp_s) if fp_s.isdigit() else None
            txn.second_percent = int(sp_s) if sp_s.isdigit() else None
            txn.third_percent  = int(tp_s) if tp_s.isdigit() else None
            pcts = [txn.first_percent, txn.second_percent, txn.third_percent]

        if pcts and sum(pcts) == 100:
            t = txn.total_amount
            txn.first_installment  = (t * pcts[0]) / 100
            txn.second_installment = (t * pcts[1]) / 100
            txn.third_installment  = (t * pcts[2]) / 100
        else:
            txn.first_installment = txn.second_installment = txn.third_installment = None

    else:
        # full payment → clear part-payment fields
        txn.payment_duration     = txn.custom_duration = None
        txn.first_percent        = txn.second_percent = txn.third_percent = None
        txn.first_installment    = txn.second_installment = txn.third_installment = None

    if allocation.payment_type == "full":
        # write payment_method
        txn.payment_method = pm or None
    else:
        txn.payment_method = None

    # --- 4) Save & respond ---
    try:
        txn.save()
        company = getattr(request.user, 'company_profile', None)
        if company:
            from django.urls import reverse
            return redirect(reverse('tenant-management', kwargs={'company_slug': company.slug}))
        return redirect("login")
    except Exception as e:
        company = getattr(request.user, 'company_profile', None)
        all_clients = CustomUser.objects.filter(
            role='client',
            company_profile=company
        ) if company else CustomUser.objects.none()
        return render(request, "admin_side/management-dashboard.html", {
            "all_clients": all_clients,
            "error":       str(e),
            "posted":      request.POST,
        })


@login_required
def ajax_client_marketer(request):
    client_id = request.GET.get('client_id')
    if not client_id:
        return JsonResponse({'error': 'Missing client_id'}, status=400)
    
    # Company isolation - get user's company
    user_company = getattr(request.user, 'company_profile', None)
    if not user_company:
        return JsonResponse({'error': 'No company access'}, status=403)
    
    try:
        # Ensure client belongs to user's company
        client = ClientUser.objects.filter(
            pk=client_id,
            company_profile=user_company
        ).first()
        
        if not client:
            return JsonResponse({'error': 'Client not found or access denied'}, status=404)

        marketer = None
        try:
            marketer = client.get_assigned_marketer(company=user_company)
        except Exception:
            marketer = getattr(client, 'assigned_marketer', None)

        return JsonResponse({
            'marketer_name': marketer.full_name if marketer else 'No marketer assigned'
        })
    except ClientUser.DoesNotExist:
        return JsonResponse({'error': 'Client not found'}, status=404)

@login_required
def ajax_client_allocations(request):
    client_id = request.GET.get('client_id')
    if not client_id:
        return JsonResponse({'error': 'Missing client_id'}, status=400)
    
    # Company isolation - get user's company
    user_company = getattr(request.user, 'company_profile', None)
    if not user_company:
        return JsonResponse({'error': 'No company access'}, status=403)
    
    try:
        # Filter allocations by company through estate
        allocations = PlotAllocation.objects.filter(
            client_id=client_id,
            estate__company=user_company  # Company isolation
        ).select_related('estate', 'plot_size_unit__plot_size')
        
        data = [{
            'id': alloc.id,
            'estate_name': alloc.estate.name,
            'plot_size': alloc.plot_size_for_transaction,
            'payment_type': alloc.payment_type,
            'date_allocated': alloc.date_allocated.strftime('%Y-%m-%d'),
        } for alloc in allocations]
        
        return JsonResponse(data, safe=False)
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)

@login_required
@require_GET
def ajax_allocation_info(request):
    alloc_id = request.GET.get("allocation_id")
    if not alloc_id:
        return HttpResponseBadRequest("Missing allocation_id")

    # Company isolation - get user's company
    user_company = getattr(request.user, 'company_profile', None)
    if not user_company:
        return JsonResponse({'error': 'No company access'}, status=403)

    # Ensure allocation belongs to user's company through estate
    alloc = get_object_or_404(
        PlotAllocation.objects.select_related(
            'estate',
            'plot_size_unit__plot_size',
            'client__assigned_marketer'
        ).filter(estate__company=user_company),  # Company isolation
        pk=alloc_id
    )

    # Filter transaction by company as well
    txn = Transaction.objects.filter(
        allocation=alloc,
        company=user_company  # Company isolation
    ).order_by('-transaction_date').first()

    # Check if presale price is set for this estate and plot size (company-isolated)
    presale_price_set = False
    presale_price_value = 0
    try:
        property_price = PropertyPrice.objects.filter(
            estate=alloc.estate,
            plot_unit=alloc.plot_size_unit,
            company=user_company  # Company isolation
        ).first()
        if property_price and property_price.presale and property_price.presale > 0:
            presale_price_set = True
            presale_price_value = float(property_price.presale)
    except Exception:
        pass

    payload = {
        "plot_size": alloc.plot_size_unit.plot_size.size,
        "payment_type": alloc.payment_type,
        "marketer_name": alloc.client.assigned_marketer.full_name if alloc.client.assigned_marketer else "",
        "transaction_date": alloc.date_allocated.strftime("%Y-%m-%d"),
        "total_amount": "",
        "special_notes": "",
        "installment_plan": "",
        "first_percent": "",
        "second_percent": "",
        "third_percent": "",
        "payment_duration": "",
        "custom_duration": "",
        "presale_price_set": presale_price_set,
        "presale_price": presale_price_value,
        "estate_name": alloc.estate.name
    }

    if txn:
        payload.update({
            "transaction_date": txn.transaction_date.strftime("%Y-%m-%d"),
            "total_amount": str(txn.total_amount),
            "special_notes": txn.special_notes or "",
            "installment_plan": txn.installment_plan or "",
            "first_percent": txn.first_percent or "",
            "second_percent": txn.second_percent or "",
            "third_percent": txn.third_percent or "",
            "payment_duration": txn.payment_duration or "",
            "custom_duration": txn.custom_duration or ""
        })

    return JsonResponse(payload)


@login_required
@require_GET
def ajax_companies_for_user(request):
    """Return JSON list of companies relevant to the current user.

    - Clients: companies where they have allocations
    - Marketers: companies where they have transactions
    - Admins: their company (if any)
    """
    user = request.user
    companies = []

    try:
        if getattr(user, 'role', None) == 'client':
            client_id = getattr(user, 'id', user)
            company_ids = (
                PlotAllocation.objects.filter(client_id=client_id)
                .values_list('estate__company', flat=True)
                .distinct()
            )
            qs = Company.objects.filter(id__in=[c for c in company_ids if c is not None])

        elif getattr(user, 'role', None) == 'marketer':
            company_ids = (
                Transaction.objects.filter(marketer=user)
                .values_list('company', flat=True)
                .distinct()
            )
            qs = Company.objects.filter(id__in=[c for c in company_ids if c is not None])

        else:
            # Admins: show their company if set
            comp = getattr(user, 'company_profile', None)
            qs = Company.objects.filter(id=comp.id) if comp else Company.objects.none()

        for comp in qs:
            alloc_count = PlotAllocation.objects.filter(client_id=getattr(user, 'id', user), estate__company=comp).count() if getattr(user, 'role', None) == 'client' else Transaction.objects.filter(marketer=user, company=comp).count() if getattr(user, 'role', None) == 'marketer' else 0
            total_invested = (
                Transaction.objects.filter(client_id=getattr(user, 'id', user), company=comp)
                .aggregate(total=Coalesce(Sum('total_amount'), Value(0, output_field=DecimalField())))['total']
                if getattr(user, 'role', None) == 'client' else 0
            )
            companies.append({
                'id': comp.id,
                'company_name': comp.company_name,
                'allocations': alloc_count,
                'total_invested': float(total_invested) if total_invested is not None else 0,
                'portfolio_url': reverse('my-company-portfolio', kwargs={'company_id': comp.id})
            })

        return JsonResponse({'companies': companies})

    except Exception as e:
        logger.exception('Failed to build companies list')
        return JsonResponse({'error': str(e)}, status=500)


@login_required
@require_GET
def ajax_company_portfolio(request, company_id):
    """Return rendered HTML snippet of the client's portfolio for the company."""
    user = request.user
    client_id = getattr(user, 'id', user)
    company = get_object_or_404(Company, id=company_id)

    allocations = (
        PlotAllocation.objects.filter(client_id=client_id, estate__company=company)
        .select_related('estate', 'plot_size', 'plot_number')
        .order_by('-date_allocated')
    )

    transactions = (
        Transaction.objects.filter(client_id=client_id, company=company)
        .select_related('allocation__estate', 'allocation__plot_size')
        .order_by('-transaction_date')
    )

    total_invested = transactions.aggregate(total=Coalesce(Sum('total_amount'), Value(0, output_field=DecimalField())))['total']

    html = render_to_string('client_side/_company_portfolio_panel.html', {
        'company': company,
        'allocations': allocations,
        'transactions': transactions,
        'total_invested': total_invested,
        'request': request,
    })

    return JsonResponse({'html': html})


@require_GET
def ajax_get_unpaid_installments(request):
    txn_id = request.GET.get("transaction_id")
    if not txn_id:
        return JsonResponse({"error": "Missing transaction_id"}, status=400)
    txn = get_object_or_404(Transaction, pk=txn_id)

    label_map = {
        1: "First",
        2: "Second",
        3: "Third",
    }

    payload = []
    for inst in txn.payment_installments:
        n = inst.get("n")
        payload.append({
            "n":         n,
            "label":     inst.get("label", label_map.get(n, f"Installment {n}")),
            "due":       f"{inst.get('due', 0):.2f}",
            "remaining": f"{inst.get('remaining', 0):.2f}",
        })

    return JsonResponse({
        "installments": payload,
        "total_balance": float(txn.balance),
        "allocation_id": txn.allocation.id if txn.allocation else None,
        "is_part_payment": txn.allocation.payment_type == 'part' if txn.allocation else False,
        "has_plot_number": bool(txn.allocation.plot_number) if txn.allocation else False
    })


@login_required
def generate_receipt_pdf(request, transaction_id):
    transaction = get_object_or_404(Transaction.objects.select_related(
        'client', 'allocation__estate', 'marketer'
    ), pk=transaction_id)
    
    # Generate a unique receipt ID
    receipt_id = f"REC-{uuid.uuid4().hex[:8].upper()}"
    today = datetime.date.today().strftime("%d %b %Y")
    
    context = {
        'transaction': transaction,
        'receipt_id': receipt_id,
        'date': today,
        'company': {
            'name': "NeuraLens Properties",
            'address': "123 Estate Avenue, City",
            'phone': "+234 800 000 0000",
            'email': "info@neuralensproperties.com",
            'logo': settings.STATIC_URL + "img/logo.png"
        }
    }
    
    template = get_template('admin_side/management_page_sections/payment_reciept.html')
    html = template.render(context)
    
    result = BytesIO()
    pdf = pisa.pisaDocument(BytesIO(html.encode("UTF-8")), result)
    
    if not pdf.err:
        response = HttpResponse(result.getvalue(), content_type='application/pdf')
        filename = f"Receipt_{receipt_id}_{transaction.client.full_name}.pdf"
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        return response
    
    return HttpResponse("Error generating PDF", status=400)


@login_required
@require_GET
def ajax_payment_history(request):
    """
    Return payment history for a transaction.
    
    SECURITY: Role-based access control:
    - Clients: Can only view payment history for their own transactions
    - Admins/Company Admins/Marketers/Support: Can view payment history within their company
    """
    txn_id = request.GET.get("transaction_id")
    if not txn_id:
        return JsonResponse({"error": "Missing transaction_id"}, status=400)

    user = request.user
    user_role = getattr(user, 'role', None)
    
    # Apply role-based filtering to fetch the transaction
    if user_role == 'client':
        # Clients can only view their own transactions
        txn = get_object_or_404(Transaction, pk=txn_id, client_id=user.id)
    elif user_role in ('admin', 'company_admin', 'marketer', 'support'):
        # Admins/Company Admins/Marketers/Support can view transactions within their company
        user_company = getattr(user, 'company_profile', None)
        if user_company:
            txn = get_object_or_404(Transaction, pk=txn_id, company=user_company)
        else:
            # Fallback: super admins without company can view any transaction
            txn = get_object_or_404(Transaction, pk=txn_id)
    else:
        # Unknown role - try to find by client ownership as fallback
        txn = Transaction.objects.filter(pk=txn_id, client_id=user.id).first()
        if not txn:
            return JsonResponse({'error': 'Unauthorized or transaction not found'}, status=403)

    # FULL PAYMENT → 1 line straight off Transaction
    if txn.allocation.payment_type == 'full':
        return JsonResponse({
            "payments": [{
                "date": txn.transaction_date.strftime("%d %b %Y"),
                "amount": str(txn.total_amount.quantize(Decimal('0.01'))),
                "method": txn.get_payment_method_display() or "",
                "installment": "Full Payment",
                "reference": txn.reference_code or ""
            }]
        })

    # PART‐PAYMENT → group by receipt
    qs = txn.payment_records.order_by('-payment_date')
    grouped = {}
    for p in qs:
        key = p.reference_code
        if key not in grouped:
            grouped[key] = {
                "date": p.payment_date.strftime("%d %b %Y"),
                "amount": Decimal('0.00'),
                "method": p.get_payment_method_display(),
                "installment": p.get_selected_installment_display() or "",  # Use selected installment
                "reference": key,
            }
        grouped[key]["amount"] += p.amount_paid

    payments = []
    for rec in grouped.values():
        rec["amount"] = str(rec["amount"].quantize(Decimal('0.01')))
        payments.append(rec)

    return JsonResponse({"payments": payments})

@require_POST
@transaction.atomic
def ajax_record_payment(request):
    txn_id = request.POST.get("transaction_id")
    inst_no = request.POST.get("installment")
    amt_s = request.POST.get("amount_paid")
    pay_date = request.POST.get("payment_date")
    method = request.POST.get("payment_method")

    if not all([txn_id, amt_s, pay_date, method]):
        return HttpResponseBadRequest("Missing required fields")

    txn = get_object_or_404(Transaction, pk=txn_id)
    total = Decimal(amt_s)
    
    # Cap payment at remaining balance
    balance = txn.balance
    if total > balance:
        total = balance

    # Generate reference code with company-specific prefix
    # CRITICAL: Company MUST be present - payment reference is legally sensitive
    company = txn.company or (txn.allocation.estate.company if txn.allocation and txn.allocation.estate else None)
    if not company:
        return JsonResponse({
            "success": False,
            "error": "Cannot record payment: Transaction company is missing. Payment reference generation requires valid company information for compliance."
        }, status=400)
    
    prefix = company._company_prefix()
    date_str = timezone.now().strftime("%Y%m%d")
    plot_raw = str(txn.allocation.plot_size)
    m = re.search(r'\d+', plot_raw)
    size_num = m.group(0) if m else plot_raw
    suffix = f"{random.randint(0, 9999):04d}"
    reference_code = f"{prefix}{date_str}-{size_num}-{suffix}"

    # Get installments in order
    installments = txn.payment_installments
    remaining = total
    records = []

    # Always allocate in installment order (1, 2, 3)
    for inst in installments:
        if remaining <= 0:
            break
            
        # Skip already paid installments
        if inst['remaining'] <= 0:
            continue
            
        # Calculate allocation amount
        slice_amt = min(inst['remaining'], remaining)
        
        records.append(
            PaymentRecord(
                transaction=txn,
                installment=inst['n'],
                amount_paid=slice_amt,
                payment_date=pay_date,
                payment_method=method,
                reference_code=reference_code,
                selected_installment=int(inst_no) if inst_no else None
            )
        )
        remaining -= slice_amt

    # Create all records at once
    PaymentRecord.objects.bulk_create(records)
    
    # Check if payment is now complete
    payment_complete = txn.balance <= 0
    allocation = txn.allocation
    needs_plot_allocation = payment_complete and allocation.payment_type == 'part' and not allocation.plot_number

    return JsonResponse({
        "success": True,
        "reference_code": reference_code,
        "payment_complete": payment_complete,
        "needs_plot_allocation": needs_plot_allocation,
        "transaction_id": txn.id,
        "allocation_id": allocation.id if allocation else None
    })

@require_http_methods(["GET"])
def ajax_get_available_plots(request):
    """
    Fetch available plot numbers for a specific allocation.
    Used when payment is complete and plot needs to be assigned.
    """
    allocation_id = request.GET.get('allocation_id')
    
    if not allocation_id:
        return JsonResponse({"success": False, "error": "Missing allocation_id"}, status=400)
    
    try:
        allocation = PlotAllocation.objects.get(pk=allocation_id)
        company = allocation.estate.company
        
        # Get the estate plot for this allocation's estate
        estate_plot = EstatePlot.objects.filter(estate=allocation.estate).first()
        
        if not estate_plot:
            return JsonResponse({
                "success": False, 
                "error": "No estate plot configuration found for this estate"
            }, status=404)
        
        # Find available plots for this estate that are linked to the estate plot
        # and not already allocated
        allocated_plot_ids = PlotAllocation.objects.filter(
            estate=allocation.estate,
            plot_number__isnull=False
        ).values_list('plot_number_id', flat=True)
        
        available_plots = PlotNumber.objects.filter(
            estates=estate_plot,
            company=company
        ).exclude(
            id__in=allocated_plot_ids
        ).order_by('number')
        
        plots_data = [{
            'id': plot.id,
            'number': plot.number
        } for plot in available_plots]
        
        return JsonResponse({
            "success": True,
            "plots": plots_data,
            "estate_name": allocation.estate.name,
            "plot_size": str(allocation.plot_size)
        })
    except PlotAllocation.DoesNotExist:
        return JsonResponse({"success": False, "error": "Allocation not found"}, status=404)
    except Exception as e:
        return JsonResponse({"success": False, "error": str(e)}, status=500)


@require_POST
@transaction.atomic
def ajax_assign_plot_after_payment(request):
    """
    Assign plot number to allocation after part payment is completed.
    This maintains transaction as 'part payment' but marks allocation as 'full' with plot number.
    """
    allocation_id = request.POST.get('allocation_id')
    plot_number_id = request.POST.get('plot_number_id')
    
    if not all([allocation_id, plot_number_id]):
        return JsonResponse({"success": False, "error": "Missing required fields"}, status=400)
    
    try:
        allocation = PlotAllocation.objects.get(pk=allocation_id)
        plot_number = PlotNumber.objects.get(pk=plot_number_id, company=allocation.estate.company)
        
        # Verify plot is still available
        if PlotAllocation.objects.filter(estate=allocation.estate, plot_number=plot_number).exists():
            return JsonResponse({
                "success": False,
                "error": f"Plot {plot_number.number} is already allocated to another client."
            }, status=400)
        
        # Assign plot and change allocation status to full
        allocation.plot_number = plot_number
        allocation.payment_type = 'full'  # Client is now considered fully allocated
        allocation.save()
        
        return JsonResponse({
            "success": True,
            "plot_number": plot_number.number,
            "message": f"Plot {plot_number.number} successfully allocated to {allocation.client.full_name}"
        })
    except PlotAllocation.DoesNotExist:
        return JsonResponse({"success": False, "error": "Allocation not found"}, status=404)
    except PlotNumber.DoesNotExist:
        return JsonResponse({"success": False, "error": "Plot number not found"}, status=404)
    except Exception as e:
        return JsonResponse({"success": False, "error": str(e)}, status=500)


@login_required
def ajax_get_transaction_details(request):
    """Fetch transaction details for editing"""
    try:
        transaction_id = request.GET.get('transaction_id')
        
        if not transaction_id:
            return JsonResponse({"success": False, "error": "Transaction ID required"}, status=400)
        
        # Get company - handle both direct company FK and through allocation
        try:
            company = request.user.company
            company_id = company.id if hasattr(company, 'id') else company
        except AttributeError:
            return JsonResponse({"success": False, "error": "User has no associated company"}, status=403)
        
        # Get transaction with company isolation - use all_objects to bypass CompanyAwareManager
        txn = Transaction.all_objects.select_related(
            'allocation__estate__company',
            'allocation__plot_size_unit',
            'allocation__plot_number',
            'marketer',
            'client',
            'company'
        ).get(id=int(transaction_id), company_id=company_id)
        
        alloc = txn.allocation
        
        # Get payment records for this transaction - use all_objects to bypass CompanyAwareManager
        payment_records = PaymentRecord.all_objects.filter(
            transaction=txn
        ).order_by('installment', 'payment_date').values(
            'id', 'installment', 'amount_paid', 'payment_date', 'payment_method'
        )
        
        # Get client name
        client_name = txn.client.get_full_name() if hasattr(txn.client, 'get_full_name') else str(txn.client)
        
        # Build response data
        data = {
            "success": True,
            "transaction": {
                "id": txn.id,
                "client": client_name,
                "marketer": txn.marketer.get_full_name() if txn.marketer else "N/A",
                "transaction_date": txn.transaction_date.strftime('%Y-%m-%d'),
                "total_amount": str(txn.total_amount),
                "payment_method": txn.payment_method or ""
            },
            "allocation": {
                "id": alloc.id,
                "estate": alloc.estate.name,
                "plot_size": alloc.plot_size,
                "payment_type": alloc.payment_type,
                "plot_number": alloc.plot_number.plot_number if alloc.plot_number else None,
                "plot_number_id": alloc.plot_number.id if alloc.plot_number else None,
                "first_installment": str(alloc.first_installment) if alloc.first_installment else None,
                "second_installment": str(alloc.second_installment) if alloc.second_installment else None,
                "third_installment": str(alloc.third_installment) if alloc.third_installment else None
            },
            "payment_records": [
                {
                    "id": pr['id'],
                    "installment": pr['installment'],
                    "amount_paid": str(pr['amount_paid']),
                    "payment_date": pr['payment_date'].strftime('%Y-%m-%d'),
                    "payment_method": pr['payment_method']
                }
                for pr in payment_records
            ]
        }
        
        return JsonResponse(data)
        
    except Transaction.DoesNotExist:
        return JsonResponse({"success": False, "error": "Transaction not found"}, status=404)
    except Exception as e:
        return JsonResponse({"success": False, "error": str(e)}, status=500)


@login_required
@require_http_methods(["POST"])
def ajax_update_transaction(request):
    """Update transaction details"""
    try:
        transaction_id = request.POST.get('transaction_id')
        
        if not transaction_id:
            return JsonResponse({"success": False, "error": "Transaction ID required"}, status=400)
        
        # Get company
        try:
            company = request.user.company
            company_id = company.id if hasattr(company, 'id') else company
        except AttributeError:
            return JsonResponse({"success": False, "error": "User has no associated company"}, status=403)
        
        # Get transaction with company isolation - use all_objects to bypass CompanyAwareManager
        txn = Transaction.all_objects.select_related(
            'allocation__estate__company',
            'allocation__plot_number',
            'company'
        ).get(id=int(transaction_id), company_id=company_id)
        alloc = txn.allocation
        
        # Update transaction fields
        txn.transaction_date = request.POST.get('transaction_date')
        txn.total_amount = Decimal(request.POST.get('total_amount', 0))
        txn.payment_method = request.POST.get('payment_method', '')
        
        # Update installment fields if part payment
        if alloc.payment_type == 'part':
            first_inst = request.POST.get('first_installment')
            second_inst = request.POST.get('second_installment')
            third_inst = request.POST.get('third_installment')
            
            if first_inst:
                alloc.first_installment = Decimal(first_inst)
            if second_inst:
                alloc.second_installment = Decimal(second_inst)
            if third_inst:
                alloc.third_installment = Decimal(third_inst)
            
            # Update individual payment records
            for key, value in request.POST.items():
                if key.startswith('payment_record_') and value:
                    payment_id = key.replace('payment_record_', '')
                    try:
                        payment_record = PaymentRecord.objects.get(
                            id=payment_id,
                            transaction=txn
                        )
                        payment_record.amount_paid = Decimal(value)
                        payment_record.save()
                    except PaymentRecord.DoesNotExist:
                        pass
                    except ValueError:
                        pass
        
        # Handle plot number change
        new_plot_id = request.POST.get('plot_number_id')
        if new_plot_id:
            # Release current plot
            if alloc.plot_number:
                old_plot = alloc.plot_number
                old_plot.status = 'available'
                old_plot.save()
            
            # Assign new plot - validate it belongs to same company
            new_plot = PlotNumber.objects.get(
                id=new_plot_id,
                company=alloc.estate.company,
                status='available'
            )
            new_plot.status = 'sold'
            new_plot.save()
            
            alloc.plot_number = new_plot
        
        # Save changes
        alloc.save()
        txn.save()
        
        return JsonResponse({
            "success": True,
            "message": "Transaction updated successfully"
        })
        
    except Transaction.DoesNotExist:
        return JsonResponse({"success": False, "error": "Transaction not found"}, status=404)
    except PlotNumber.DoesNotExist:
        return JsonResponse({"success": False, "error": "Selected plot is not available"}, status=400)
    except ValueError as e:
        return JsonResponse({"success": False, "error": f"Invalid number format: {str(e)}"}, status=400)
    except Exception as e:
        return JsonResponse({"success": False, "error": str(e)}, status=500)


def payment_receipt(request, reference_code):
    # For full payments
    if Transaction.objects.filter(reference_code=reference_code).exists():
        txn = Transaction.objects.get(reference_code=reference_code)
        payment = None
        payments = []
        payments_total = txn.total_amount
    else:
        # For installment payments
        payments = PaymentRecord.objects.filter(reference_code=reference_code)
        if not payments.exists():
            return HttpResponse("Receipt not found", status=404)
        
        txn = payments.first().transaction
        payment = payments.first()
        payments_total = sum(p.amount_paid for p in payments)
    
    # Resolve company from transaction allocation (tenant-aware)
    company_obj = None
    try:
        company_obj = txn.allocation.estate.company
    except Exception:
        company_obj = None

    # Generate a company-scoped sequential receipt number when missing.
    receipt_number = None
    if company_obj:
        from .models import CompanySequence

        if payment:
            if not payment.receipt_number:
                seq = CompanySequence.get_next(company_obj, 'receipt')
                receipt_number = f"{company_obj._company_prefix()}-RC-{seq:06d}"
                payment.receipt_number = receipt_number
                payment.receipt_generated = True
                payment.receipt_date = timezone.now()
                payment.save()
            else:
                receipt_number = payment.receipt_number
        else:
            # Full payment (no PaymentRecord): generate a non-persistent receipt number for display
            seq = CompanySequence.get_next(company_obj, 'receipt')
            receipt_number = f"{company_obj._company_prefix()}-RC-{seq:06d}"

    # Build company dict for template (use company fields when available)
    company_context = {
        'name': (company_obj.company_name if (company_obj and getattr(company_obj, 'company_name', None)) else '-'),
        'office_address': (company_obj.office_address if (company_obj and getattr(company_obj, 'office_address', None)) else '-'),
        'address': (company_obj.location if (company_obj and getattr(company_obj, 'location', None)) else '-'),
        'phone': (company_obj.phone if (company_obj and getattr(company_obj, 'phone', None)) else '-'),
        'email': (company_obj.email if (company_obj and getattr(company_obj, 'email', None)) else '-'),
        'website': (company_obj.custom_domain if (company_obj and getattr(company_obj, 'custom_domain', None)) else '-'),
        'logo_url': (company_obj.logo.url if (company_obj and getattr(company_obj, 'logo', None)) else None),
        'cashier_name': (company_obj.cashier_name if (company_obj and getattr(company_obj, 'cashier_name', None)) else '-'),
        'cashier_signature_url': (company_obj.cashier_signature.url if (company_obj and getattr(company_obj, 'cashier_signature', None)) else None),
    }

    # Ensure numeric totals are Decimal and compute formatted displays
    try:
        payments_total_dec = Decimal(payments_total) if payments_total is not None else Decimal('0.00')
    except Exception:
        payments_total_dec = Decimal('0.00')

    try:
        total_amount_dec = Decimal(txn.total_amount) if txn and getattr(txn, 'total_amount', None) is not None else Decimal('0.00')
    except Exception:
        total_amount_dec = Decimal('0.00')

    # Calculate current payment amount (for display in receipt)
    current_payment_amount = Decimal('0.00')
    total_paid_so_far = Decimal('0.00')  # Cumulative: current + all previous
    is_first_installment = False
    
    if payment:
        current_payment_amount = Decimal(payment.amount_paid)
        # Calculate total paid INCLUDING this current payment (cumulative)
        all_payments_for_transaction = PaymentRecord.objects.filter(
            transaction=txn
        ).order_by('payment_date', 'id')
        
        total_paid_so_far = Decimal('0.00')
        for p in all_payments_for_transaction:
            total_paid_so_far += Decimal(p.amount_paid)
            if p.id == payment.id:
                # Stop after processing the current payment
                break
        
        # Check if this is the first installment (no previous payments before this one)
        is_first_installment = (total_paid_so_far == current_payment_amount)
    else:
        current_payment_amount = payments_total_dec
        total_paid_so_far = payments_total_dec

    # Outstanding balance: never negative for display
    # Calculate as Total Price MINUS Total Amount Paid So Far
    outstanding_dec = total_amount_dec - total_paid_so_far
    if outstanding_dec < Decimal('0'):
        outstanding_dec = Decimal('0.00')

    # Format money with thousand separators and 2 decimals
    def money_fmt(d: Decimal) -> str:
        try:
            return f"₦ {d:,.2f}"
        except Exception:
            return f"₦ {Decimal(d):,.2f}"

    # Convert amount to words (supports integers and two-decimal kobo)
    def _num_to_words(n: int) -> str:
        units = ["", "One", "Two", "Three", "Four", "Five", "Six", "Seven", "Eight", "Nine", "Ten", "Eleven", "Twelve", "Thirteen", "Fourteen", "Fifteen", "Sixteen", "Seventeen", "Eighteen", "Nineteen"]
        tens = ["", "", "Twenty", "Thirty", "Forty", "Fifty", "Sixty", "Seventy", "Eighty", "Ninety"]

        if n < 20:
            return units[n]
        if n < 100:
            return tens[n // 10] + ('' if n % 10 == 0 else ' ' + units[n % 10])
        if n < 1000:
            return units[n // 100] + ' Hundred' + ('' if n % 100 == 0 else ' and ' + _num_to_words(n % 100))

        for idx, word in enumerate(['Thousand', 'Million', 'Billion', 'Trillion'], 1):
            unit = 1000 ** idx
            if n < unit * 1000:
                major = n // unit
                rem = n % unit
                return _num_to_words(major) + ' ' + word + ('' if rem == 0 else ' ' + _num_to_words(rem))

    def amount_in_words(amount: Decimal) -> str:
        amt = amount.quantize(Decimal('0.01'), rounding=ROUND_HALF_UP)
        naira = int(amt // 1)
        kobo = int((amt - Decimal(naira)) * 100)
        if naira == 0 and kobo == 0:
            return 'Zero Naira'
        parts = []
        if naira > 0:
            parts.append((_num_to_words(naira) + ' Naira'))
        if kobo > 0:
            parts.append((_num_to_words(kobo) + ' Kobo'))
        return ' and '.join(parts)

    amount_words = amount_in_words(payments_total_dec)

    context = {
        'transaction': txn,
        'payment': payment,
        'payments': payments,
        'payments_total': payments_total_dec,
        'payments_total_display': money_fmt(payments_total_dec),
        'total_amount_display': money_fmt(total_amount_dec),
        'balance_display': money_fmt(outstanding_dec),
        'subtotal_display': money_fmt(total_amount_dec),
        'paid_display': money_fmt(payments_total_dec),
        'current_payment_display': money_fmt(current_payment_amount),
        'total_paid_so_far_display': money_fmt(total_paid_so_far),
        'outstanding_display': money_fmt(outstanding_dec),
        'amount_in_words': amount_words,
        'today': timezone.now().date(),
        'company': company_context,
        'receipt_number': receipt_number,
        'is_part_payment': payment is not None,  # True if this is a part/installment payment
        'is_first_installment': is_first_installment,  # True if this is the FIRST payment
    }
    
    # Prepare payments for display (add display_amount) and render HTML
    payments_list = []
    if payments is not None:
        try:
            for p in payments:
                # attach a display value for each payment
                try:
                    p.display_amount = money_fmt(Decimal(p.amount_paid))
                except Exception:
                    p.display_amount = money_fmt(Decimal('0.00'))
                payments_list.append(p)
        except Exception:
            payments_list = list(payments) if hasattr(payments, '__iter__') else []

    context['payments'] = payments_list
    template = 'admin_side/management_page_sections/absolute_payment_reciept.html'
    html_string = render_to_string(template, context)

    # If the rendered template already contains a full HTML document (doctype/html tags),
    # pass it directly to the PDF renderer to preserve the template's own CSS and structure.
    rendered_lower = html_string.lower()
    if '<html' in rendered_lower or '<!doctype' in rendered_lower:
        html_to_render = html_string
    else:
        # Otherwise, wrap with a minimal HTML document and safe defaults.
        html_to_render = f"""
        <html>
        <head>
            <style>
                @page {{ size: A4; margin: 20mm; }}
                html, body {{ font-family: 'Poppins', sans-serif !important; font-size: 12px; margin: 0; padding: 0; }}
            </style>
        </head>
        <body>
            {html_string}
        </body>
        </html>
        """

    # Preprocess HTML to convert CSS units (px, mm) to points for reportlab compatibility
    def _convert_units_to_pt(s: str) -> str:
        # px -> pt (1px = 0.75pt), mm -> pt (1mm = 2.834645669pt)
        def _px(m):
            v = float(m.group(1))
            pt = round(v * 0.75, 2)
            if float(pt).is_integer():
                pt = int(pt)
            return f"{pt}pt"

        def _mm(m):
            v = float(m.group(1))
            pt = round(v * 2.834645669, 2)
            if float(pt).is_integer():
                pt = int(pt)
            return f"{pt}pt"

        s = re.sub(r"(\d+(?:\.\d+)?)px", _px, s)
        s = re.sub(r"(\d+(?:\.\d+)?)mm", _mm, s)
        return s

    html_processed = _convert_units_to_pt(html_to_render)

    # Debug mode: return the processed HTML used to generate the PDF for inspection
    if request.GET.get('debug') == 'html':
        return HttpResponse(html_processed, content_type='text/html')

    # Create PDF using xhtml2pdf with fallback and diagnostics
    result = BytesIO()
    try:
        pdf = pisa.pisaDocument(BytesIO(html_processed.encode("UTF-8")), result)
    except Exception as e:
        logger.exception("PDF generation failed for receipt %s: %s", reference_code, str(e))
        # Return the processed HTML and error for debugging when in debug mode
        if settings.DEBUG:
            return HttpResponse(f"<h3>PDF generation error:</h3><pre>{str(e)}</pre>\n\n" + html_processed, content_type='text/html')
        return HttpResponse("Error generating PDF", status=500)

    if not pdf.err:
        # Create HTTP response
        response = HttpResponse(content_type='application/pdf')
        filename = f"receipt_{reference_code}.pdf"
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        response.write(result.getvalue())
        return response
    else:
        logger.error("xhtml2pdf reported errors while generating receipt %s", reference_code)
        if settings.DEBUG:
            return HttpResponse("Error generating PDF", status=500)
        return HttpResponse("Error generating PDF", status=500)

@login_required
@require_GET
def ajax_transaction_details(request, transaction_id):
    """
    Return transaction details for modal display.
    
    SECURITY: Role-based access control:
    - Clients: Can only view their own transactions
    - Admins/Company Admins/Marketers/Support: Can view transactions within their company
    """
    user = request.user
    user_role = getattr(user, 'role', None)
    
    # Build base queryset with related objects
    base_qs = Transaction.objects.select_related(
        'allocation__estate',
        'allocation__plot_size',
        'client',
        'marketer',
        'company'
    )
    
    # Apply role-based filtering
    if user_role == 'client':
        # Clients can only view their own transactions
        txn = get_object_or_404(base_qs, pk=transaction_id, client_id=user.id)
    elif user_role in ('admin', 'company_admin', 'marketer', 'support'):
        # Admins/Company Admins/Marketers/Support can view transactions within their company
        user_company = getattr(user, 'company_profile', None)
        if user_company:
            txn = get_object_or_404(base_qs, pk=transaction_id, company=user_company)
        else:
            # Fallback: super admins without company can view any transaction
            txn = get_object_or_404(base_qs, pk=transaction_id)
    else:
        # Unknown role - try to find by client ownership as fallback
        # This handles edge cases where role might be None or unexpected
        txn = base_qs.filter(pk=transaction_id, client_id=user.id).first()
        if not txn:
            return JsonResponse({'error': 'Unauthorized or transaction not found'}, status=403)

    data = {
        'client': txn.client.full_name,
        'marketer': txn.marketer.full_name if txn.marketer else '',
        'transaction_date': txn.transaction_date.strftime("%d %b %Y"),
        'total_amount': str(txn.total_amount),
        'status': txn.status,
        'allocation': {
            'estate': {
                'name': txn.allocation.estate.name
            },
            'plot_size': str(txn.allocation.plot_size) if txn.allocation.plot_size else 'N/A',
            'payment_type': txn.allocation.payment_type
        },
        'payment_duration': txn.payment_duration,
        'custom_duration': txn.custom_duration,
        'installment_plan': txn.installment_plan,
        'first_percent': txn.first_percent,
        'second_percent': txn.second_percent,
        'third_percent': txn.third_percent,
        'first_installment': str(txn.first_installment) if txn.first_installment else None,
        'second_installment': str(txn.second_installment) if txn.second_installment else None,
        'third_installment': str(txn.third_installment) if txn.third_installment else None,
    }

    return JsonResponse(data)

@login_required
@require_GET
def ajax_existing_transaction(request):
    """
    Given client_id & allocation_id, return the existing transaction's
    payment_method and reference_code (if any), so the form can pre-fill.
    Company-isolated for multi-tenant security.
    """
    # Company isolation - get user's company
    user_company = getattr(request.user, 'company_profile', None)
    if not user_company:
        return JsonResponse({'error': 'No company access'}, status=403)
    
    client_id     = request.GET.get('client_id')
    allocation_id = request.GET.get('allocation_id')

    # Company isolation - filter by company
    txn = (
        Transaction.objects
        .filter(client_id=client_id, allocation_id=allocation_id, company=user_company)
        .first()
    )

    if not txn:
        return JsonResponse({'payment_method': None, 'reference_code': None})

    return JsonResponse({
        'payment_method': txn.payment_method,
        'reference_code': txn.reference_code,
    })

@login_required
@require_POST
def ajax_send_receipt(request):
    # Company isolation - get user's company
    user_company = getattr(request.user, 'company_profile', None)
    if not user_company:
        return JsonResponse({"error": "No company access"}, status=403)
    
    txn_id = request.POST.get("transaction_id")
    if not txn_id:
        return JsonResponse({"error":"Missing transaction_id"}, status=400)

    try:
        # Company isolation - filter by company
        txn = Transaction.objects.get(pk=txn_id, company=user_company)
        # Here you would implement your actual email sending logic
        # For now we'll just simulate success
        return JsonResponse({"success": True, "message": "Receipt sent successfully"})
    except Transaction.DoesNotExist:
        return JsonResponse({"error": "Transaction not found"}, status=404)
    except Exception as e:
        return JsonResponse({"error": str(e)}, status=500)


class CustomLogoutView(LogoutView):
    """
    Custom logout view that handles tenant-aware redirects.
    Ensures logout redirects to the appropriate tenant login page.
    """
    
    http_method_names = ['get', 'post', 'options']  # Allow GET requests for logout
    
    def get(self, request, *args, **kwargs):
        """
        Handle GET requests by logging out and redirecting.
        """
        from django.contrib.auth import logout
        from django.shortcuts import redirect
        
        # Log the user out
        logout(request)
        
        # Determine the redirect URL
        next_page = self.get_redirect_url()
        
        # Redirect to the next page
        return redirect(next_page)
    
    def get_redirect_url(self):
        """
        Determine the appropriate redirect URL after logout.
        Always redirect to the main login page.
        """
        # Check for next parameter in URL
        next_url = self.request.GET.get('next')
        if next_url:
            return next_url
        
        # Always redirect to main login page
        return reverse('login')


# Sales Volume
class MarketersAPI(View):
    def get(self, request):
        # SECURITY: Only return marketers from the same company as the requesting admin
        if not request.user.is_authenticated or request.user.role != 'admin':
            return JsonResponse([], safe=False)
        
        # Get marketers from both company_profile and MarketerAffiliation
        company = request.user.company_profile
        if not company:
            return JsonResponse([], safe=False)
        
        # Get marketers from company_profile
        company_marketers = MarketerUser.objects.filter(company_profile=company).values('id', 'full_name')
        
        # Get marketers from MarketerAffiliation
        affiliation_marketer_ids = list(
            MarketerAffiliation.objects.filter(company=company).values_list('marketer_id', flat=True).distinct()
        )
        affiliation_marketers = MarketerUser.objects.filter(
            id__in=affiliation_marketer_ids
        ).exclude(
            id__in=company_marketers.values_list('id', flat=True)
        ).values('id', 'full_name')
        
        # Combine both lists
        all_marketers = list(company_marketers) + list(affiliation_marketers)
        
        return JsonResponse(all_marketers, safe=False)

class MarketerPerformanceView(View):
    """Redirects to Company Profile with Analytics tab active"""
    
    def get(self, request):
        # Redirect to company profile page with analytics tab
        from django.shortcuts import redirect
        return redirect('/company-profile/?tab=analytics')
    
    def post(self, request):
        return JsonResponse({'status': 'error'}, status=400)


class PerformanceDataAPI(View):
    def get(self, request):
        period_type     = request.GET.get('period_type', 'monthly')
        specific_period = request.GET.get('specific_period')
        if not specific_period:
            return JsonResponse({'error': 'Specific period required'}, status=400)

        today = timezone.now().date()
        start_date, end_date = self.get_date_range(period_type, specific_period)

        # 1) Gather all metrics first
        response_data    = []
        record_payloads  = []

        # SECURITY: Filter by company to prevent cross-tenant leakage
        company = request.user.company_profile
        
        # Get marketers from both company_profile and MarketerAffiliation
        company_marketers_qs = MarketerUser.objects.filter(company_profile=company) if company else MarketerUser.objects.none()
        affiliation_marketer_ids = list(
            MarketerAffiliation.objects.filter(company=company).values_list('marketer_id', flat=True).distinct()
        ) if company else []
        affiliation_marketers = MarketerUser.objects.filter(id__in=affiliation_marketer_ids).exclude(
            id__in=company_marketers_qs.values_list('pk', flat=True)
        ) if affiliation_marketer_ids else MarketerUser.objects.none()
        
        company_marketers = list(company_marketers_qs) + list(affiliation_marketers)
        
        for marketer in company_marketers:
            # transactions in period
            txns = Transaction.objects.filter(
                marketer=marketer,
                company=company,
                transaction_date__range=(start_date, end_date)
            )
            closed_deals  = txns.count()
            total_sales   = txns.aggregate(total=Sum('total_amount'))['total'] or 0

            # commission lookup (COMPANY-SCOPED)
            # Get the most recent commission rate for this marketer in this specific company
            # COMMISSIONS ARE STRICTLY COMPANY-BASED AND MARKETER-SPECIFIC - NO GLOBAL COMMISSIONS
            commission = (
                MarketerCommission.objects
                .filter(
                    marketer=marketer,  # ✅ Only marketer-specific commissions
                    company=company,    # ✅ Company-scoped
                    effective_date__lte=today
                )
                .order_by('-effective_date')
                .first()
            )
            rate               = commission.rate if commission else 0
            commission_earned  = total_sales * rate / 100

            # target lookup: individual marketer target first, then fall back to company-wide target
            # 1. Try marketer-specific target for this period
            specific_tgt = MarketerTarget.objects.filter(
                marketer=marketer,  # ✅ Marketer-specific only
                company=company,    # ✅ Company-scoped
                period_type=period_type,
                specific_period=specific_period
            ).first()
            
            # 2. If no marketer-specific target, try company-wide target for this period
            if not specific_tgt:
                specific_tgt = MarketerTarget.objects.filter(
                    marketer=None,      # Company-wide target (no specific marketer)
                    company=company,
                    period_type=period_type,
                    specific_period=specific_period
                ).first()
            
            # 3. If still no target, try to get a general annual target for the year
            if not specific_tgt and specific_period:
                year = specific_period.split('-')[0] if '-' in specific_period else specific_period
                # Try marketer-specific annual target
                specific_tgt = MarketerTarget.objects.filter(
                    marketer=marketer,
                    company=company,
                    period_type='annual',
                    specific_period=year
                ).first()
                # If not, try company-wide annual target
                if not specific_tgt:
                    specific_tgt = MarketerTarget.objects.filter(
                        marketer=None,
                        company=company,
                        period_type='annual',
                        specific_period=year
                    ).first()
            
            tgt_amt = specific_tgt.target_amount if specific_tgt else 0

            target_percent = round((total_sales / tgt_amt * 100), 1) if tgt_amt else 0

            # Get profile image URL
            profile_image_url = None
            if marketer.profile_image:
                try:
                    profile_image_url = marketer.profile_image.url
                except Exception:
                    profile_image_url = None

            # stash for JSON
            response_data.append({
                'marketer_id':       marketer.id,
                'marketer_name':     marketer.full_name,
                'profile_image':     profile_image_url,
                'closed_deals':      closed_deals,
                'total_sales':       float(total_sales),
                'commission_rate':   float(rate),
                'commission_earned': float(commission_earned),
                'target_amount':     float(tgt_amt),
                'target_percent':    target_percent,
            })

            # stash for DB write
            record_payloads.append({
                'marketer':         marketer,
                'period_type':      period_type,
                'specific_period':  specific_period,
                'closed_deals':     closed_deals,
                'total_sales':      total_sales,
                'commission_earned': commission_earned,
            })

        # 2) Write performance records in their own small transactions
        for rec in record_payloads:
            try:
                with transaction.atomic():
                    MarketerPerformanceRecord.objects.update_or_create(
                        marketer=rec['marketer'],
                        period_type=rec['period_type'],
                        specific_period=rec['specific_period'],
                        defaults={
                            'closed_deals':      rec['closed_deals'],
                            'total_sales':       rec['total_sales'],
                            'commission_earned': rec['commission_earned'],
                        }
                    )
            except DatabaseError:
                # Skip on DB lock or other error
                continue

        # 3) Get company-wide target for this period
        company_target_record = MarketerTarget.objects.filter(
            marketer=None,
            company=company,
            period_type=period_type,
            specific_period=specific_period
        ).first()
        company_target = float(company_target_record.target_amount) if company_target_record else 0

        # Return both performance data and company target
        return JsonResponse({
            'performance_data': response_data,
            'company_target': company_target
        }, safe=False)

    def get_date_range(self, period_type, specific_period):
        today = timezone.now().date()

        if period_type == 'monthly':
            year, month = map(int, specific_period.split('-'))
            start = datetime(year, month, 1).date()
            end   = (start + relativedelta(months=1)) - timedelta(days=1)

        elif period_type == 'quarterly':
            y_str, q_str = specific_period.split('-Q')
            y, q = int(y_str), int(q_str)
            start_month = 3*(q-1) + 1
            start = datetime(y, start_month, 1).date()
            if q == 4:
                next_start = datetime(y+1, 1, 1).date()
            else:
                next_start = datetime(y, start_month+3, 1).date()
            end = next_start - timedelta(days=1)

        else:  # annual
            y = int(specific_period)
            start = datetime(y, 1, 1).date()
            end   = datetime(y, 12, 31).date()

        return start, end


class SetTargetAPI(View):
    @method_decorator(csrf_exempt)
    def dispatch(self, *args, **kwargs):
        return super().dispatch(*args, **kwargs)
        
    def get(self, request):
        return HttpResponseNotAllowed(['POST'])

    def post(self, request):
        # Parse payload
        try:
            data = json.loads(request.body)
        except json.JSONDecodeError:
            return JsonResponse({'status':'error','message':'Invalid JSON'}, status=400)

        period_type     = data.get('period_type')
        specific_period = data.get('specific_period')
        target_amount   = data.get('target_amount')
        marketer_id     = data.get('marketer')  # Can be 'all' or a specific marketer ID

        # Validate
        if not period_type or not specific_period:
            return JsonResponse({'status':'error','message':'period_type & specific_period required'}, status=400)
        if target_amount is None:
            return JsonResponse({'status':'error','message':'target_amount required'}, status=400)

        # SECURITY: Company-scoped target setting
        company = request.user.company_profile
        
        if marketer_id == 'all' or not marketer_id:
            # Set company-wide target (marketer=None)
            MarketerTarget.objects.update_or_create(
                marketer=None,
                company=company,
                period_type=period_type,
                specific_period=specific_period,
                defaults={'target_amount': target_amount}
            )
        else:
            # Set individual marketer target
            # SECURITY: Check both direct company_profile AND MarketerAffiliation
            try:
                marketer_id_int = int(marketer_id)
                
                # First try direct company_profile match
                marketer = MarketerUser.objects.filter(
                    id=marketer_id_int, 
                    company_profile=company
                ).first()
                
                # If not found, check MarketerAffiliation
                if not marketer:
                    affiliation = MarketerAffiliation.objects.filter(
                        marketer_id=marketer_id_int,
                        company=company
                    ).first()
                    if affiliation:
                        marketer = affiliation.marketer
                
                if not marketer:
                    return JsonResponse({'status':'error','message':'Invalid marketer ID or not in your company'}, status=400)
                
                MarketerTarget.objects.update_or_create(
                    marketer=marketer,
                    company=company,
                    period_type=period_type,
                    specific_period=specific_period,
                    defaults={'target_amount': target_amount}
                )
            except ValueError:
                return JsonResponse({'status':'error','message':'Invalid marketer ID'}, status=400)

        return JsonResponse({'status':'success'})


class GetGlobalTargetAPI(View):
    def get(self, request):
        period_type     = request.GET.get('period_type')
        specific_period = request.GET.get('specific_period')
        marketer_id     = request.GET.get('marketer')  # Can be 'all' or a specific marketer ID

        if not period_type or not specific_period:
            return JsonResponse({'status':'error','message':'period_type & specific_period required'}, status=400)

        company = request.user.company_profile
        
        if marketer_id == 'all' or not marketer_id:
            # Get company-wide target
            record = MarketerTarget.objects.filter(
                marketer=None,
                company=company,
                period_type=period_type,
                specific_period=specific_period
            ).first()
        else:
            # Get individual marketer target
            # SECURITY: Check both direct company_profile AND MarketerAffiliation
            try:
                marketer_id_int = int(marketer_id)
                
                # First try direct company_profile match
                marketer = MarketerUser.objects.filter(
                    id=marketer_id_int, 
                    company_profile=company
                ).first()
                
                # If not found, check MarketerAffiliation
                if not marketer:
                    affiliation = MarketerAffiliation.objects.filter(
                        marketer_id=marketer_id_int,
                        company=company
                    ).first()
                    if affiliation:
                        marketer = affiliation.marketer
                
                if not marketer:
                    return JsonResponse({'target_amount': None})
                
                record = MarketerTarget.objects.filter(
                    marketer=marketer,
                    company=company,
                    period_type=period_type,
                    specific_period=specific_period
                ).first()
            except ValueError:
                return JsonResponse({'target_amount': None})

        return JsonResponse({
            'target_amount': float(record.target_amount) if record else None
        })


class SetCommissionAPI(View):
    @method_decorator(csrf_exempt)
    def dispatch(self, *args, **kwargs):
        return super().dispatch(*args, **kwargs)
        
    def get(self, request):
        return HttpResponseNotAllowed(['POST'])

    def post(self, request):
        try:
            payload = json.loads(request.body)
        except json.JSONDecodeError:
            return JsonResponse({'status':'error','message':'Invalid JSON'}, status=400)

        marketer_id    = payload.get('marketer')
        rate           = payload.get('commission_rate')
        effective_date = payload.get('effective_date')

        # Basic validation
        if marketer_id is None or rate is None or not effective_date:
            return JsonResponse({'status':'error','message':'Missing parameters'}, status=400)

        # SECURITY: Verify user is authenticated and has admin privileges
        if not request.user.is_authenticated:
            return JsonResponse({'status':'error','message':'Authentication required'}, status=401)
        
        if request.user.role != 'admin':
            return JsonResponse({'status':'error','message':'Admin access required'}, status=403)

        # Fetch exactly one marketer
        try:
            marketer = MarketerUser.objects.get(id=marketer_id)
        except MarketerUser.DoesNotExist:
            return JsonResponse({'status':'error','message':'Marketer not found'}, status=404)

        # SECURITY: Verify the admin has permission to set commissions for this company
        # Only company admins (not system admins) should be able to set commissions for their company
        if request.user.admin_level != 'company':
            return JsonResponse({'status':'error','message':'Only company admins can set commissions'}, status=403)

        # SECURITY: Verify marketer is affiliated with the admin's company
        # Check if marketer has direct company_profile OR MarketerAffiliation with this company
        marketer_company = marketer.company_profile
        marketer_affiliation = MarketerAffiliation.objects.filter(
            marketer=marketer,
            company=request.user.company_profile
        ).first()
        
        if marketer_company != request.user.company_profile and not marketer_affiliation:
            return JsonResponse({'status':'error','message':'Access denied: Marketer not in your company'}, status=403)

        # ALWAYS create a new commission record for this company
        # This ensures complete isolation - each company maintains their own commission rates
        # regardless of what other companies have set for the same marketer
        MarketerCommission.objects.create(
            marketer=marketer,
            rate=rate,
            effective_date=effective_date,
            company=request.user.company_profile  # ✅ Use admin's company for isolation
        )

        return JsonResponse({'status':'success'})

class GetCommissionAPI(View):
    def get(self, request):
        marketer_id = request.GET.get('marketer')
        if not marketer_id:
            return JsonResponse({'status':'error','message':'marketer param required'}, status=400)

        try:
            marketer = MarketerUser.objects.get(id=marketer_id)
        except MarketerUser.DoesNotExist:
            return JsonResponse({'status':'error','message':'Marketer not found'}, status=404)

        # SECURITY: Verify marketer is affiliated with the requesting user's company
        # This prevents unauthorized access to marketer commission data
        company = request.user.company_profile
        marketer_in_company = (
            marketer.company_profile == company or
            MarketerAffiliation.objects.filter(marketer=marketer, company=company).exists()
        )
        
        if not marketer_in_company:
            return JsonResponse({'status':'error','message':'Access denied: Marketer not in your company'}, status=403)

        # SECURITY: Use request.user.company_profile to ensure company isolation
        # This prevents cross-company data leakage when a marketer works for multiple companies
        commission = MarketerCommission.objects.filter(
            marketer=marketer,
            company=request.user.company_profile  # ✅ Use requesting user's company for isolation
        ).order_by('-effective_date').first()

        if not commission:
            # no commission yet
            return JsonResponse({'commission_rate': None, 'effective_date': None})

        return JsonResponse({
            'commission_rate': commission.rate,
            'effective_date': commission.effective_date.isoformat()
        })


class GetAvailableYearsAPI(View):
    """
    API to get dynamically available years from the database.
    Returns years from:
    1. Transactions (earliest to current year)
    2. Company creation date
    3. Performance records
    This allows companies to view data from any year they have records.
    """
    def get(self, request):
        if not request.user.is_authenticated:
            return JsonResponse({'error': 'Authentication required'}, status=401)
        
        company = request.user.company_profile
        if not company:
            return JsonResponse({'error': 'Company profile not found'}, status=404)
        
        current_year = timezone.now().year
        years_set = set()
        
        # Always include current year and next year (for planning)
        years_set.add(current_year)
        years_set.add(current_year + 1)
        
        # 1. Get years from company creation date
        if company.created_at:
            company_start_year = company.created_at.year
            # Add all years from company creation to current
            for y in range(company_start_year, current_year + 1):
                years_set.add(y)
        
        # 2. Get years from transactions
        try:
            transaction_years = Transaction.objects.filter(
                company=company
            ).dates('transaction_date', 'year').distinct()
            
            for date_obj in transaction_years:
                years_set.add(date_obj.year)
        except Exception:
            pass
        
        # 3. Get years from performance records
        try:
            from .models import PerformanceRecord
            performance_periods = PerformanceRecord.objects.filter(
                marketer__company_profile=company
            ).values_list('specific_period', flat=True).distinct()
            
            for period in performance_periods:
                if period and '-' in period:
                    year = int(period.split('-')[0])
                    years_set.add(year)
                elif period and period.isdigit():
                    years_set.add(int(period))
        except Exception:
            pass
        
        # 4. Get years from targets
        try:
            target_periods = MarketerTarget.objects.filter(
                company=company
            ).values_list('specific_period', flat=True).distinct()
            
            for period in target_periods:
                if period and '-' in period:
                    year = int(period.split('-')[0])
                    years_set.add(year)
                elif period and period.isdigit():
                    years_set.add(int(period))
        except Exception:
            pass
        
        # Sort years descending (most recent first)
        years_list = sorted(list(years_set), reverse=True)
        
        # Also return some metadata
        earliest_year = min(years_list) if years_list else current_year
        latest_year = max(years_list) if years_list else current_year
        
        return JsonResponse({
            'years': years_list,
            'earliest_year': earliest_year,
            'latest_year': latest_year,
            'current_year': current_year,
            'total_years': len(years_list)
        })


class PaymentHealthAPI(View):
    """
    API endpoint for Payment Health Analytics.
    Provides comprehensive payment collection data for company growth tracking:
    - Total sales, collected, outstanding, overdue amounts
    - Payment status distribution
    - Monthly collection trends
    - Estate-wise sales breakdown
    - Overdue clients list
    """
    def get(self, request):
        if not request.user.is_authenticated:
            return JsonResponse({'error': 'Authentication required'}, status=401)
        
        company = request.user.company_profile
        if not company:
            return JsonResponse({'error': 'Company profile not found'}, status=404)
        
        try:
            # Get all transactions for this company
            transactions = Transaction.objects.filter(company=company)
            
            # Calculate totals
            total_sales = Decimal('0')
            collected = Decimal('0')
            outstanding = Decimal('0')
            overdue = Decimal('0')
            
            status_counts = {
                'Fully Paid': 0,
                'Paid Complete': 0,
                'Part Payment': 0,
                'Pending': 0,
                'Overdue': 0
            }
            
            overdue_clients = []
            estate_sales = {}
            today = timezone.now().date()
            
            for txn in transactions:
                total_sales += txn.total_amount
                
                # Get status
                status = txn.status
                status_counts[status] = status_counts.get(status, 0) + 1
                
                # Calculate collected and outstanding
                txn_paid = txn.total_paid
                txn_balance = txn.balance
                collected += txn_paid
                
                if txn_balance > 0:
                    outstanding += txn_balance
                    
                    # Check if overdue
                    if status == 'Overdue' or (txn.due_date and today > txn.due_date):
                        overdue += txn_balance
                        days_overdue = (today - txn.due_date).days if txn.due_date else 0
                        
                        # Get client profile picture
                        profile_pic = None
                        if hasattr(txn.client, 'profile_picture') and txn.client.profile_picture:
                            profile_pic = txn.client.profile_picture.url
                        
                        overdue_clients.append({
                            'transaction_id': txn.id,
                            'client_name': txn.client.full_name if txn.client else 'Unknown',
                            'profile_picture': profile_pic,
                            'estate': txn.allocation.estate.name if txn.allocation and txn.allocation.estate else '-',
                            'outstanding': float(txn_balance),
                            'days_overdue': max(0, days_overdue)
                        })
                
                # Estate sales aggregation
                estate_name = txn.allocation.estate.name if txn.allocation and txn.allocation.estate else 'Other'
                if estate_name not in estate_sales:
                    estate_sales[estate_name] = Decimal('0')
                estate_sales[estate_name] += txn.total_amount
            
            # Sort overdue clients by days overdue (most overdue first)
            overdue_clients.sort(key=lambda x: x['days_overdue'], reverse=True)
            
            # Get monthly collection trend (last 6 months)
            from django.db.models import Sum
            from django.db.models.functions import TruncMonth
            from dateutil.relativedelta import relativedelta
            
            six_months_ago = today - relativedelta(months=6)
            
            # Get payment records grouped by month
            monthly_collections = PaymentRecord.objects.filter(
                transaction__company=company,
                payment_date__gte=six_months_ago
            ).annotate(
                month=TruncMonth('payment_date')
            ).values('month').annotate(
                amount=Sum('amount_paid')
            ).order_by('month')
            
            monthly_collection_list = []
            for item in monthly_collections:
                if item['month']:
                    monthly_collection_list.append({
                        'month': item['month'].strftime('%b %Y'),
                        'amount': float(item['amount'] or 0)
                    })
            
            # Estate sales list
            estate_sales_list = [
                {'estate': estate, 'sales': float(amount)} 
                for estate, amount in sorted(estate_sales.items(), key=lambda x: x[1], reverse=True)
            ][:8]  # Top 8 estates
            
            # Combine Fully Paid and Paid Complete for display
            combined_paid = status_counts.get('Fully Paid', 0) + status_counts.get('Paid Complete', 0)
            status_distribution = {
                'Fully Paid': combined_paid,
                'Part Payment': status_counts.get('Part Payment', 0),
                'Pending': status_counts.get('Pending', 0),
                'Overdue': status_counts.get('Overdue', 0)
            }
            
            return JsonResponse({
                'total_sales': float(total_sales),
                'collected': float(collected),
                'outstanding': float(outstanding),
                'overdue': float(overdue),
                'total_count': transactions.count(),
                'pending_count': status_counts.get('Part Payment', 0) + status_counts.get('Pending', 0),
                'overdue_count': status_counts.get('Overdue', 0),
                'status_distribution': status_distribution,
                'monthly_collection': monthly_collection_list,
                'estate_sales': estate_sales_list,
                'overdue_clients': overdue_clients[:10]  # Limit to 10
            })
            
        except Exception as e:
            import traceback
            traceback.print_exc()
            return JsonResponse({'error': str(e)}, status=500)


class TransactionExportFiltersAPI(View):
    """API to get available filter options for transaction export."""
    def get(self, request):
        if not request.user.is_authenticated:
            return JsonResponse({'error': 'Authentication required'}, status=401)
        
        company = request.user.company_profile
        if not company:
            return JsonResponse({'error': 'Company profile not found'}, status=404)
        
        # Get unique estates from transactions
        estates = list(Transaction.objects.filter(company=company)
            .values_list('allocation__estate__name', flat=True)
            .distinct()
            .order_by('allocation__estate__name'))
        
        # Get marketers
        marketers = []
        marketer_qs = MarketerUser.objects.filter(company_profile=company)
        for m in marketer_qs:
            marketers.append({'id': m.id, 'name': m.full_name})
        
        return JsonResponse({
            'estates': [e for e in estates if e],
            'marketers': marketers
        })


class TransactionExportCountAPI(View):
    """API to get count and total value of transactions matching filters."""
    def get(self, request):
        if not request.user.is_authenticated:
            return JsonResponse({'error': 'Authentication required'}, status=401)
        
        company = request.user.company_profile
        if not company:
            return JsonResponse({'error': 'Company profile not found'}, status=404)
        
        transactions = self._get_filtered_transactions(request, company)
        
        total_value = sum(t.total_amount for t in transactions)
        
        return JsonResponse({
            'count': transactions.count(),
            'total_value': float(total_value)
        })
    
    def _get_filtered_transactions(self, request, company):
        transactions = Transaction.objects.filter(company=company)
        
        from_date = request.GET.get('from_date')
        to_date = request.GET.get('to_date')
        estate = request.GET.get('estate')
        status = request.GET.get('status')
        marketer = request.GET.get('marketer')
        
        if from_date:
            transactions = transactions.filter(transaction_date__gte=from_date)
        if to_date:
            transactions = transactions.filter(transaction_date__lte=to_date)
        if estate:
            transactions = transactions.filter(allocation__estate__name=estate)
        if marketer:
            transactions = transactions.filter(marketer_id=marketer)
        
        # Filter by computed status
        if status:
            # Since status is a property, we need to filter in Python
            transactions = [t for t in transactions if t.status == status]
            return transactions
        
        return transactions


class TransactionExportPreviewAPI(View):
    """API to get preview of transactions for export."""
    def get(self, request):
        if not request.user.is_authenticated:
            return JsonResponse({'error': 'Authentication required'}, status=401)
        
        company = request.user.company_profile
        if not company:
            return JsonResponse({'error': 'Company profile not found'}, status=404)
        
        transactions = self._get_filtered_transactions(request, company)
        
        # Limit to first 50 for preview
        if hasattr(transactions, '__iter__') and not hasattr(transactions, 'count'):
            transactions = transactions[:50]
        else:
            transactions = transactions[:50]
        
        result = []
        for txn in transactions:
            result.append({
                'id': txn.id,
                'date': txn.transaction_date.strftime('%d %b %Y'),
                'client': txn.client.full_name if txn.client else 'Unknown',
                'estate': txn.allocation.estate.name if txn.allocation and txn.allocation.estate else 'N/A',
                'plot_size': str(txn.allocation.plot_size_unit.plot_size.size) if txn.allocation and txn.allocation.plot_size_unit else 'N/A',
                'marketer': txn.marketer.full_name if txn.marketer else 'N/A',
                'total_amount': float(txn.total_amount),
                'paid': float(txn.total_paid),
                'balance': float(txn.balance),
                'status': txn.status
            })
        
        return JsonResponse({'transactions': result})
    
    def _get_filtered_transactions(self, request, company):
        transactions = Transaction.objects.filter(company=company).select_related(
            'client', 'marketer', 'allocation__estate', 'allocation__plot_size_unit__plot_size'
        ).order_by('-transaction_date')
        
        from_date = request.GET.get('from_date')
        to_date = request.GET.get('to_date')
        estate = request.GET.get('estate')
        status = request.GET.get('status')
        marketer = request.GET.get('marketer')
        
        if from_date:
            transactions = transactions.filter(transaction_date__gte=from_date)
        if to_date:
            transactions = transactions.filter(transaction_date__lte=to_date)
        if estate:
            transactions = transactions.filter(allocation__estate__name=estate)
        if marketer:
            transactions = transactions.filter(marketer_id=marketer)
        
        # Filter by computed status
        if status:
            transactions = [t for t in transactions if t.status == status]
        
        return transactions


class TransactionExportAPI(View):
    """API to export transactions to Excel, CSV, or PDF."""
    def get(self, request):
        if not request.user.is_authenticated:
            return JsonResponse({'error': 'Authentication required'}, status=401)
        
        company = request.user.company_profile
        if not company:
            return JsonResponse({'error': 'Company profile not found'}, status=404)
        
        export_format = request.GET.get('format', 'excel')
        transactions = self._get_filtered_transactions(request, company)
        
        if export_format == 'excel':
            return self._export_excel(transactions, company)
        elif export_format == 'csv':
            return self._export_csv(transactions, company)
        elif export_format == 'pdf':
            return self._export_pdf(transactions, company)
        else:
            return JsonResponse({'error': 'Invalid format'}, status=400)
    
    def _get_filtered_transactions(self, request, company):
        transactions = Transaction.objects.filter(company=company).select_related(
            'client', 'marketer', 'allocation__estate', 'allocation__plot_size_unit__plot_size'
        ).order_by('-transaction_date')
        
        from_date = request.GET.get('from_date')
        to_date = request.GET.get('to_date')
        estate = request.GET.get('estate')
        status = request.GET.get('status')
        marketer = request.GET.get('marketer')
        
        if from_date:
            transactions = transactions.filter(transaction_date__gte=from_date)
        if to_date:
            transactions = transactions.filter(transaction_date__lte=to_date)
        if estate:
            transactions = transactions.filter(allocation__estate__name=estate)
        if marketer:
            transactions = transactions.filter(marketer_id=marketer)
        
        # Filter by computed status
        if status:
            transactions = [t for t in transactions if t.status == status]
        
        return transactions
    
    def _export_excel(self, transactions, company):
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from io import BytesIO
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Transactions'
        
        # Header styling
        header_fill = PatternFill(start_color='667eea', end_color='667eea', fill_type='solid')
        header_font = Font(bold=True, color='FFFFFF')
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Headers
        headers = ['Date', 'Client', 'Estate', 'Plot Size', 'Marketer', 
                   'Total Amount (₦)', 'Paid (₦)', 'Balance (₦)', 'Status', 'Reference']
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center')
            cell.border = thin_border
        
        # Data rows
        row = 2
        for txn in transactions:
            ws.cell(row=row, column=1, value=txn.transaction_date.strftime('%Y-%m-%d'))
            ws.cell(row=row, column=2, value=txn.client.full_name if txn.client else 'Unknown')
            ws.cell(row=row, column=3, value=txn.allocation.estate.name if txn.allocation and txn.allocation.estate else 'N/A')
            ws.cell(row=row, column=4, value=str(txn.allocation.plot_size_unit.plot_size.size) if txn.allocation and txn.allocation.plot_size_unit else 'N/A')
            ws.cell(row=row, column=5, value=txn.marketer.full_name if txn.marketer else 'N/A')
            ws.cell(row=row, column=6, value=float(txn.total_amount))
            ws.cell(row=row, column=7, value=float(txn.total_paid))
            ws.cell(row=row, column=8, value=float(txn.balance))
            ws.cell(row=row, column=9, value=txn.status)
            ws.cell(row=row, column=10, value=txn.reference_code or '')
            
            for col in range(1, 11):
                ws.cell(row=row, column=col).border = thin_border
            
            row += 1
        
        # Summary row
        if hasattr(transactions, 'count'):
            total_count = transactions.count()
        else:
            total_count = len(transactions)
        
        ws.cell(row=row + 1, column=1, value=f'Total: {total_count} transactions')
        ws.cell(row=row + 1, column=1).font = Font(bold=True)
        
        # Auto-adjust column widths
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            ws.column_dimensions[column].width = min(max_length + 2, 50)
        
        # Save to buffer
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        
        response = HttpResponse(
            buffer.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename=transactions_{company.company_name}_{timezone.now().strftime("%Y%m%d")}.xlsx'
        return response
    
    def _export_csv(self, transactions, company):
        import csv
        from io import StringIO
        
        buffer = StringIO()
        writer = csv.writer(buffer)
        
        # Headers
        writer.writerow(['Date', 'Client', 'Estate', 'Plot Size', 'Marketer', 
                        'Total Amount', 'Paid', 'Balance', 'Status', 'Reference'])
        
        # Data rows
        for txn in transactions:
            writer.writerow([
                txn.transaction_date.strftime('%Y-%m-%d'),
                txn.client.full_name if txn.client else 'Unknown',
                txn.allocation.estate.name if txn.allocation and txn.allocation.estate else 'N/A',
                str(txn.allocation.plot_size_unit.plot_size.size) if txn.allocation and txn.allocation.plot_size_unit else 'N/A',
                txn.marketer.full_name if txn.marketer else 'N/A',
                float(txn.total_amount),
                float(txn.total_paid),
                float(txn.balance),
                txn.status,
                txn.reference_code or ''
            ])
        
        response = HttpResponse(buffer.getvalue(), content_type='text/csv')
        response['Content-Disposition'] = f'attachment; filename=transactions_{company.company_name}_{timezone.now().strftime("%Y%m%d")}.csv'
        return response
    
    def _export_pdf(self, transactions, company):
        # For PDF, we'll return a simple text response indicating PDF generation
        # In production, you'd use reportlab or weasyprint
        from django.http import HttpResponse
        
        # Basic PDF using reportlab if available
        try:
            from reportlab.lib import colors
            from reportlab.lib.pagesizes import letter, landscape
            from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
            from reportlab.lib.styles import getSampleStyleSheet
            from io import BytesIO
            
            buffer = BytesIO()
            doc = SimpleDocTemplate(buffer, pagesize=landscape(letter))
            elements = []
            
            styles = getSampleStyleSheet()
            elements.append(Paragraph(f'Transaction Report - {company.company_name}', styles['Heading1']))
            elements.append(Paragraph(f'Generated: {timezone.now().strftime("%Y-%m-%d %H:%M")}', styles['Normal']))
            elements.append(Spacer(1, 20))
            
            # Table data
            data = [['Date', 'Client', 'Estate', 'Amount', 'Paid', 'Balance', 'Status']]
            
            for txn in transactions:
                data.append([
                    txn.transaction_date.strftime('%Y-%m-%d'),
                    (txn.client.full_name if txn.client else 'Unknown')[:20],
                    (txn.allocation.estate.name if txn.allocation and txn.allocation.estate else 'N/A')[:15],
                    f'₦{float(txn.total_amount):,.2f}',
                    f'₦{float(txn.total_paid):,.2f}',
                    f'₦{float(txn.balance):,.2f}',
                    txn.status
                ])
            
            table = Table(data)
            table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#667eea')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 10),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), colors.white),
                ('TEXTCOLOR', (0, 1), (-1, -1), colors.black),
                ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
                ('FONTSIZE', (0, 1), (-1, -1), 8),
                ('GRID', (0, 0), (-1, -1), 1, colors.black),
                ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f8f9fa')])
            ]))
            
            elements.append(table)
            doc.build(elements)
            
            buffer.seek(0)
            response = HttpResponse(buffer.getvalue(), content_type='application/pdf')
            response['Content-Disposition'] = f'attachment; filename=transactions_{company.company_name}_{timezone.now().strftime("%Y%m%d")}.pdf'
            return response
            
        except ImportError:
            # Fallback if reportlab is not installed
            return JsonResponse({
                'error': 'PDF generation requires reportlab library. Please install it or use Excel/CSV format.'
            }, status=400)


class MarketerExportAPI(View):
    """
    API to export all marketers to Excel or PDF.
    SECURITY: Only exports marketers within the authenticated user's company.
    """
    def get(self, request):
        if not request.user.is_authenticated:
            return JsonResponse({'error': 'Authentication required'}, status=401)
        
        company = request.user.company_profile
        if not company:
            return JsonResponse({'error': 'Company profile not found'}, status=404)
        
        export_format = request.GET.get('format', 'excel')
        marketers = self._get_company_marketers(company)
        
        if export_format == 'excel':
            return self._export_excel(marketers, company)
        elif export_format == 'pdf':
            return self._export_pdf(marketers, company)
        else:
            return JsonResponse({'error': 'Invalid format. Use excel or pdf.'}, status=400)
    
    def _get_company_marketers(self, company):
        """Get all marketers (primary + affiliated) for a company."""
        # Primary marketers (direct company assignment)
        primary_marketers = MarketerUser.objects.filter(company_profile=company)
        
        # Affiliated marketers (added via MarketerAffiliation)
        affiliation_marketer_ids = list(
            MarketerAffiliation.objects.filter(company=company)
            .values_list('marketer_id', flat=True)
            .distinct()
        )
        affiliated_marketers = MarketerUser.objects.filter(
            id__in=affiliation_marketer_ids
        ).exclude(id__in=primary_marketers.values_list('pk', flat=True))
        
        # Combine and sort by registration date
        marketers = list(primary_marketers) + list(affiliated_marketers)
        marketers.sort(key=lambda u: getattr(u, 'date_registered', None) or timezone.now(), reverse=True)
        
        # Calculate client counts for each marketer
        for marketer in marketers:
            assign_client_ids = set(
                ClientMarketerAssignment.objects.filter(
                    company=company, 
                    marketer_id=marketer.id
                ).values_list('client_id', flat=True)
            )
            direct_client_ids = set(
                ClientUser.objects.filter(
                    company_profile=company, 
                    assigned_marketer_id=marketer.id
                ).values_list('pk', flat=True)
            )
            marketer.client_count = len(assign_client_ids.union(direct_client_ids))
        
        return marketers
    
    def _export_excel(self, marketers, company):
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from io import BytesIO
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Marketers'
        
        # Header styling
        header_fill = PatternFill(start_color='28a745', end_color='28a745', fill_type='solid')
        header_font = Font(bold=True, color='FFFFFF')
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Headers
        headers = ['S/N', 'Full Name', 'Email', 'Phone', 'Clients', 'Date Registered']
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center')
            cell.border = thin_border
        
        # Data rows
        for idx, marketer in enumerate(marketers, 1):
            row = idx + 1
            ws.cell(row=row, column=1, value=idx)
            ws.cell(row=row, column=2, value=marketer.full_name or 'N/A')
            ws.cell(row=row, column=3, value=marketer.email or 'N/A')
            ws.cell(row=row, column=4, value=getattr(marketer, 'phone', '') or 'N/A')
            ws.cell(row=row, column=5, value=getattr(marketer, 'client_count', 0))
            ws.cell(row=row, column=6, value=marketer.date_registered.strftime('%Y-%m-%d') if hasattr(marketer, 'date_registered') and marketer.date_registered else 'N/A')
            
            for col in range(1, 7):
                ws.cell(row=row, column=col).border = thin_border
        
        # Summary row
        summary_row = len(marketers) + 3
        ws.cell(row=summary_row, column=1, value=f'Total Marketers: {len(marketers)}')
        ws.cell(row=summary_row, column=1).font = Font(bold=True)
        
        # Auto-adjust column widths
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            ws.column_dimensions[column].width = min(max_length + 2, 50)
        
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        
        response = HttpResponse(
            buffer.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename=marketers_{company.company_name}_{timezone.now().strftime("%Y%m%d")}.xlsx'
        return response
    
    def _export_pdf(self, marketers, company):
        try:
            from reportlab.lib import colors
            from reportlab.lib.pagesizes import letter, landscape
            from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
            from reportlab.lib.styles import getSampleStyleSheet
            from io import BytesIO
            
            buffer = BytesIO()
            doc = SimpleDocTemplate(buffer, pagesize=landscape(letter))
            elements = []
            
            styles = getSampleStyleSheet()
            elements.append(Paragraph(f'Marketers Report - {company.company_name}', styles['Heading1']))
            elements.append(Paragraph(f'Generated: {timezone.now().strftime("%Y-%m-%d %H:%M")}', styles['Normal']))
            elements.append(Spacer(1, 20))
            
            # Table data
            data = [['S/N', 'Full Name', 'Email', 'Phone', 'Clients', 'Status']]
            
            for idx, marketer in enumerate(marketers, 1):
                data.append([
                    str(idx),
                    (marketer.full_name or 'N/A')[:25],
                    (marketer.email or 'N/A')[:30],
                    (getattr(marketer, 'phone', '') or 'N/A')[:15],
                    str(getattr(marketer, 'client_count', 0)),
                    'Active' if marketer.is_active else 'Inactive'
                ])
            
            table = Table(data)
            table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#28a745')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 10),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('GRID', (0, 0), (-1, -1), 1, colors.black),
                ('FONTSIZE', (0, 1), (-1, -1), 8),
                ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f8f9fa')])
            ]))
            
            elements.append(table)
            elements.append(Spacer(1, 20))
            elements.append(Paragraph(f'Total Marketers: {len(marketers)}', styles['Normal']))
            
            doc.build(elements)
            buffer.seek(0)
            
            response = HttpResponse(buffer.getvalue(), content_type='application/pdf')
            response['Content-Disposition'] = f'attachment; filename=marketers_{company.company_name}_{timezone.now().strftime("%Y%m%d")}.pdf'
            return response
            
        except ImportError:
            return JsonResponse({
                'error': 'PDF generation requires reportlab library. Please install it or use Excel format.'
            }, status=400)


class ClientExportAPI(View):
    """
    API to export all clients to Excel or PDF.
    SECURITY: Only exports clients within the authenticated user's company.
    """
    def get(self, request):
        if not request.user.is_authenticated:
            return JsonResponse({'error': 'Authentication required'}, status=401)
        
        company = request.user.company_profile
        if not company:
            return JsonResponse({'error': 'Company profile not found'}, status=404)
        
        export_format = request.GET.get('format', 'excel')
        clients = self._get_company_clients(company)
        
        if export_format == 'excel':
            return self._export_excel(clients, company)
        elif export_format == 'pdf':
            return self._export_pdf(clients, company)
        else:
            return JsonResponse({'error': 'Invalid format. Use excel or pdf.'}, status=400)
    
    def _get_company_clients(self, company):
        """Get all clients (primary + affiliated) for a company."""
        # Primary clients (direct company assignment)
        primary_clients = ClientUser.objects.filter(company_profile=company)
        
        # Affiliated clients (added via ClientMarketerAssignment)
        affiliation_client_ids = list(
            ClientMarketerAssignment.objects.filter(company=company)
            .values_list('client_id', flat=True)
            .distinct()
        )
        affiliated_clients = ClientUser.objects.filter(
            id__in=affiliation_client_ids
        ).exclude(id__in=primary_clients.values_list('pk', flat=True))
        
        # Combine and sort by registration date
        clients = list(primary_clients) + list(affiliated_clients)
        clients.sort(key=lambda u: getattr(u, 'date_registered', None) or timezone.now(), reverse=True)
        
        # Get marketer info for each client
        for client in clients:
            # Check for assigned marketer
            assignment = ClientMarketerAssignment.objects.filter(
                company=company, 
                client_id=client.id
            ).first()
            if assignment:
                try:
                    marketer = MarketerUser.objects.get(id=assignment.marketer_id)
                    client.assigned_marketer_name = marketer.full_name
                except MarketerUser.DoesNotExist:
                    client.assigned_marketer_name = 'N/A'
            elif hasattr(client, 'assigned_marketer') and client.assigned_marketer:
                client.assigned_marketer_name = client.assigned_marketer.full_name
            else:
                client.assigned_marketer_name = 'N/A'
        
        return clients
    
    def _export_excel(self, clients, company):
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from io import BytesIO
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Clients'
        
        # Header styling
        header_fill = PatternFill(start_color='007bff', end_color='007bff', fill_type='solid')
        header_font = Font(bold=True, color='FFFFFF')
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Headers
        headers = ['S/N', 'Full Name', 'Email', 'Phone', 'Assigned Marketer', 'Date Registered']
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center')
            cell.border = thin_border
        
        # Data rows
        for idx, client in enumerate(clients, 1):
            row = idx + 1
            ws.cell(row=row, column=1, value=idx)
            ws.cell(row=row, column=2, value=client.full_name or 'N/A')
            ws.cell(row=row, column=3, value=client.email or 'N/A')
            ws.cell(row=row, column=4, value=getattr(client, 'phone', '') or 'N/A')
            ws.cell(row=row, column=5, value=getattr(client, 'assigned_marketer_name', 'N/A'))
            ws.cell(row=row, column=6, value=client.date_registered.strftime('%Y-%m-%d') if hasattr(client, 'date_registered') and client.date_registered else 'N/A')
            
            for col in range(1, 7):
                ws.cell(row=row, column=col).border = thin_border
        
        # Summary row
        summary_row = len(clients) + 3
        ws.cell(row=summary_row, column=1, value=f'Total Clients: {len(clients)}')
        ws.cell(row=summary_row, column=1).font = Font(bold=True)
        
        # Auto-adjust column widths
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            ws.column_dimensions[column].width = min(max_length + 2, 50)
        
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        
        response = HttpResponse(
            buffer.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename=clients_{company.company_name}_{timezone.now().strftime("%Y%m%d")}.xlsx'
        return response
    
    def _export_pdf(self, clients, company):
        try:
            from reportlab.lib import colors
            from reportlab.lib.pagesizes import letter, landscape
            from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
            from reportlab.lib.styles import getSampleStyleSheet
            from io import BytesIO
            
            buffer = BytesIO()
            doc = SimpleDocTemplate(buffer, pagesize=landscape(letter))
            elements = []
            
            styles = getSampleStyleSheet()
            elements.append(Paragraph(f'Clients Report - {company.company_name}', styles['Heading1']))
            elements.append(Paragraph(f'Generated: {timezone.now().strftime("%Y-%m-%d %H:%M")}', styles['Normal']))
            elements.append(Spacer(1, 20))
            
            # Table data
            data = [['S/N', 'Full Name', 'Email', 'Phone', 'Marketer', 'Status']]
            
            for idx, client in enumerate(clients, 1):
                data.append([
                    str(idx),
                    (client.full_name or 'N/A')[:25],
                    (client.email or 'N/A')[:30],
                    (getattr(client, 'phone', '') or 'N/A')[:15],
                    (getattr(client, 'assigned_marketer_name', 'N/A') or 'N/A')[:20],
                    'Active' if client.is_active else 'Inactive'
                ])
            
            table = Table(data)
            table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#007bff')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 10),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('GRID', (0, 0), (-1, -1), 1, colors.black),
                ('FONTSIZE', (0, 1), (-1, -1), 8),
                ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f8f9fa')])
            ]))
            
            elements.append(table)
            elements.append(Spacer(1, 20))
            elements.append(Paragraph(f'Total Clients: {len(clients)}', styles['Normal']))
            
            doc.build(elements)
            buffer.seek(0)
            
            response = HttpResponse(buffer.getvalue(), content_type='application/pdf')
            response['Content-Disposition'] = f'attachment; filename=clients_{company.company_name}_{timezone.now().strftime("%Y%m%d")}.pdf'
            return response
            
        except ImportError:
            return JsonResponse({
                'error': 'PDF generation requires reportlab library. Please install it or use Excel format.'
            }, status=400)


class ExportPerformanceAPI(View):
    def get(self, request):
        period_type = request.GET.get('period_type')
        specific_period = request.GET.get('specific_period')
        format = request.GET.get('format', 'csv')
        
        # This would generate a file in real implementation
        # For now, just return success
        return JsonResponse({
            'status': 'success',
            'message': f'Exported {period_type} {specific_period} in {format.upper()} format'
        })


class SendMarketerMessageAPI(View):
    """
    API endpoint to send messages to marketers via Email, SMS, and/or In-App notifications.
    Supports single marketer, multiple marketers, or bulk sending.
    SECURITY: Only allows sending to marketers within the admin's company.
    """
    @method_decorator(csrf_exempt)
    def dispatch(self, *args, **kwargs):
        return super().dispatch(*args, **kwargs)
    
    def post(self, request):
        # SECURITY: Verify user is authenticated
        if not request.user.is_authenticated:
            return JsonResponse({'status': 'error', 'message': 'Authentication required'}, status=401)
        
        # SECURITY: Verify user is an admin
        if request.user.role != 'admin':
            return JsonResponse({'status': 'error', 'message': 'Admin access required'}, status=403)
        
        # SECURITY: Verify user has a company profile
        company = request.user.company_profile
        if not company:
            return JsonResponse({'status': 'error', 'message': 'No company associated with your account'}, status=403)
        
        try:
            data = json.loads(request.body)
        except json.JSONDecodeError:
            return JsonResponse({'status': 'error', 'message': 'Invalid JSON'}, status=400)
        
        marketer_id = data.get('marketer_id')  # Single marketer
        marketer_ids = data.get('marketer_ids', [])  # Multiple marketers
        recipient_type = data.get('recipient_type', 'single')  # single, no_sales, all
        subject = data.get('subject', 'Message from Management')
        message = data.get('message', '')
        channels = data.get('channels', {})
        
        if not message.strip():
            return JsonResponse({'status': 'error', 'message': 'Message content required'}, status=400)
        
        # Import utility functions
        from adminSupport.utils import send_email_message, send_sms_message, send_inapp_message
        
        # Determine which marketers to send to
        if recipient_type == 'single':
            if not marketer_id:
                return JsonResponse({'status': 'error', 'message': 'Marketer ID required'}, status=400)
            
            try:
                marketer = MarketerUser.objects.get(id=int(marketer_id))
            except (MarketerUser.DoesNotExist, ValueError):
                return JsonResponse({'status': 'error', 'message': 'Marketer not found'}, status=404)
            
            # Check if marketer is in admin's company
            marketer_in_company = (
                marketer.company_profile == company or
                MarketerAffiliation.objects.filter(marketer=marketer, company=company).exists()
            )
            
            if not marketer_in_company:
                return JsonResponse({'status': 'error', 'message': 'Access denied: Marketer not in your company'}, status=403)
            
            # Send to single marketer
            result = self._send_to_marketer(marketer, subject, message, channels, send_email_message, send_sms_message, send_inapp_message)
            
            return JsonResponse({
                'status': 'success',
                'message': 'Message sent successfully',
                'email_sent': result['email_sent'],
                'sms_sent': result['sms_sent'],
                'app_sent': result['app_sent'],
                'errors': result['errors'] if result['errors'] else None
            })
        
        else:
            # Bulk send to multiple marketers
            if not marketer_ids:
                return JsonResponse({'status': 'error', 'message': 'No marketers specified'}, status=400)
            
            # Get all marketers in admin's company (filter for security)
            company_marketer_ids = set()
            
            # Direct company profile marketers
            direct_marketers = MarketerUser.objects.filter(
                id__in=marketer_ids,
                company_profile=company
            ).values_list('id', flat=True)
            company_marketer_ids.update(direct_marketers)
            
            # Affiliated marketers
            affiliated_marketers = MarketerAffiliation.objects.filter(
                marketer_id__in=marketer_ids,
                company=company
            ).values_list('marketer_id', flat=True)
            company_marketer_ids.update(affiliated_marketers)
            
            # Get the actual marketer objects
            marketers = MarketerUser.objects.filter(id__in=company_marketer_ids)
            
            if not marketers.exists():
                return JsonResponse({'status': 'error', 'message': 'No valid marketers found'}, status=400)
            
            # Send to all marketers
            total_count = marketers.count()
            sent_count = 0
            emails_sent = 0
            sms_count = 0
            app_count = 0
            all_errors = []
            
            for marketer in marketers:
                result = self._send_to_marketer(marketer, subject, message, channels, send_email_message, send_sms_message, send_inapp_message)
                
                if result['email_sent'] or result['sms_sent'] or result['app_sent']:
                    sent_count += 1
                
                if result['email_sent']:
                    emails_sent += 1
                if result['sms_sent']:
                    sms_count += 1
                if result['app_sent']:
                    app_count += 1
                
                if result['errors']:
                    all_errors.extend(result['errors'])
            
            return JsonResponse({
                'status': 'success',
                'message': f'Message sent to {sent_count} of {total_count} marketers',
                'total_count': total_count,
                'sent_count': sent_count,
                'emails_sent': emails_sent,
                'sms_count': sms_count,
                'app_count': app_count,
                'errors': all_errors[:10] if all_errors else None  # Limit errors to first 10
            })
    
    def _send_to_marketer(self, marketer, subject, message, channels, send_email_message, send_sms_message, send_inapp_message):
        """Helper method to send message to a single marketer"""
        # Personalize message with placeholders
        first_name = marketer.full_name.split()[0] if marketer.full_name else ''
        personalized_message = message.replace('{firstName}', first_name)
        personalized_message = personalized_message.replace('{fullName}', marketer.full_name or '')
        personalized_message = personalized_message.replace('{email}', marketer.email or '')
        
        email_sent = False
        sms_sent = False
        app_sent = False
        errors = []
        
        # Send via Email
        if channels.get('email') and marketer.email:
            try:
                success, error = send_email_message(subject, personalized_message, marketer.email)
                email_sent = success
                if error:
                    errors.append(f'{marketer.full_name} - Email: {error}')
            except Exception as e:
                errors.append(f'{marketer.full_name} - Email: {str(e)}')
        
        # Send via SMS
        if channels.get('sms') and hasattr(marketer, 'phone') and marketer.phone:
            try:
                success, error = send_sms_message(personalized_message, marketer.phone)
                sms_sent = success
                if error:
                    errors.append(f'{marketer.full_name} - SMS: {error}')
            except Exception as e:
                errors.append(f'{marketer.full_name} - SMS: {str(e)}')
        
        # Send via In-App notification
        if channels.get('app'):
            try:
                success, error = send_inapp_message(personalized_message, marketer)
                app_sent = success
                if error:
                    errors.append(f'{marketer.full_name} - In-App: {error}')
            except Exception as e:
                errors.append(f'{marketer.full_name} - In-App: {str(e)}')
        
        return {
            'email_sent': email_sent,
            'sms_sent': sms_sent,
            'app_sent': app_sent,
            'errors': errors
        }


# VALUE REGULATION
def is_admin(user):
    return user.is_authenticated and user.role == "admin"

@require_http_methods(["GET"])
@login_required
@user_passes_test(is_admin)
def estate_plot_sizes(request, estate_id):
    # Company isolation - ensure estate belongs to user's company
    user_company = getattr(request.user, 'company_profile', None)
    if not user_company:
        return JsonResponse({'error': 'No company access'}, status=403)
    
    estate = get_object_or_404(Estate, pk=estate_id, company=user_company)
    units = PlotSizeUnits.objects.filter(estate_plot__estate=estate).select_related('plot_size')
    
    plot_units = [{
        "id": u.id,
        "size": u.plot_size.size,
        "available": u.available_units
    } for u in units]

    location = estate.location
    plot_unit_id = request.GET.get("plot_unit_id")
    existing = None
    
    if plot_unit_id:
        try:
            pp = PropertyPrice.objects.get(estate=estate, plot_unit__id=plot_unit_id)
            existing = {
                "presale": str(pp.presale),
                "previous": str(pp.previous),
                "current": str(pp.current),
                "effective": pp.effective.isoformat(),
                "notes": pp.notes or ""
            }
        except PropertyPrice.DoesNotExist:
            pass

    return JsonResponse({
        "plot_units": plot_units,
        "location": location,
        "existing_price": existing,
    })



@login_required
@user_passes_test(is_admin)
def estate_bulk_price_data(request, estate_id):
    """
    Returns estate info with all plot units and their current prices for bulk updating.
    Company-isolated for multi-tenant security.
    """
    # Company isolation - ensure estate belongs to user's company
    user_company = getattr(request.user, 'company_profile', None)
    if not user_company:
        return JsonResponse({'error': 'No company access'}, status=403)
    
    estate = get_object_or_404(Estate, pk=estate_id, company=user_company)
    
    # Get all plot size units for this estate
    units = PlotSizeUnits.objects.filter(
        estate_plot__estate=estate
    ).select_related('plot_size').distinct()
    
    plot_units = []
    for unit in units:
        # Try to get existing price
        try:
            pp = PropertyPrice.objects.get(estate=estate, plot_unit=unit)
            presale = float(pp.presale)
            previous = float(pp.previous)
            current = float(pp.current)
        except PropertyPrice.DoesNotExist:
            presale = 0
            previous = 0
            current = 0
        
        plot_units.append({
            "id": unit.id,
            "size": unit.plot_size.size,
            "available": unit.available_units,
            "total": unit.total_units,
            "presale": presale,
            "previous": previous,
            "current": current,
        })
    
    return JsonResponse({
        "estate": {
            "id": estate.id,
            "name": estate.name,
            "location": estate.location,
        },
        "plot_units": plot_units,
    })


def send_bulk_price_update_notification(estate, price_changes, unlaunched_plots, effective_date, notes):
    """
    Send comprehensive notifications to clients and marketers about bulk price updates.
    Uses the SAME notification system as the "Send Notification" modal.
    
    Args:
        estate: Estate object
        price_changes: List of dicts with plot_size, previous_price, new_price, pct_change, available, total, is_sold_out, changed
        unlaunched_plots: List of plot sizes (strings) that have no PropertyPrice yet
        effective_date: Date string for when prices become effective
        notes: Additional notes about the price update
    
    Returns:
        dict with notification counts
    """
    from datetime import datetime
    
    # Get User model
    User = get_user_model()
    
    # Categorize plots
    changed_plots = [p for p in price_changes if p['changed'] and not p['is_sold_out']]
    unchanged_plots = [p for p in price_changes if not p['changed'] and not p['is_sold_out']]
    sold_out_plots = [p for p in price_changes if p['is_sold_out']]
    
    # Format effective date
    try:
        eff_date = datetime.strptime(effective_date, '%Y-%m-%d').strftime('%B %d, %Y')
    except:
        eff_date = effective_date
    
    # Build HTML notification message
    message_parts = []
    
    # Header
    message_parts.append(f"""
    <div style="font-family: Arial, sans-serif; max-width: 800px; margin: 0 auto;">
        <div style="background: linear-gradient(135deg, #667eea 0%, #764ba2 100%); color: white; padding: 20px; border-radius: 8px 8px 0 0;">
            <h2 style="margin: 0; font-size: 24px;">🏘️ NEW PRICE REVIEW FOR {estate.name}</h2>
            <p style="margin: 10px 0 0 0; font-size: 14px; opacity: 0.9;">Effective Date: {eff_date}</p>
        </div>
        <div style="background: #f8f9fa; padding: 20px; border: 1px solid #e0e0e0; border-top: none;">
    """)
    
    # Changed Plots Section (Price Increased or Decreased)
    if changed_plots:
        message_parts.append("""
        <div style="margin-bottom: 20px;">
            <h3 style="color: #333; font-size: 18px; margin-bottom: 15px; border-bottom: 2px solid #667eea; padding-bottom: 5px;">
                📊 Price Updates
            </h3>
            <table style="width: 100%; border-collapse: collapse; background: white; border-radius: 8px; overflow: hidden; box-shadow: 0 2px 4px rgba(0,0,0,0.1);">
                <thead>
                    <tr style="background: #667eea; color: white;">
                        <th style="padding: 12px; text-align: left; border-bottom: 2px solid #5568d3;">Plot Size</th>
                        <th style="padding: 12px; text-align: right; border-bottom: 2px solid #5568d3;">Presale Price</th>
                        <th style="padding: 12px; text-align: right; border-bottom: 2px solid #5568d3;">Previous Price</th>
                        <th style="padding: 12px; text-align: right; border-bottom: 2px solid #5568d3;">New Price</th>
                        <th style="padding: 12px; text-align: center; border-bottom: 2px solid #5568d3;">Change</th>
                        <th style="padding: 12px; text-align: center; border-bottom: 2px solid #5568d3;">Total % from Presale</th>
                    </tr>
                </thead>
                <tbody>
        """)
        
        for plot in changed_plots:
            pct_change = plot['pct_change']
            total_pct = plot['total_pct']
            presale_price = plot.get('presale', 0)
            
            change_color = '#28a745' if pct_change > 0 else '#dc3545' if pct_change < 0 else '#6c757d'
            change_icon = '📈' if pct_change > 0 else '📉' if pct_change < 0 else '➖'
            badge_color = 'background: #d4edda; color: #155724; border: 1px solid #c3e6cb;' if pct_change > 0 else 'background: #f8d7da; color: #721c24; border: 1px solid #f5c6cb;'
            
            # Total % styling
            total_color = '#28a745' if total_pct > 0 else '#dc3545' if total_pct < 0 else '#6c757d'
            total_icon = '🚀' if total_pct > 0 else '⬇️' if total_pct < 0 else '➖'
            total_badge_color = 'background: #d1f2eb; color: #0c5460; border: 1px solid #bee5eb; font-weight: 700;' if total_pct > 0 else 'background: #f8d7da; color: #721c24; border: 1px solid #f5c6cb;'
            
            message_parts.append(f"""
                    <tr style="border-bottom: 1px solid #e0e0e0;">
                        <td style="padding: 12px; font-weight: 600; color: #333;">{plot['plot_size']}</td>
                        <td style="padding: 12px; text-align: right; color: #666; font-style: italic;">₦{presale_price:,.0f}</td>
                        <td style="padding: 12px; text-align: right; color: #666;">₦{plot['previous_price']:,.0f}</td>
                        <td style="padding: 12px; text-align: right; font-weight: 700; color: {change_color};">₦{plot['new_price']:,.0f}</td>
                        <td style="padding: 12px; text-align: center;">
                            <span style="{badge_color} padding: 4px 10px; border-radius: 12px; font-size: 12px; font-weight: 600; display: inline-block;">
                                {change_icon} {pct_change:+.1f}%
                            </span>
                        </td>
                        <td style="padding: 12px; text-align: center;">
                            <span style="{total_badge_color} padding: 6px 12px; border-radius: 12px; font-size: 13px; display: inline-block;">
                                {total_icon} {total_pct:+.1f}%
                            </span>
                        </td>
                    </tr>
            """)
        
        message_parts.append("""
                </tbody>
            </table>
        </div>
        """)
    
    # Sold Out Plots Section
    if sold_out_plots:
        message_parts.append("""
        <div style="margin-bottom: 20px;">
            <h3 style="color: #dc3545; font-size: 16px; margin-bottom: 10px;">
                🚫 Sold Out Plots
            </h3>
            <div style="background: white; padding: 15px; border-radius: 8px; border-left: 4px solid #dc3545;">
        """)
        
        for plot in sold_out_plots:
            message_parts.append(f"""
                <div style="padding: 8px 0; border-bottom: 1px dashed #e0e0e0; display: flex; justify-content: space-between; align-items: center;">
                    <span style="font-weight: 600; color: #333;">{plot['plot_size']}</span>
                    <span style="background: #f8d7da; color: #721c24; padding: 4px 12px; border-radius: 12px; font-size: 12px; font-weight: 600;">SOLD OUT</span>
                </div>
            """)
        
        message_parts.append("""
            </div>
        </div>
        """)
    
    # Not Launched Yet Section
    if unlaunched_plots:
        message_parts.append("""
        <div style="margin-bottom: 20px;">
            <h3 style="color: #ffc107; font-size: 16px; margin-bottom: 10px;">
                🔜 Coming Soon
            </h3>
            <div style="background: white; padding: 15px; border-radius: 8px; border-left: 4px solid #ffc107;">
        """)
        
        for plot_size in unlaunched_plots:
            message_parts.append(f"""
                <div style="padding: 8px 0; border-bottom: 1px dashed #e0e0e0; display: flex; justify-content: space-between; align-items: center;">
                    <span style="font-weight: 600; color: #333;">{plot_size}</span>
                    <span style="background: #fff3cd; color: #856404; padding: 4px 12px; border-radius: 12px; font-size: 12px; font-weight: 600;">NOT LAUNCHED YET</span>
                </div>
            """)
        
        message_parts.append("""
            </div>
        </div>
        """)
    
    # Additional Notes
    if notes and notes.strip():
        message_parts.append(f"""
        <div style="background: white; padding: 15px; border-radius: 8px; margin-bottom: 20px; border-left: 4px solid #17a2b8;">
            <h4 style="color: #17a2b8; font-size: 14px; margin: 0 0 8px 0;">📝 Additional Notes:</h4>
            <p style="margin: 0; color: #666; font-size: 14px; line-height: 1.6;">{notes}</p>
        </div>
        """)
    
    # Call to Action
    message_parts.append("""
        <div style="background: linear-gradient(135deg, #f093fb 0%, #f5576c 100%); color: white; padding: 20px; border-radius: 8px; text-align: center; margin-top: 20px;">
            <h3 style="margin: 0 0 10px 0; font-size: 20px;">🚀 HURRY UP TO BAG THIS OPPORTUNITY!</h3>
            <p style="margin: 0; font-size: 14px; opacity: 0.95;">Don't miss out on these updated prices. Contact us today to secure your plot!</p>
        </div>
        </div>
    </div>
    """)
    
    full_message = ''.join(message_parts)
    subject = f'Price Update: {estate.name}'
    
    # Create notification for CLIENTS (using SAME approach as notify_clients_marketer)
    client_notif = Notification.objects.create(
        title=subject,
        message=full_message,
        notification_type=Notification.CLIENT_ANNOUNCEMENT
    )
    
    # Create notification for MARKETERS
    marketer_notif = Notification.objects.create(
        title=subject,
        message=full_message,
        notification_type=Notification.MARKETER_ANNOUNCEMENT
    )
    
    # Get users by role (SAME as notify_clients_marketer function)
    clients = User.objects.filter(role='client')
    marketers = User.objects.filter(role='marketer')
    
    # Link notifications to users (SAME approach as notify_clients_marketer)
    clients_count = 0
    for user in clients:
        UserNotification.objects.get_or_create(user=user, notification=client_notif)
        clients_count += 1
    
    marketers_count = 0
    for user in marketers:
        UserNotification.objects.get_or_create(user=user, notification=marketer_notif)
        marketers_count += 1
    
    return {
        'clients_notified': clients_count,
        'marketers_notified': marketers_count,
        'total_notified': clients_count + marketers_count
    }


@csrf_exempt
@login_required
@user_passes_test(is_admin)
@require_http_methods(["POST"])
def property_price_bulk_update(request):
    """
    Bulk update property prices for multiple plot units.
    Company-isolated for multi-tenant security.
    """
    # Company isolation - get user's company
    user_company = getattr(request.user, 'company_profile', None)
    if not user_company:
        return JsonResponse({'error': 'No company access'}, status=403)
    
    try:
        payload = json.loads(request.body)
        estate_id = payload["estate_id"]
        effective = payload["effective"]
        notes = payload.get("notes", "")
        notify = payload.get("notify", False)
        updates = payload["updates"]  # List of {plot_unit_id, new_price}
        
        # Company isolation - ensure estate belongs to user's company
        estate = Estate.objects.get(pk=estate_id, company=user_company)
    except (KeyError, Estate.DoesNotExist, json.JSONDecodeError, ValueError) as e:
        return JsonResponse({"status": "error", "message": f"Invalid payload: {str(e)}"}, status=400)
    
    if not updates:
        return JsonResponse({"status": "error", "message": "No updates provided"}, status=400)
    
    updated_rows = []
    updated_count = 0
    price_changes = []  # Track all price changes for notification
    
    with transaction.atomic():
        # Get all plot units for this estate to categorize them
        all_units = PlotSizeUnits.objects.filter(estate_plot__estate=estate).select_related('plot_size')
        updated_unit_ids = [u["plot_unit_id"] for u in updates]
        
        for update_data in updates:
            try:
                plot_unit_id = update_data["plot_unit_id"]
                new_price = Decimal(str(update_data["new_price"]))
                
                unit = PlotSizeUnits.objects.get(pk=plot_unit_id)
                
                # Get or create PropertyPrice
                try:
                    pp = PropertyPrice.objects.get(estate=estate, plot_unit=unit)
                    previous_current = pp.current
                    presale = pp.presale
                    
                    # Calculate percentage change
                    if previous_current > 0:
                        pct_change = ((new_price - previous_current) / previous_current * 100)
                    else:
                        pct_change = Decimal(0)
                    
                    # Calculate total % from presale
                    if presale and presale > 0:
                        total_pct = ((new_price - presale) / presale) * 100
                    else:
                        total_pct = Decimal(0)
                    
                    # Update existing
                    pp.previous = previous_current
                    pp.current = new_price
                    pp.effective = effective
                    pp.notes = notes
                    pp.save()
                    
                except PropertyPrice.DoesNotExist:
                    # Create new with presale = new_price
                    pp = PropertyPrice.objects.create(
                        estate=estate,
                        plot_unit=unit,
                        presale=new_price,
                        previous=new_price,
                        current=new_price,
                        effective=effective,
                        notes=notes
                    )
                    previous_current = new_price
                    presale = new_price
                    pct_change = Decimal(0)
                
                # Calculate total % from presale
                if presale and presale > 0:
                    total_pct = ((new_price - presale) / presale) * 100
                else:
                    total_pct = Decimal(0)
                
                # Track price change info
                price_changes.append({
                    'plot_size': unit.plot_size.size,
                    'previous_price': float(previous_current),
                    'new_price': float(new_price),
                    'pct_change': float(pct_change),
                    'total_pct': float(total_pct),
                    'presale': float(presale),
                    'available': unit.available_units,
                    'total': unit.total_units,
                    'is_sold_out': unit.available_units == 0,
                    'changed': previous_current != new_price
                })
                
                # Create history entry
                PriceHistory.objects.create(
                    price=pp,
                    presale=pp.presale,
                    previous=previous_current,
                    current=pp.current,
                    effective=effective,
                    notes=notes
                )
                
                row_key = f"{estate_id}-{plot_unit_id}"
                updated_rows.append(row_key)
                updated_count += 1
                
            except (KeyError, PlotSizeUnits.DoesNotExist, ValueError) as e:
                # Skip invalid updates
                continue
        
        # Check for unlaunched plots (no PropertyPrice yet)
        unlaunched_plots = []
        for unit in all_units:
            if unit.id not in updated_unit_ids:
                try:
                    PropertyPrice.objects.get(estate=estate, plot_unit=unit)
                except PropertyPrice.DoesNotExist:
                    unlaunched_plots.append(unit.plot_size.size)
    
    # Send notifications if requested
    if notify and price_changes:
        try:
            notification_result = send_bulk_price_update_notification(
                estate=estate,
                price_changes=price_changes,
                unlaunched_plots=unlaunched_plots,
                effective_date=effective,
                notes=notes
            )
        except Exception as e:
            # Log error but don't fail the update
            print(f"Notification error: {str(e)}")
    
    return JsonResponse({
        "status": "ok",
        "updated_count": updated_count,
        "updated_rows": updated_rows,
        "notification_sent": notify and len(price_changes) > 0
    })


@csrf_exempt
@login_required
@user_passes_test(is_admin)
@require_http_methods(["POST"])
def property_price_add(request):
    # Company isolation - get user's company
    user_company = getattr(request.user, 'company_profile', None)
    if not user_company:
        return JsonResponse({'error': 'No company access'}, status=403)
    
    try:
        payload = json.loads(request.body)
        estate_id = payload["estate_id"]
        plot_unit_id = payload["plot_unit_id"]
        presale = Decimal(payload["presale"])
        effective = payload["effective"]
        notes = payload.get("notes", "")
        
        # Company isolation - ensure estate belongs to user's company
        estate = Estate.objects.get(pk=estate_id, company=user_company)
        unit = PlotSizeUnits.objects.get(pk=plot_unit_id, estate_plot__estate__company=user_company)

    except (KeyError, Estate.DoesNotExist, PlotSizeUnits.DoesNotExist, 
            ValueError, json.JSONDecodeError) as e:
        return HttpResponseBadRequest("Invalid payload")

    with transaction.atomic():
        obj, created = PropertyPrice.objects.update_or_create(
            estate=estate,
            plot_unit=unit,
            defaults={
                "presale": presale,
                "previous": presale,
                "current": presale,
                "effective": effective,
                "notes": notes,
            }
        )
        
        PriceHistory.objects.create(
            price=obj,
            presale=presale,
            previous=obj.previous,
            current=obj.current,
            effective=effective,
            notes=notes
        )

    return JsonResponse({
        "status": "ok", 
        "id": obj.id, 
        "created": created,
        "row_key": f"{estate_id}-{plot_unit_id}"
    })


@require_POST
@login_required
def property_price_bulk_presale(request):
    """
    Handle bulk presale price setting for multiple plot sizes at once.
    Expects JSON payload with:
    - estate_id: ID of the estate
    - effective: Effective date for all presale prices
    - notes: Optional notes
    - notify: Boolean for notification
    - presales: Array of {plot_unit_id, presale} objects
    """
    # Company isolation - get user's company
    user_company = getattr(request.user, 'company_profile', None)
    if not user_company:
        return JsonResponse({'error': 'No company access'}, status=403)
    
    try:
        payload = json.loads(request.body)
        estate_id = payload["estate_id"]
        effective = payload["effective"]
        notes = payload.get("notes", "")
        notify = payload.get("notify", True)
        presales = payload.get("presales", [])
        
        if not presales:
            return HttpResponseBadRequest("No presale prices provided")
        
        # Company isolation - ensure estate belongs to user's company
        estate = Estate.objects.get(pk=estate_id, company=user_company)

    except (KeyError, Estate.DoesNotExist, ValueError, json.JSONDecodeError) as e:
        return HttpResponseBadRequest(f"Invalid payload: {str(e)}")

    saved_count = 0
    created_count = 0
    updated_count = 0
    errors = []
    
    with transaction.atomic():
        for item in presales:
            try:
                plot_unit_id = item["plot_unit_id"]
                presale_value = Decimal(str(item["presale"]))
                
                # Company isolation - ensure plot unit belongs to user's company
                unit = PlotSizeUnits.objects.get(
                    pk=plot_unit_id, 
                    estate_plot__estate__company=user_company
                )
                
                # Check if this is an update or new entry
                existing = PropertyPrice.objects.filter(
                    estate=estate,
                    plot_unit=unit
                ).first()
                
                if existing:
                    # Update existing - only update presale if it changed
                    if existing.presale != presale_value:
                        existing.presale = presale_value
                        existing.effective = effective
                        existing.notes = notes
                        existing.save()
                        
                        PriceHistory.objects.create(
                            price=existing,
                            presale=presale_value,
                            previous=existing.previous,
                            current=existing.current,
                            effective=effective,
                            notes=f"Presale updated: {notes}" if notes else "Presale price updated"
                        )
                        updated_count += 1
                else:
                    # Create new PropertyPrice
                    obj = PropertyPrice.objects.create(
                        estate=estate,
                        plot_unit=unit,
                        presale=presale_value,
                        previous=presale_value,
                        current=presale_value,
                        effective=effective,
                        notes=notes
                    )
                    
                    PriceHistory.objects.create(
                        price=obj,
                        presale=presale_value,
                        previous=presale_value,
                        current=presale_value,
                        effective=effective,
                        notes=notes or "Initial presale price set"
                    )
                    created_count += 1
                
                saved_count += 1
                
            except (KeyError, PlotSizeUnits.DoesNotExist, ValueError) as e:
                errors.append(f"Plot {plot_unit_id}: {str(e)}")
                continue

    response_data = {
        "status": "ok",
        "saved_count": saved_count,
        "created_count": created_count,
        "updated_count": updated_count,
    }
    
    if errors:
        response_data["errors"] = errors
    
    return JsonResponse(response_data)


def send_single_price_update_notification(estate, plot_size, presale_price, previous_price, new_price, effective_date, notes):
    """
    Send notification to clients and marketers about a single property price update.
    
    Args:
        estate: Estate object
        plot_size: String (e.g., "500 sqm")
        presale_price: Float - original launch price
        previous_price: Float - price before update
        new_price: Float - updated current price
        effective_date: Date string for when price becomes effective
        notes: Additional notes about the price update
    """
    from datetime import datetime
    
    # Get User model
    User = get_user_model()
    
    # Calculate percentages
    if previous_price > 0:
        pct_change = ((new_price - previous_price) / previous_price) * 100
    else:
        pct_change = 0
    
    if presale_price > 0:
        total_pct = ((new_price - presale_price) / presale_price) * 100
    else:
        total_pct = 0
    
    # Format effective date
    try:
        eff_date = datetime.strptime(effective_date, '%Y-%m-%d').strftime('%B %d, %Y')
    except:
        eff_date = effective_date
    
    # Determine change direction
    change_icon = '📈' if pct_change > 0 else '📉' if pct_change < 0 else '➖'
    change_word = 'INCREASE' if pct_change > 0 else 'DECREASE' if pct_change < 0 else 'NO CHANGE'
    change_color = '#28a745' if pct_change > 0 else '#dc3545' if pct_change < 0 else '#6c757d'
    badge_color = 'background: #d4edda; color: #155724; border: 1px solid #c3e6cb;' if pct_change > 0 else 'background: #f8d7da; color: #721c24; border: 1px solid #f5c6cb;'
    
    # Total % styling
    total_icon = '🚀' if total_pct > 0 else '⬇️' if total_pct < 0 else '➖'
    total_badge_color = 'background: #d1f2eb; color: #0c5460; border: 1px solid #bee5eb; font-weight: 700;' if total_pct > 0 else 'background: #f8d7da; color: #721c24; border: 1px solid #f5c6cb;'
    
    # Build HTML notification message
    message = f"""
    <div style="font-family: Arial, sans-serif; max-width: 700px; margin: 0 auto;">
        <div style="background: linear-gradient(135deg, #667eea 0%, #764ba2 100%); color: white; padding: 25px; border-radius: 12px 12px 0 0; text-align: center;">
            <h2 style="margin: 0 0 10px 0; font-size: 26px; font-weight: 800;">🏘️ PRICE ALERT!</h2>
            <h3 style="margin: 0; font-size: 20px; font-weight: 700; opacity: 0.95;">{estate.name}</h3>
            <p style="margin: 8px 0 0 0; font-size: 14px; opacity: 0.9;">📍 {estate.location}</p>
        </div>
        
        <div style="background: #f8f9fa; padding: 25px; border: 1px solid #e0e0e0; border-top: none;">
            <!-- Plot Size Banner -->
            <div style="background: white; padding: 20px; border-radius: 10px; margin-bottom: 20px; border-left: 5px solid #667eea; box-shadow: 0 2px 8px rgba(0,0,0,0.08);">
                <div style="font-size: 14px; color: #666; margin-bottom: 5px; text-transform: uppercase; letter-spacing: 1px; font-weight: 600;">Plot Size</div>
                <div style="font-size: 28px; font-weight: 800; color: #667eea;">{plot_size}</div>
            </div>
            
            <!-- Price Comparison Table -->
            <div style="background: white; border-radius: 10px; overflow: hidden; box-shadow: 0 2px 8px rgba(0,0,0,0.08); margin-bottom: 20px;">
                <table style="width: 100%; border-collapse: collapse;">
                    <thead>
                        <tr style="background: linear-gradient(135deg, #667eea 0%, #764ba2 100%); color: white;">
                            <th style="padding: 15px; text-align: left; font-size: 13px; font-weight: 700; text-transform: uppercase; letter-spacing: 0.5px;">Price Point</th>
                            <th style="padding: 15px; text-align: right; font-size: 13px; font-weight: 700; text-transform: uppercase; letter-spacing: 0.5px;">Amount</th>
                        </tr>
                    </thead>
                    <tbody>
                        <tr style="border-bottom: 1px solid #f0f0f0;">
                            <td style="padding: 15px; color: #666; font-size: 14px; font-weight: 600; font-style: italic;">Presale Price (Launch)</td>
                            <td style="padding: 15px; text-align: right; color: #666; font-size: 16px; font-weight: 600;">₦{presale_price:,.0f}</td>
                        </tr>
                        <tr style="border-bottom: 1px solid #f0f0f0;">
                            <td style="padding: 15px; color: #666; font-size: 14px; font-weight: 600;">Previous Price</td>
                            <td style="padding: 15px; text-align: right; color: #666; font-size: 16px; font-weight: 600;">₦{previous_price:,.0f}</td>
                        </tr>
                        <tr style="background: rgba(102, 126, 234, 0.05);">
                            <td style="padding: 15px; color: #333; font-size: 15px; font-weight: 700;">NEW PRICE</td>
                            <td style="padding: 15px; text-align: right; font-size: 20px; font-weight: 800; color: {change_color};">₦{new_price:,.0f}</td>
                        </tr>
                    </tbody>
                </table>
            </div>
            
            <!-- Change Badges -->
            <div style="display: flex; gap: 15px; margin-bottom: 20px; flex-wrap: wrap;">
                <div style="flex: 1; min-width: 200px; background: white; padding: 20px; border-radius: 10px; text-align: center; box-shadow: 0 2px 8px rgba(0,0,0,0.08);">
                    <div style="font-size: 12px; color: #666; margin-bottom: 8px; text-transform: uppercase; letter-spacing: 1px; font-weight: 600;">Recent Change</div>
                    <div style="{badge_color} padding: 10px 15px; border-radius: 8px; font-size: 18px; font-weight: 700; display: inline-block;">
                        {change_icon} {pct_change:+.1f}%
                    </div>
                    <div style="font-size: 13px; color: #666; margin-top: 8px; font-weight: 600;">{abs(pct_change):.1f}% {change_word}</div>
                </div>
                
                <div style="flex: 1; min-width: 200px; background: white; padding: 20px; border-radius: 10px; text-align: center; box-shadow: 0 2px 8px rgba(0,0,0,0.08);">
                    <div style="font-size: 12px; color: #666; margin-bottom: 8px; text-transform: uppercase; letter-spacing: 1px; font-weight: 600;">Total Growth</div>
                    <div style="{total_badge_color} padding: 10px 15px; border-radius: 8px; font-size: 18px; font-weight: 700; display: inline-block;">
                        {total_icon} {total_pct:+.1f}%
                    </div>
                    <div style="font-size: 13px; color: #666; margin-top: 8px; font-weight: 600;">Since Launch</div>
                </div>
            </div>
            
            <!-- Effective Date -->
            <div style="background: white; padding: 18px; border-radius: 10px; margin-bottom: 20px; border-left: 4px solid #17a2b8; box-shadow: 0 2px 8px rgba(0,0,0,0.08);">
                <div style="display: flex; align-items: center; gap: 12px;">
                    <div style="background: #17a2b8; color: white; width: 45px; height: 45px; border-radius: 50%; display: flex; align-items: center; justify-content: center; font-size: 20px;">📅</div>
                    <div style="flex: 1;">
                        <div style="font-size: 12px; color: #666; margin-bottom: 3px; text-transform: uppercase; letter-spacing: 0.5px; font-weight: 600;">Effective Date</div>
                        <div style="font-size: 16px; color: #17a2b8; font-weight: 700;">{eff_date}</div>
                    </div>
                </div>
            </div>
            
            {f'''
            <!-- Notes -->
            <div style="background: white; padding: 18px; border-radius: 10px; margin-bottom: 20px; border-left: 4px solid #ffc107; box-shadow: 0 2px 8px rgba(0,0,0,0.08);">
                <div style="font-size: 12px; color: #666; margin-bottom: 8px; text-transform: uppercase; letter-spacing: 1px; font-weight: 600;">📝 Additional Notes:</div>
                <div style="color: #333; font-size: 14px; line-height: 1.6;">{notes}</div>
            </div>
            ''' if notes and notes.strip() else ''}
            
            <!-- Call to Action -->
            <div style="background: linear-gradient(135deg, #f093fb 0%, #f5576c 100%); color: white; padding: 30px 25px; border-radius: 12px; text-align: center; box-shadow: 0 4px 15px rgba(240, 147, 251, 0.4);">
                <h3 style="margin: 0 0 12px 0; font-size: 24px; font-weight: 800;">🚀 HURRY UP TO BAG THIS OPPORTUNITY!</h3>
                <p style="margin: 0; font-size: 15px; opacity: 0.95; line-height: 1.5;">
                    {f"Don't miss out on this {abs(pct_change):.1f}% price {'increase' if pct_change > 0 else 'adjustment'}!" if pct_change != 0 else "Secure your plot at this competitive price!"}
                    <br>Contact us today to secure your investment!
                </p>
            </div>
        </div>
    </div>
    """
    
    subject = f'🏘️ Price Alert: {estate.name} - {plot_size}'
    
    # Create notification for CLIENTS
    client_notif = Notification.objects.create(
        title=subject,
        message=message,
        notification_type=Notification.CLIENT_ANNOUNCEMENT
    )
    
    # Create notification for MARKETERS
    marketer_notif = Notification.objects.create(
        title=subject,
        message=message,
        notification_type=Notification.MARKETER_ANNOUNCEMENT
    )
    
    # Get users by role
    clients = User.objects.filter(role='client')
    marketers = User.objects.filter(role='marketer')
    
    # Link notifications to users
    for user in clients:
        UserNotification.objects.get_or_create(user=user, notification=client_notif)
    
    for user in marketers:
        UserNotification.objects.get_or_create(user=user, notification=marketer_notif)
    
    return {
        'clients_count': clients.count(),
        'marketers_count': marketers.count()
    }


@csrf_exempt
@login_required
@user_passes_test(is_admin)
@require_http_methods(["POST", "PUT"])
def property_price_edit(request, pk):
    """
    Updates an existing PropertyPrice (only current/effective/notes)
    and appends a PriceHistory entry using the unchanged presale.
    Optionally sends notifications to clients and marketers.
    Company-isolated for multi-tenant security.
    """
    # Company isolation - ensure property price belongs to user's company
    user_company = getattr(request.user, 'company_profile', None)
    if not user_company:
        return JsonResponse({'error': 'No company access'}, status=403)
    
    pp = get_object_or_404(PropertyPrice, pk=pk, estate__company=user_company)

    try:
        payload = json.loads(request.body)
        current = Decimal(payload["current"])
        effective = payload["effective"]
        notes = payload.get("notes", "")
        notify = payload.get("notify", False)
    except (KeyError, json.JSONDecodeError, ValueError) as e:
        return HttpResponseBadRequest(f"Invalid payload: {str(e)}")

    previous_current = pp.current
    presale = pp.presale

    pp.previous = previous_current
    pp.current = current
    pp.effective = effective
    pp.notes = notes
    pp.save()

    PriceHistory.objects.create(
        price=pp,
        presale=pp.presale,
        previous=previous_current,
        current=pp.current,
        effective=pp.effective,
        notes=pp.notes
    )

    # Send notifications if requested
    notification_sent = False
    if notify:
        try:
            send_single_price_update_notification(
                estate=pp.estate,
                plot_size=pp.plot_unit.plot_size.size,
                presale_price=float(presale),
                previous_price=float(previous_current),
                new_price=float(current),
                effective_date=effective,
                notes=notes
            )
            notification_sent = True
        except Exception as e:
            print(f"Notification error: {str(e)}")

    row_key = f"{pp.estate_id}-{pp.plot_unit_id}"
    return JsonResponse({
        "status": "ok",
        "row_key": row_key,
        "notification_sent": notification_sent
    })

@login_required
@require_POST
def property_price_save(request):
    # Company isolation - get user's company
    user_company = getattr(request.user, 'company_profile', None)
    if not user_company:
        return JsonResponse({'error': 'No company access'}, status=403)
    
    estate_id = request.POST.get('estate')
    plot_unit_id = request.POST.get('plot_unit')
    presale = request.POST.get('presale')
    previous = request.POST.get('previous')
    current = request.POST.get('current')
    effective = request.POST.get('effective')
    notes = request.POST.get('notes', '')

    # Company isolation - ensure estate and unit belong to user's company
    estate = get_object_or_404(Estate, id=estate_id, company=user_company)
    unit = get_object_or_404(PlotSizeUnits, id=plot_unit_id, estate_plot__estate__company=user_company)
    created = False

    try:
        pp, created = PropertyPrice.objects.update_or_create(
            estate=estate, plot_unit=unit,
            defaults={
                'presale': presale,
                'previous': previous,
                'current': current,
                'effective': effective,
                'notes': notes
            }
        )
        response = {
            'success': True,
            'created': created,
            'price': {
                'id': pp.id,
                'estate': estate.name,
                'location': estate.location,
                'plot_unit': unit.name,
                'presale': str(pp.presale),
                'previous': str(pp.previous),
                'current': str(pp.current),
                'effective': pp.effective.isoformat(),
                'notes': pp.notes,
            }
        }
        return JsonResponse(response)

    except IntegrityError:
        return JsonResponse({'success': False, 'message': 'Duplicate estate + plot unit.'}, status=400)
    except Exception as e:
        return JsonResponse({'success': False, 'message': str(e)}, status=500)

@login_required
def property_row_html(request, row_key):
    # Company isolation - get user's company
    user_company = getattr(request.user, 'company_profile', None)
    if not user_company:
        return JsonResponse({'error': 'No company access'}, status=403)
    
    estate_id, unit_id = row_key.split('-')
    today = date.today()
    
    try:
        # Company isolation - ensure property price belongs to user's company
        pp = PropertyPrice.objects.select_related(
            'estate', 'plot_unit__plot_size'
        ).get(
            estate_id=estate_id, 
            plot_unit_id=unit_id,
            estate__company=user_company
        )
        
        # Check for active promo
        active_promo = PromotionalOffer.objects.filter(
            estates__id=estate_id,
            start__lte=today,
            end__gte=today
        ).first()
        
        # Calculate display values (same logic as management_dashboard)
        current_price = float(pp.current)
        
        if active_promo:
            discount_factor = float(1 - active_promo.discount / 100)
            discounted_price = Decimal(str(current_price * discount_factor))
        else:
            discounted_price = pp.current
        
        # Calculate percentages
        if pp.previous and pp.previous > 0:
            percent_change = (float(discounted_price) - float(pp.previous)) / float(pp.previous) * 100
            pp.percent_change = Decimal(str(percent_change))
        else:
            pp.percent_change = None
            
        if pp.presale and pp.presale > 0:
            overtime = (float(discounted_price) - float(pp.presale)) / float(pp.presale) * 100
            pp.overtime = Decimal(str(overtime))
        else:
            pp.overtime = None
        
        pp.display_current = discounted_price
        pp.active_promo = active_promo
        
        html = render_to_string(
            "admin_side/price_row.html", 
            {'row': pp, 'today': str(today)}, 
            request
        )
    except PropertyPrice.DoesNotExist:
        estate = Estate.objects.select_related().get(id=estate_id)
        unit = PlotSizeUnits.objects.select_related('plot_size').get(id=unit_id)
        
        # Check for active promo even for non-existent price
        active_promo = PromotionalOffer.objects.filter(
            estates__id=estate_id,
            start__lte=today,
            end__gte=today
        ).first()
        
        class DummyPrice:
            def __init__(self, estate, unit, active_promo):
                self.estate = estate
                self.plot_unit = unit
                self.id = None
                self.presale = None
                self.previous = None
                self.current = None
                self.percent_change = None
                self.overtime = None
                self.display_current = None
                self.effective = None
                self.notes = None
                self.active_promo = active_promo
                
        html = render_to_string(
            "admin_side/price_row.html", 
            {'row': DummyPrice(estate, unit, active_promo), 'today': str(today)}, 
            request
        )
    
    return JsonResponse({'html': html})

@login_required
@require_GET
def property_price_prefill(request):
    # Company isolation - get user's company
    user_company = getattr(request.user, 'company_profile', None)
    if not user_company:
        return JsonResponse({'error': 'No company access'}, status=403)
    
    estate_id = request.GET.get('estate')
    plot_unit_id = request.GET.get('plot_unit')
    try:
        # Company isolation - ensure property price belongs to user's company
        pp = PropertyPrice.objects.get(estate_id=estate_id, plot_unit_id=plot_unit_id, estate__company=user_company)
        return JsonResponse({
            'exists': True,
            'presale': str(pp.presale),
            'previous': str(pp.previous),
            'current': str(pp.current),
            'effective': pp.effective.isoformat(),
            'notes': pp.notes,
        })
    except PropertyPrice.DoesNotExist:
        return JsonResponse({'exists': False})

# Add to imports
from django.utils.dateparse import parse_date as django_parse_date
from decimal import Decimal

@login_required
@user_passes_test(is_admin)
@require_GET
def property_price_history(request, pk):
    # Company isolation - ensure property price belongs to user's company
    user_company = getattr(request.user, 'company_profile', None)
    if not user_company:
        return JsonResponse({'error': 'No company access'}, status=403)
    
    pp = get_object_or_404(PropertyPrice, pk=pk, estate__company=user_company)
    history = []

    # Build real history from PriceHistory records
    for h in pp.history.order_by('effective'):
        history.append({
            "presale": float(h.presale),
            "previous": float(h.previous),
            "current": float(h.current),
            "effective": h.effective.isoformat(),
            "notes": h.notes or "",
            "is_promo": False  # Mark as real price record
        })

    # Get ALL promos for this estate (active and expired)
    promos = PromotionalOffer.objects.filter(estates=pp.estate).order_by('start')
    
    # Create list to track promo injection points
    promo_events = []
    
    # For each promo, create start and end events
    for promo in promos:
        # Find base price at promo start
        base_price = None
        for h in history:
            if django_parse_date(h['effective']) <= promo.start:
                base_price = h['current']
        
        if base_price is None:
            continue
            
        # Calculate discounted price
        discounted = base_price * (100 - promo.discount) / 100
        discounted = round(discounted, 2)
        
        # Add promo start event
        promo_events.append({
            "effective": promo.start.isoformat(),
            "current": discounted,
            "previous": base_price,
            "presale": float(pp.presale),
            "notes": f"PROMO START: {promo.name} ({promo.discount}% OFF)",
            "is_promo": True
        })
        
        # Add promo end event
        promo_events.append({
            "effective": promo.end.isoformat(),
            "current": base_price,  # Revert to base price
            "previous": discounted,
            "presale": float(pp.presale),
            "notes": f"PROMO END: {promo.name} expired",
            "is_promo": True
        })
    
    # Combine real history with promo events
    combined_history = history + promo_events
    combined_history.sort(key=lambda x: x['effective'])
    
    # Calculate price changes
    if combined_history:
        initial = combined_history[0]
        latest = combined_history[-1]
        presale_val = Decimal(initial["presale"])
        latest_val = Decimal(latest["current"])
        
        # Find previous price (skip promo events for change calculation)
        prev_record = None
        for record in reversed(combined_history):
            if not record['is_promo']:
                prev_record = record
                break
        
        prev_val = Decimal(prev_record["current"]) if prev_record else presale_val
        
        current_change = float((latest_val - prev_val) / prev_val * 100) if prev_val > 0 else 0.0
        total_change = float((latest_val - presale_val) / presale_val * 100) if presale_val > 0 else 0.0
    else:
        current_change = 0.0
        total_change = 0.0

    return JsonResponse({
        "history": combined_history,
        "current_change": round(current_change, 2),
        "total_change": round(total_change, 2),
        "has_promos": len(promos) > 0
    })


@csrf_exempt
@login_required
@user_passes_test(is_admin)
@require_http_methods(["POST"])
def promo_create(request):
    # Company isolation - get user's company
    user_company = getattr(request.user, 'company_profile', None)
    if not user_company:
        return JsonResponse({'error': 'No company access'}, status=403)
    
    try:
        data = json.loads(request.body)

        name = data.get("name", "").strip()
        discount = float(data.get("discount", 0))
        start = parse_date(data.get("start"))
        end = parse_date(data.get("end"))
        description = data.get("description", "")
        estate_ids = data.get("estates", [])

        if not name or not start or not end or not estate_ids:
            return JsonResponse({"status": "error", "message": "Missing required fields."}, status=400)

        promo = PromotionalOffer.objects.create(
            name=name, discount=discount,
            start=start, end=end, description=description,
            company=user_company  # Company isolation
        )
        # Company isolation - only set estates that belong to user's company
        promo.estates.set(Estate.objects.filter(id__in=estate_ids, company=user_company))

        return JsonResponse({
            "status": "ok",
            "id": promo.id,
            "discount": promo.discount,
            "start": promo.start.isoformat(),
            "end": promo.end.isoformat(),
        })

    except Exception as e:
        return JsonResponse({"status": "error", "message": str(e)}, status=500)

@csrf_exempt
@login_required
@user_passes_test(is_admin)
@require_http_methods(["PUT"])
def promo_update(request, promo_id):
    # Company isolation - get user's company
    user_company = getattr(request.user, 'company_profile', None)
    if not user_company:
        return JsonResponse({'error': 'No company access'}, status=403)
    
    try:
        data = json.loads(request.body)
        # Company isolation - ensure promo belongs to user's company
        promo = get_object_or_404(PromotionalOffer, id=promo_id, company=user_company)

        promo.name = data.get("name", promo.name).strip()
        promo.discount = float(data.get("discount", promo.discount))
        promo.start = parse_date(data.get("start")) or promo.start
        promo.end = parse_date(data.get("end")) or promo.end
        promo.description = data.get("description", promo.description)
        estate_ids = data.get("estates", [])

        if not promo.name or not promo.start or not promo.end or not estate_ids:
            return JsonResponse({"status": "error", "message": "Missing required fields."}, status=400)

        promo.save()
        # Company isolation - only set estates that belong to user's company
        promo.estates.set(Estate.objects.filter(id__in=estate_ids, company=user_company))

        return JsonResponse({
            "status": "ok",
            "id": promo.id,
            "discount": promo.discount,
            "start": promo.start.isoformat(),
            "end": promo.end.isoformat(),
        })

    except Exception as e:
        return JsonResponse({"status": "error", "message": str(e)}, status=500)

@require_GET
@login_required
@user_passes_test(is_admin)
def get_active_promo_for_estate(request, estate_id):
    # Company isolation - get user's company
    user_company = getattr(request.user, 'company_profile', None)
    if not user_company:
        return JsonResponse({'error': 'No company access'}, status=403)
    
    try:
        # Company isolation - ensure estate belongs to user's company
        estate = Estate.objects.get(pk=estate_id, company=user_company)
        # Only return promo that is still active
        promo = PromotionalOffer.objects.filter(
            estates__id=estate_id,
            start__lte=now(),
            end__gte=now()
        ).order_by('-start').first()

        if not promo:
            return JsonResponse({'status': 'no_active_promo'})

        return JsonResponse({
            'status': 'ok',
            'data': {
                'name': promo.name,
                'discount': promo.discount,
                'start': promo.start.isoformat(),
                'end': promo.end.isoformat(),
                'description': promo.description,
            }
        })

    except PromotionalOffer.DoesNotExist:
        return JsonResponse({'status': 'no_active_promo'})


# # MANAGEMENT NOTIFICATIONS

@csrf_exempt
def notify_clients_marketer(request):
    if request.method != 'POST':
        return JsonResponse({'status':'error','message':'Invalid request method'}, status=405)

    try:
        data = json.loads(request.body)
        subject     = data['subject']
        message     = data['message']
        notify_type = data['type']
        estate_ids  = data.get('estate_ids', [])
        send_inapp  = data.get('send_inapp', False)

        User = get_user_model()
        
        if notify_type == 'client_update':
            users     = User.objects.filter(role='client')
            ntype     = Notification.CLIENT_ANNOUNCEMENT
        elif notify_type == 'marketer_update':
            users     = User.objects.filter(role='marketer')
            ntype     = Notification.MARKETER_ANNOUNCEMENT
        elif notify_type == 'general_notification':
            users     = User.objects.exclude(role='admin')
            ntype     = Notification.ANNOUNCEMENT
        else:
            users     = User.objects.filter(estate__id__in=estate_ids).distinct()
            ntype     = Notification.ANNOUNCEMENT

        recipients = list(dict.fromkeys(users.values_list('id', flat=True)))

        # create one Notification record
        notif = Notification.objects.create(
            title=subject,
            message=message,
            notification_type=ntype
        )

        dispatched = False
        dispatch_payload = None
        if send_inapp and recipients:
            total_recipients = len(recipients)
            total_batches = ceil(total_recipients / BATCH_SIZE) if total_recipients else 0
            dispatch = NotificationDispatch.objects.create(
                notification=notif,
                total_recipients=total_recipients,
                total_batches=total_batches,
            )

            celery_ready = is_celery_worker_available(timeout=2.0)

            if celery_ready:
                try:
                    dispatch_notification_stream.delay(dispatch.id, notif.id, recipients)
                except Exception as exc:
                    logger.warning(
                        "Falling back to synchronous notification dispatch (dispatch_id=%s): %s",
                        dispatch.id,
                        exc,
                        exc_info=True,
                    )
                    try:
                        dispatch_notification_stream_sync(dispatch.id, notif.id, recipients)
                    except Exception:
                        dispatch.refresh_from_db()
                        raise
                    else:
                        dispatch.refresh_from_db()
                else:
                    dispatched = True
            else:
                logger.info(
                    "Celery worker unavailable within timeout; running synchronous dispatch (dispatch_id=%s)",
                    dispatch.id,
                )
                try:
                    dispatch_notification_stream_sync(dispatch.id, notif.id, recipients)
                except Exception:
                    dispatch.refresh_from_db()
                    raise
                else:
                    dispatch.refresh_from_db()

            dispatch_payload = dispatch.as_dict()

        response_status = 'queued' if dispatched else 'success'

        return JsonResponse({
            'status': response_status,
            'recipients': len(recipients),
            'notification_id': notif.id,
            'dispatched': dispatched,
            'dispatch': dispatch_payload,
            'queue_message': None if dispatched else 'No queue',
        })

    except Exception as exc:
        logger.exception("Failed to dispatch notification", exc_info=exc)
        return JsonResponse(
            {
                'status': 'error',
                'message': 'Failed to dispatch notification.',
                'detail': str(exc),
            },
            status=500,
        )

    except Exception as e:
        return JsonResponse({'status':'error','message':str(e)}, status=400)


@login_required
@require_GET
def notification_dispatch_status(request, dispatch_id: int):
    try:
        dispatch = NotificationDispatch.objects.get(pk=dispatch_id)
    except NotificationDispatch.DoesNotExist:
        return JsonResponse({'ok': False, 'error': 'Dispatch not found'}, status=404)

    return JsonResponse({'ok': True, 'dispatch': dispatch.as_dict()})


@login_required
def notifications_all(request):
    """
    Display notifications for the logged-in user.
    For marketers: Show notifications from ALL companies they're connected to
    (primary company, affiliations, profiles, and transaction history).
    """
    user = request.user
    
    if getattr(user, 'role', None) == 'marketer':
        # Get ALL company IDs the marketer is connected to
        # 1. Companies where marketer has transactions
        transaction_company_ids = set(
            Transaction.objects.filter(marketer=user)
            .values_list('company', flat=True)
            .distinct()
        )
        
        # 2. Companies where marketer has affiliations
        affiliation_company_ids = set(
            MarketerAffiliation.objects.filter(marketer=user)
            .values_list('company_id', flat=True)
            .distinct()
        )
        
        # 3. Companies where marketer has CompanyMarketerProfile records
        profile_company_ids = set(
            CompanyMarketerProfile.objects.filter(marketer=user)
            .values_list('company_id', flat=True)
            .distinct()
        )
        
        # 4. The marketer's own company_profile (primary company)
        own_company_id = getattr(user, 'company_profile_id', None)
        
        # Combine all company IDs
        all_company_ids = transaction_company_ids | affiliation_company_ids | profile_company_ids
        if own_company_id:
            all_company_ids.add(own_company_id)
        
        # Remove None values
        all_company_ids = [c for c in all_company_ids if c is not None]
        
        # Get notifications for this user that belong to their affiliated companies OR are global (company=None)
        qs = user.notifications.select_related('notification', 'notification__company').filter(
            Q(notification__company_id__in=all_company_ids) | Q(notification__company__isnull=True)
        ).order_by('-notification__created_at')
    elif getattr(user, 'role', None) == 'client':
        # Get ALL company IDs the client is connected to
        # 1. The client's own company_profile (primary company)
        own_company_id = getattr(user, 'company_profile_id', None)
        
        # 2. Companies where client has CompanyClientProfile records (affiliations)
        profile_company_ids = set(
            CompanyClientProfile.objects.filter(client=user)
            .values_list('company_id', flat=True)
            .distinct()
        )
        
        # Combine all company IDs
        all_company_ids = profile_company_ids.copy()
        if own_company_id:
            all_company_ids.add(own_company_id)
        
        # Remove None values
        all_company_ids = [c for c in all_company_ids if c is not None]
        
        # Get notifications for this user that belong to their affiliated companies OR are global (company=None)
        qs = user.notifications.select_related('notification', 'notification__company').filter(
            Q(notification__company_id__in=all_company_ids) | Q(notification__company__isnull=True)
        ).order_by('-notification__created_at')
    else:
        # For other roles, show all their notifications
        qs = user.notifications.select_related('notification').order_by('-notification__created_at')
    
    unread = qs.filter(read=False)
    read = qs.filter(read=True)
    tpl = f"{user.role}_side/notification.html"
    
    return render(request, tpl, {
        'unread_list': unread,
        'read_list': read,
        'home_url': f'{user.role}-dashboard' if user.role else 'login'
    })

@login_required
def notification_detail(request, un_id):
    """
    Display notification detail.
    For marketers: Verify they have access to this notification based on company affiliation.
    """
    user = request.user
    un = get_object_or_404(UserNotification, pk=un_id, user=user)
    
    # For marketers, verify they have access to this notification's company
    if getattr(user, 'role', None) == 'marketer' and un.notification.company:
        # Get all company IDs the marketer is connected to
        transaction_company_ids = set(
            Transaction.objects.filter(marketer=user)
            .values_list('company', flat=True)
            .distinct()
        )
        affiliation_company_ids = set(
            MarketerAffiliation.objects.filter(marketer=user)
            .values_list('company_id', flat=True)
            .distinct()
        )
        profile_company_ids = set(
            CompanyMarketerProfile.objects.filter(marketer=user)
            .values_list('company_id', flat=True)
            .distinct()
        )
        own_company_id = getattr(user, 'company_profile_id', None)
        
        all_company_ids = transaction_company_ids | affiliation_company_ids | profile_company_ids
        if own_company_id:
            all_company_ids.add(own_company_id)
        
        # Check if user has access to this notification's company
        if un.notification.company_id not in all_company_ids:
            messages.error(request, "You don't have access to this notification.")
            return redirect('notifications_all')
    
    # For clients, verify they have access to this notification's company
    elif getattr(user, 'role', None) == 'client' and un.notification.company:
        # Get all company IDs the client is connected to
        own_company_id = getattr(user, 'company_profile_id', None)
        profile_company_ids = set(
            CompanyClientProfile.objects.filter(client=user)
            .values_list('company_id', flat=True)
            .distinct()
        )
        
        all_company_ids = profile_company_ids.copy()
        if own_company_id:
            all_company_ids.add(own_company_id)
        
        # Check if user has access to this notification's company
        if un.notification.company_id not in all_company_ids:
            messages.error(request, "You don't have access to this notification.")
            return redirect('notifications_all')
    
    if not un.read:
        un.read = True
        un.save(update_fields=['read'])
    
    # Determine home URL based on role
    home_url = f'{user.role}-dashboard' if user.role else 'login'
    
    return render(
        request,
        f"{user.role}_side/notification_detail.html",
        {
            'notification': un.notification,
            'back_url': 'notifications_all',
            'home_url': home_url
        }
    )

@login_required
def mark_notification_read(request, un_id):
    un = get_object_or_404(UserNotification, pk=un_id, user=request.user)
    un.read = True
    un.save(update_fields=['read'])
    return redirect('notification_detail', un_id=un.id)


@login_required
@user_passes_test(is_admin)
@require_GET
def property_price_detail(request, pk):
    """
    Returns JSON detail of a PropertyPrice including an active promo adjustment.
    Company-isolated for multi-tenant security.
    """
    # Company isolation - ensure property price belongs to user's company
    user_company = getattr(request.user, 'company_profile', None)
    if not user_company:
        return JsonResponse({'error': 'No company access'}, status=403)
    
    pp = get_object_or_404(PropertyPrice, pk=pk, estate__company=user_company)
    today = now().date()

    promo = PromotionalOffer.objects.filter(
        estates=pp.estate,
        start__lte=today,
        end__gte=today
    ).order_by('-discount').first()

    original = pp.current
    adjusted = original
    if promo:
        adjusted = (original * (Decimal(100) - promo.discount) / Decimal(100)).quantize(Decimal('0.01'))

    return JsonResponse({
        "id": pp.id,
        "estate":    {"id": pp.estate.id, "name": pp.estate.name},
        "plot_unit": {"id": pp.plot_unit.id, "size": pp.plot_unit.plot_size.size},
        "presale":   str(pp.presale),
        "previous":  str(pp.previous),
        "current":   str(adjusted),
        "original":  str(original),
        "promo_applied":  bool(promo),
        "promo_discount": promo.discount if promo else 0,
        "promo_name":     promo.name if promo else "",
        "promo_expires":  promo.end.isoformat() if promo else None,
        "effective": pp.effective.isoformat(),
        "notes":     pp.notes,
    })



# MESSAGING AND BIRTHDAY.


@login_required
def edit_admin_user(request, user_id):
    """
    AJAX endpoint to edit admin/support user details and password.
    Only master admin can edit other admins/support users.
    """
    if request.method != 'POST':
        return JsonResponse({'ok': False, 'error': 'Method not allowed'}, status=405)

    try:
        # Get the target user - try AdminUser first, then SupportUser
        try:
            target = AdminUser.objects.get(id=user_id)
        except AdminUser.DoesNotExist:
            try:
                target = SupportUser.objects.get(id=user_id)
            except SupportUser.DoesNotExist:
                return JsonResponse({'ok': False, 'error': 'User not found'}, status=404)

        # Security checks
        current_user = request.user
        
        # Support dynamic slug parameter
        try:
            company = get_user_company_from_slug(request)
        except HttpResponseForbidden:
            return JsonResponse({'ok': False, 'error': 'Forbidden'}, status=403)
        
        if not company:
            company = current_user.company_profile
        
        if current_user.admin_level != 'system':
            return JsonResponse({'ok': False, 'error': 'Only master admin can edit users'}, status=403)

        # Validate master admin password
        master_password = request.POST.get('master_admin_password', '').strip()
        if not master_password:
            return JsonResponse({'ok': False, 'error': 'Master admin password is required'}, status=400)
        
        if not current_user.check_password(master_password):
            return JsonResponse({'ok': False, 'error': 'Invalid master admin password'}, status=403)

        # Cannot Update master admin
        if target.admin_level == 'system':
            return JsonResponse({'ok': False, 'error': 'Cannot Update master admin'}, status=403)

        # Must be same company
        if target.company_profile != current_user.company_profile:
            return JsonResponse({'ok': False, 'error': 'Cannot Update users from different company'}, status=403)

        # Get form data
        full_name = request.POST.get('full_name', '').strip()
        email = request.POST.get('email', '').strip()
        phone = request.POST.get('phone', '').strip()
        address = request.POST.get('address', '').strip()
        new_password = request.POST.get('new_password', '').strip()

        # Validate required fields
        if not full_name or not email:
            return JsonResponse({'ok': False, 'error': 'Full name and email are required'}, status=400)

        # Check email uniqueness conditionally
        if target.role in ['admin', 'support']:
            # For admin/support, allow duplicate emails only if not master admin
            existing = CustomUser.objects.filter(email=email).exclude(id=target.id)
            if existing.exists():
                master_admin = existing.filter(admin_level='system').first()
                if master_admin:
                    return JsonResponse({'ok': False, 'error': 'Email already used by master admin'}, status=400)
        else:
            # For other roles, enforce strict uniqueness
            if CustomUser.objects.filter(email=email).exclude(id=target.id).exists():
                return JsonResponse({'ok': False, 'error': 'Email already exists'}, status=400)

        # Update user details
        target.full_name = full_name
        target.email = email
        target.phone = phone
        target.address = address

        if new_password:
            target.set_password(new_password)

        target.save()

        return JsonResponse({
            'ok': True,
            'message': f'{target.role.title()} user updated successfully',
            'user': {
                'id': target.id,
                'full_name': target.full_name,
                'email': target.email,
                'phone': target.phone,
                'address': target.address,
            }
        })

    except Exception as e:
        logger.error(f'Error editing admin user: {str(e)}')
        return JsonResponse({'ok': False, 'error': 'Internal server error'}, status=500)


@login_required
def edit_support_user(request, user_id):
    """
    AJAX endpoint to edit support user details and password.
    Only master admin can edit support users.
    """
    if request.method != 'POST':
        return JsonResponse({'ok': False, 'error': 'Method not allowed'}, status=405)

    try:
        # Get the target user - use SupportUser model directly
        target = get_object_or_404(SupportUser, id=user_id)

        # Security checks
        current_user = request.user
        
        # Support dynamic slug parameter
        try:
            company = get_user_company_from_slug(request)
        except HttpResponseForbidden:
            return JsonResponse({'ok': False, 'error': 'Forbidden'}, status=403)
        
        if not company:
            company = current_user.company_profile
        
        if current_user.admin_level != 'system':
            return JsonResponse({'ok': False, 'error': 'Only master admin can edit users'}, status=403)

        # Validate master admin password
        master_password = request.POST.get('master_admin_password', '').strip()
        if not master_password:
            return JsonResponse({'ok': False, 'error': 'Master admin password is required'}, status=400)
        
        if not current_user.check_password(master_password):
            return JsonResponse({'ok': False, 'error': 'Invalid master admin password'}, status=403)

        # Must be same company
        if target.company_profile != current_user.company_profile:
            return JsonResponse({'ok': False, 'error': 'Cannot edit users from different company'}, status=403)

        # Get form data
        full_name = request.POST.get('full_name', '').strip()
        email = request.POST.get('email', '').strip()
        phone = request.POST.get('phone', '').strip()
        address = request.POST.get('address', '').strip()
        new_password = request.POST.get('new_password', '').strip()

        # Validate required fields
        if not full_name or not email:
            return JsonResponse({'ok': False, 'error': 'Full name and email are required'}, status=400)

        # Check email uniqueness conditionally
        if target.role in ['admin', 'support']:
            # For admin/support, allow duplicate emails only if not master admin
            existing = CustomUser.objects.filter(email=email).exclude(id=target.id)
            if existing.exists():
                master_admin = existing.filter(admin_level='system').first()
                if master_admin:
                    return JsonResponse({'ok': False, 'error': 'Email already used by master admin'}, status=400)
        else:
            # For other roles, enforce strict uniqueness
            if CustomUser.objects.filter(email=email).exclude(id=target.id).exists():
                return JsonResponse({'ok': False, 'error': 'Email already exists'}, status=400)

        # Update user details
        target.full_name = full_name
        target.email = email
        target.phone = phone
        target.address = address

        if new_password:
            target.set_password(new_password)

        target.save()

        return JsonResponse({
            'ok': True,
            'message': 'Support user updated successfully',
            'user': {
                'id': target.id,
                'full_name': target.full_name,
                'email': target.email,
                'phone': target.phone,
                'address': target.address,
            }
        })

    except SupportUser.DoesNotExist:
        return JsonResponse({'ok': False, 'error': 'User not found'}, status=404)
    except Exception as e:
        logger.error(f'Error editing support user: {str(e)}')
        return JsonResponse({'ok': False, 'error': 'Internal server error'}, status=500)
